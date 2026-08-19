from __future__ import annotations

# 版本：V28.7｜賈維斯固定口令可靠版

import base64
import hashlib
import html
import json
import math
import os
import re
import time
import tempfile
import unicodedata
import uuid
from datetime import datetime
from difflib import SequenceMatcher
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dispatch_core import (
    SHIFT_COLUMNS,
    available_sources,
    parse_route,
)


APP_VERSION = "V28.8"
APP_VERSION_NAME = "測試版"
APP_BUILD_DATE = "2026-08-19"
DEFAULT_BATTERY_THRESHOLD = 89
DEFAULT_BATTERY_PRIORITY_THRESHOLD = 40
SMART_DISPATCH_CANDIDATE_LIMIT = 10

# 賈維斯測試版：庫存場站平常不列入自動推薦；只有本站已接近空站或滿站時解鎖。
INVENTORY_STATION_NAMES = ("臺東縣政府文化處圖書館", "臺東轉運站")
INVENTORY_STATION_ALERT_EDGE = 1
JARVIS_WAKE_WORD = "賈維斯"


def is_mobile_browser() -> bool:
    """依瀏覽器 User-Agent 判斷是否為手機或平板。"""
    try:
        user_agent = st.context.headers.get("User-Agent", "")
    except Exception:
        # 舊版 Streamlit 沒有 st.context 時，改由側邊欄手動切換。
        return False

    mobile_tokens = ("android", "iphone", "ipad", "ipod", "mobile", "windows phone")
    normalized_user_agent = user_agent.lower()
    return any(token in normalized_user_agent for token in mobile_tokens)


def safe_nonnegative_int(value) -> int:
    """把 Excel 儲存格值安全轉成不小於 0 的整數。"""
    try:
        return max(0, int(value))
    except (TypeError, ValueError, OverflowError):
        return 0


CURRENT_STATUS_COLUMNS = ("2.0 現況", "2.0E 現況")
STATION_LIVE_COLUMNS = ("總車數", "空位數", "總柱數", "服務狀態")
PERSISTED_STATUS_COLUMNS = (*CURRENT_STATUS_COLUMNS, *STATION_LIVE_COLUMNS)
STATUS_UNAVAILABLE_TEXT = "資料未取得"

CONFIGURATION_TYPES = ("暑假配置", "平日配置", "假日配置")
CONFIGURATION_TYPE_KEYWORDS = {
    "暑假配置": ("暑假", "夏季", "summer"),
    "平日配置": ("平日", "weekday", "週間"),
    "假日配置": ("假日", "週末", "周末", "holiday", "weekend"),
}


def normalize_current_status(value) -> int | None:
    """將現況值正規化；空白與無法轉換的內容保留為 None，不再自動變成 0。"""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    text = str(value).strip()
    if not text or text == STATUS_UNAVAILABLE_TEXT:
        return None

    try:
        return max(0, int(float(text)))
    except (TypeError, ValueError, OverflowError):
        return None


def blank_current_status(status_df: pd.DataFrame) -> pd.DataFrame:
    """建立現況與即時場站欄位預設為空白的資料表。"""
    blank_df = status_df.copy()
    for column in PERSISTED_STATUS_COLUMNS:
        blank_df[column] = pd.Series([pd.NA] * len(blank_df), dtype="Int64")
    return blank_df



def coerce_nullable_current_status(status_df: pd.DataFrame) -> pd.DataFrame:
    """讓現況欄維持可輸入空白的 nullable integer 型別；使用向量化避免逐格 Python 迴圈。"""
    normalized_df = status_df.copy()
    for column in CURRENT_STATUS_COLUMNS:
        if column not in normalized_df.columns:
            normalized_df[column] = pd.Series(pd.NA, index=normalized_df.index, dtype="Int64")
            continue
        numeric = pd.to_numeric(
            normalized_df[column].replace(STATUS_UNAVAILABLE_TEXT, pd.NA),
            errors="coerce",
        )
        normalized_df[column] = pd.Series(
            np.trunc(numeric).clip(lower=0),
            index=normalized_df.index,
        ).astype("Int64")
    return normalized_df


def coerce_nullable_station_status(status_df: pd.DataFrame) -> pd.DataFrame:
    """統一現況、總車數、空位、總柱與服務狀態為可空白整數。"""
    normalized_df = coerce_nullable_current_status(status_df)
    for column in STATION_LIVE_COLUMNS:
        if column not in normalized_df.columns:
            normalized_df[column] = pd.Series(pd.NA, index=normalized_df.index, dtype="Int64")
            continue
        numeric = pd.to_numeric(normalized_df[column], errors="coerce")
        normalized_df[column] = pd.Series(
            np.trunc(numeric).clip(lower=0),
            index=normalized_df.index,
        ).astype("Int64")
    return normalized_df


def optional_count_text(value, *, suffix: str = "") -> str:
    """把可空白數量轉成顯示文字。"""
    normalized = normalize_current_status(value)
    return "—" if normalized is None else f"{normalized}{suffix}"


def station_total_bikes(row: pd.Series | dict) -> int | None:
    """優先採用官網總車數；缺少時由 2.0 與 2.0E 安全加總。"""
    live_total = normalize_current_status(row.get("總車數"))
    if live_total is not None:
        return live_total
    bike = normalize_current_status(row.get("2.0 現況"))
    ebike = normalize_current_status(row.get("2.0E 現況"))
    if bike is None or ebike is None:
        return None
    return bike + ebike


def station_empty_spaces(row: pd.Series | dict) -> int | None:
    """取得官網總空位；缺少時不推算，避免把配置標準誤當總柱數。"""
    return normalize_current_status(row.get("空位數"))


def vehicle_balance_text(current_value, standard_value) -> str:
    """顯示目前、標準與缺多狀態。"""
    return (
        f"目前 {optional_count_text(current_value)}／標準 {safe_nonnegative_int(standard_value)}／"
        f"{format_dispatch_status(current_value, standard_value).replace(' 台', '')}"
    )


def format_dispatch_status(current_value, standard_value) -> str:
    """現況為空白時顯示資料未取得，其餘依標準值計算缺車或多車。"""
    current = normalize_current_status(current_value)
    if current is None:
        return STATUS_UNAVAILABLE_TEXT

    standard = safe_nonnegative_int(standard_value)
    difference = current - standard
    if difference > 0:
        return f"多 {difference} 台"
    if difference < 0:
        return f"缺 {abs(difference)} 台"
    return "符合"



@st.cache_data(show_spinner=False, max_entries=64)
def build_analysis_result(status_df: pd.DataFrame) -> pd.DataFrame:
    """建立含現況、標準、缺多、空位與服務狀態的完整分析結果。"""
    result = status_df[["行政區", "場站名稱"]].astype(str).copy()
    detail_columns = (
        "路線區域", "2.0 現況", "2.0 標準", "2.0E 現況", "2.0E 標準",
        "總車數", "空位數", "總柱數", "服務狀態",
    )
    for column in detail_columns:
        result[column] = status_df[column] if column in status_df.columns else pd.NA

    computed_total = (
        pd.to_numeric(result["2.0 現況"], errors="coerce")
        + pd.to_numeric(result["2.0E 現況"], errors="coerce")
    )
    existing_total = pd.to_numeric(result["總車數"], errors="coerce")
    result["總車數"] = existing_total.fillna(computed_total).astype("Int64")

    def build_status_series(current_column: str, standard_column: str) -> pd.Series:
        current = pd.Series(
            np.trunc(pd.to_numeric(status_df[current_column], errors="coerce")),
            index=status_df.index,
        ).clip(lower=0)
        standard = pd.Series(
            np.trunc(pd.to_numeric(status_df[standard_column], errors="coerce")),
            index=status_df.index,
        ).fillna(0).clip(lower=0)
        difference = current - standard
        output = pd.Series(STATUS_UNAVAILABLE_TEXT, index=status_df.index, dtype="object")
        valid = current.notna()
        output.loc[valid & difference.eq(0)] = "符合"
        output.loc[valid & difference.gt(0)] = (
            "多 " + difference.loc[valid & difference.gt(0)].astype(int).astype(str) + " 台"
        )
        output.loc[valid & difference.lt(0)] = (
            "缺 " + difference.loc[valid & difference.lt(0)].abs().astype(int).astype(str) + " 台"
        )
        return output

    result["2.0 缺／多幾台"] = build_status_series("2.0 現況", "2.0 標準")
    result["2.0E 缺／多幾台"] = build_status_series("2.0E 現況", "2.0E 標準")
    return result



def calculate_totals_ignoring_missing(
    status_df: pd.DataFrame,
    current_column: str,
    standard_column: str,
) -> tuple[int, int]:
    """缺／多合計只計算有效現況；使用向量化加速。"""
    current = pd.Series(
        np.trunc(pd.to_numeric(status_df[current_column], errors="coerce")),
        index=status_df.index,
    ).clip(lower=0)
    standard = pd.Series(
        np.trunc(pd.to_numeric(status_df[standard_column], errors="coerce")),
        index=status_df.index,
    ).fillna(0).clip(lower=0)
    difference = (current - standard).dropna()
    short_total = int((-difference[difference.lt(0)]).sum())
    extra_total = int(difference[difference.gt(0)].sum())
    return short_total, extra_total




def calculate_inventory_summary(
    status_df: pd.DataFrame,
    current_column: str,
    standard_column: str,
) -> dict[str, int | str | None]:
    """計算配置總數、目前總數與完整資料下的整體差額。"""
    standard = pd.Series(
        np.trunc(pd.to_numeric(status_df[standard_column], errors="coerce")),
        index=status_df.index,
    ).fillna(0).clip(lower=0)
    current = pd.Series(
        np.trunc(pd.to_numeric(status_df[current_column], errors="coerce")),
        index=status_df.index,
    ).clip(lower=0)

    configured_total = int(standard.sum())
    current_total = int(current.dropna().sum())
    missing_count = int(current.isna().sum())
    station_count = int(len(status_df))

    # 只在所有場站都有現況資料時顯示整體缺／多，避免把空白誤判成缺車。
    difference: int | None = None
    state = "pending"
    state_label = "資料未完整"
    difference_text = f"待補 {missing_count} 筆"
    signed_difference_text = "—"

    if missing_count == 0:
        difference = current_total - configured_total
        if difference > 0:
            state = "extra"
            state_label = "多車"
            difference_text = f"多 {difference} 台"
            signed_difference_text = f"+{difference} 台"
        elif difference < 0:
            state = "short"
            state_label = "缺車"
            difference_text = f"缺 {abs(difference)} 台"
            signed_difference_text = f"−{abs(difference)} 台"
        else:
            state = "balanced"
            state_label = "符合配置"
            difference_text = "剛好 0 台"
            signed_difference_text = "0 台"

    return {
        "configured_total": configured_total,
        "current_total": current_total,
        "missing_count": missing_count,
        "station_count": station_count,
        "difference": difference,
        "state": state,
        "state_label": state_label,
        "difference_text": difference_text,
        "signed_difference_text": signed_difference_text,
    }


def _summary_illustration(vehicle_type: str) -> str:
    """回傳不需額外圖片檔的內嵌 SVG 小插畫。"""
    if vehicle_type == "bike":
        return """
        <svg class="fleet-card-svg" viewBox="0 0 150 120" aria-hidden="true">
          <circle cx="49" cy="79" r="23" fill="none" stroke="currentColor" stroke-width="7"/>
          <circle cx="111" cy="79" r="23" fill="none" stroke="currentColor" stroke-width="7"/>
          <path d="M49 79 L67 45 L86 79 L49 79 M67 45 L94 45 L111 79 M65 45 L58 31 M53 31 H70 M85 79 L101 57"
                fill="none" stroke="currentColor" stroke-width="7" stroke-linecap="round" stroke-linejoin="round"/>
          <path d="M91 39 H111 L116 52 H96 Z" fill="currentColor" opacity=".28"/>
          <path d="M18 105 H132" stroke="currentColor" stroke-width="5" stroke-linecap="round" opacity=".28"/>
          <circle cx="123" cy="24" r="11" fill="currentColor" opacity=".18"/>
        </svg>
        """
    return """
    <svg class="fleet-card-svg" viewBox="0 0 150 120" aria-hidden="true">
      <circle cx="75" cy="51" r="39" fill="currentColor" opacity=".17"/>
      <path d="M83 17 L53 60 H72 L62 93 L101 45 H80 Z" fill="currentColor"/>
      <path d="M18 105 H132" stroke="currentColor" stroke-width="5" stroke-linecap="round" opacity=".25"/>
      <rect x="22" y="72" width="19" height="33" rx="3" fill="currentColor" opacity=".18"/>
      <rect x="45" y="63" width="17" height="42" rx="3" fill="currentColor" opacity=".22"/>
      <rect x="108" y="69" width="20" height="36" rx="3" fill="currentColor" opacity=".16"/>
    </svg>
    """


def render_inventory_summary_card(
    title: str,
    vehicle_type: str,
    summary: dict[str, int | str | None],
) -> None:
    """渲染配置／現況／差額三欄總覽卡。"""
    state = str(summary["state"])
    illustration = _summary_illustration(vehicle_type)
    station_count = safe_nonnegative_int(summary["station_count"])
    configured_total = safe_nonnegative_int(summary["configured_total"])
    current_total = safe_nonnegative_int(summary["current_total"])
    state_label = html.escape(str(summary["state_label"]))
    difference_text = html.escape(str(summary["difference_text"]))
    signed_difference_text = html.escape(str(summary["signed_difference_text"]))

    st.markdown(
        f"""
        <section class="fleet-summary-card fleet-theme-{vehicle_type}">
          <div class="fleet-card-illustration">{illustration}</div>
          <div class="fleet-card-content">
            <div class="fleet-card-heading">
              <div>
                <div class="fleet-card-title">{html.escape(title)}</div>
                <div class="fleet-card-subtitle">目前篩選共 {station_count} 個場站</div>
              </div>
              <span class="fleet-state-badge fleet-state-{state}">{state_label}</span>
            </div>
            <div class="fleet-card-metrics">
              <div class="fleet-metric-block">
                <div class="fleet-metric-label">配置總數</div>
                <div class="fleet-metric-value">{configured_total}<span>台</span></div>
              </div>
              <div class="fleet-metric-block">
                <div class="fleet-metric-label">目前總數</div>
                <div class="fleet-metric-value">{current_total}<span>台</span></div>
              </div>
              <div class="fleet-metric-block fleet-difference-block">
                <div class="fleet-metric-label">差額</div>
                <div class="fleet-difference-chip fleet-difference-{state}">
                  <strong>{difference_text}</strong>
                  <small>{signed_difference_text}</small>
                </div>
              </div>
            </div>
          </div>
        </section>
        """,
        unsafe_allow_html=True,
    )



def render_region_inventory_overview(region_name: str, status_df: pd.DataFrame) -> None:
    """在每個行政區標題下方顯示 2.0／2.0E 的區域車輛總覽。"""
    bike_summary = calculate_inventory_summary(status_df, "2.0 現況", "2.0 標準")
    ebike_summary = calculate_inventory_summary(status_df, "2.0E 現況", "2.0E 標準")

    def metric_html(label: str, summary: dict[str, int | str | None]) -> str:
        state = html.escape(str(summary.get("state") or "pending"))
        state_label = html.escape(str(summary.get("state_label") or "資料未完整"))
        configured_total = safe_nonnegative_int(summary.get("configured_total"))
        current_total = safe_nonnegative_int(summary.get("current_total"))
        difference_text = html.escape(str(summary.get("difference_text") or "—"))
        return f"""
          <div class="region-fleet-metric region-fleet-{state}">
            <div class="region-fleet-metric-head">
              <strong>{html.escape(label)}</strong>
              <span>{state_label}</span>
            </div>
            <div class="region-fleet-numbers">
              <div><small>配置</small><b>{configured_total}<em>台</em></b></div>
              <div><small>目前</small><b>{current_total}<em>台</em></b></div>
              <div><small>差額</small><b>{difference_text}</b></div>
            </div>
          </div>
        """

    st.markdown(
        f"""
        <section class="region-fleet-overview">
          <div class="region-fleet-title">{html.escape(region_name)}｜行政區車輛總覽</div>
          <div class="region-fleet-grid">
            {metric_html("2.0", bike_summary)}
            {metric_html("2.0E", ebike_summary)}
          </div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def render_new_window_download_panel(
    *,
    csv_data: bytes,
    csv_filename: str,
    excel_data: bytes,
    excel_filename: str,
) -> None:
    """以新視窗／新分頁開啟下載，避免 iOS App 被檔案預覽頁取代後無法返回。"""
    file_payload = {
        "csv": {
            "name": csv_filename,
            "mime": "text/csv;charset=utf-8",
            "data": base64.b64encode(csv_data).decode("ascii"),
        },
        "excel": {
            "name": excel_filename,
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "data": base64.b64encode(excel_data).decode("ascii"),
        },
    }
    payload_json = json.dumps(file_payload, ensure_ascii=False).replace("</", "<\\/")
    components.html(
        f"""
        <!doctype html><html lang="zh-Hant"><head><meta charset="utf-8">
        <meta name="viewport" content="width=device-width,initial-scale=1">
        <style>
          *{{box-sizing:border-box}} body{{margin:0;background:transparent;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif}}
          .download-note{{font-size:12px;line-height:1.45;color:#64748b;margin:0 0 8px}}
          .download-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}
          button{{width:100%;border:0;border-radius:12px;padding:12px 10px;font-size:15px;font-weight:800;cursor:pointer}}
          .csv{{background:#e8f3ff;color:#075fb8}} .excel{{background:#e7f8ef;color:#087f5b}}
        </style></head><body>
          <p class="download-note">下載會另開新頁或新分頁；看完檔案後關閉下載頁，即可直接回到原本分析畫面。</p>
          <div class="download-grid">
            <button class="csv" type="button" onclick="openDownload('csv')">下載 CSV</button>
            <button class="excel" type="button" onclick="openDownload('excel')">下載 Excel</button>
          </div>
        <script>
          const files = {payload_json};
          function openDownload(key) {{
            const file = files[key];
            const binary = atob(file.data);
            const bytes = new Uint8Array(binary.length);
            for (let i = 0; i < binary.length; i += 1) bytes[i] = binary.charCodeAt(i);
            const blob = new Blob([bytes], {{type:file.mime}});
            const url = URL.createObjectURL(blob);
            const popup = window.open('', '_blank');
            if (popup) {{
              popup.document.title = `下載 ${{file.name}}`;
              popup.document.body.style.margin = '0';
              popup.document.body.style.padding = '24px';
              popup.document.body.style.fontFamily = '-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif';
              const title = popup.document.createElement('h2');
              title.textContent = '報表下載';
              const note = popup.document.createElement('p');
              note.textContent = '檔案已開始下載。完成後請按下方按鈕關閉此頁，返回原本分析畫面。';
              const downloadLink = popup.document.createElement('a');
              downloadLink.href = url;
              downloadLink.download = file.name;
              downloadLink.textContent = `再次下載：${{file.name}}`;
              downloadLink.style.display = 'block';
              downloadLink.style.margin = '20px 0';
              const closeButton = popup.document.createElement('button');
              closeButton.type = 'button';
              closeButton.textContent = '關閉下載頁，返回系統';
              closeButton.style.padding = '12px 18px';
              closeButton.style.border = '0';
              closeButton.style.borderRadius = '12px';
              closeButton.style.fontSize = '16px';
              closeButton.style.fontWeight = '800';
              closeButton.onclick = () => popup.close();
              popup.document.body.append(title, note, downloadLink, closeButton);
              downloadLink.click();
            }} else {{
              const anchor = document.createElement('a');
              anchor.href = url;
              anchor.download = file.name;
              anchor.target = '_blank';
              anchor.rel = 'noopener';
              document.body.appendChild(anchor);
              anchor.click();
              anchor.remove();
            }}
            window.setTimeout(() => URL.revokeObjectURL(url), 120000);
          }}
        </script></body></html>
        """,
        height=106,
        scrolling=False,
    )


def render_missing_data_notice(missing_bike_count: int, missing_ebike_count: int) -> None:
    """渲染資料不完整提示；有空白時暫停顯示整體缺／多差額。"""
    st.markdown(
        f"""
        <div class="fleet-data-notice">
          <div class="fleet-notice-icon" aria-hidden="true">!</div>
          <div class="fleet-notice-copy">
            <div class="fleet-notice-title">資料提醒</div>
            <div class="fleet-notice-text">
              尚有識別錯誤／空白資料：2.0 共 <strong>{missing_bike_count}</strong> 筆，
              2.0E 共 <strong>{missing_ebike_count}</strong> 筆。<br>
              配置總數與目前已取得的總數仍會顯示，但整體差額會等資料完整後再計算。
            </div>
          </div>
          <div class="fleet-notice-decoration" aria-hidden="true">📋</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_dispatch_legend() -> None:
    st.markdown(
        """
        <div class="fleet-legend" aria-label="狀態圖例">
          <span><i class="fleet-legend-dot fleet-legend-extra"></i>多車</span>
          <span class="fleet-legend-divider">｜</span>
          <span><i class="fleet-legend-dot fleet-legend-short"></i>缺車</span>
          <span class="fleet-legend-divider">｜</span>
          <span><i class="fleet-legend-dot fleet-legend-balanced"></i>符合</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def add_dispatch_indicator(value) -> str:
    """在分析結果後方加上容易辨識的缺車／多車圖案。"""
    text = str(value)
    if "多" in text:
        return f"{text} 🔴"
    if "缺" in text or "少" in text:
        return f"{text} 🟠"
    return text


def extract_dispatch_count(value) -> int | None:
    """只擷取缺車／多車的台數；「符合」不納入排序條件。"""
    text = str(value)
    if not any(status in text for status in ("多", "缺", "少")):
        return None

    match = re.search(r"\d+", text)
    return int(match.group()) if match else 0


SORT_FIELD_OPTIONS = {
    "最大缺／多台數": "max",
    "缺／多總台數": "total",
    "2.0 缺／多台數": "bike",
    "2.0E 缺／多台數": "ebike",
    "距離目前位置最近": "distance",
    "預估行車時間最短": "drive_time",
    "場站名稱": "station",
    "Excel 原始順序": "original",
}



def sort_dispatch_results(result_df, sort_field: str, descending: bool):
    """依使用者選擇排序；以向量化字串擷取取代逐格正規表示式。"""
    if result_df.empty:
        return result_df

    sorted_df = result_df.copy()
    bike_counts = pd.to_numeric(
        sorted_df["2.0 缺／多幾台"].astype(str).str.extract(r"(\d+)", expand=False),
        errors="coerce",
    )
    ebike_counts = pd.to_numeric(
        sorted_df["2.0E 缺／多幾台"].astype(str).str.extract(r"(\d+)", expand=False),
        errors="coerce",
    )
    bike_valid = sorted_df["2.0 缺／多幾台"].astype(str).str.contains("多|缺|少", regex=True)
    ebike_valid = sorted_df["2.0E 缺／多幾台"].astype(str).str.contains("多|缺|少", regex=True)
    bike_counts = bike_counts.where(bike_valid)
    ebike_counts = ebike_counts.where(ebike_valid)

    count_frame = pd.concat([bike_counts, ebike_counts], axis=1)
    sorted_df["_排序最大台數"] = count_frame.max(axis=1, skipna=True).fillna(-1).astype(int)
    sorted_df["_排序總台數"] = count_frame.fillna(0).sum(axis=1).astype(int)
    sorted_df["_排序2.0台數"] = bike_counts.fillna(-1).astype(int)
    sorted_df["_排序2.0E台數"] = ebike_counts.fillna(-1).astype(int)
    distance_source = (
        sorted_df["距離目前位置 (km)"]
        if "距離目前位置 (km)" in sorted_df.columns
        else pd.Series(np.nan, index=sorted_df.index, dtype="float64")
    )
    drive_time_source = (
        sorted_df["預估行車時間 (分)"]
        if "預估行車時間 (分)" in sorted_df.columns
        else pd.Series(np.nan, index=sorted_df.index, dtype="float64")
    )
    sorted_df["_排序距離"] = pd.to_numeric(distance_source, errors="coerce").fillna(float("inf"))
    sorted_df["_排序時間"] = pd.to_numeric(drive_time_source, errors="coerce").fillna(float("inf"))
    sorted_df["_原始順序"] = np.arange(len(sorted_df), dtype=int)

    sort_key = SORT_FIELD_OPTIONS.get(sort_field, "max")
    ascending = not descending

    if sort_key == "max":
        by = ["_排序最大台數", "_排序總台數", "_排序2.0台數", "_排序2.0E台數", "_原始順序"]
        ascending_values = [ascending, ascending, ascending, ascending, True]
    elif sort_key == "total":
        by = ["_排序總台數", "_排序最大台數", "_排序2.0台數", "_排序2.0E台數", "_原始順序"]
        ascending_values = [ascending, ascending, ascending, ascending, True]
    elif sort_key == "bike":
        by = ["_排序2.0台數", "_排序2.0E台數", "_排序總台數", "_原始順序"]
        ascending_values = [ascending, ascending, ascending, True]
    elif sort_key == "ebike":
        by = ["_排序2.0E台數", "_排序2.0台數", "_排序總台數", "_原始順序"]
        ascending_values = [ascending, ascending, ascending, True]
    elif sort_key == "distance":
        by = ["_排序距離", "_排序時間", "_原始順序"]
        ascending_values = [True, True, True]
    elif sort_key == "drive_time":
        by = ["_排序時間", "_排序距離", "_原始順序"]
        ascending_values = [True, True, True]
    elif sort_key == "station":
        by = ["場站名稱", "_原始順序"]
        ascending_values = [ascending, True]
    else:
        by = ["_原始順序"]
        ascending_values = [ascending]

    sorted_df = sorted_df.sort_values(by=by, ascending=ascending_values, kind="mergesort")
    return sorted_df.drop(
        columns=[
            "_排序最大台數", "_排序總台數", "_排序2.0台數",
            "_排序2.0E台數", "_排序距離", "_排序時間", "_原始順序",
        ]
    ).reset_index(drop=True)


def make_colored_export_df(result_df: pd.DataFrame) -> pd.DataFrame:
    """建立含紅／橘圖案的匯出資料，CSV 可直接辨識缺車與多車。"""
    export_df = result_df[
        ["行政區", "場站名稱", "2.0 缺／多幾台", "2.0E 缺／多幾台"]
    ].copy()
    for status_column in ("2.0 缺／多幾台", "2.0E 缺／多幾台"):
        export_df[status_column] = export_df[status_column].map(add_dispatch_indicator)
    return export_df


@st.cache_data(show_spinner=False, max_entries=32)
def build_colored_excel(export_df: pd.DataFrame) -> bytes:
    """輸出真正帶有儲存格底色的 Excel 分析表。"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "調度分析"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    orange_fill = PatternFill("solid", fgColor="FCE5CD")
    green_fill = PatternFill("solid", fgColor="D9EAD3")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")

    headers = list(export_df.columns)
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for record in export_df.itertuples(index=False, name=None):
        worksheet.append(list(record))

    status_columns = {3, 4}
    for row in worksheet.iter_rows(min_row=2):
        for column_index, cell in enumerate(row, start=1):
            cell.alignment = Alignment(
                horizontal="center" if column_index in status_columns else "left",
                vertical="center",
            )
            if column_index not in status_columns:
                continue
            text = str(cell.value or "")
            if "多" in text:
                cell.fill = red_fill
            elif "缺" in text or "少" in text:
                cell.fill = orange_fill
            elif STATUS_UNAVAILABLE_TEXT in text:
                cell.fill = yellow_fill
            elif "符合" in text:
                cell.fill = green_fill

    widths = [14, 34, 20, 20]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()



def normalize_station_text(value) -> str:
    """統一全形字、空白與常見符號，便於場站名稱比對。"""
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("｜", "|").replace("—", "-").replace("–", "-")
    return re.sub(r"\s+", " ", text).strip()



@lru_cache(maxsize=8192)
def _normalize_station_key_cached(text: str) -> str:
    normalized = normalize_station_text(text).lower()
    return re.sub(r"[^0-9a-z\u3400-\u9fff]", "", normalized)


def normalize_station_key(value) -> str:
    """建立只保留中英數字的場站比對鍵，並快取重複字串。"""
    return _normalize_station_key_cached(str(value or ""))


def is_inventory_station_name(value) -> bool:
    """辨識 D1 的兩個庫存場站。"""
    key = normalize_station_key(value)
    return key in {normalize_station_key(name) for name in INVENTORY_STATION_NAMES}


def inventory_station_needs_attention(plan: dict | pd.Series) -> bool:
    """庫存場站只有接近空站或滿站時才解除自動推薦抑制。

    先沿用 v27.5 已有的場站警示定義：總車數 0/1 台，或空位 0/1 格。
    """
    total_bikes = normalize_current_status(plan.get("total_bikes"))
    if total_bikes is None:
        total_bikes = station_total_bikes(plan)
    empty_spaces = normalize_current_status(plan.get("empty_spaces"))
    if empty_spaces is None:
        empty_spaces = station_empty_spaces(plan)
    edge = max(0, safe_nonnegative_int(INVENTORY_STATION_ALERT_EDGE))
    return (
        total_bikes is not None and total_bikes <= edge
    ) or (
        empty_spaces is not None and empty_spaces <= edge
    )


YOUBIKE_STATION_CATALOG_URL = "https://apis.youbike.com.tw/json/station-min-yb2.json"
YOUBIKE_PARKING_INFO_URL = "https://apis.youbike.com.tw/tw2/parkingInfo"
YOUBIKE_REQUEST_TIMEOUT_SECONDS = 25
YOUBIKE_HTTP_MAX_ATTEMPTS = 3
YOUBIKE_STATION_BATCH_SIZE = 100
YOUBIKE_MATCH_THRESHOLD = 0.82
TAIPEI_TIMEZONE = ZoneInfo("Asia/Taipei")


class YouBikeDataError(RuntimeError):
    """YouBike 官網公開資料連線或格式異常。"""


YOUBIKE_BROWSER_COMPONENT_HTML = r"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <style>
    :root { color-scheme: light dark; }
    * { box-sizing: border-box; }
    html, body {
      width: 1px; height: 1px; margin: 0; padding: 0; overflow: hidden;
      background: transparent; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    /* 同步元件留在背景執行；手動操作改由主頁右側懸浮按鈕觸發。 */
    #syncButton, #status { display: none !important; }
    .error { color: #c62828 !important; }
  </style>
</head>
<body>
  <button id="syncButton" type="button">🔄 由手機／瀏覽器取得 YouBike 即時車數</button>
  <div id="status"></div>
<script>
(() => {
  const API_VERSION = 1;
  const button = document.getElementById("syncButton");
  const statusNode = document.getElementById("status");
  let args = {};
  let busy = false;
  let autoTimer = null;
  const SIGNATURE_STORAGE_KEY = "ubike-live-count-signature-v1";
  const UNCHANGED_HEARTBEAT_MS = 5 * 60 * 1000;

  function signatureStorageKey() {
    return `${SIGNATURE_STORAGE_KEY}:${String(args.signature_scope || "default")}`;
  }

  function readDeliveredState() {
    try {
      const parsed = JSON.parse(window.sessionStorage.getItem(signatureStorageKey()) || "{}");
      return {
        signature: String(parsed.signature || ""),
        deliveredAt: Number(parsed.deliveredAt || 0),
      };
    } catch (_) {
      return { signature: "", deliveredAt: 0 };
    }
  }

  function writeDeliveredState(signature, deliveredAt) {
    try {
      window.sessionStorage.setItem(
        signatureStorageKey(),
        JSON.stringify({ signature, deliveredAt })
      );
    } catch (_) {}
  }

  function send(type, data = {}) {
    window.parent.postMessage({ isStreamlitMessage: true, type, ...data }, "*");
  }
  function sendHostSyncState(state, detail = {}) {
    window.parent.postMessage({
      source: "ubike-browser-sync",
      type: "ubike:sync-state",
      state,
      ...detail,
    }, "*");
  }
  function setHeight() {
    send("streamlit:setFrameHeight", { height: 1 });
  }
  function setValue(value) {
    send("streamlit:setComponentValue", { value, dataType: "json" });
  }
  function setStatus(text, isError = false) {
    statusNode.textContent = text || "";
    statusNode.className = isError ? "error" : "";
    setHeight();
  }
  function intOrNull(value) {
    if (value === null || value === undefined || value === "") return null;
    const number = Number(value);
    return Number.isFinite(number) ? Math.max(0, Math.trunc(number)) : null;
  }
  function firstNonempty(...values) {
    for (const value of values) {
      if (value === null || value === undefined) continue;
      if (typeof value === "string" && !value.trim()) continue;
      return value;
    }
    return null;
  }
  function extractItems(payload) {
    if (Array.isArray(payload)) return payload.filter(item => item && typeof item === "object");
    if (!payload || typeof payload !== "object") return [];
    const candidates = [payload.data, payload.result, payload.stations, payload.retVal];
    for (const candidate of candidates) {
      if (Array.isArray(candidate)) return candidate.filter(item => item && typeof item === "object");
      if (candidate && typeof candidate === "object" && Array.isArray(candidate.data)) {
        return candidate.data.filter(item => item && typeof item === "object");
      }
    }
    return [];
  }
  function isTaitung(item) {
    const locationText = [
      item.county_tw, item.city_tw, item.scity, item.district_tw,
      item.address_tw, item.name_tw, item.sarea, item.ar, item.sna
    ].map(value => String(value || "")).join(" ").replaceAll("臺", "台");
    if (locationText.includes("台東縣")) return true;
    const lat = Number(firstNonempty(item.lat, item.latitude));
    const lng = Number(firstNonempty(item.lng, item.longitude));
    return Number.isFinite(lat) && Number.isFinite(lng)
      && lat >= 21.85 && lat <= 23.60 && lng >= 120.70 && lng <= 122.20;
  }
  function sleep(milliseconds) {
    return new Promise(resolve => setTimeout(resolve, Math.max(0, milliseconds)));
  }
  async function fetchJson(url, options = {}, maxAttempts = 3) {
    let lastError = null;
    for (let attempt = 1; attempt <= maxAttempts; attempt += 1) {
      const controller = new AbortController();
      const timeout = setTimeout(() => controller.abort(), 25000);
      try {
        const response = await fetch(url, {
          cache: "no-store",
          credentials: "omit",
          ...options,
          signal: controller.signal,
        });
        if (!response.ok) {
          const error = new Error(`HTTP ${response.status}`);
          error.status = response.status;
          throw error;
        }
        const responseText = await response.text();
        try { return JSON.parse(responseText); }
        catch (_) { throw new Error("官網回傳的內容不是 JSON"); }
      } catch (error) {
        lastError = error;
        const status = Number(error && error.status);
        const retryable = error && (
          error.name === "AbortError" || status === 429 || status >= 500 || !Number.isFinite(status)
        );
        if (!retryable || attempt >= maxAttempts) throw error;
        // 加入少量隨機退避，避免多個並行請求在同一時間再次撞上官網限制。
        await sleep(180 * attempt + Math.floor(Math.random() * 140));
      } finally {
        clearTimeout(timeout);
      }
    }
    throw lastError || new Error("官網請求失敗");
  }
  function batched(values, size) {
    const output = [];
    for (let index = 0; index < values.length; index += size) output.push(values.slice(index, index + size));
    return output;
  }
  function taipeiTimeText() {
    return new Intl.DateTimeFormat("zh-TW", {
      timeZone: "Asia/Taipei", year: "numeric", month: "2-digit", day: "2-digit",
      hour: "2-digit", minute: "2-digit", second: "2-digit", hour12: false,
    }).format(new Date()).replaceAll("-", "/");
  }
  function eventId() {
    if (globalThis.crypto && typeof globalThis.crypto.randomUUID === "function") return globalThis.crypto.randomUUID();
    return `${Date.now()}-${Math.random().toString(16).slice(2)}`;
  }

  function scheduleAutoSync() {
    if (autoTimer !== null) {
      clearTimeout(autoTimer);
      autoTimer = null;
    }
    if (!args.auto_refresh) return;
    const seconds = Math.max(5, Math.min(60, Number(args.auto_refresh_seconds || 60)));
    autoTimer = setTimeout(() => {
      autoTimer = null;
      if (busy) scheduleAutoSync();
      else runSync();
    }, seconds * 1000);
  }

  async function runSync({ forceDelivery = false } = {}) {
    if (busy) return;
    if (autoTimer !== null) {
      clearTimeout(autoTimer);
      autoTimer = null;
    }
    busy = true;
    const startedAt = performance.now();
    sendHostSyncState("busy");
    button.disabled = true;
    button.textContent = "⏳ 正在高速分批讀取 YouBike 官網……";
    setStatus("將以最大批次、有限並行及只補漏站的方式取得資料，不經 TDX。", false);

    try {
      const catalogUrl = args.catalog_url || "https://apis.youbike.com.tw/json/station-min-yb2.json";
      const parkingUrl = args.parking_url || "https://apis.youbike.com.tw/tw2/parkingInfo";
      const batchSize = Math.max(1, Math.min(50, Number(args.batch_size || 20)));
      const concurrency = Math.max(1, Math.min(6, Number(args.request_concurrency || 4)));
      const maxBatchRounds = Math.max(1, Math.min(8, Number(args.max_batch_rounds || 4)));
      const maxSingleRounds = Math.max(0, Math.min(4, Number(args.max_single_rounds || 2)));
      const waveDelayMs = Math.max(0, Math.min(1000, Number(args.wave_delay_ms || 70)));

      const catalogPayload = await fetchJson(catalogUrl, {
        method: "GET",
        headers: { "Accept": "application/json, text/plain, */*" },
      });
      const catalog = extractItems(catalogPayload).filter(isTaitung).map(item => {
        const stationId = String(firstNonempty(item.station_no, item.sno, item.station_id) || "").trim();
        const stationName = String(firstNonempty(item.name_tw, item.sna, item.station_name) || "").trim();
        return {
          station_uid: stationId,
          station_id: stationId,
          station_name: stationName,
          service_status: intOrNull(firstNonempty(item.status, item.act, 1)) ?? 1,
          source_update_time: String(firstNonempty(item.updated_at, item.mday, item.time) || "").trim(),
          latitude: firstNonempty(item.lat, item.latitude),
          longitude: firstNonempty(item.lng, item.longitude),
        };
      }).filter(item => item.station_id && item.station_name);

      if (!catalog.length) throw new Error("官網站點清單中找不到臺東候選場站");

      const requestedStationIds = [...new Set(catalog.map(item => item.station_id))];
      const requestedStationIdSet = new Set(requestedStationIds);
      const parkingMap = new Map();
      let requestCount = 0;
      let failedRequestCount = 0;
      let batchRoundCount = 0;
      let singleRoundCount = 0;

      function stationIdOf(item) {
        return String(firstNonempty(item && item.station_no, item && item.sno) || "").trim();
      }

      function mergeParkingItems(items) {
        let addedCount = 0;
        for (const item of items) {
          const stationId = stationIdOf(item);
          if (!stationId || !requestedStationIdSet.has(stationId)) continue;
          if (!parkingMap.has(stationId)) addedCount += 1;
          parkingMap.set(stationId, item);
        }
        return addedCount;
      }

      function currentMissingIds() {
        return requestedStationIds.filter(stationId => !parkingMap.has(stationId));
      }

      async function requestParkingGroup(stationIds) {
        const payload = await fetchJson(parkingUrl, {
          method: "POST",
          headers: {
            "Accept": "application/json, text/plain, */*",
            "Content-Type": "application/json;charset=UTF-8",
          },
          body: JSON.stringify({ station_no: stationIds }),
        });
        return extractItems(payload);
      }

      async function runGroups(groups, phaseText, workerLimit = concurrency) {
        if (!groups.length) return { addedCount: 0, failedGroups: [] };
        let nextIndex = 0;
        let addedCount = 0;
        let completedCount = 0;
        const failedGroups = [];

        async function worker() {
          while (true) {
            const index = nextIndex;
            nextIndex += 1;
            if (index >= groups.length) return;
            const stationIds = groups[index];
            try {
              requestCount += 1;
              const items = await requestParkingGroup(stationIds);
              addedCount += mergeParkingItems(items);
            } catch (error) {
              failedRequestCount += 1;
              failedGroups.push({ stationIds, error: String(error && error.message ? error.message : error) });
            } finally {
              completedCount += 1;
              const missingCount = currentMissingIds().length;
              setStatus(
                `${phaseText}：完成 ${completedCount}／${groups.length} 批，已取得 ` +
                `${parkingMap.size}／${requestedStationIds.length} 站，尚缺 ${missingCount} 站`,
                false,
              );
            }
          }
        }

        const workerCount = Math.min(Math.max(1, workerLimit), groups.length);
        await Promise.all(Array.from({ length: workerCount }, () => worker()));
        return { addedCount, failedGroups };
      }

      let missingStationIds = currentMissingIds();
      let previousMissingCount = missingStationIds.length + 1;

      // 主階段：每一輪都使用設定的最大批次，並行查完後只保留仍缺少的場站進入下一輪。
      for (let round = 1; round <= maxBatchRounds && missingStationIds.length; round += 1) {
        batchRoundCount = round;
        const groups = batched(missingStationIds, batchSize);
        setStatus(
          `高速批次第 ${round} 輪：${missingStationIds.length} 站，分成 ${groups.length} 批並行讀取……`,
          false,
        );
        const result = await runGroups(groups, `高速批次第 ${round} 輪`);
        missingStationIds = currentMissingIds();

        if (!missingStationIds.length) break;
        // 這一輪完全沒有新增資料時，繼續重送相同批次沒有速度效益，立即改走單站補查。
        if (result.addedCount <= 0 || missingStationIds.length >= previousMissingCount) break;
        previousMissingCount = missingStationIds.length;
        if (waveDelayMs) await sleep(waveDelayMs);
      }

      // 最後階段：只對殘留漏站做單站並行查詢，避免一個異常站拖累同批其他場站。
      missingStationIds = currentMissingIds();
      for (let round = 1; round <= maxSingleRounds && missingStationIds.length; round += 1) {
        singleRoundCount = round;
        const singleGroups = missingStationIds.map(stationId => [stationId]);
        setStatus(
          `單站補查第 ${round} 輪：正在並行補齊最後 ${missingStationIds.length} 個場站……`,
          false,
        );
        const beforeCount = parkingMap.size;
        await runGroups(singleGroups, `單站補查第 ${round} 輪`, Math.min(6, concurrency + 1));
        missingStationIds = currentMissingIds();
        if (!missingStationIds.length) break;
        // 即使這一輪暫時沒有新增，也保留後續重試機會；官網可能只是短暫漏回或限流。
        const noProgressDelay = parkingMap.size <= beforeCount ? 260 * round : waveDelayMs + 60;
        if (noProgressDelay) await sleep(noProgressDelay);
      }

      const sourceTimes = [];
      const records = [];
      for (const station of catalog) {
        const parking = parkingMap.get(station.station_id);
        if (!parking) continue;
        let detail = firstNonempty(parking.available_spaces_detail, parking.sbi_detail);
        if (!detail || typeof detail !== "object") detail = {};
        const sourceTime = String(firstNonempty(
          parking.updated_at, parking.mday, parking.time, station.source_update_time
        ) || "").trim();
        if (sourceTime) sourceTimes.push(sourceTime);
        records.push({
          ...station,
          service_status: intOrNull(firstNonempty(parking.status, parking.act, station.service_status, 1)) ?? 1,
          general_bikes: intOrNull(detail.yb2),
          electric_bikes: intOrNull(detail.eyb),
          available_spaces: intOrNull(firstNonempty(parking.available_spaces, parking.sbi)),
          empty_spaces: intOrNull(firstNonempty(parking.empty_spaces, parking.bemp)),
          parking_spaces: intOrNull(firstNonempty(parking.parking_spaces, parking.tot)),
          source_update_time: sourceTime,
        });
      }

      if (!records.length) throw new Error("官網沒有回傳臺東場站即時車數");
      missingStationIds = currentMissingIds();

      // 每次仍照常向官網取得資料；只有車數／營運狀態／漏站清單真的變動時，
      // 才把值送回 Streamlit 觸發整頁重算。資料完全相同時最多五分鐘送一次心跳。
      const signature = records
        .map(record => [
          record.station_id,
          record.general_bikes ?? "",
          record.electric_bikes ?? "",
          record.empty_spaces ?? "",
          record.parking_spaces ?? "",
          record.service_status ?? "",
        ].join(":"))
        .join("|") + `|missing:${missingStationIds.join(",")}`;
      const deliveredState = readDeliveredState();
      const nowMilliseconds = Date.now();
      const shouldDeliver = (
        forceDelivery
        || Boolean(args.force_initial_delivery)
        || signature !== deliveredState.signature
        || nowMilliseconds - deliveredState.deliveredAt >= UNCHANGED_HEARTBEAT_MS
      );

      if (shouldDeliver) {
        setValue({
          ok: true,
          event_id: eventId(),
          records,
          fetched_at: taipeiTimeText(),
          latest_source_time: sourceTimes.length ? sourceTimes.sort().at(-1) : "",
          station_count: records.length,
          requested_station_count: requestedStationIds.length,
          missing_station_count: missingStationIds.length,
          missing_station_ids: missingStationIds,
          request_batch_count: requestCount,
          request_count: requestCount,
          failed_request_count: failedRequestCount,
          batch_round_count: batchRoundCount,
          single_round_count: singleRoundCount,
          batch_size: batchSize,
          request_concurrency: concurrency,
          elapsed_ms: Math.max(0, Math.round(performance.now() - startedAt)),
          source: "YouBike 官網公開接口（高速循環補查，由使用者瀏覽器直接取得，免 TDX）",
        });
        writeDeliveredState(signature, nowMilliseconds);
      }

      const missingText = missingStationIds.length ? `，仍缺 ${missingStationIds.length} 個` : "，已全數取得";
      const deliveryText = shouldDeliver
        ? "，資料有更新，正在寫入分析系統……"
        : "，車數無變化，已略過整頁重算";
      setStatus(
        `已取得 ${records.length}／${requestedStationIds.length} 個場站${missingText}，共送出 ${requestCount} 次請求${deliveryText}`,
        false,
      );
      sendHostSyncState("success", { station_count: records.length, changed: shouldDeliver });
    } catch (error) {
      const message = error && error.name === "AbortError"
        ? "連線逾時，請檢查手機網路後再試"
        : String(error && error.message ? error.message : error);
      setValue({ ok: false, event_id: eventId(), error: message });
      setStatus(`同步失敗：${message}`, true);
      sendHostSyncState("error", { message });
    } finally {
      busy = false;
      button.disabled = false;
      button.textContent = args.button_label || "🔄 由手機／瀏覽器取得 YouBike 即時車數";
      setHeight();
      scheduleAutoSync();
    }
  }

  button.addEventListener("click", () => runSync({ forceDelivery: true }));
  window.addEventListener("message", event => {
    if (!event.data) return;
    if (event.data.type === "ubike:manual-sync") {
      runSync({ forceDelivery: true });
      return;
    }
    if (event.data.type !== "streamlit:render") return;
    args = event.data.args || {};
    button.textContent = args.button_label || "🔄 手動更新即時車數";
    button.disabled = Boolean(event.data.disabled) || busy;
    scheduleAutoSync();
    setHeight();
  });

  send("streamlit:componentReady", { apiVersion: API_VERSION });
  setHeight();
})();
</script>
</body>
</html>
"""


_YOUBIKE_BROWSER_SYNC_COMPONENT = None


def get_youbike_browser_sync_component():
    """建立雙向 Streamlit 元件，讓請求從使用者瀏覽器發出以避開雲端主機 503。"""
    global _YOUBIKE_BROWSER_SYNC_COMPONENT
    if _YOUBIKE_BROWSER_SYNC_COMPONENT is not None:
        return _YOUBIKE_BROWSER_SYNC_COMPONENT

    component_dir = Path(tempfile.gettempdir()) / "youbike_browser_sync_component_v2"
    component_dir.mkdir(parents=True, exist_ok=True)
    index_path = component_dir / "index.html"
    try:
        if not index_path.exists() or index_path.read_text(encoding="utf-8") != YOUBIKE_BROWSER_COMPONENT_HTML:
            index_path.write_text(YOUBIKE_BROWSER_COMPONENT_HTML, encoding="utf-8")
    except OSError as exc:
        raise YouBikeDataError(f"無法建立瀏覽器同步元件：{exc}") from exc

    _YOUBIKE_BROWSER_SYNC_COMPONENT = components.declare_component(
        "youbike_browser_sync_v2",
        path=str(component_dir),
    )
    return _YOUBIKE_BROWSER_SYNC_COMPONENT


def normalize_browser_live_payload(payload) -> dict:
    """驗證瀏覽器回傳資料並補齊 Python 端配對所需欄位。"""
    if not isinstance(payload, dict):
        raise YouBikeDataError("瀏覽器沒有回傳有效資料。")
    if not payload.get("ok"):
        raise YouBikeDataError(str(payload.get("error") or "瀏覽器同步失敗。"))

    raw_records = payload.get("records")
    if not isinstance(raw_records, list):
        raise YouBikeDataError("瀏覽器回傳的場站資料格式不正確。")

    records: list[dict] = []
    for raw_record in raw_records:
        if not isinstance(raw_record, dict):
            continue
        station_id = str(raw_record.get("station_id") or raw_record.get("station_uid") or "").strip()
        station_name = str(raw_record.get("station_name") or "").strip()
        if not station_id or not station_name:
            continue
        records.append(
            {
                **raw_record,
                "station_uid": station_id,
                "station_id": station_id,
                "station_name": station_name,
                "station_key": normalize_youbike_station_key(station_name),
                "service_status": safe_nonnegative_int(raw_record.get("service_status", 1)),
                "general_bikes": normalize_current_status(raw_record.get("general_bikes")),
                "electric_bikes": normalize_current_status(raw_record.get("electric_bikes")),
                "available_spaces": normalize_current_status(raw_record.get("available_spaces")),
                "empty_spaces": normalize_current_status(raw_record.get("empty_spaces")),
                "parking_spaces": normalize_current_status(raw_record.get("parking_spaces")),
            }
        )

    if not records:
        raise YouBikeDataError("瀏覽器沒有回傳可用的臺東場站即時車數。")

    return {
        "records": records,
        "fetched_at": str(payload.get("fetched_at") or datetime.now(TAIPEI_TIMEZONE).strftime("%Y/%m/%d %H:%M:%S")),
        "latest_source_time": str(payload.get("latest_source_time") or "").strip(),
        "station_count": len(records),
        "requested_station_count": safe_nonnegative_int(payload.get("requested_station_count")),
        "missing_station_count": safe_nonnegative_int(payload.get("missing_station_count")),
        "request_batch_count": safe_nonnegative_int(payload.get("request_batch_count")),
        "request_count": safe_nonnegative_int(payload.get("request_count")),
        "failed_request_count": safe_nonnegative_int(payload.get("failed_request_count")),
        "batch_round_count": safe_nonnegative_int(payload.get("batch_round_count")),
        "single_round_count": safe_nonnegative_int(payload.get("single_round_count")),
        "batch_size": safe_nonnegative_int(payload.get("batch_size")),
        "request_concurrency": safe_nonnegative_int(payload.get("request_concurrency")),
        "elapsed_ms": safe_nonnegative_int(payload.get("elapsed_ms")),
        "missing_station_ids": [
            str(value).strip() for value in payload.get("missing_station_ids", [])
            if str(value).strip()
        ] if isinstance(payload.get("missing_station_ids"), list) else [],
        "source": str(payload.get("source") or "YouBike 官網公開接口（高速循環補查，由瀏覽器直接取得，免 TDX）"),
        "event_id": str(payload.get("event_id") or "").strip(),
    }


def _first_nonempty(*values):
    """回傳第一個不是 None／空字串的值。"""
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return None


def _youbike_http_json(
    url: str,
    *,
    method: str = "GET",
    json_body: dict | None = None,
):
    """讀取 YouBike 官網公開 JSON；免 TDX、免帳號、免 API 金鑰。"""
    request_headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.7",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/126.0 Safari/537.36"
        ),
        "Referer": "https://www.youbike.com.tw/region/taitung/stations/",
        "Origin": "https://www.youbike.com.tw",
    }

    encoded_body = None
    if json_body is not None:
        encoded_body = json.dumps(json_body, ensure_ascii=False).encode("utf-8")
        request_headers["Content-Type"] = "application/json;charset=UTF-8"

    last_error: Exception | None = None
    for attempt in range(1, YOUBIKE_HTTP_MAX_ATTEMPTS + 1):
        request = Request(
            url,
            data=encoded_body,
            headers=request_headers,
            method=method,
        )

        try:
            with urlopen(request, timeout=YOUBIKE_REQUEST_TIMEOUT_SECONDS) as response:
                raw_body = response.read().decode("utf-8-sig")

            try:
                payload = json.loads(raw_body)
            except json.JSONDecodeError as exc:
                raise YouBikeDataError("YouBike 官網回傳內容不是有效的 JSON。") from exc

            if isinstance(payload, dict):
                ret_code = payload.get("retCode")
                if ret_code not in (None, 1, "1", True):
                    ret_message = str(payload.get("retMsg") or "官方資料服務回傳失敗").strip()
                    raise YouBikeDataError(f"YouBike 官網資料服務錯誤：{ret_message}")
            return payload

        except HTTPError as exc:
            last_error = exc
            try:
                detail = exc.read().decode("utf-8", errors="replace")[:240]
            except Exception:
                detail = ""

            # 429 與 5xx 通常是暫時性錯誤，先短暫退避後自動重試。
            if exc.code == 429 or 500 <= exc.code <= 599:
                if attempt < YOUBIKE_HTTP_MAX_ATTEMPTS:
                    time.sleep(0.8 * attempt)
                    continue
                if exc.code == 429:
                    raise YouBikeDataError(
                        "官方資料請求過於頻繁，請等候約 1 分鐘再試。"
                    ) from exc

            if exc.code in (401, 403):
                raise YouBikeDataError(
                    "YouBike 官網暫時拒絕此主機連線，請稍後再試。"
                ) from exc
            raise YouBikeDataError(
                f"YouBike 官網資料回傳 HTTP {exc.code}。{detail or '請稍後再試。'}"
            ) from exc
        except (URLError, TimeoutError) as exc:
            last_error = exc
            if attempt < YOUBIKE_HTTP_MAX_ATTEMPTS:
                time.sleep(0.8 * attempt)
                continue

    if isinstance(last_error, URLError):
        reason = getattr(last_error, "reason", last_error)
        raise YouBikeDataError(f"無法連線至 YouBike 官網資料服務：{reason}") from last_error
    if isinstance(last_error, TimeoutError):
        raise YouBikeDataError("連線 YouBike 官網資料服務逾時，請稍後重試。") from last_error
    raise YouBikeDataError("YouBike 官網資料服務暫時無法使用，請稍後重試。")


def _extract_youbike_station_items(payload) -> list[dict]:
    """相容清單陣列、data 包裝，以及 retVal.data 等常見官方格式。"""
    containers = [payload]
    if isinstance(payload, dict):
        containers.append(payload.get("retVal"))

    for container in containers:
        if isinstance(container, list):
            return [item for item in container if isinstance(item, dict)]
        if not isinstance(container, dict):
            continue
        for key in ("data", "result", "stations", "retVal"):
            values = container.get(key)
            if isinstance(values, list):
                return [item for item in values if isinstance(item, dict)]
            if isinstance(values, dict):
                nested_data = values.get("data")
                if isinstance(nested_data, list):
                    return [item for item in nested_data if isinstance(item, dict)]
    return []


@lru_cache(maxsize=8192)
def _normalize_youbike_station_key_cached(raw_text: str) -> str:
    text = normalize_station_text(raw_text).lower().replace("臺", "台")
    text = re.sub(
        r"^(?:youbike|ubike)\s*2\s*[.．]?\s*0\s*e?\s*[_\-－—:：]*\s*",
        "",
        text,
        flags=re.IGNORECASE,
    )
    text = text.replace("公共自行車租賃站", "")
    return re.sub(r"[^0-9a-z\u3400-\u9fff]", "", text)


def normalize_youbike_station_key(value) -> str:
    """正規化 Excel 與官網站名；常見重複站名會直接從快取取用。"""
    return _normalize_youbike_station_key_cached(str(value or ""))


def _youbike_station_similarity(excel_name: str, api_name: str) -> float:
    excel_key = normalize_youbike_station_key(excel_name)
    api_key = normalize_youbike_station_key(api_name)
    if not excel_key or not api_key:
        return 0.0
    if excel_key == api_key:
        return 1.0
    if min(len(excel_key), len(api_key)) >= 4 and (
        excel_key in api_key or api_key in excel_key
    ):
        return 0.96
    return SequenceMatcher(None, excel_key, api_key, autojunk=False).ratio()


def _looks_like_taitung_station(record: dict) -> bool:
    """以官方欄位及經緯度範圍篩出臺東縣候選場站。"""
    location_text = " ".join(
        str(record.get(key) or "")
        for key in (
            "county_tw", "city_tw", "scity", "district_tw", "address_tw",
            "name_tw", "sarea", "ar", "sna",
        )
    ).replace("臺", "台")
    if "台東縣" in location_text:
        return True

    try:
        latitude = float(_first_nonempty(record.get("lat"), record.get("latitude")))
        longitude = float(_first_nonempty(record.get("lng"), record.get("longitude")))
    except (TypeError, ValueError):
        return False

    # 包含臺東本島、綠島與蘭嶼。後續仍會以場站名稱做一對一安全配對。
    return 21.85 <= latitude <= 23.60 and 120.70 <= longitude <= 122.20


@st.cache_data(show_spinner=False, ttl=21600, max_entries=4)
def fetch_youbike_taitung_station_catalog() -> list[dict]:
    """取得 YouBike 全臺站點清單並留下臺東縣候選場站。"""
    payload = _youbike_http_json(YOUBIKE_STATION_CATALOG_URL)
    items = _extract_youbike_station_items(payload)
    records: list[dict] = []

    for item in items:
        if not _looks_like_taitung_station(item):
            continue
        station_no = str(
            _first_nonempty(item.get("station_no"), item.get("sno"), item.get("station_id"))
            or ""
        ).strip()
        station_name = str(
            _first_nonempty(item.get("name_tw"), item.get("sna"), item.get("station_name"))
            or ""
        ).strip()
        if not station_no or not station_name:
            continue

        raw_status = _first_nonempty(item.get("status"), item.get("act"), 1)
        records.append(
            {
                "station_uid": station_no,
                "station_id": station_no,
                "station_name": station_name,
                "station_key": normalize_youbike_station_key(station_name),
                "service_status": safe_nonnegative_int(raw_status),
                "source_update_time": str(
                    _first_nonempty(item.get("updated_at"), item.get("mday"), item.get("time"))
                    or ""
                ).strip(),
                "latitude": _first_nonempty(item.get("lat"), item.get("latitude")),
                "longitude": _first_nonempty(item.get("lng"), item.get("longitude")),
            }
        )

    if not records:
        raise YouBikeDataError("YouBike 官網沒有回傳可辨識的臺東場站清單。")
    return records


def _batched_station_numbers(station_numbers: list[str]):
    """將場站編號分批，避免單次 POST 過大而被官方服務拒絕。"""
    for start_index in range(0, len(station_numbers), YOUBIKE_STATION_BATCH_SIZE):
        yield station_numbers[start_index : start_index + YOUBIKE_STATION_BATCH_SIZE]


@st.cache_data(show_spinner=False, ttl=60, max_entries=8)
def fetch_youbike_taitung_bike_data(refresh_bucket: int) -> dict:
    """取得臺東縣 YouBike 2.0／2.0E 即時可借車數；完全不使用 TDX。"""
    del refresh_bucket  # 讓快取依傳入分鐘批次更新，同分鐘內避免重複打官方接口。
    catalog = fetch_youbike_taitung_station_catalog()
    station_numbers = [record["station_id"] for record in catalog]

    parking_items: list[dict] = []
    station_batches = list(_batched_station_numbers(station_numbers))
    for batch_index, station_batch in enumerate(station_batches):
        payload = _youbike_http_json(
            YOUBIKE_PARKING_INFO_URL,
            method="POST",
            json_body={"station_no": station_batch},
        )
        parking_items.extend(_extract_youbike_station_items(payload))
        if batch_index < len(station_batches) - 1:
            time.sleep(0.15)

    parking_by_station = {
        str(_first_nonempty(item.get("station_no"), item.get("sno")) or "").strip(): item
        for item in parking_items
        if str(_first_nonempty(item.get("station_no"), item.get("sno")) or "").strip()
    }

    records: list[dict] = []
    source_times: list[str] = []
    for station in catalog:
        parking = parking_by_station.get(station["station_id"])
        if not isinstance(parking, dict):
            continue

        detail = _first_nonempty(
            parking.get("available_spaces_detail"),
            parking.get("sbi_detail"),
        )
        if not isinstance(detail, dict):
            detail = {}

        general_bikes = normalize_current_status(detail.get("yb2"))
        electric_bikes = normalize_current_status(detail.get("eyb"))
        source_update_time = str(
            _first_nonempty(
                parking.get("updated_at"),
                parking.get("mday"),
                parking.get("time"),
                station.get("source_update_time"),
            )
            or ""
        ).strip()
        if source_update_time:
            source_times.append(source_update_time)

        raw_service_status = _first_nonempty(
            parking.get("status"),
            parking.get("act"),
            station.get("service_status"),
            1,
        )
        records.append(
            {
                **station,
                "service_status": safe_nonnegative_int(raw_service_status),
                "general_bikes": general_bikes,
                "electric_bikes": electric_bikes,
                "available_spaces": normalize_current_status(
                    _first_nonempty(parking.get("available_spaces"), parking.get("sbi"))
                ),
                "empty_spaces": normalize_current_status(
                    _first_nonempty(parking.get("empty_spaces"), parking.get("bemp"))
                ),
                "parking_spaces": normalize_current_status(
                    _first_nonempty(parking.get("parking_spaces"), parking.get("tot"))
                ),
                "source_update_time": source_update_time,
            }
        )

    if not records:
        raise YouBikeDataError(
            "YouBike 官網沒有回傳臺東場站即時車數，可能是官方資料服務暫時異常。"
        )

    fetched_at = datetime.now(TAIPEI_TIMEZONE).strftime("%Y/%m/%d %H:%M:%S")
    return {
        "records": records,
        "fetched_at": fetched_at,
        "latest_source_time": max(source_times) if source_times else "",
        "station_count": len(records),
        "request_batch_count": len(station_batches),
        "source": "YouBike 官網公開接口（免 TDX）",
    }


def build_youbike_match_index(live_records: list[dict]) -> dict[str, object]:
    """預先建立精確站名索引，供同一批即時資料重複配對。"""
    exact: dict[str, list[dict]] = {}
    prepared_records: list[tuple[dict, str]] = []
    for record in live_records:
        station_key = str(record.get("station_key") or "").strip()
        if not station_key:
            station_key = normalize_youbike_station_key(record.get("station_name", ""))
        exact.setdefault(station_key, []).append(record)
        prepared_records.append((record, str(record.get("station_name", ""))))
    return {"exact": exact, "prepared_records": prepared_records}


def match_youbike_station(
    excel_name: str,
    live_records: list[dict],
    match_index: dict[str, object] | None = None,
) -> tuple[dict | None, float, bool]:
    """配對 Excel 與官網站名；使用共用索引並避免每站完整排序。"""
    excel_key = normalize_youbike_station_key(excel_name)
    if not excel_key:
        return None, 0.0, False

    index = match_index if isinstance(match_index, dict) else build_youbike_match_index(live_records)
    exact_map = index.get("exact", {})
    exact_matches = exact_map.get(excel_key, []) if isinstance(exact_map, dict) else []
    if len(exact_matches) == 1:
        return exact_matches[0], 1.0, False
    if len(exact_matches) > 1:
        return None, 1.0, True

    prepared_records = index.get("prepared_records", [])
    if not isinstance(prepared_records, list) or not prepared_records:
        return None, 0.0, False

    best_score = -1.0
    second_score = -1.0
    best_record: dict | None = None
    for item in prepared_records:
        if not isinstance(item, tuple) or len(item) != 2:
            continue
        record, station_name = item
        score = _youbike_station_similarity(excel_name, station_name)
        if score > best_score:
            second_score = best_score
            best_score = score
            best_record = record
        elif score > second_score:
            second_score = score

    if best_record is None or best_score < YOUBIKE_MATCH_THRESHOLD:
        return None, max(0.0, best_score), False

    ambiguous = best_score < 0.96 and second_score >= best_score - 0.035
    if ambiguous:
        return None, best_score, True
    return best_record, best_score, False


def apply_youbike_updates_to_dataframe(
    base_df: pd.DataFrame,
    live_records: list[dict],
    *,
    match_index: dict[str, object] | None = None,
    match_cache: dict[str, tuple[dict | None, float, bool]] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """將 YouBike 官網即時資料寫入配置；同一輪同步共用站名配對結果。"""
    updated_df = coerce_nullable_station_status(base_df.copy())
    report_rows: list[dict] = []
    used_station_ids: set[str] = set()
    matched_count = 0
    skipped_count = 0
    unmatched_count = 0
    resolved_match_index = (
        match_index
        if isinstance(match_index, dict)
        else build_youbike_match_index(live_records)
    )
    resolved_match_cache = match_cache if isinstance(match_cache, dict) else {}

    for row_index, raw_excel_name in updated_df["場站名稱"].items():
        excel_name = str(raw_excel_name or "").strip()
        cached_match = resolved_match_cache.get(excel_name)
        if cached_match is None:
            cached_match = match_youbike_station(
                excel_name,
                live_records,
                resolved_match_index,
            )
            resolved_match_cache[excel_name] = cached_match
        matched_record, score, ambiguous = cached_match

        if matched_record is None:
            unmatched_count += 1
            report_rows.append(
                {
                    "Excel 場站": excel_name,
                    "YouBike 場站": "",
                    "2.0": pd.NA,
                    "2.0E": pd.NA,
                    "結果": "名稱可能重複，未寫入" if ambiguous else "找不到安全配對",
                    "相似度": round(score, 3),
                }
            )
            continue

        station_id = str(matched_record.get("station_id", "") or "")
        if station_id in used_station_ids:
            unmatched_count += 1
            report_rows.append(
                {
                    "Excel 場站": excel_name,
                    "YouBike 場站": matched_record.get("station_name", ""),
                    "2.0": pd.NA,
                    "2.0E": pd.NA,
                    "結果": "同一官網場站已配對，未重複寫入",
                    "相似度": round(score, 3),
                }
            )
            continue

        bike_count = normalize_current_status(matched_record.get("general_bikes"))
        ebike_count = normalize_current_status(matched_record.get("electric_bikes"))
        total_available = normalize_current_status(matched_record.get("available_spaces"))

        # 官網偶爾只缺其中一個車種明細；若總可借數完整，可用加總關係安全補算。
        if total_available is not None:
            if bike_count is None and ebike_count is not None and total_available >= ebike_count:
                bike_count = total_available - ebike_count
            elif ebike_count is None and bike_count is not None and total_available >= bike_count:
                ebike_count = total_available - bike_count

        service_status = normalize_current_status(matched_record.get("service_status"))
        empty_spaces = normalize_current_status(matched_record.get("empty_spaces"))
        parking_spaces = normalize_current_status(matched_record.get("parking_spaces"))
        total_bikes = total_available
        if total_bikes is None and bike_count is not None and ebike_count is not None:
            total_bikes = bike_count + ebike_count
        if parking_spaces is None and total_bikes is not None and empty_spaces is not None:
            parking_spaces = total_bikes + empty_spaces

        # 即使車種明細不完整或場站暫停服務，仍保存官網狀態供搜尋與警示使用。
        updated_df.at[row_index, "總車數"] = pd.NA if total_bikes is None else total_bikes
        updated_df.at[row_index, "空位數"] = pd.NA if empty_spaces is None else empty_spaces
        updated_df.at[row_index, "總柱數"] = pd.NA if parking_spaces is None else parking_spaces
        updated_df.at[row_index, "服務狀態"] = pd.NA if service_status is None else service_status

        service_stopped = service_status is not None and service_status != 1
        if service_stopped or bike_count is None or ebike_count is None:
            skipped_count += 1
            reason = "場站目前非正常服務" if service_stopped else "官網車種明細不完整"
            report_rows.append(
                {
                    "Excel 場站": excel_name,
                    "YouBike 場站": matched_record.get("station_name", ""),
                    "2.0": pd.NA if bike_count is None else bike_count,
                    "2.0E": pd.NA if ebike_count is None else ebike_count,
                    "結果": f"{reason}，未寫入",
                    "相似度": round(score, 3),
                }
            )
            continue

        used_station_ids.add(station_id)
        updated_df.at[row_index, "2.0 現況"] = bike_count
        updated_df.at[row_index, "2.0E 現況"] = ebike_count
        matched_count += 1
        report_rows.append(
            {
                "Excel 場站": excel_name,
                "YouBike 場站": matched_record.get("station_name", ""),
                "2.0": bike_count,
                "2.0E": ebike_count,
                "結果": "已寫入",
                "相似度": round(score, 3),
            }
        )

    report_df = pd.DataFrame(report_rows)
    summary = {
        "matched_count": matched_count,
        "skipped_count": skipped_count,
        "unmatched_count": unmatched_count,
        "total_count": len(updated_df),
    }
    return coerce_nullable_station_status(updated_df), report_df, summary


BASE_CACHE_DIR = Path(__file__).resolve().parent / ".base_cache"


BASE_TOKEN_BROWSER_STORE_HTML = r"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <style>html,body{width:1px;height:1px;margin:0;overflow:hidden;background:transparent}</style>
</head>
<body>
<script>
(() => {
  const API_VERSION = 1;
  let lastValue = "";

  function send(type, data = {}) {
    window.parent.postMessage({ isStreamlitMessage: true, type, ...data }, "*");
  }

  function setValue(token) {
    const normalized = String(token || "").trim().toLowerCase();
    if (normalized === lastValue) return;
    lastValue = normalized;
    send("streamlit:setComponentValue", {
      value: { token: normalized },
      dataType: "json",
    });
  }

  function validToken(token) {
    return /^[0-9a-f]{32}$/.test(String(token || "").trim().toLowerCase());
  }

  function handleRender(event) {
    const args = event.data.args || {};
    const storageKey = String(args.storage_key || "ubike_dispatch_active_base_token_v1");
    const currentToken = String(args.current_token || "").trim().toLowerCase();
    const clearStored = Boolean(args.clear_stored);

    try {
      if (clearStored) {
        localStorage.removeItem(storageKey);
        setValue("");
      } else if (validToken(currentToken)) {
        localStorage.setItem(storageKey, currentToken);
        setValue(currentToken);
      } else {
        const storedToken = String(localStorage.getItem(storageKey) || "").trim().toLowerCase();
        setValue(validToken(storedToken) ? storedToken : "");
      }
    } catch (_) {
      setValue(validToken(currentToken) ? currentToken : "");
    }

    send("streamlit:setFrameHeight", { height: 1 });
  }

  window.addEventListener("message", event => {
    if (event.data && event.data.type === "streamlit:render") handleRender(event);
  });
  send("streamlit:componentReady", { apiVersion: API_VERSION });
  send("streamlit:setFrameHeight", { height: 1 });
})();
</script>
</body>
</html>
"""


_BASE_TOKEN_BROWSER_STORE_COMPONENT = None


def get_base_token_browser_store_component():
    """建立瀏覽器端 token 保存元件；iOS 切背景後可由 localStorage 找回原配置。"""
    global _BASE_TOKEN_BROWSER_STORE_COMPONENT
    if _BASE_TOKEN_BROWSER_STORE_COMPONENT is not None:
        return _BASE_TOKEN_BROWSER_STORE_COMPONENT

    component_dir = Path(tempfile.gettempdir()) / "ubike_base_token_store_component_v1"
    component_dir.mkdir(parents=True, exist_ok=True)
    index_path = component_dir / "index.html"
    try:
        if not index_path.exists() or index_path.read_text(encoding="utf-8") != BASE_TOKEN_BROWSER_STORE_HTML:
            index_path.write_text(BASE_TOKEN_BROWSER_STORE_HTML, encoding="utf-8")
    except OSError:
        return None

    _BASE_TOKEN_BROWSER_STORE_COMPONENT = components.declare_component(
        "ubike_base_token_store_v1",
        path=str(component_dir),
    )
    return _BASE_TOKEN_BROWSER_STORE_COMPONENT


def valid_base_token(value) -> str | None:
    token = str(value or "").strip().lower()
    if len(token) != 32 or any(char not in "0123456789abcdef" for char in token):
        return None
    return token


def recover_base_token_from_browser(
    url_token: str | None,
    *,
    clear_stored: bool = False,
) -> str | None:
    """網址 token 遺失時，從同一瀏覽器的 localStorage 自動找回。"""
    normalized_url_token = valid_base_token(url_token)
    try:
        component = get_base_token_browser_store_component()
        if component is None:
            return normalized_url_token
        payload = component(
            current_token=normalized_url_token or "",
            storage_key="ubike_dispatch_active_base_token_v1",
            clear_stored=clear_stored,
            default=None,
            key="base_token_browser_store",
        )
    except Exception:
        return normalized_url_token

    if clear_stored:
        return normalized_url_token
    if normalized_url_token:
        return normalized_url_token
    if isinstance(payload, dict):
        return valid_base_token(payload.get("token"))
    return None


def clear_browser_base_token() -> None:
    """要求瀏覽器端移除已保存的配置 token。"""
    try:
        component = get_base_token_browser_store_component()
        if component is not None:
            component(
                current_token="",
                storage_key="ubike_dispatch_active_base_token_v1",
                clear_stored=True,
                default=None,
                key="base_token_browser_store_clear",
            )
    except Exception:
        pass


def runtime_state_cache_path(token: str) -> Path:
    """保存智慧調度執行狀態，避免手機切背景後 WebSocket 重建造成整頁重置。"""
    return BASE_CACHE_DIR / f"{token}.runtime.json"


SMART_DISPATCH_PERSISTED_SUFFIXES = (
    "::active_trip",
    "::cooldowns",
    "::decision_round",
    "::history",
    "::manual_next_station",
    "::loop_zone_order",
    "::loop_active_phase",
    "::max_capacity",
    "::truck_bike",
    "::truck_ebike",
    "::location",
)


def is_runtime_state_key_persistable(key: object, token: str) -> bool:
    if not isinstance(key, str):
        return False
    token_marker = f"::{token}::"
    if token_marker not in key and not key.endswith(f"::{token}"):
        return False
    if key.startswith("smart_dispatch::"):
        return key.endswith(SMART_DISPATCH_PERSISTED_SUFFIXES)
    if key.startswith("long_distance_settings::"):
        return not key.endswith(("::refresh_location", "::reset_loop_route"))
    if key.startswith((
        "configuration_type::", "shift::", "page_mode::", "analysis_zone::", "analysis_region::",
        "station_alert_state::", "low_battery_threshold::", "low_battery_priority_threshold::",
    )):
        return True
    return False


def json_safe_runtime_value(value):
    """只保存可安全 JSON 化的基本型別。"""
    if isinstance(value, np.generic):
        return value.item()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, list):
        return [json_safe_runtime_value(item) for item in value]
    if isinstance(value, tuple):
        return [json_safe_runtime_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): json_safe_runtime_value(item) for key, item in value.items()}
    raise TypeError(f"unsupported runtime state type: {type(value)!r}")


def restore_runtime_state(token: str) -> None:
    """新 Streamlit session 建立時，把上一個手機 session 的調度狀態套回。"""
    loaded_key = f"runtime_state_loaded::{token}"
    if st.session_state.get(loaded_key):
        return
    st.session_state[loaded_key] = True

    state_path = runtime_state_cache_path(token)
    if not state_path.exists():
        return
    try:
        payload = json.loads(state_path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError, json.JSONDecodeError):
        return
    if not isinstance(payload, dict):
        return

    for key, value in payload.items():
        if is_runtime_state_key_persistable(key, token) and key not in st.session_state:
            st.session_state[key] = value


def persist_runtime_state(token: str) -> None:
    """把重要狀態寫入暫存；內容沒變時不再讀檔或重寫磁碟。"""
    payload = {}
    for key, value in st.session_state.items():
        if not is_runtime_state_key_persistable(key, token):
            continue
        try:
            payload[key] = json_safe_runtime_value(value)
        except TypeError:
            continue

    BASE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    state_path = runtime_state_cache_path(token)
    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    digest = hashlib.sha1(encoded.encode("utf-8")).hexdigest()
    digest_key = f"runtime_state_digest::{token}"
    if st.session_state.get(digest_key) == digest and state_path.exists():
        return

    temporary_path = state_path.with_suffix(".runtime.tmp")
    try:
        temporary_path.write_text(encoded, encoding="utf-8")
        temporary_path.replace(state_path)
        st.session_state[digest_key] = digest
    except OSError:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass


def get_base_token() -> str | None:
    """從網址參數取得本次瀏覽器使用的基底識別碼。"""
    try:
        token = st.query_params.get("base")
    except Exception:
        params = st.experimental_get_query_params()
        values = params.get("base", [])
        token = values[0] if values else None

    if isinstance(token, list):
        token = token[0] if token else None

    token = str(token or "").strip()
    if len(token) != 32 or any(char not in "0123456789abcdef" for char in token.lower()):
        return None
    return token.lower()


def pop_floating_dispatch_station_request() -> str:
    """取得懸浮搜尋送出的智慧調度場站，並立即移除一次性網址參數。"""
    state_key = "floating_dispatch_station_request"
    requested_station = ""
    try:
        requested_station = st.query_params.get("dispatch_station") or ""
    except Exception:
        params = st.experimental_get_query_params()
        values = params.get("dispatch_station", [])
        requested_station = values[0] if values else ""

    if isinstance(requested_station, list):
        requested_station = requested_station[0] if requested_station else ""
    requested_station = str(requested_station or "").strip()[:200]
    if requested_station:
        # 先放進 session；若舊版 Streamlit 在清除網址參數時觸發 rerun，下一輪仍能接續處理。
        st.session_state[state_key] = requested_station
        try:
            if "dispatch_station" in st.query_params:
                del st.query_params["dispatch_station"]
        except Exception:
            params = st.experimental_get_query_params()
            params.pop("dispatch_station", None)
            st.experimental_set_query_params(**params)

    return str(st.session_state.pop(state_key, "") or "").strip()


def set_base_token(token: str) -> None:
    """把基底識別碼寫入網址，讓重新整理後仍能找到同一份基底。"""
    try:
        st.query_params["base"] = token
    except Exception:
        params = st.experimental_get_query_params()
        params["base"] = token
        st.experimental_set_query_params(**params)


def clear_base_token() -> None:
    """移除已失效的基底識別碼。"""
    try:
        if "base" in st.query_params:
            del st.query_params["base"]
    except Exception:
        params = st.experimental_get_query_params()
        params.pop("base", None)
        st.experimental_set_query_params(**params)


def base_cache_paths(token: str) -> tuple[Path, Path]:
    return BASE_CACHE_DIR / f"{token}.xlsx", BASE_CACHE_DIR / f"{token}.json"


def current_status_cache_path(token: str) -> Path:
    """取得指定基底所對應的現況暫存檔路徑。"""
    return BASE_CACHE_DIR / f"{token}.status.json"


def delete_cached_base(token: str) -> None:
    """刪除指定的暫存基底，以及與它綁定的現況與調度狀態。"""
    excel_path, metadata_path = base_cache_paths(token)
    status_path = current_status_cache_path(token)
    runtime_path = runtime_state_cache_path(token)
    for path in (excel_path, metadata_path, status_path, runtime_path):
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass
    for memory_key in (
        f"disk_base_cache::{token}",
        f"disk_status_cache::{token}",
        f"runtime_state_digest::{token}",
    ):
        st.session_state.pop(memory_key, None)


def load_cached_base(token: str | None) -> tuple[dict | None, bool]:
    """讀取已保存的配置基底；同一 Streamlit session 只從磁碟載入一次。"""
    if not token:
        return None, False

    memory_key = f"disk_base_cache::{token}"
    cached = st.session_state.get(memory_key)
    if isinstance(cached, dict) and cached.get("token") == token and isinstance(cached.get("bytes"), bytes):
        return cached, False

    excel_path, metadata_path = base_cache_paths(token)
    if not excel_path.exists() or not metadata_path.exists():
        st.session_state.pop(memory_key, None)
        return None, False

    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        file_bytes = excel_path.read_bytes()
        loaded = {
            "token": token,
            "name": str(metadata.get("name") or "配置基底.xlsx"),
            "bytes": file_bytes,
            "sha256": str(
                metadata.get("sha256") or hashlib.sha256(file_bytes).hexdigest()
            ),
            "uploaded_at": float(metadata.get("uploaded_at") or 0.0),
            "expires_at": None,
        }
        st.session_state[memory_key] = loaded
        return loaded, False
    except (OSError, ValueError, TypeError, json.JSONDecodeError):
        st.session_state.pop(memory_key, None)
        return None, False


def load_cached_status(token: str, expires_at: float | None = None) -> dict:
    """讀取現況資料；檔案未變時直接使用 session 記憶體快取。"""
    del expires_at
    status_path = current_status_cache_path(token)
    memory_key = f"disk_status_cache::{token}"
    try:
        current_mtime = status_path.stat().st_mtime_ns
    except OSError:
        empty_payload = {"contexts": {}, "metadata": {}}
        st.session_state[memory_key] = {"mtime_ns": None, "payload": empty_payload, "digest": ""}
        return empty_payload

    cached = st.session_state.get(memory_key)
    if isinstance(cached, dict) and cached.get("mtime_ns") == current_mtime:
        payload = cached.get("payload")
        if isinstance(payload, dict):
            return payload

    try:
        encoded = status_path.read_text(encoding="utf-8")
        payload = json.loads(encoded)
        contexts = payload.get("contexts", {})
        metadata = payload.get("metadata", {})
        if not isinstance(contexts, dict):
            contexts = {}
        if not isinstance(metadata, dict):
            metadata = {}
        normalized = {"contexts": contexts, "metadata": metadata}
        st.session_state[memory_key] = {
            "mtime_ns": current_mtime,
            "payload": normalized,
            "digest": hashlib.sha1(encoded.encode("utf-8")).hexdigest(),
        }
        return normalized
    except (OSError, TypeError, ValueError, json.JSONDecodeError):
        empty_payload = {"contexts": {}, "metadata": {}}
        st.session_state[memory_key] = {"mtime_ns": current_mtime, "payload": empty_payload, "digest": ""}
        return empty_payload


def save_cached_status(
    token: str,
    expires_at: float | None,
    payload: dict,
) -> None:
    """保存現況資料；以內容摘要避免每次先把整份 JSON 從磁碟讀回比較。"""
    del expires_at
    BASE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    status_path = current_status_cache_path(token)
    memory_key = f"disk_status_cache::{token}"
    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    digest = hashlib.sha1(encoded.encode("utf-8")).hexdigest()
    cached = st.session_state.get(memory_key)
    try:
        current_mtime = status_path.stat().st_mtime_ns
    except OSError:
        current_mtime = None
    if (
        isinstance(cached, dict)
        and cached.get("digest") == digest
        and cached.get("mtime_ns") == current_mtime
        and current_mtime is not None
    ):
        cached["payload"] = payload
        return

    temporary_path = status_path.with_suffix(".status.tmp")
    try:
        temporary_path.write_text(encoded, encoding="utf-8")
        temporary_path.replace(status_path)
        try:
            mtime_ns = status_path.stat().st_mtime_ns
        except OSError:
            mtime_ns = None
        st.session_state[memory_key] = {
            "mtime_ns": mtime_ns,
            "payload": payload,
            "digest": digest,
        }
    except OSError:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass


def status_context_key(sheet_name: str, route: str, shift: str) -> str:
    """為每個工作表、分區與班別建立獨立的現況保存區。"""
    return "｜".join((str(sheet_name), str(route), str(shift)))


def _apply_status_lookup(target_df: pd.DataFrame, source_df: pd.DataFrame) -> pd.DataFrame:
    """以行政區＋場站名稱套用現況與完整即時場站資料，並相容舊版暫存。"""
    result_df = coerce_nullable_station_status(target_df.copy())
    identity_columns = ["行政區", "場站名稱"]
    if source_df.empty or any(column not in source_df.columns for column in identity_columns):
        return result_df

    available_columns = [
        column for column in PERSISTED_STATUS_COLUMNS
        if column in source_df.columns
    ]
    if not available_columns:
        return result_df

    lookup_df = source_df[identity_columns + available_columns].copy()
    lookup_df["行政區"] = lookup_df["行政區"].astype(str)
    lookup_df["場站名稱"] = lookup_df["場站名稱"].astype(str)
    lookup_df = coerce_nullable_station_status(lookup_df)
    lookup_df = lookup_df.drop_duplicates(identity_columns, keep="last")
    lookup_df = lookup_df.set_index(identity_columns)

    target_keys = pd.MultiIndex.from_frame(
        result_df[identity_columns].astype(str),
        names=identity_columns,
    )
    matched_mask = target_keys.isin(lookup_df.index)
    if not bool(np.any(matched_mask)):
        return result_df

    aligned = lookup_df.reindex(target_keys)
    for column in available_columns:
        values = pd.Series(aligned[column].array, index=result_df.index, dtype="Int64")
        result_df.loc[matched_mask, column] = values.loc[matched_mask].array
    return coerce_nullable_station_status(result_df)


def restore_current_status(base_df: pd.DataFrame, saved_records) -> pd.DataFrame:
    """把先前保存的現況套回剛解析完成的基底資料。"""
    if not isinstance(saved_records, list):
        return base_df.copy()
    valid_records = [record for record in saved_records if isinstance(record, dict)]
    if not valid_records:
        return coerce_nullable_station_status(base_df.copy())
    return _apply_status_lookup(base_df, pd.DataFrame.from_records(valid_records))


def dataframe_to_status_records(status_df: pd.DataFrame) -> list[dict]:
    """將現況、空位、總柱與服務狀態轉成 JSON 紀錄。"""
    normalized_df = status_df.copy()
    for column in PERSISTED_STATUS_COLUMNS:
        if column not in normalized_df.columns:
            normalized_df[column] = pd.NA
    columns = ["行政區", "場站名稱", *PERSISTED_STATUS_COLUMNS]
    records_df = coerce_nullable_station_status(normalized_df[columns].copy())
    records_df["行政區"] = records_df["行政區"].astype(str)
    records_df["場站名稱"] = records_df["場站名稱"].astype(str)
    object_df = records_df.astype(object).where(records_df.notna(), None)
    return object_df.to_dict(orient="records")


def save_cached_base(file_name: str, file_bytes: bytes) -> dict:
    """將新上傳的 Excel 保存於伺服器，不設定自動失效時間。"""
    BASE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    token = uuid.uuid4().hex
    uploaded_at = time.time()
    digest = hashlib.sha256(file_bytes).hexdigest()
    excel_path, metadata_path = base_cache_paths(token)

    excel_path.write_bytes(file_bytes)
    metadata_path.write_text(
        json.dumps(
            {
                "name": file_name,
                "uploaded_at": uploaded_at,
                "sha256": digest,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    set_base_token(token)

    saved_base = {
        "token": token,
        "name": file_name,
        "bytes": file_bytes,
        "sha256": digest,
        "uploaded_at": uploaded_at,
        "expires_at": None,
    }
    st.session_state[f"disk_base_cache::{token}"] = saved_base
    return saved_base


def format_remaining_time(expires_at: float | None) -> str:
    """相容舊介面；目前保存期限為無時間限制。"""
    return "無時間限制"


def rerun_app() -> None:
    """相容新舊版 Streamlit 的重新執行方法。"""
    try:
        st.rerun()
    except AttributeError:
        st.experimental_rerun()


def build_blank_status_cache(source: bytes | str, options: list[tuple[str, str]]) -> dict:
    """把所有工作表、分區與班別的現況清成空白。"""
    blank_contexts: dict[str, list[dict]] = {}

    for sheet_name, route in options:
        for shift in SHIFT_COLUMNS.keys():
            try:
                route_df = cached_parse_route(source, sheet_name, route, shift)
            except Exception:
                continue

            if route_df.empty:
                continue

            context_key = status_context_key(sheet_name, route, shift)
            blank_contexts[context_key] = dataframe_to_status_records(
                blank_current_status(route_df)
            )

    return {"contexts": blank_contexts, "metadata": {}}



def on_demand_toggle(label: str, *, key: str, value: bool = False, help_text: str | None = None) -> bool:
    """只在使用者需要時建立重量級介面，避免折疊內容仍拖慢每次 Streamlit 重跑。"""
    toggle = getattr(st, "toggle", st.checkbox)
    return bool(toggle(label, value=value, key=key, help=help_text))


def fragment_if_available(function):
    """新版 Streamlit 只重跑互動區塊；舊版則安全退回原本整頁模式。"""
    fragment = getattr(st, "fragment", None)
    return fragment(function) if callable(fragment) else function


def render_app_hero() -> None:
    """以純 SVG 建立輕量首頁插畫；同步顯示目前程式版本。"""
    st.markdown(
        f"""
        <section class="dispatch-hero">
          <div class="dispatch-hero-copy">
            <div class="dispatch-kicker">TAITUNG · SMART DISPATCH</div>
            <div id="jarvis-secret-trigger" class="dispatch-version-badge" title="">測試版</div>
            <h1>臺東 YouBike 智慧調度</h1>
            <p>配置、即時車數、分析與依實際道路路網計算的 AI 路線，集中在同一套工作流程。</p>
          </div>
          <div class="dispatch-hero-art" aria-hidden="true">
            <svg viewBox="0 0 330 150" role="img">
              <circle cx="280" cy="34" r="22" class="hero-sun"/>
              <path d="M8 117 C58 74 95 119 139 83 C174 55 204 104 242 75 C270 54 302 72 330 54 V150 H8 Z" class="hero-hill hero-hill-back"/>
              <path d="M0 127 C47 99 82 130 123 104 C169 75 205 127 251 96 C282 75 309 88 330 80 V150 H0 Z" class="hero-hill hero-hill-front"/>
              <g class="hero-truck">
                <rect x="82" y="73" width="97" height="42" rx="9"/>
                <path d="M179 86 H213 L229 104 V115 H179 Z"/>
                <rect x="190" y="91" width="20" height="12" rx="2" class="hero-window"/>
                <circle cx="108" cy="118" r="13" class="hero-wheel"/>
                <circle cx="203" cy="118" r="13" class="hero-wheel"/>
                <path d="M99 71 C112 55 143 53 158 71" class="hero-route"/>
              </g>
              <g class="hero-bike" transform="translate(236 88)">
                <circle cx="18" cy="29" r="15"/>
                <circle cx="63" cy="29" r="15"/>
                <path d="M18 29 L31 6 L44 29 L18 29 M31 6 H50 L63 29 M30 6 L25 -4 M22 -4 H35 M43 29 L55 13"/>
              </g>
            </svg>
          </div>
        </section>
        """,
        unsafe_allow_html=True,
    )

def render_context_strip(
    *,
    route: str,
    shift: str,
    station_count: int,
    page_mode: str,
    live_meta: dict | None = None,
) -> None:
    """顯示目前工作情境，讓使用者不用回頭確認選項。"""
    live_meta = live_meta if isinstance(live_meta, dict) else {}
    fetched_at = html.escape(str(live_meta.get("fetched_at") or "尚未同步"))
    mode_label = "智慧調度" if page_mode == "智慧調度" else "一般分析"
    st.markdown(
        f"""
        <div class="dispatch-context-strip">
          <span><b>範圍</b>{html.escape(route)}</span>
          <span><b>班別</b>{html.escape(shift)}</span>
          <span><b>場站</b>{safe_nonnegative_int(station_count)} 站</span>
          <span><b>模式</b>{html.escape(mode_label)}</span>
          <span class="dispatch-live-time"><b>即時資料</b>{fetched_at}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def low_battery_threshold_state_keys(active_base_token: str) -> tuple[str, str]:
    """回傳一般分析與智慧調度共用的兩道電量門檻狀態鍵。"""
    token = str(active_base_token or "default").strip() or "default"
    return (
        f"low_battery_threshold::{token}",
        f"low_battery_priority_threshold::{token}",
    )


def get_low_battery_thresholds(active_base_token: str) -> tuple[int, int]:
    """安全取得兩道門檻；第二門檻永遠不會高於第一門檻。"""
    threshold_key, priority_key = low_battery_threshold_state_keys(active_base_token)
    threshold = min(
        100,
        max(0, safe_nonnegative_int(st.session_state.get(threshold_key, DEFAULT_BATTERY_THRESHOLD))),
    )
    priority_threshold = min(
        threshold,
        max(
            0,
            safe_nonnegative_int(
                st.session_state.get(priority_key, DEFAULT_BATTERY_PRIORITY_THRESHOLD)
            ),
        ),
    )
    return threshold, priority_threshold


def render_low_battery_threshold_controls(
    active_base_token: str,
    *,
    page_mode: str,
) -> tuple[int, int]:
    """在工作頁上方顯示共用雙門檻；只控制柱號顯示，不改智慧推薦分數。"""
    threshold_key, priority_key = low_battery_threshold_state_keys(active_base_token)
    threshold, priority_threshold = get_low_battery_thresholds(active_base_token)
    # widget 建立前先校正既有值，避免第一門檻調低後第二門檻超過 max_value。
    st.session_state[threshold_key] = threshold
    st.session_state[priority_key] = priority_threshold

    with st.container(border=True):
        st.markdown(
            f"**⚡ {page_mode}｜低電量柱號門檻**  "
            "\n第二門檻只做紅色警示，不會改變智慧調度推薦分數或順序。"
        )
        threshold_col, priority_col = st.columns(2)
        with threshold_col:
            threshold = int(
                st.number_input(
                    "第一門檻（低於或等於）",
                    min_value=0,
                    max_value=100,
                    step=1,
                    key=threshold_key,
                    help="預設 89%；符合者會列出柱號。",
                )
            )

        # 第一欄在本輪已回傳新值，先再次校正第二門檻再建立第二個 widget。
        if safe_nonnegative_int(st.session_state.get(priority_key)) > threshold:
            st.session_state[priority_key] = threshold
        with priority_col:
            priority_threshold = int(
                st.number_input(
                    "第二門檻（紅色警示）",
                    min_value=0,
                    max_value=threshold,
                    step=1,
                    key=priority_key,
                    help="預設 40%；符合者的柱號會以紅色發光顯示。",
                )
            )

    return threshold, min(threshold, priority_threshold)


@st.cache_data(show_spinner=False, max_entries=128)
def _build_analysis_result_table_html(rows: tuple[tuple, ...]) -> str:
    """快取分析表 HTML；手機版以分行呈現目前、標準與缺多。"""
    html_rows: list[str] = []
    for (
        row_index, station_name_raw, route_zone_raw,
        bike_current_raw, bike_standard_raw, bike_status_raw,
        ebike_current_raw, ebike_standard_raw, ebike_status_raw,
        total_bikes_raw, empty_spaces_raw, distance_raw, drive_minutes_raw,
    ) in rows:
        station_name = html.escape(str(station_name_raw))
        station_name_attr = html.escape(str(station_name_raw), quote=True)
        route_zone = html.escape(str(route_zone_raw or ""))
        bike_status = add_dispatch_indicator(bike_status_raw)
        ebike_status = add_dispatch_indicator(ebike_status_raw)
        bike_class = dispatch_status_class(bike_status)
        ebike_class = dispatch_status_class(ebike_status)

        station_meta = [
            f"總車 {html.escape(str(total_bikes_raw))}",
            f"空位 {html.escape(str(empty_spaces_raw))}",
        ]
        if str(distance_raw) != "—":
            station_meta.append(f"{html.escape(str(distance_raw))} km")
        if str(drive_minutes_raw) != "—":
            station_meta.append(f"約 {html.escape(str(drive_minutes_raw))} 分")
        zone_html = f'<small class="analysis-zone-label">{route_zone}</small>' if route_zone else ""

        html_rows.append(
            f'<tr id="analysis-result-anchor-{row_index}" class="analysis-result-row" '
            f'data-ubike-station-name="{station_name_attr}">'
            f'<td class="analysis-station-cell"><strong>{station_name}</strong>{zone_html}'
            f'<small>{"｜".join(station_meta)}</small></td>'
            f'<td class="analysis-status-cell {bike_class}"><strong>{html.escape(bike_status)}</strong>'
            f'<small>目前 {html.escape(str(bike_current_raw))}／標準 {html.escape(str(bike_standard_raw))}</small></td>'
            f'<td class="analysis-status-cell {ebike_class}"><strong>{html.escape(ebike_status)}</strong>'
            f'<small>目前 {html.escape(str(ebike_current_raw))}／標準 {html.escape(str(ebike_standard_raw))}</small></td>'
            "</tr>"
        )

    return (
        '<div class="analysis-result-table-wrap">'
        '<table class="analysis-result-table">'
        '<colgroup><col class="analysis-col-station" />'
        '<col class="analysis-col-status" /><col class="analysis-col-status" /></colgroup>'
        '<thead><tr><th>場站／總量</th><th>2.0 現況／標準</th><th>2.0E 現況／標準</th></tr></thead>'
        f'<tbody>{"".join(html_rows)}</tbody></table></div>'
    )

def dispatch_status_class(value) -> str:
    """回傳分析結果儲存格的視覺狀態類別。"""
    text = str(value)
    if "多" in text:
        return "analysis-status-extra"
    if "缺" in text or "少" in text:
        return "analysis-status-short"
    if STATUS_UNAVAILABLE_TEXT in text:
        return "analysis-status-error"
    return "analysis-status-ok"


def render_analysis_result_table(region_df: pd.DataFrame) -> None:
    """完整顯示分析結果；數值欄一次向量化，避免逐列建立暫存 Series。"""
    display_df = region_df.copy()
    for column in (
        "路線區域", "2.0 現況", "2.0 標準", "2.0E 現況", "2.0E 標準",
        "總車數", "空位數", "距離目前位置 (km)", "預估行車時間 (分)",
    ):
        if column not in display_df.columns:
            display_df[column] = pd.NA

    distance_values = pd.to_numeric(
        display_df["距離目前位置 (km)"],
        errors="coerce",
    ).to_numpy()
    minute_values = pd.to_numeric(
        display_df["預估行車時間 (分)"],
        errors="coerce",
    ).to_numpy()
    table_columns = [
        "場站名稱", "路線區域",
        "2.0 現況", "2.0 標準", "2.0 缺／多幾台",
        "2.0E 現況", "2.0E 標準", "2.0E 缺／多幾台",
        "總車數", "空位數",
    ]
    rows = []
    for position, values in enumerate(
        display_df[table_columns].itertuples(index=False, name=None)
    ):
        (
            station_name, route_zone_raw,
            bike_current, bike_standard, bike_status,
            ebike_current, ebike_standard, ebike_status,
            total_bikes, empty_spaces,
        ) = values
        distance = distance_values[position]
        minutes = minute_values[position]
        route_zone = "" if pd.isna(route_zone_raw) else str(route_zone_raw)
        rows.append((
            int(display_df.index[position]), str(station_name), route_zone,
            optional_count_text(bike_current), safe_nonnegative_int(bike_standard),
            str(bike_status),
            optional_count_text(ebike_current), safe_nonnegative_int(ebike_standard),
            str(ebike_status),
            optional_count_text(total_bikes), optional_count_text(empty_spaces),
            "—" if pd.isna(distance) else f"{float(distance):.1f}",
            "—" if pd.isna(minutes) else f"{float(minutes):.0f}",
        ))
    st.markdown(_build_analysis_result_table_html(tuple(rows)), unsafe_allow_html=True)

@st.cache_data(show_spinner=False, max_entries=64)
def _build_floating_station_search_html(
    result_df: pd.DataFrame,
    mobile_mode: bool,
    *,
    page_mode: str = "一般分析",
) -> str:
    """快取懸浮搜尋元件 HTML；資料沒變時不重組整段樣式與腳本。"""
    dispatch_mode = page_mode == "智慧調度"
    stations = []
    station_columns = ["場站名稱", "行政區", "2.0 缺／多幾台", "2.0E 缺／多幾台"]
    for row_index, (station_name_raw, region_raw, bike_raw, ebike_raw) in enumerate(
        result_df[station_columns].itertuples(index=False, name=None)
    ):
        station_name = str(station_name_raw).strip()
        if not station_name:
            continue
        stations.append(
            {
                "name": station_name,
                "region": str(region_raw).strip(),
                "bike": str(bike_raw).strip(),
                "ebike": str(ebike_raw).strip(),
                "anchor": "" if dispatch_mode else f"analysis-result-anchor-{row_index}",
            }
        )

    station_payload = json.dumps(stations, ensure_ascii=False).replace("</", "<\\/")
    display_mode = "mobile" if mobile_mode else "desktop"
    page_mode_payload = json.dumps(page_mode, ensure_ascii=False)
    analysis_button_html = (
        '<button class="uft-button uft-analysis" type="button" '
        'title="跳到調度分析結果">分析</button>'
        if not dispatch_mode else ""
    )
    search_title = "🔎 搜尋智慧調度場站" if dispatch_mode else "🔎 搜尋分析結果"
    search_hint = (
        "搜尋目前智慧調度配置中的全部場站｜點選後設為目前選擇｜快捷鍵 Ctrl + K"
        if dispatch_mode
        else "只搜尋目前的調度分析結果｜快捷鍵 Ctrl + K"
    )
    floating_fingerprint = hashlib.sha1(
        f"{display_mode}|{page_mode}|{station_payload}".encode("utf-8")
    ).hexdigest()

    return f"""
        <script>
        (() => {{
            const stations = {station_payload};
            const displayMode = {json.dumps(display_mode)};
            const pageMode = {page_mode_payload};
            const dispatchMode = pageMode === "智慧調度";
            const fingerprint = {json.dumps(floating_fingerprint)};
            const doc = window.parent.document;
            const win = window.parent;

            // 分析資料沒變時沿用已建立的懸浮工具，避免每次 Streamlit rerun 都拆掉重建 DOM。
            if (win.__ubikeFloatingFingerprint === fingerprint && doc.getElementById("ubike-float-tools")) {{
                if (typeof win.__ubikeFloatingResizeHandler === "function") {{
                    win.__ubikeFloatingResizeHandler();
                }}
                return;
            }}
            win.__ubikeFloatingFingerprint = fingerprint;

            const oldRoot = doc.getElementById("ubike-float-tools");
            if (oldRoot) oldRoot.remove();
            const oldStyle = doc.getElementById("ubike-float-tools-style");
            if (oldStyle) oldStyle.remove();

            const style = doc.createElement("style");
            style.id = "ubike-float-tools-style";
            style.textContent = `
                #ubike-float-tools {{
                    position: fixed;
                    right: 18px;
                    bottom: 22px;
                    z-index: 2147483000;
                    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI",
                                 "Microsoft JhengHei", sans-serif;
                }}
                #ubike-float-tools * {{ box-sizing: border-box; }}
                #ubike-float-tools .uft-actions {{
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    gap: 8px;
                }}
                #ubike-float-tools .uft-button {{
                    width: 56px;
                    height: 56px;
                    flex: 0 0 56px;
                    border: 0;
                    border-radius: 50%;
                    color: #151515;
                    font-weight: 850;
                    cursor: pointer;
                    box-shadow: 0 8px 28px rgba(0, 0, 0, 0.28);
                    transition: transform 0.16s ease, box-shadow 0.16s ease;
                }}
                #ubike-float-tools .uft-button:hover {{
                    transform: translateY(-2px);
                    box-shadow: 0 11px 32px rgba(0, 0, 0, 0.34);
                }}
                #ubike-float-tools .uft-top {{ background: #f1f3f5; font-size: 12px; }}
                #ubike-float-tools .uft-analysis {{ background: #8bd3ff; font-size: 13px; }}
                #ubike-float-tools .uft-search {{ background: #ffbf00; font-size: 24px; }}
                #ubike-float-tools .uft-refresh {{ background: #6ee7b7; font-size: 13px; }}
                #ubike-float-tools .uft-refresh.is-syncing {{
                    cursor: wait;
                    animation: ubikeRefreshPulse 0.9s ease-in-out infinite;
                }}
                @keyframes ubikeRefreshPulse {{
                    0%, 100% {{ transform: scale(1); }}
                    50% {{ transform: scale(0.91); }}
                }}
                #ubike-float-tools .uft-panel {{
                    position: absolute;
                    right: 66px;
                    bottom: 0;
                    width: min(340px, calc(100vw - 96px));
                    padding: 13px;
                    border: 1px solid rgba(0, 0, 0, 0.13);
                    border-radius: 16px;
                    background: rgba(255, 255, 255, 0.98);
                    color: #171717;
                    box-shadow: 0 16px 46px rgba(0, 0, 0, 0.28);
                    backdrop-filter: blur(8px);
                }}
                #ubike-float-tools .uft-panel[hidden] {{ display: none !important; }}
                #ubike-float-tools .uft-title-row {{
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    margin-bottom: 9px;
                }}
                #ubike-float-tools .uft-title {{ font-size: 15px; font-weight: 800; }}
                #ubike-float-tools .uft-close {{
                    border: 0;
                    background: transparent;
                    color: #555;
                    font-size: 20px;
                    cursor: pointer;
                }}
                #ubike-float-tools .uft-input {{
                    width: 100%;
                    min-height: 44px;
                    padding: 10px 12px;
                    border: 2px solid #e0e0e0;
                    border-radius: 11px;
                    outline: none;
                    font-size: 16px;
                    color: #111;
                    background: #fff;
                }}
                #ubike-float-tools .uft-input:focus {{
                    border-color: #ffbf00;
                    box-shadow: 0 0 0 3px rgba(255, 191, 0, 0.18);
                }}
                #ubike-float-tools .uft-hint {{
                    margin: 7px 2px 8px;
                    color: #666;
                    font-size: 12px;
                }}
                #ubike-float-tools .uft-results {{
                    display: flex;
                    flex-direction: column;
                    gap: 5px;
                    max-height: min(310px, 48vh);
                    overflow-y: auto;
                    overflow-x: hidden;
                }}
                #ubike-float-tools .uft-result {{
                    width: 100%;
                    padding: 9px 10px;
                    border: 1px solid #e4e4e4;
                    border-radius: 10px;
                    background: #fff;
                    text-align: left;
                    cursor: pointer;
                }}
                #ubike-float-tools .uft-result:hover,
                #ubike-float-tools .uft-result:focus {{
                    border-color: #ffbf00;
                    background: #fff9df;
                    outline: none;
                }}
                #ubike-float-tools .uft-result-name {{
                    display: block;
                    color: #111;
                    font-size: 14px;
                    font-weight: 750;
                }}
                #ubike-float-tools .uft-result-region,
                #ubike-float-tools .uft-result-status {{
                    display: block;
                    margin-top: 2px;
                    color: #777;
                    font-size: 12px;
                    line-height: 1.35;
                }}
                #ubike-float-tools .uft-result-status {{ color: #444; }}
                #ubike-float-tools .uft-empty {{
                    padding: 12px 4px 5px;
                    color: #777;
                    font-size: 13px;
                    text-align: center;
                }}
                #ubike-search-toast {{
                    position: fixed;
                    left: 50%;
                    bottom: var(--ubike-float-toast-bottom, 210px);
                    z-index: 2147483001;
                    transform: translateX(-50%);
                    padding: 10px 15px;
                    border-radius: 999px;
                    background: rgba(20, 20, 20, 0.92);
                    color: #fff;
                    font: 700 14px/1.25 -apple-system, BlinkMacSystemFont, "Segoe UI",
                          "Microsoft JhengHei", sans-serif;
                    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.28);
                    pointer-events: none;
                }}
                .ubike-analysis-row-focus {{
                    position: relative;
                    z-index: 2;
                    outline: 3px solid #ff9f00;
                    outline-offset: -3px;
                    animation: ubikeAnalysisPulse 0.72s ease-in-out 3;
                }}
                @keyframes ubikeAnalysisPulse {{
                    0%, 100% {{ filter: brightness(1); }}
                    50% {{ filter: brightness(1.16); }}
                }}
                /* v27.5：延續原版格局，只更新懸浮工具配色。 */
                #ubike-float-tools .uft-button {{
                    color: #f3ffff;
                    box-shadow: 0 0 18px rgba(85, 246, 255, .18), 0 8px 28px rgba(0, 0, 0, .34);
                    text-shadow: 0 0 8px rgba(255,255,255,.22);
                }}
                #ubike-float-tools .uft-top {{ background: linear-gradient(135deg, #15263b, #263653); }}
                #ubike-float-tools .uft-analysis {{ background: linear-gradient(135deg, #0b7180, #164c8f); }}
                #ubike-float-tools .uft-search {{ background: linear-gradient(135deg, #ff3fcf, #7a27d8); }}
                #ubike-float-tools .uft-refresh {{ background: linear-gradient(135deg, #0d8c70, #17699a); }}
                #ubike-float-tools .uft-panel {{
                    color: #effcff;
                    border-color: rgba(85,246,255,.42);
                    background: linear-gradient(135deg, rgba(7,18,34,.98), rgba(29,8,38,.98));
                    box-shadow: 0 0 26px rgba(85,246,255,.12), 0 16px 46px rgba(0,0,0,.42);
                }}
                #ubike-float-tools .uft-title {{ color:#55f6ff; text-shadow:0 0 10px rgba(85,246,255,.38); }}
                #ubike-float-tools .uft-close {{ color:#ff83df; }}
                #ubike-float-tools .uft-input {{
                    color:#effcff;
                    border-color:rgba(85,246,255,.36);
                    background:#071225;
                    box-shadow:inset 0 0 12px rgba(85,246,255,.04);
                }}
                #ubike-float-tools .uft-input:focus {{
                    border-color:#f4ff57;
                    box-shadow:0 0 0 3px rgba(244,255,87,.12), 0 0 16px rgba(85,246,255,.12);
                }}
                #ubike-float-tools .uft-hint,
                #ubike-float-tools .uft-result-region,
                #ubike-float-tools .uft-empty {{ color:#9bb7c4; }}
                #ubike-float-tools .uft-result {{
                    color:#effcff;
                    border-color:rgba(85,246,255,.20);
                    background:rgba(8,22,38,.92);
                }}
                #ubike-float-tools .uft-result:hover,
                #ubike-float-tools .uft-result:focus {{
                    border-color:#f4ff57;
                    background:rgba(244,255,87,.08);
                }}
                #ubike-float-tools .uft-result-name {{ color:#effcff; }}
                #ubike-float-tools .uft-result-status {{ color:#c5dce4; }}
                #ubike-search-toast {{
                    color:#071018;
                    background:#55f6ff;
                    box-shadow:0 0 22px rgba(85,246,255,.35), 0 8px 24px rgba(0,0,0,.32);
                }}
                .ubike-analysis-row-focus {{ outline-color:#f4ff57; }}
                @media (max-width: 700px) {{
                    #ubike-float-tools {{
                        right: 10px;
                        bottom: calc(72px + env(safe-area-inset-bottom, 0px));
                    }}
                    #ubike-float-tools .uft-button {{
                        width: 52px;
                        height: 52px;
                        flex-basis: 52px;
                    }}
                    #ubike-float-tools .uft-panel {{
                        right: 60px;
                        bottom: 0;
                        width: min(305px, calc(100vw - 82px));
                        max-height: 72vh;
                    }}
                }}
            `;
            doc.head.appendChild(style);

            const root = doc.createElement("div");
            root.id = "ubike-float-tools";
            root.innerHTML = `
                <div class="uft-panel" hidden>
                    <div class="uft-title-row">
                        <div class="uft-title">{html.escape(search_title)}</div>
                        <button class="uft-close" type="button" aria-label="關閉搜尋">×</button>
                    </div>
                    <input class="uft-input" type="search"
                           placeholder="輸入場站名稱或行政區" autocomplete="off" />
                    <div class="uft-hint">{html.escape(search_hint)}</div>
                    <div class="uft-results"></div>
                </div>
                <div class="uft-actions">
                    <button class="uft-button uft-top" type="button" title="回到頁面最上方">TOP</button>
                    {analysis_button_html}
                    <button class="uft-button uft-search" type="button" title="搜尋場站（Ctrl + K）">🔎</button>
                    <button class="uft-button uft-refresh" type="button" title="手動更新 YouBike 即時車數">更新</button>
                </div>
            `;
            doc.body.appendChild(root);

            const topButton = root.querySelector(".uft-top");
            const analysisButton = root.querySelector(".uft-analysis");
            const searchButton = root.querySelector(".uft-search");
            const refreshButton = root.querySelector(".uft-refresh");
            const panel = root.querySelector(".uft-panel");
            const closeButton = root.querySelector(".uft-close");
            const input = root.querySelector(".uft-input");
            const results = root.querySelector(".uft-results");

            function isMobileLayout() {{
                return displayMode === "mobile" || win.matchMedia("(max-width: 700px)").matches;
            }}

            function updateFloatingPosition() {{
                // 固定成本：不再掃描整個頁面 DOM。手機保留底部安全距離即可。
                const bottomGap = isMobileLayout() ? 72 : 22;
                root.style.bottom = isMobileLayout()
                    ? `calc(${{bottomGap}}px + env(safe-area-inset-bottom, 0px))`
                    : `${{bottomGap}}px`;
                doc.documentElement.style.setProperty(
                    "--ubike-float-toast-bottom",
                    `${{bottomGap + 252}}px`,
                );
            }}

            function setRefreshButtonState(syncing) {{
                refreshButton.disabled = Boolean(syncing);
                refreshButton.classList.toggle("is-syncing", Boolean(syncing));
                refreshButton.textContent = syncing ? "更新中" : "更新";
                refreshButton.title = syncing
                    ? "正在更新 YouBike 即時車數"
                    : "手動更新 YouBike 即時車數";
            }}

            function requestManualSync() {{
                let postedCount = 0;
                for (const frame of doc.querySelectorAll("iframe")) {{
                    try {{
                        if (!frame.contentWindow) continue;
                        const frameTitle = String(frame.getAttribute("title") || "").toLowerCase();
                        const frameSource = String(frame.getAttribute("src") || "").toLowerCase();
                        let isSyncFrame = frameTitle.includes("youbike_browser_sync")
                            || frameSource.includes("youbike_browser_sync");
                        try {{
                            isSyncFrame = isSyncFrame
                                || Boolean(frame.contentDocument?.getElementById("syncButton"));
                        }} catch (_accessError) {{
                            // 跨來源時改以 title／src 判斷。
                        }}
                        if (!isSyncFrame) continue;
                        frame.contentWindow.postMessage({{ type: "ubike:manual-sync" }}, "*");
                        postedCount += 1;
                    }} catch (_error) {{
                        // 略過無法存取的其他 iframe。
                    }}
                }}
                if (!postedCount) {{
                    showToast("同步元件尚未準備完成，請稍後再按一次");
                    return;
                }}
                setRefreshButtonState(true);
                showToast("正在手動更新 YouBike 即時車數");
                // 防止外部網路錯誤造成按鈕永久鎖住；元件回報時會更早解除。
                win.clearTimeout(win.__ubikeManualSyncFallbackTimer);
                win.__ubikeManualSyncFallbackTimer = win.setTimeout(() => {{
                    setRefreshButtonState(false);
                }}, 45000);
            }}

            function getScrollTargets() {{
                const candidates = [
                    doc.scrollingElement,
                    doc.documentElement,
                    doc.body,
                    doc.querySelector('[data-testid="stAppViewContainer"]'),
                    doc.querySelector('[data-testid="stMain"]'),
                    doc.querySelector('section.main'),
                    doc.querySelector('.stApp'),
                ];
                return Array.from(new Set(candidates.filter(Boolean)));
            }}

            function scrollPageToTop() {{
                setOpen(false);
                const topAnchor = doc.getElementById("ubike-page-top-anchor");
                if (topAnchor) topAnchor.scrollIntoView({{ behavior: "smooth", block: "start" }});

                for (const target of getScrollTargets()) {{
                    try {{
                        target.scrollTo({{ top: 0, left: 0, behavior: "smooth" }});
                    }} catch (_error) {{
                        target.scrollTop = 0;
                        target.scrollLeft = 0;
                    }}
                }}
                try {{
                    win.scrollTo({{ top: 0, left: 0, behavior: "smooth" }});
                }} catch (_error) {{
                    win.scrollTo(0, 0);
                }}
                win.setTimeout(() => {{
                    for (const target of getScrollTargets()) {{
                        target.scrollTop = 0;
                        target.scrollLeft = 0;
                    }}
                    win.scrollTo(0, 0);
                }}, 420);
                showToast("已回到頁面最上方");
            }}

            function showToast(message) {{
                const previous = doc.getElementById("ubike-search-toast");
                if (previous) previous.remove();
                const toast = doc.createElement("div");
                toast.id = "ubike-search-toast";
                toast.textContent = message;
                doc.body.appendChild(toast);
                win.setTimeout(() => toast.remove(), 2100);
            }}

            function setOpen(open) {{
                panel.hidden = !open;
                if (open) {{
                    input.value = "";
                    renderResults("");
                    win.setTimeout(() => input.focus(), 30);
                }}
            }}

            function jumpToStation(station) {{
                setOpen(false);
                if (dispatchMode) {{
                    const targetUrl = new URL(win.location.href);
                    targetUrl.searchParams.set("dispatch_station", station.name);
                    showToast(`正在套用智慧調度場站：${{station.name}}`);
                    win.location.assign(targetUrl.toString());
                    return;
                }}
                const anchor = doc.getElementById(station.anchor);
                if (!anchor) {{
                    showToast("找不到這個分析結果，請重新整理後再試");
                    return;
                }}

                anchor.scrollIntoView({{ behavior: "smooth", block: "center" }});
                anchor.classList.remove("ubike-analysis-row-focus");
                void anchor.offsetWidth;
                anchor.classList.add("ubike-analysis-row-focus");
                win.setTimeout(() => anchor.classList.remove("ubike-analysis-row-focus"), 2400);
                showToast(`已找到分析結果：${{station.name}}`);
            }}

            function createResultButton(station) {{
                const button = doc.createElement("button");
                button.type = "button";
                button.className = "uft-result";

                const name = doc.createElement("span");
                name.className = "uft-result-name";
                name.textContent = station.name;
                button.appendChild(name);

                if (station.region) {{
                    const region = doc.createElement("span");
                    region.className = "uft-result-region";
                    region.textContent = station.region;
                    button.appendChild(region);
                }}

                const status = doc.createElement("span");
                status.className = "uft-result-status";
                status.textContent = `2.0：${{station.bike || "—"}}｜2.0E：${{station.ebike || "—"}}`;
                button.appendChild(status);

                button.addEventListener("click", () => jumpToStation(station));
                return button;
            }}

            function renderResults(query) {{
                const keyword = String(query || "").trim().toLocaleLowerCase("zh-TW");
                const matched = stations
                    .filter((station) => {{
                        const haystack = `${{station.name}} ${{station.region}} ${{station.bike}} ${{station.ebike}}`
                            .toLocaleLowerCase("zh-TW");
                        return !keyword || haystack.includes(keyword);
                    }})
                    .slice(0, 12);

                results.replaceChildren();
                if (!matched.length) {{
                    const empty = doc.createElement("div");
                    empty.className = "uft-empty";
                    empty.textContent = stations.length
                        ? (dispatchMode ? "智慧調度配置中查無符合的場站" : "分析結果中查無符合的場站")
                        : (dispatchMode ? "目前沒有可搜尋的智慧調度場站" : "目前沒有需要調度的分析結果");
                    results.appendChild(empty);
                    return;
                }}
                matched.forEach((station) => results.appendChild(createResultButton(station)));
            }}

            topButton.addEventListener("click", scrollPageToTop);
            if (analysisButton) {{
                analysisButton.addEventListener("click", () => {{
                    setOpen(false);
                    const anchor = doc.getElementById("analysis-results-anchor");
                    if (!anchor) {{
                        showToast("找不到分析結果區");
                        return;
                    }}
                    anchor.scrollIntoView({{ behavior: "smooth", block: "start" }});
                }});
            }}
            searchButton.addEventListener("click", () => setOpen(panel.hidden));
            refreshButton.addEventListener("click", requestManualSync);
            closeButton.addEventListener("click", () => setOpen(false));
            input.addEventListener("input", (event) => renderResults(event.target.value));
            input.addEventListener("keydown", (event) => {{
                if (event.key === "Enter") {{
                    const firstResult = results.querySelector(".uft-result");
                    if (firstResult) firstResult.click();
                }}
                if (event.key === "Escape") setOpen(false);
            }});

            if (win.__ubikeSyncStateHandler) {{
                win.removeEventListener("message", win.__ubikeSyncStateHandler);
            }}
            win.__ubikeSyncStateHandler = (event) => {{
                const data = event.data || {{}};
                if (data.source !== "ubike-browser-sync" || data.type !== "ubike:sync-state") return;
                if (data.state === "busy") {{
                    setRefreshButtonState(true);
                    return;
                }}
                win.clearTimeout(win.__ubikeManualSyncFallbackTimer);
                setRefreshButtonState(false);
                if (data.state === "success") {{
                    const countText = Number(data.station_count) > 0 ? `（${{Number(data.station_count)}} 站）` : "";
                    showToast(`即時數據更新完成${{countText}}`);
                }} else if (data.state === "error") {{
                    showToast(`即時數據更新失敗：${{String(data.message || "請稍後再試")}}`);
                }}
            }};
            win.addEventListener("message", win.__ubikeSyncStateHandler);

            if (win.__ubikeSearchKeyHandler) {{
                win.removeEventListener("keydown", win.__ubikeSearchKeyHandler);
            }}
            win.__ubikeSearchKeyHandler = (event) => {{
                if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === "k") {{
                    event.preventDefault();
                    setOpen(true);
                }}
                if (event.key === "Escape" && !panel.hidden) setOpen(false);
            }};
            win.addEventListener("keydown", win.__ubikeSearchKeyHandler);

            if (win.__ubikeFloatingResizeHandler) {{
                win.removeEventListener("resize", win.__ubikeFloatingResizeHandler);
                if (win.visualViewport) {{
                    win.visualViewport.removeEventListener("resize", win.__ubikeFloatingResizeHandler);
                    win.visualViewport.removeEventListener("scroll", win.__ubikeFloatingResizeHandler);
                }}
            }}
            win.__ubikeFloatingResizeHandler = () => updateFloatingPosition();
            win.addEventListener("resize", win.__ubikeFloatingResizeHandler, {{ passive: true }});
            if (win.visualViewport) {{
                win.visualViewport.addEventListener("resize", win.__ubikeFloatingResizeHandler, {{ passive: true }});
                win.visualViewport.addEventListener("scroll", win.__ubikeFloatingResizeHandler, {{ passive: true }});
            }}

            // 舊版若已建立 MutationObserver，先關閉；之後只在視窗尺寸改變時更新。
            if (win.__ubikeFloatingObserver) {{
                win.__ubikeFloatingObserver.disconnect();
                win.__ubikeFloatingObserver = null;
            }}

            updateFloatingPosition();
            win.setTimeout(updateFloatingPosition, 350);
            renderResults("");
        }})();
        </script>
        """


def render_floating_station_search(
    result_df: pd.DataFrame,
    mobile_mode: bool,
    *,
    page_mode: str = "一般分析",
) -> None:
    """建立共用場站搜尋與頁面工具；分析鍵只在一般分析頁顯示。"""
    station_columns = ["場站名稱", "行政區", "2.0 缺／多幾台", "2.0E 缺／多幾台"]
    if result_df.empty:
        search_df = pd.DataFrame(columns=station_columns)
    else:
        search_df = result_df.reindex(columns=station_columns).copy()
    component_html = _build_floating_station_search_html(
        search_df,
        mobile_mode,
        page_mode=page_mode,
    )
    components.html(component_html, height=0, scrolling=False)

st.set_page_config(
    page_title=f"臺東 YouBike 智慧調度｜{APP_VERSION_NAME}",
    page_icon="🚚",
    layout="wide",
)

st.markdown(
    """
    <style>
    html, body, .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"] {
        max-width: 100%;
        overflow-x: hidden !important;
    }
    .block-container {
        width: 100%;
        max-width: 100%;
        padding-top: 1.3rem;
        padding-bottom: 3rem;
        overflow-x: clip;
    }
    [data-testid="stMetricValue"] {font-size: 1.65rem;}
    div[data-testid="stDataEditor"] {border: 1px solid #dddddd; border-radius: 10px;}
    div[data-testid="stNumberInput"] input {text-align: center;}

    /* 輕量插畫首頁：純 SVG，不發出額外網路請求。 */
    .dispatch-hero {
        position: relative;
        display: grid;
        grid-template-columns: minmax(0, 1.45fr) minmax(250px, .75fr);
        align-items: center;
        gap: 1rem;
        min-height: 150px;
        margin: 0 0 1rem;
        padding: 1.15rem 1.35rem;
        overflow: hidden;
        border: 1px solid #d9e4f0;
        border-radius: 24px;
        background: linear-gradient(135deg, #f8fbff 0%, #edf7ff 55%, #fff7df 100%);
        box-shadow: 0 12px 32px rgba(31, 70, 104, .08);
    }
    .dispatch-hero::after {
        content: "";
        position: absolute;
        width: 180px;
        height: 180px;
        right: -75px;
        top: -90px;
        border-radius: 50%;
        background: rgba(255, 190, 44, .14);
        pointer-events: none;
    }
    .dispatch-hero-copy {position: relative; z-index: 1;}
    .dispatch-kicker {
        margin-bottom: .28rem;
        color: #1e6ca5;
        font-size: .72rem;
        font-weight: 900;
        letter-spacing: .12em;
    }
    .dispatch-version-badge {
        display: inline-flex;
        align-items: center;
        min-height: 25px;
        margin: 0 0 .48rem;
        padding: .18rem .55rem;
        border: 1px solid rgba(30,108,165,.18);
        border-radius: 999px;
        color: #145e91;
        background: rgba(255,255,255,.72);
        font-size: .7rem;
        font-weight: 850;
        box-shadow: 0 4px 12px rgba(31,70,104,.05);
    }
    .dispatch-hero h1 {
        margin: 0;
        color: #142a3b;
        font-size: clamp(1.65rem, 3vw, 2.45rem);
        font-weight: 950;
        line-height: 1.08;
    }
    .dispatch-hero p {
        max-width: 620px;
        margin: .55rem 0 0;
        color: #5e7080;
        font-size: .95rem;
        font-weight: 650;
    }
    .dispatch-hero-art {min-width: 0; align-self: stretch;}
    .dispatch-hero-art svg {width: 100%; height: 100%; min-height: 125px;}
    .hero-sun {fill: #ffc44d;}
    .hero-hill-back {fill: #b7ddc9;}
    .hero-hill-front {fill: #79bd9b;}
    .hero-truck rect, .hero-truck path {fill: #f2a900;}
    .hero-truck .hero-window {fill: #dff4ff;}
    .hero-wheel {fill: #263746;}
    .hero-route {fill: none !important; stroke: #2d77a8; stroke-width: 5; stroke-linecap: round; stroke-dasharray: 7 9;}
    .hero-bike circle, .hero-bike path {fill: none; stroke: #f8fbff; stroke-width: 5; stroke-linecap: round; stroke-linejoin: round;}

    /* 行政區總覽樣式只注入一次，避免每個行政區重複建立相同 CSS 節點。 */
    .region-fleet-overview {margin:.35rem 0 .72rem;padding:.72rem;border:1px solid rgba(148,163,184,.24);border-radius:16px;background:rgba(248,250,252,.74);}
    .region-fleet-title {font-size:.78rem;font-weight:850;opacity:.68;margin:0 0 .48rem .08rem;}
    .region-fleet-grid {display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:.48rem;}
    .region-fleet-metric {padding:.62rem .66rem;border-radius:13px;background:rgba(255,255,255,.82);border:1px solid rgba(148,163,184,.18);}
    .region-fleet-metric-head {display:flex;justify-content:space-between;align-items:center;gap:.4rem;}
    .region-fleet-metric-head strong {font-size:.93rem;}
    .region-fleet-metric-head span {font-size:.68rem;font-weight:800;padding:.2rem .4rem;border-radius:999px;background:rgba(148,163,184,.12);}
    .region-fleet-extra .region-fleet-metric-head span {color:#d9363e;background:rgba(244,63,94,.12);}
    .region-fleet-short .region-fleet-metric-head span {color:#b96b00;background:rgba(245,158,11,.15);}
    .region-fleet-balanced .region-fleet-metric-head span {color:#087f5b;background:rgba(16,185,129,.13);}
    .region-fleet-numbers {display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:.3rem;margin-top:.48rem;}
    .region-fleet-numbers div {text-align:center;min-width:0;}
    .region-fleet-numbers small {display:block;font-size:.62rem;opacity:.6;}
    .region-fleet-numbers b {display:block;font-size:.9rem;margin-top:.12rem;white-space:nowrap;}
    .region-fleet-numbers em {font-size:.62rem;font-style:normal;margin-left:.08rem;opacity:.65;}

    .dispatch-context-strip {
        display: flex;
        align-items: center;
        flex-wrap: wrap;
        gap: .45rem;
        margin: .5rem 0 .8rem;
        padding: .62rem .72rem;
        border: 1px solid #e2e7ed;
        border-radius: 14px;
        background: #fff;
        box-shadow: 0 5px 16px rgba(31, 41, 55, .04);
    }
    .dispatch-context-strip span {
        display: inline-flex;
        align-items: center;
        gap: .33rem;
        min-height: 30px;
        padding: .25rem .58rem;
        color: #344454;
        border-radius: 999px;
        background: #f4f7fa;
        font-size: .82rem;
        font-weight: 700;
    }
    .dispatch-context-strip b {color: #1470aa; font-size: .7rem;}
    .dispatch-live-time {margin-left: auto;}

    /* 將操作型元件收斂成明確區塊，減少畫面干擾。 */
    [data-testid="stVerticalBlockBorderWrapper"] {
        border-radius: 16px !important;
        border-color: #e0e6ec !important;
    }
    div[data-testid="stExpander"] {
        border-color: #e1e7ed;
        border-radius: 14px;
        overflow: hidden;
    }
    div[data-testid="stExpander"] summary {font-weight: 800;}

    .analysis-result-table-wrap {
        width: 100%;
        max-width: 100%;
        margin: 0.15rem 0 0.75rem;
        overflow: visible;
    }
    .analysis-result-table {
        width: 100%;
        max-width: 100%;
        table-layout: fixed;
        border-collapse: separate;
        border-spacing: 0;
        border: 1px solid #d9dee5;
        border-radius: 10px;
        overflow: hidden;
        background: white;
    }
    .analysis-result-table .analysis-col-station {width: 42%;}
    .analysis-result-table .analysis-col-status {width: 29%;}
    .analysis-result-table th,
    .analysis-result-table td {
        padding: 8px 6px;
        border-right: 1px solid #e4e8ed;
        border-bottom: 1px solid #e4e8ed;
        vertical-align: middle;
        line-height: 1.3;
        white-space: normal;
        overflow-wrap: anywhere;
        word-break: break-word;
    }
    .analysis-result-table th:last-child,
    .analysis-result-table td:last-child {border-right: 0;}
    .analysis-result-table tbody tr:last-child td {border-bottom: 0;}
    .analysis-result-table th {
        background: #f4f6f8;
        color: #222;
        font-size: 0.88rem;
        font-weight: 800;
        text-align: center;
    }
    .analysis-result-table td {
        font-size: 0.9rem;
    }
    .analysis-station-cell {
        font-weight: 750;
        text-align: left;
    }
    .analysis-station-cell strong,
    .analysis-status-cell strong {display:block;}
    .analysis-station-cell small,
    .analysis-status-cell small {display:block; margin-top:.2rem; font-size:.68rem; font-weight:650; opacity:.68; line-height:1.3;}
    .analysis-zone-label {display:inline-block !important; margin-left:.35rem !important; color:#1677ff; opacity:.9 !important;}
    .analysis-status-cell {
        text-align: center;
        font-weight: 750;
    }
    .analysis-status-extra {background: #fce1e1;}
    .analysis-status-short {background: #fff0df;}
    .analysis-status-error {background: #fff7cc;}
    .analysis-status-ok {background: #e7f5e9;}

    .fleet-summary-card {
        --fleet-accent: #d88900;
        --fleet-soft: #fff8e8;
        --fleet-border: #f5d88f;
        position: relative;
        display: grid;
        grid-template-columns: 118px minmax(0, 1fr);
        gap: 0.7rem;
        width: 100%;
        margin: 0.8rem 0;
        padding: 1rem 1.05rem;
        overflow: hidden;
        border: 1px solid var(--fleet-border);
        border-radius: 20px;
        background:
            radial-gradient(circle at 12% 20%, rgba(255,255,255,.95), rgba(255,255,255,0) 35%),
            linear-gradient(135deg, var(--fleet-soft), #ffffff 70%);
        box-shadow: 0 10px 28px rgba(31, 41, 55, 0.07);
    }
    .fleet-theme-ebike {
        --fleet-accent: #2468c9;
        --fleet-soft: #eef5ff;
        --fleet-border: #bdd4fb;
    }
    .fleet-card-illustration {
        min-width: 0;
        display: flex;
        align-items: center;
        justify-content: center;
        color: var(--fleet-accent);
        border-radius: 18px;
        background: color-mix(in srgb, var(--fleet-accent) 10%, white);
    }
    .fleet-card-svg {
        width: 100%;
        max-width: 110px;
        height: auto;
        filter: drop-shadow(0 7px 9px rgba(31, 41, 55, .08));
    }
    .fleet-card-content {min-width: 0;}
    .fleet-card-heading {
        display: flex;
        align-items: flex-start;
        justify-content: space-between;
        gap: 0.8rem;
        margin-bottom: 0.7rem;
    }
    .fleet-card-title {
        color: var(--fleet-accent);
        font-size: 1.55rem;
        font-weight: 900;
        line-height: 1.1;
        letter-spacing: .01em;
    }
    .fleet-card-subtitle {
        margin-top: .22rem;
        color: #7b8494;
        font-size: .78rem;
        font-weight: 650;
    }
    .fleet-state-badge {
        flex: 0 0 auto;
        display: inline-flex;
        align-items: center;
        min-height: 34px;
        padding: .32rem .72rem;
        border-radius: 999px;
        font-size: .85rem;
        font-weight: 850;
        white-space: nowrap;
        border: 1px solid transparent;
    }
    .fleet-state-short {
        color: #d46700;
        background: #fff1d7;
        border-color: #ffc96c;
    }
    .fleet-state-extra {
        color: #d7243f;
        background: #ffe5e9;
        border-color: #ffabb7;
    }
    .fleet-state-balanced {
        color: #187642;
        background: #e4f7ec;
        border-color: #9fd8b8;
    }
    .fleet-state-pending {
        color: #806000;
        background: #fff7cc;
        border-color: #ead27a;
    }
    .fleet-card-metrics {
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        align-items: stretch;
    }
    .fleet-metric-block {
        min-width: 0;
        padding: 0 .8rem;
        text-align: center;
        border-left: 1px dashed #cfd5df;
    }
    .fleet-metric-block:first-child {
        padding-left: 0;
        border-left: 0;
    }
    .fleet-metric-block:last-child {padding-right: 0;}
    .fleet-metric-label {
        min-height: 1.4rem;
        color: #555f6f;
        font-size: .84rem;
        font-weight: 750;
    }
    .fleet-metric-value {
        margin-top: .32rem;
        color: #151922;
        font-size: clamp(1.55rem, 4vw, 2.25rem);
        font-weight: 900;
        line-height: 1;
        white-space: nowrap;
    }
    .fleet-metric-value span {
        margin-left: .12rem;
        font-size: .56em;
        font-weight: 800;
    }
    .fleet-difference-chip {
        display: flex;
        min-height: 58px;
        margin-top: .25rem;
        padding: .4rem .3rem;
        flex-direction: column;
        justify-content: center;
        border-radius: 13px;
        line-height: 1.06;
    }
    .fleet-difference-chip strong {
        font-size: clamp(1.05rem, 3.4vw, 1.55rem);
        font-weight: 950;
        white-space: nowrap;
    }
    .fleet-difference-chip small {
        margin-top: .22rem;
        font-size: .72rem;
        font-weight: 750;
        opacity: .72;
    }
    .fleet-difference-short {color: #ef7200; background: #fff0d5;}
    .fleet-difference-extra {color: #df203b; background: #ffe1e6;}
    .fleet-difference-balanced {color: #197a45; background: #e2f6ea;}
    .fleet-difference-pending {color: #876900; background: #fff7d4;}

    .fleet-data-notice {
        position: relative;
        display: grid;
        grid-template-columns: 44px minmax(0, 1fr) 56px;
        align-items: center;
        gap: .7rem;
        margin: 1rem 0 .8rem;
        padding: .95rem 1rem;
        overflow: hidden;
        border: 1px solid #f1d67a;
        border-radius: 17px;
        background: linear-gradient(135deg, #fff9dc, #fffdf2);
        box-shadow: 0 7px 20px rgba(91, 67, 0, .05);
    }
    .fleet-notice-icon {
        display: grid;
        width: 40px;
        height: 40px;
        place-items: center;
        color: white;
        font-size: 1.45rem;
        font-weight: 950;
        border-radius: 13px 13px 17px 17px;
        background: #f0a300;
        clip-path: polygon(50% 0, 100% 100%, 0 100%);
        padding-top: 9px;
    }
    .fleet-notice-title {
        color: #8d6100;
        font-size: 1.02rem;
        font-weight: 900;
    }
    .fleet-notice-text {
        margin-top: .2rem;
        color: #7b5a0b;
        font-size: .88rem;
        line-height: 1.55;
    }
    .fleet-notice-decoration {
        font-size: 2.25rem;
        text-align: right;
        filter: saturate(.8);
        opacity: .75;
    }
    .fleet-legend {
        display: flex;
        align-items: center;
        flex-wrap: wrap;
        gap: .35rem;
        margin: .6rem 0 .4rem;
        color: #707887;
        font-size: .9rem;
        font-weight: 700;
    }
    .fleet-legend span {display: inline-flex; align-items: center; gap: .3rem;}
    .fleet-legend-dot {
        display: inline-block;
        width: 16px;
        height: 16px;
        border-radius: 50%;
        box-shadow: inset 0 1px 2px rgba(255,255,255,.7), 0 2px 4px rgba(31,41,55,.14);
    }
    .fleet-legend-extra {background: linear-gradient(#ff8a96, #ef4457);}
    .fleet-legend-short {background: linear-gradient(#ffd07e, #f1a43a);}
    .fleet-legend-balanced {background: linear-gradient(#8cd6a9, #39a96b);}
    .fleet-legend-divider {opacity: .48;}

    /* 手機版：縮小頁面留白，並放大數字輸入框，方便直接叫出九宮格鍵盤。 */
    @media (max-width: 900px) {
        .region-fleet-grid {grid-template-columns:1fr;}
        .block-container {
            padding-left: 0.22rem;
            padding-right: 0.22rem;
            padding-top: 0.7rem;
        }

        .analysis-result-table-wrap {
            margin-left: 0;
            margin-right: 0;
        }
        .analysis-result-table .analysis-col-station {width: 40%;}
        .analysis-result-table .analysis-col-status {width: 30%;}
        .analysis-result-table th,
        .analysis-result-table td {
            padding: 7px 3px;
            font-size: 0.78rem;
            line-height: 1.24;
        }
        .analysis-result-table th {
            font-size: 0.75rem;
        }

        .fleet-summary-card {
            grid-template-columns: 78px minmax(0, 1fr);
            gap: .48rem;
            margin: .65rem 0;
            padding: .8rem .72rem;
            border-radius: 16px;
        }
        .fleet-card-illustration {border-radius: 14px;}
        .fleet-card-svg {max-width: 72px;}
        .fleet-card-heading {gap: .42rem; margin-bottom: .62rem;}
        .fleet-card-title {font-size: 1.25rem;}
        .fleet-card-subtitle {font-size: .67rem;}
        .fleet-state-badge {min-height: 29px; padding: .25rem .5rem; font-size: .72rem;}
        .fleet-metric-block {padding: 0 .32rem;}
        .fleet-metric-label {font-size: .7rem; min-height: 1.2rem;}
        .fleet-metric-value {font-size: clamp(1.22rem, 6vw, 1.7rem);}
        .fleet-difference-chip {min-height: 49px; padding: .3rem .14rem; border-radius: 10px;}
        .fleet-difference-chip strong {font-size: clamp(.84rem, 4vw, 1.12rem);}
        .fleet-difference-chip small {font-size: .61rem;}
        .fleet-data-notice {
            grid-template-columns: 35px minmax(0, 1fr) 38px;
            gap: .48rem;
            padding: .78rem .72rem;
            border-radius: 14px;
        }
        .fleet-notice-icon {width: 34px; height: 34px; font-size: 1.15rem; padding-top: 8px;}
        .fleet-notice-title {font-size: .92rem;}
        .fleet-notice-text {font-size: .77rem; line-height: 1.45;}
        .fleet-notice-decoration {font-size: 1.7rem;}

        .dispatch-hero {
            grid-template-columns: minmax(0, 1fr) 112px;
            min-height: 112px;
            gap: .35rem;
            margin-bottom: .65rem;
            padding: .82rem .78rem;
            border-radius: 17px;
        }
        .dispatch-kicker {font-size: .58rem; letter-spacing: .08em;}
        .dispatch-version-badge {min-height:22px; margin-bottom:.35rem; padding:.12rem .42rem; font-size:.58rem;}
        .dispatch-hero h1 {font-size: 1.35rem;}
        .dispatch-hero p {margin-top: .35rem; font-size: .72rem; line-height: 1.35;}
        .dispatch-hero-art svg {min-height: 88px;}
        .dispatch-context-strip {gap: .3rem; padding: .48rem; margin-bottom: .58rem;}
        .dispatch-context-strip span {min-height: 27px; padding: .2rem .42rem; font-size: .7rem;}
        .dispatch-context-strip b {font-size: .61rem;}
        .dispatch-live-time {width: 100%; margin-left: 0;}

        div[data-testid="stDataEditor"] {
            font-size: 0.78rem;
        }

        div[data-testid="stDataEditor"] [role="columnheader"],
        div[data-testid="stDataEditor"] [role="gridcell"] {
            padding-left: 2px !important;
            padding-right: 2px !important;
        }

        div[data-testid="stNumberInput"] input {
            min-height: 44px;
            font-size: 18px !important;
            font-weight: 700;
        }
    }

    /* v27.5 CYBER_SKIN_BEGIN
       僅改色彩、背景、邊框與光影；不覆寫任何排版、尺寸、間距或定位。 */
    :root {
        --neon-bg: #050811;
        --neon-panel: #0a1222;
        --neon-panel-soft: rgba(10, 22, 38, .92);
        --neon-cyan: #55f6ff;
        --neon-pink: #ff3fcf;
        --neon-yellow: #f4ff57;
        --neon-green: #53f6ad;
        --neon-text: #effcff;
        --neon-muted: #9bb7c4;
        --neon-line: rgba(85, 246, 255, .30);
    }
    body,
    .stApp,
    [data-testid="stAppViewContainer"],
    [data-testid="stMain"] {
        color: var(--neon-text);
        background:
            linear-gradient(rgba(85,246,255,.035) 1px, transparent 1px),
            linear-gradient(90deg, rgba(85,246,255,.035) 1px, transparent 1px),
            radial-gradient(circle at 12% 10%, rgba(255,63,207,.15), transparent 28rem),
            radial-gradient(circle at 88% 20%, rgba(85,246,255,.12), transparent 30rem),
            linear-gradient(145deg, #050811, #091326 52%, #100918) !important;
        background-size: 36px 36px, 36px 36px, auto, auto, auto !important;
    }
    ::selection {color:#071018; background:var(--neon-cyan);}
    [data-testid="stHeader"] {
        background: rgba(5, 8, 17, .88) !important;
        border-bottom-color: rgba(85,246,255,.22) !important;
        box-shadow: 0 8px 24px rgba(0,0,0,.28) !important;
    }
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #110b22, #071522) !important;
        border-right-color: rgba(85,246,255,.28) !important;
        box-shadow: 10px 0 30px rgba(0,0,0,.32) !important;
    }
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] [data-testid="stCaptionContainer"] {color:var(--neon-text) !important;}
    .stApp h1,
    .stApp h2,
    .stApp h3,
    .stApp h4,
    .stApp h5,
    .stApp h6 {color:var(--neon-text) !important; text-shadow:0 0 16px rgba(85,246,255,.16);}
    .stApp p,
    .stApp label,
    .stApp [data-testid="stCaptionContainer"] {color:var(--neon-muted);}
    .stApp a {color:var(--neon-cyan) !important;}
    .stApp hr {border-color:rgba(85,246,255,.18) !important;}

    .dispatch-hero {
        border-color: rgba(85,246,255,.56);
        background:
            linear-gradient(rgba(85,246,255,.055) 1px, transparent 1px),
            linear-gradient(90deg, rgba(85,246,255,.055) 1px, transparent 1px),
            linear-gradient(125deg, rgba(7,14,31,.98), rgba(34,9,43,.94) 55%, rgba(4,31,41,.96));
        background-size: 24px 24px, 24px 24px, auto;
        box-shadow: 0 0 0 1px rgba(255,63,207,.13), 0 16px 40px rgba(0,0,0,.34), 0 0 28px rgba(85,246,255,.10);
    }
    .dispatch-hero::after {background:rgba(255,63,207,.18); box-shadow:0 0 45px rgba(255,63,207,.20);}
    .dispatch-kicker {color:var(--neon-cyan); text-shadow:0 0 12px rgba(85,246,255,.68);}
    .dispatch-version-badge {
        color:#ffd7f5;
        border-color:rgba(255,63,207,.52);
        background:rgba(255,63,207,.10);
        box-shadow:inset 0 0 14px rgba(255,63,207,.06), 0 0 12px rgba(255,63,207,.10);
    }
    .dispatch-hero h1 {color:#f7ffff !important; text-shadow:2px 2px 0 rgba(255,63,207,.24), -1px -1px 0 rgba(85,246,255,.22);}
    .dispatch-hero p {color:#b6cfda;}
    .hero-sun {fill:var(--neon-yellow); filter:drop-shadow(0 0 8px rgba(244,255,87,.72));}
    .hero-hill-back {fill:#17324b;}
    .hero-hill-front {fill:#0e5362;}
    .hero-truck rect,
    .hero-truck path {fill:var(--neon-pink);}
    .hero-truck .hero-window {fill:#bffcff;}
    .hero-wheel {fill:#071018; stroke:var(--neon-cyan);}
    .hero-route {stroke:var(--neon-cyan); filter:drop-shadow(0 0 5px rgba(85,246,255,.65));}
    .hero-bike circle,
    .hero-bike path {stroke:var(--neon-yellow); filter:drop-shadow(0 0 5px rgba(244,255,87,.48));}

    .dispatch-context-strip {
        border-color:rgba(85,246,255,.34);
        background:linear-gradient(100deg, rgba(8,22,37,.96), rgba(27,9,36,.92));
        box-shadow:inset 3px 0 var(--neon-cyan), 0 8px 22px rgba(0,0,0,.24);
    }
    .dispatch-context-strip span {color:#dffbff; background:rgba(85,246,255,.07);}
    .dispatch-context-strip b {color:var(--neon-yellow);}
    [data-testid="stVerticalBlockBorderWrapper"] {
        border-color:rgba(85,246,255,.27) !important;
        background:linear-gradient(135deg, rgba(7,18,34,.92), rgba(25,9,36,.84)) !important;
        box-shadow:0 10px 26px rgba(0,0,0,.24), inset 0 0 18px rgba(85,246,255,.025) !important;
    }
    div[data-testid="stExpander"] {
        border-color:rgba(255,63,207,.27) !important;
        background:rgba(8,17,32,.90) !important;
        box-shadow:inset 0 0 16px rgba(255,63,207,.025);
    }
    div[data-testid="stExpander"] summary {color:var(--neon-cyan) !important;}
    [data-testid="stTabs"] [data-baseweb="tab-list"] {border-bottom-color:rgba(85,246,255,.22) !important;}
    [data-testid="stTabs"] button[role="tab"] {color:#9db8c6 !important;}
    [data-testid="stTabs"] button[aria-selected="true"] {color:var(--neon-yellow) !important; border-bottom-color:var(--neon-pink) !important;}

    .stButton > button,
    .stDownloadButton > button,
    button[kind="primary"] {
        color:#f4ffff !important;
        border-color:rgba(85,246,255,.58) !important;
        background:linear-gradient(110deg, #12344d, #52195b) !important;
        box-shadow:inset 0 0 14px rgba(85,246,255,.07), 0 0 14px rgba(85,246,255,.10) !important;
    }
    .stButton > button:hover,
    .stDownloadButton > button:hover {
        color:var(--neon-yellow) !important;
        border-color:var(--neon-yellow) !important;
        box-shadow:0 0 20px rgba(244,255,87,.18), 0 8px 20px rgba(0,0,0,.28) !important;
    }
    div[data-baseweb="select"] > div,
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input,
    div[data-testid="stTextArea"] textarea {
        color:var(--neon-text) !important;
        border-color:rgba(85,246,255,.34) !important;
        background:#071225 !important;
        box-shadow:inset 0 0 12px rgba(85,246,255,.035) !important;
    }
    div[data-testid="stNumberInput"] button {color:var(--neon-cyan) !important; background:#0a1930 !important; border-color:rgba(85,246,255,.20) !important;}
    div[data-baseweb="select"] svg {fill:var(--neon-cyan) !important;}
    div[data-baseweb="popover"],
    [role="listbox"] {color:var(--neon-text) !important; background:#0a1224 !important;}
    [role="option"] {color:var(--neon-text) !important; background:#0a1224 !important;}
    [role="option"]:hover {background:rgba(85,246,255,.13) !important;}
    [data-testid="stFileUploaderDropzone"] {
        color:var(--neon-text) !important;
        border-color:rgba(85,246,255,.35) !important;
        background:rgba(8,20,36,.88) !important;
    }
    [data-testid="stAlert"] {
        color:var(--neon-text) !important;
        border-color:rgba(244,255,87,.32) !important;
        background:linear-gradient(110deg, rgba(45,40,7,.82), rgba(9,19,34,.94)) !important;
        box-shadow:0 7px 20px rgba(0,0,0,.20) !important;
    }
    [data-testid="stAlert"] p {color:#eef4ba !important;}
    [data-testid="stMetric"] {
        border-color:rgba(85,246,255,.24) !important;
        background:rgba(7,17,32,.88) !important;
        box-shadow:inset 3px 0 rgba(85,246,255,.42), 0 8px 20px rgba(0,0,0,.20) !important;
    }
    [data-testid="stMetricLabel"] {color:var(--neon-muted) !important;}
    [data-testid="stMetricValue"] {color:var(--neon-yellow) !important; text-shadow:0 0 12px rgba(244,255,87,.22);}
    div[data-testid="stDataEditor"] {border-color:rgba(85,246,255,.32) !important; box-shadow:0 0 18px rgba(85,246,255,.07);}

    .region-fleet-overview {
        border-color:rgba(85,246,255,.26);
        background:rgba(7,18,34,.88);
        box-shadow:inset 3px 0 rgba(85,246,255,.50);
    }
    .region-fleet-title {color:#b9d5df;}
    .region-fleet-metric {
        border-color:rgba(255,63,207,.20);
        background:linear-gradient(135deg, rgba(10,25,42,.94), rgba(31,9,39,.88));
    }
    .region-fleet-metric-head strong,
    .region-fleet-numbers b {color:var(--neon-text);}
    .region-fleet-metric-head span {background:rgba(85,246,255,.08);}
    .region-fleet-extra .region-fleet-metric-head span {color:#ff86df; background:rgba(255,63,207,.12);}
    .region-fleet-short .region-fleet-metric-head span {color:#ffd27a; background:rgba(255,178,55,.12);}
    .region-fleet-balanced .region-fleet-metric-head span {color:#7bffc2; background:rgba(83,246,173,.11);}

    .fleet-summary-card {
        --fleet-accent: var(--neon-pink);
        --fleet-soft: rgba(38,7,36,.96);
        --fleet-border: rgba(255,63,207,.38);
        background:radial-gradient(circle at 12% 20%, rgba(255,63,207,.10), transparent 35%), linear-gradient(135deg, #0a1728, #26091f 72%);
        box-shadow:0 13px 30px rgba(0,0,0,.28), inset 0 0 20px rgba(255,63,207,.025);
    }
    .fleet-theme-ebike {
        --fleet-accent: var(--neon-cyan);
        --fleet-soft: rgba(5,29,38,.96);
        --fleet-border: rgba(85,246,255,.40);
    }
    .fleet-card-illustration {background:rgba(85,246,255,.07); box-shadow:inset 0 0 16px rgba(85,246,255,.04);}
    .fleet-card-title {text-shadow:0 0 13px currentColor;}
    .fleet-card-subtitle,
    .fleet-metric-label {color:var(--neon-muted);}
    .fleet-metric-value {color:var(--neon-text);}
    .fleet-metric-block {border-left-color:rgba(85,246,255,.18);}
    .fleet-state-short,
    .fleet-difference-short {color:#ffd27a; background:rgba(255,157,47,.13); border-color:rgba(255,157,47,.42);}
    .fleet-state-extra,
    .fleet-difference-extra {color:#ff91df; background:rgba(255,63,207,.13); border-color:rgba(255,63,207,.42);}
    .fleet-state-balanced,
    .fleet-difference-balanced {color:#78ffc1; background:rgba(83,246,173,.12); border-color:rgba(83,246,173,.38);}
    .fleet-state-pending,
    .fleet-difference-pending {color:var(--neon-yellow); background:rgba(244,255,87,.10); border-color:rgba(244,255,87,.35);}
    .fleet-data-notice {
        border-color:rgba(244,255,87,.44);
        background:linear-gradient(135deg, rgba(48,42,5,.88), rgba(9,18,34,.96));
        box-shadow:0 0 22px rgba(244,255,87,.07), 0 8px 22px rgba(0,0,0,.22);
    }
    .fleet-notice-icon {color:#091018; background:var(--neon-yellow); box-shadow:0 0 15px rgba(244,255,87,.28);}
    .fleet-notice-title {color:var(--neon-yellow);}
    .fleet-notice-text {color:#e0e4a6;}
    .fleet-legend {color:var(--neon-muted);}

    .analysis-result-table {
        border-color:rgba(85,246,255,.32);
        background:#081426;
        box-shadow:0 9px 22px rgba(0,0,0,.24);
    }
    .analysis-result-table th {
        color:var(--neon-cyan);
        border-color:rgba(85,246,255,.14);
        background:linear-gradient(90deg, #0b1c32, #24102f);
        text-shadow:0 0 8px rgba(85,246,255,.28);
    }
    .analysis-result-table td {
        color:#ecfbff;
        border-color:rgba(85,246,255,.12);
        background:rgba(8,18,34,.94);
    }
    .analysis-zone-label {color:var(--neon-yellow) !important;}
    .analysis-status-extra {background:rgba(255,63,207,.20) !important;}
    .analysis-status-short {background:rgba(255,157,47,.18) !important;}
    .analysis-status-error {background:rgba(244,255,87,.14) !important;}
    .analysis-status-ok {background:rgba(83,246,173,.13) !important;}

    .stApp .dispatch-plan-card,
    .stApp .dispatch-truck-status,
    .stApp .dispatch-candidate-card {
        color:var(--neon-text);
        border-color:rgba(85,246,255,.26);
        background:linear-gradient(135deg, rgba(8,21,38,.94), rgba(28,8,38,.88));
        box-shadow:0 9px 22px rgba(0,0,0,.22);
    }
    .stApp .dispatch-plan-grid div,
    .stApp .dispatch-truck-compare,
    .stApp .dispatch-truck-row,
    .stApp .dispatch-truck-status-grid div {
        color:var(--neon-text);
        background:rgba(85,246,255,.055);
    }
    .stApp .dispatch-plan-action {color:var(--neon-cyan); background:rgba(85,246,255,.10);}
    .stApp .dispatch-plan-kicker,
    .stApp .dispatch-candidate-rank {color:var(--neon-cyan);}
    .stApp .dispatch-candidate-count {color:#78ffc1; background:rgba(83,246,173,.11);}
    /* v27.5 CYBER_SKIN_END */
    </style>
    """,
    unsafe_allow_html=True,
)




# 跳過不再詢問原因，也不再以原因倍率改變 AI 排名。
DISPATCH_IGNORE_ROUNDS = 2
DISPATCH_ESTIMATED_SPEED_KMH = 32.0
DISPATCH_ROAD_DISTANCE_FACTOR = 1.22  # 僅供道路服務暫時失效時的備援估算。
DISPATCH_OPERATION_BASE_MINUTES = 2.0
DISPATCH_OPERATION_MINUTES_PER_BIKE = 0.75

# 實際道路路網：以 OSRM／OpenStreetMap 的可行駛道路時間與距離為主要依據。
ROAD_ROUTER_BASE_URL = os.getenv("OSRM_BASE_URL", "https://router.project-osrm.org").rstrip("/")
ROAD_ROUTER_PROFILE = os.getenv("OSRM_PROFILE", "driving").strip() or "driving"
ROAD_ROUTER_TIMEOUT_SECONDS = 12
ROAD_ROUTER_MAX_ATTEMPTS = 2
ROAD_ROUTER_BATCH_SIZE = 36
ROAD_ROUTER_CACHE_TTL_SECONDS = 300
ROAD_ROUTER_ORIGIN_PRECISION = 4   # 約 11 公尺；降低 GPS 微小飄移造成的重複查詢。
ROAD_ROUTER_STATION_PRECISION = 5  # 約 1 公尺；保留場站道路定位精度。
ROAD_ROUTER_LOOKAHEAD_OPTIONS = 3
ROAD_ROUTER_LOOKAHEAD_STOPS = 3


DISPATCH_GEOLOCATION_COMPONENT_HTML = r"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <style>
    * { box-sizing: border-box; }
    html, body { margin: 0; padding: 0; background: transparent; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; }
    body.compact #locateButton, body.compact #status { display: none !important; }
    #locateButton {
      width: 100%; min-height: 50px; border: 0; border-radius: 13px;
      padding: 10px 14px; font-size: 16px; font-weight: 750;
      color: #fff; background: #1677ff; cursor: pointer;
      -webkit-tap-highlight-color: transparent; touch-action: manipulation;
    }
    #locateButton:disabled { opacity: .68; cursor: wait; }
    #status { min-height: 18px; margin-top: 6px; padding: 0 3px; font-size: 13px; color: #6b7280; }
    .error { color: #c62828 !important; }
  </style>
</head>
<body>
  <button id="locateButton" type="button">📍 取得／更新目前位置</button>
  <div id="status">第一次使用時，瀏覽器會詢問定位權限。</div>
<script>
(() => {
  const API_VERSION = 1;
  const button = document.getElementById("locateButton");
  const statusNode = document.getElementById("status");
  let args = {};
  let busy = false;
  let lastAutoRequestToken = "";
  let autoTimer = null;
  let autoStarted = false;
  let lastDeliveredLocation = null;
  const LOCATION_HEARTBEAT_MS = 5 * 60 * 1000;

  function distanceMeters(lat1, lon1, lat2, lon2) {
    const earthRadius = 6371008.8;
    const radians = value => Number(value) * Math.PI / 180;
    const phi1 = radians(lat1);
    const phi2 = radians(lat2);
    const deltaPhi = radians(lat2 - lat1);
    const deltaLambda = radians(lon2 - lon1);
    const a = Math.sin(deltaPhi / 2) ** 2
      + Math.cos(phi1) * Math.cos(phi2) * Math.sin(deltaLambda / 2) ** 2;
    return earthRadius * 2 * Math.atan2(Math.sqrt(a), Math.sqrt(Math.max(0, 1 - a)));
  }

  function send(type, data = {}) {
    window.parent.postMessage({ isStreamlitMessage: true, type, ...data }, "*");
  }
  function setHeight() {
    const compact = Boolean(args.compact);
    send("streamlit:setFrameHeight", { height: compact ? 1 : Math.max(78, document.body.scrollHeight + 2) });
  }
  function clearAutoTimer() {
    if (autoTimer !== null) {
      window.clearTimeout(autoTimer);
      autoTimer = null;
    }
  }
  function scheduleAutoLocate() {
    clearAutoTimer();
    if (!args.auto_refresh) return;
    const seconds = Math.max(10, Math.min(300, Number(args.auto_refresh_seconds || 30)));
    autoTimer = window.setTimeout(() => {
      autoTimer = null;
      if (busy) scheduleAutoLocate();
      else runLocate({ automatic: true });
    }, seconds * 1000);
  }
  function setValue(value) {
    send("streamlit:setComponentValue", { value, dataType: "json" });
  }
  function eventId() {
    if (globalThis.crypto && typeof globalThis.crypto.randomUUID === "function") return globalThis.crypto.randomUUID();
    return `${Date.now()}-${Math.random().toString(16).slice(2)}`;
  }
  function setStatus(text, isError = false) {
    statusNode.textContent = text || "";
    statusNode.className = isError ? "error" : "";
    setHeight();
  }
  function runLocate({ automatic = false, forceDelivery = false } = {}) {
    if (busy) return;
    clearAutoTimer();
    if (!navigator.geolocation) {
      setStatus("此瀏覽器不支援定位。", true);
      setValue({ ok: false, event_id: eventId(), error: "此瀏覽器不支援定位" });
      return;
    }
    busy = true;
    button.disabled = true;
    button.textContent = "⏳ 正在取得目前位置……";
    setStatus(
      automatic
        ? "正在更新目前位置；完成後會自動重新計算距離與路線。"
        : "請允許瀏覽器使用定位；室外或靠近窗邊通常較準。",
      false,
    );
    navigator.geolocation.getCurrentPosition(
      position => {
        const payload = {
          ok: true,
          event_id: eventId(),
          request_token: String(args.request_token || ""),
          latitude: Number(position.coords.latitude),
          longitude: Number(position.coords.longitude),
          accuracy: Number(position.coords.accuracy || 0),
          timestamp: Number(position.timestamp || Date.now()),
        };
        const nowMilliseconds = Date.now();
        const movementThreshold = Math.max(20, Math.min(60, payload.accuracy * 0.55));
        const movedMeters = lastDeliveredLocation
          ? distanceMeters(
              lastDeliveredLocation.latitude,
              lastDeliveredLocation.longitude,
              payload.latitude,
              payload.longitude,
            )
          : Number.POSITIVE_INFINITY;
        const accuracyImproved = Boolean(
          lastDeliveredLocation
          && payload.accuracy + 15 < Number(lastDeliveredLocation.accuracy || 0)
        );
        const heartbeatDue = Boolean(
          lastDeliveredLocation
          && nowMilliseconds - Number(lastDeliveredLocation.deliveredAt || 0) >= LOCATION_HEARTBEAT_MS
        );
        const shouldDeliver = forceDelivery || !automatic || !lastDeliveredLocation
          || movedMeters >= movementThreshold || accuracyImproved || heartbeatDue;

        if (shouldDeliver) {
          setValue(payload);
          lastDeliveredLocation = {
            latitude: payload.latitude,
            longitude: payload.longitude,
            accuracy: payload.accuracy,
            deliveredAt: nowMilliseconds,
          };
        }
        const movementText = Number.isFinite(movedMeters) ? `，位移約 ${Math.round(movedMeters)} 公尺` : "";
        setStatus(
          shouldDeliver
            ? `定位完成，誤差約 ${Math.round(payload.accuracy)} 公尺${movementText}。`
            : `定位已檢查，位置無明顯變化，略過整頁重算。`,
          false,
        );
        busy = false;
        button.disabled = false;
        button.textContent = "📍 重新取得目前位置";
        scheduleAutoLocate();
      },
      error => {
        const message = error && error.message ? error.message : "定位失敗";
        setValue({
          ok: false,
          event_id: eventId(),
          request_token: String(args.request_token || ""),
          error: message,
        });
        setStatus(`定位失敗：${message}`, true);
        busy = false;
        button.disabled = false;
        button.textContent = "📍 再試一次";
        scheduleAutoLocate();
      },
      { enableHighAccuracy: true, timeout: 15000, maximumAge: 0 }
    );
  }
  button.addEventListener("click", () => runLocate());
  window.addEventListener("message", event => {
    if (!event.data || event.data.type !== "streamlit:render") return;
    args = event.data.args || {};
    document.body.classList.toggle("compact", Boolean(args.compact));
    button.disabled = Boolean(event.data.disabled) || busy;
    const requestToken = String(args.request_token || "");
    if (requestToken && requestToken !== lastAutoRequestToken) {
      lastAutoRequestToken = requestToken;
      window.setTimeout(() => runLocate({ automatic: true, forceDelivery: true }), 0);
    } else if (args.auto_start && !autoStarted) {
      autoStarted = true;
      window.setTimeout(() => runLocate({ automatic: true }), 0);
    } else {
      scheduleAutoLocate();
    }
    setHeight();
  });
  send("streamlit:componentReady", { apiVersion: API_VERSION });
  setHeight();
})();
</script>
</body>
</html>
"""


_DISPATCH_GEOLOCATION_COMPONENT = None


LOW_BATTERY_CLIENT_CORE_JS = r"""
      function ensureUbikeBatteryService(win) {
        const serviceVersion = "v27.5.1";
        const existing = win.__ubikeBatteryService;
        if (existing && existing.version === serviceVersion) return existing;

        const catalogUrl = "https://apis.youbike.com.tw/json/station-min-yb2.json";
        const batteryUrl = "https://apis.youbike.com.tw/api/front/bike/lists";
        let catalogCache = null;
        let catalogPromise = null;
        const batteryCache = new Map();
        const batteryInflight = new Map();

        function wait(milliseconds) {
          return new Promise(resolve => win.setTimeout(resolve, milliseconds));
        }
        function normalizeStationName(value) {
          return String(value || "")
            .toLowerCase()
            .replaceAll("臺", "台")
            .replace(/^(?:youbike|ubike)\s*2\s*[.．]?\s*0\s*e?\s*[_\-－—:：]*\s*/i, "")
            .replaceAll("公共自行車租賃站", "")
            .replace(/[^0-9a-z㐀-鿿]/g, "");
        }
        function extractItems(payload) {
          if (Array.isArray(payload)) return payload.filter(item => item && typeof item === "object");
          if (!payload || typeof payload !== "object") return [];
          for (const candidate of [payload.data, payload.result, payload.stations, payload.retVal]) {
            if (Array.isArray(candidate)) return candidate.filter(item => item && typeof item === "object");
            if (candidate && typeof candidate === "object" && Array.isArray(candidate.data)) {
              return candidate.data.filter(item => item && typeof item === "object");
            }
          }
          return [];
        }
        function isTaitungStation(item) {
          const text = [item.county_tw, item.city_tw, item.district_tw, item.address_tw, item.name_tw]
            .map(value => String(value || ""))
            .join(" ")
            .replaceAll("臺", "台");
          return String(item.area_code || "") === "15" || text.includes("台東");
        }
        function matchCatalogStation(stationName, catalog) {
          const wantedKey = normalizeStationName(stationName);
          const exact = catalog.filter(
            item => normalizeStationName(item.name_tw || item.sna || item.station_name) === wantedKey,
          );
          if (exact.length === 1) return exact[0];
          const partial = catalog
            .map(item => {
              const candidateKey = normalizeStationName(item.name_tw || item.sna || item.station_name);
              const includes = wantedKey.length >= 4 && candidateKey.length >= 4
                && (wantedKey.includes(candidateKey) || candidateKey.includes(wantedKey));
              const score = includes
                ? Math.min(wantedKey.length, candidateKey.length) / Math.max(wantedKey.length, candidateKey.length)
                : 0;
              return { item, score };
            })
            .filter(record => record.score >= .72)
            .sort((left, right) => right.score - left.score);
          if (!partial.length) return null;
          if (partial.length > 1 && partial[1].score >= partial[0].score - .02) return null;
          return partial[0].item;
        }
        async function fetchJson(url, { attempts = 2, timeoutMs = 14000 } = {}) {
          let lastError = null;
          for (let attempt = 1; attempt <= Math.max(1, attempts); attempt += 1) {
            const controller = new win.AbortController();
            const timeout = win.setTimeout(() => controller.abort(), timeoutMs);
            try {
              const response = await win.fetch(url, {
                cache: "no-store",
                credentials: "omit",
                signal: controller.signal,
              });
              if (!response.ok) {
                const error = new Error(`HTTP ${response.status}`);
                error.retryable = [408, 425, 429, 500, 502, 503, 504].includes(response.status);
                throw error;
              }
              return await response.json();
            } catch (error) {
              lastError = error;
              const networkFailure = error?.name === "AbortError"
                || error instanceof TypeError
                || /load failed|failed to fetch|networkerror|network request failed/i.test(
                  String(error?.message || error),
                );
              if (attempt >= attempts || (!networkFailure && error?.retryable !== true)) throw error;
              await wait(320 * attempt + Math.floor(Math.random() * 260));
            } finally {
              win.clearTimeout(timeout);
            }
          }
          throw lastError || new Error("網路連線失敗");
        }
        async function getCatalog({ force = false, attempts = 2, timeoutMs = 15000 } = {}) {
          if (catalogCache && !force) return catalogCache;
          if (catalogPromise && !force) return catalogPromise;
          catalogPromise = fetchJson(catalogUrl, { attempts, timeoutMs })
            .then(payload => {
              const catalog = extractItems(payload).filter(isTaitungStation);
              if (!catalog.length) throw new Error("找不到臺東場站清單");
              catalogCache = catalog;
              return catalogCache;
            })
            .finally(() => {
              catalogPromise = null;
            });
          return catalogPromise;
        }
        function normalizeBatteryRecords(payload) {
          const records = Array.isArray(payload?.retVal) ? payload.retVal : extractItems(payload);
          return records.map(record => {
            const batteryPower = Number(record.battery_power);
            return {
              bike_no: String(record.bike_no || "").trim(),
              pillar_no: String(record.pillar_no || "").trim(),
              battery_power: Number.isFinite(batteryPower)
                ? Math.max(0, Math.min(100, Math.trunc(batteryPower)))
                : null,
            };
          }).filter(record => record.bike_no && Number.isFinite(record.battery_power));
        }
        async function getBatteryListByStationNo(
          stationNo,
          { force = false, ttlMs = 45000, attempts = 2, timeoutMs = 14000 } = {},
        ) {
          const key = String(stationNo || "").trim();
          if (!key) throw new Error("場站編號不存在");
          const cached = batteryCache.get(key);
          if (!force && cached && Date.now() - cached.fetchedAt <= ttlMs) return cached.bikes;
          if (batteryInflight.has(key)) {
            try {
              const inflightResult = await batteryInflight.get(key);
              if (!force) return inflightResult;
            } catch (_) {}
          }
          const promise = fetchJson(
            `${batteryUrl}?station_no=${encodeURIComponent(key)}`,
            { attempts, timeoutMs },
          ).then(payload => {
            const bikes = normalizeBatteryRecords(payload);
            batteryCache.set(key, { fetchedAt: Date.now(), bikes });
            return bikes;
          }).finally(() => {
            if (batteryInflight.get(key) === promise) batteryInflight.delete(key);
          });
          batteryInflight.set(key, promise);
          return promise;
        }
        async function queryStationByName(stationName, options = {}) {
          const catalog = await getCatalog({
            attempts: options.attempts,
            timeoutMs: options.timeoutMs,
          });
          const matched = matchCatalogStation(stationName, catalog);
          if (!matched) throw new Error("官方清單找不到此場站");
          const stationNo = String(
            matched.station_no || matched.sno || matched.station_id || "",
          ).trim();
          if (!stationNo) throw new Error("官方場站缺少站號");
          const bikes = await getBatteryListByStationNo(stationNo, options);
          return {
            matched: true,
            stationNo,
            stationName: String(stationName || "").trim(),
            bikes,
          };
        }

        const service = {
          version: serviceVersion,
          normalizeStationName,
          getCatalog,
          matchCatalogStation,
          getBatteryListByStationNo,
          queryStationByName,
          clearBatteryCache() { batteryCache.clear(); },
        };
        win.__ubikeBatteryService = service;
        return service;
      }
"""

JARVIS_TRIGGER_BOOTSTRAP_HTML = r'''<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<style>html,body{width:1px;height:1px;margin:0;overflow:hidden;background:transparent}</style></head><body>
<script>
(() => {
  const win = window.parent;
  const doc = win.document;
  const STORAGE_KEY = "taitung-jarvis-enabled-v1";
  const GLOBAL_KEY = "__taitungJarvisTriggerBootstrapV3";

  // 若同一個瀏覽器頁面仍殘留 V28.1 的五連點 state，讓它永遠不會累積到五下。
  // 正常重新載入頁面時不會遇到，但 Streamlit 熱重載時可以避免舊監聽器重複切換。
  const legacyState = win.__taitungJarvisTriggerBootstrapV2;
  if (legacyState && typeof legacyState === "object") legacyState.count = -1000000;

  function ensureStyle(){
    if(doc.getElementById("jarvis-trigger-bootstrap-style")) return;
    const style=doc.createElement("style");
    style.id="jarvis-trigger-bootstrap-style";
    style.textContent=`
      #jarvis-secret-trigger{touch-action:manipulation!important;user-select:none!important;-webkit-user-select:none!important;cursor:pointer!important}
      #jarvis-voice-indicator{position:fixed;right:12px;top:max(72px,calc(env(safe-area-inset-top,0px) + 58px));bottom:auto;z-index:2147483000;display:none;
        align-items:center;gap:7px;padding:7px 10px;border:1px solid rgba(85,246,255,.35);border-radius:999px;
        background:rgba(4,16,29,.88);color:#dffcff;box-shadow:0 8px 24px rgba(0,0,0,.24);
        font:800 12px/1.2 -apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft JhengHei",sans-serif;backdrop-filter:blur(8px)}
      #jarvis-voice-indicator.on{display:flex} #jarvis-voice-indicator .dot{width:8px;height:8px;border-radius:50%;background:#55f6ff;box-shadow:0 0 10px #55f6ff}
      #jarvis-voice-indicator.busy .dot{background:#ff5da2;box-shadow:0 0 12px #ff5da2}
    `;
    doc.head.appendChild(style);
  }
  function readEnabled(){ try{return win.sessionStorage.getItem(STORAGE_KEY)==="1";}catch(_){return false;} }
  function writeEnabled(value){ try{win.sessionStorage.setItem(STORAGE_KEY,value?"1":"0");}catch(_){} }
  function isStandalone(){
    try{return Boolean(win.navigator.standalone) || win.matchMedia("(display-mode: standalone)").matches;}catch(_){return false;}
  }
  function speechRecognitionCtor(){ return win.SpeechRecognition || win.webkitSpeechRecognition || null; }
  function preferredZhVoice(){
    try{
      const voices=win.speechSynthesis?.getVoices?.() || [];
      return voices.find(v=>/^zh-TW$/i.test(v.lang)) || voices.find(v=>/^zh/i.test(v.lang)) || null;
    }catch(_){return null;}
  }
  function sayActivation(message){
    try{
      if(!win.speechSynthesis || !win.SpeechSynthesisUtterance) return;
      win.speechSynthesis.cancel();
      const u=new win.SpeechSynthesisUtterance(message); u.lang="zh-TW"; u.rate=1.02;
      const voice=preferredZhVoice(); if(voice) u.voice=voice;
      win.speechSynthesis.speak(u);
    }catch(_){}
  }
  function primeMicrophone(){
    // 這個函式一定由使用者點擊直接觸發，讓 iOS 有機會在 user gesture 內要求權限。
    const R=speechRecognitionCtor();
    if(!R){ reflect(true, isStandalone()?"此模式不支援語音，請用 Safari 開啟":"瀏覽器不支援語音辨識"); return; }
    try{
      const probe=new R(); probe.lang="zh-TW"; probe.continuous=false; probe.interimResults=false;
      let stopped=false;
      probe.onstart=()=>{ win.setTimeout(()=>{if(!stopped){stopped=true;try{probe.stop();}catch(_){}}},180); };
      probe.onerror=(event)=>{
        const err=String(event?.error || "");
        if(err==="not-allowed"||err==="service-not-allowed") reflect(true,"請開啟麥克風／Siri 聽寫權限");
      };
      probe.start();
    }catch(_){}
  }
  function ensureIndicator(){
    ensureStyle();
    let node=doc.getElementById("jarvis-voice-indicator");
    if(!node){
      node=doc.createElement("div"); node.id="jarvis-voice-indicator";
      node.innerHTML='<span class="dot"></span><span class="label">賈維斯測試已開啟</span>';
      doc.body.appendChild(node);
    }
    return node;
  }
  function reflect(enabled, label){
    const node=ensureIndicator();
    node.classList.toggle("on", Boolean(enabled));
    node.classList.remove("busy");
    const labelNode=node.querySelector(".label");
    if(labelNode) labelNode.textContent=label || (enabled?"賈維斯測試已開啟":"賈維斯已關閉");
  }
  function toggle(){
    const enabled=!readEnabled();
    writeEnabled(enabled);
    reflect(enabled, enabled?"賈維斯測試已開啟":"賈維斯已關閉");
    if(enabled){ primeMicrophone(); sayActivation("賈維斯已啟動"); }
    else { try{win.speechSynthesis?.cancel?.();}catch(_){} }
    win.dispatchEvent(new CustomEvent("taitung:jarvis-enabled-change", {detail:{enabled}}));
    if(!enabled){
      win.setTimeout(()=>{ const node=doc.getElementById("jarvis-voice-indicator"); if(node && !readEnabled()) node.classList.remove("on"); },700);
    }
  }

  ensureStyle();
  reflect(readEnabled(), readEnabled()?"賈維斯測試已開啟":"賈維斯已關閉");
  if(!readEnabled()) ensureIndicator().classList.remove("on");

  if(win[GLOBAL_KEY]) return;
  const state={lastTouchAt:0};
  function registerTap(event){
    const target=event.target && event.target.closest ? event.target.closest("#jarvis-secret-trigger") : null;
    if(!target) return;
    const now=Date.now();
    if(event.type==="touchend"){
      state.lastTouchAt=now;
      event.preventDefault();
      toggle();
      return;
    }
    // iOS touchend 之後通常還會補一個 click；這個 click 必須忽略，否則會立刻開了又關。
    if(event.type==="click" && now-state.lastTouchAt<800) return;
    event.preventDefault();
    toggle();
  }
  // iPhone/iPad 用 touchend；桌機／其他裝置用 click。事件委派可承受 Streamlit 重畫 DOM。
  doc.addEventListener("touchend", registerTap, {capture:true, passive:false});
  doc.addEventListener("click", registerTap, true);
  doc.addEventListener("contextmenu", event=>{
    const target=event.target && event.target.closest ? event.target.closest("#jarvis-secret-trigger") : null;
    if(target) event.preventDefault();
  }, true);
  win[GLOBAL_KEY]=state;
})();
</script></body></html>'''


def render_jarvis_trigger_bootstrap() -> None:
    """頁面一載入就掛上賈維斯單擊入口；不等待 GPS、推薦或語音元件。"""
    components.html(JARVIS_TRIGGER_BOOTSTRAP_HTML, height=1, scrolling=False)


JARVIS_BROWSER_COMPONENT_HTML = r'''<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<style>html,body{width:1px;height:1px;margin:0;overflow:hidden;background:transparent}</style></head><body>
<script>
(() => {
  __LOW_BATTERY_CLIENT_CORE__
  const API_VERSION = 1;
  const win = window.parent;
  const doc = win.document;
  const service = ensureUbikeBatteryService(win);
  const STORAGE_KEY = "taitung-jarvis-enabled-v1";
  let args = {};
  let recognition = null;
  let listening = false;
  let enabled = false;
  let speaking = false;
  let lastSpeech = "";
  let lastHeard = "";
  let wakeWindowUntil = 0;
  let awaitingActual = false;
  let awaitingConfirm = false;
  let pendingActual = null;
  let currentPlan = null;
  let candidates = [];
  let contextStatus = "updating";
  let contextMessage = "";
  let lastAutoAnnounceToken = "";
  let pressTimer = null;

  function send(type, data = {}) {
    win.postMessage({isStreamlitMessage:true, type, ...data}, "*");
  }
  function setHeight(){ send("streamlit:setFrameHeight", {height:1}); }
  function emit(type, detail = {}) {
    send("streamlit:setComponentValue", {
      value: {type, event_id:`${Date.now()}-${Math.random().toString(16).slice(2)}`, ...detail},
      dataType:"json",
    });
  }
  function emitAction(type, detail = {}) {
    // Streamlit rerun 前先把動作送出；並把去重狀態放在 parent window，避免元件重建後同一句又執行一次。
    const signature=JSON.stringify([type,text(detail.station_name),text(detail.target_text),
      num(detail.unload_bike),num(detail.unload_ebike),num(detail.pickup_bike),num(detail.pickup_ebike),text(detail.heard)]);
    const now=Date.now();
    let last=null; try{last=win.__taitungJarvisLastAction||null;}catch(_){}
    if(last && last.signature===signature && now-Number(last.at||0)<2500){
      setIndicator("已忽略重複口令");
      return false;
    }
    try{win.__taitungJarvisLastAction={signature,at:now};}catch(_){}
    emit(type,detail);
    return true;
  }
  function text(value){ return String(value ?? "").trim(); }
  function num(value){ const n=Number(value); return Number.isFinite(n)?Math.max(0,Math.trunc(n)):0; }

  // V28.8：語音動作先直接控制目前 Streamlit 畫面。
  // 這條 DOM bridge 不依賴 custom component value 是否觸發 fragment rerun，
  // 因此手機 Safari 上也能立刻看到按鈕被按下或數字欄位被改變。
  function uiText(value){ return text(value).replace(/[\s\uFE0F]/g,"").replace(/[✅⏭️❌🧭]/g,""); }
  function elementUsable(node){
    if(!node || node.disabled) return false;
    const style=win.getComputedStyle ? win.getComputedStyle(node) : null;
    if(style && (style.display==="none" || style.visibility==="hidden")) return false;
    return true;
  }
  function findButtonByText(labels){
    const wanted=(Array.isArray(labels)?labels:[labels]).map(uiText).filter(Boolean);
    const buttons=Array.from(doc.querySelectorAll("button"));
    for(const button of buttons){
      if(!elementUsable(button)) continue;
      const label=uiText(button.innerText || button.textContent || "");
      if(wanted.some(item=>label===item || label.includes(item))) return button;
    }
    return null;
  }
  function clickUiButton(labels){
    const button=findButtonByText(labels);
    if(!button) return false;
    try{ button.scrollIntoView({block:"center",behavior:"auto"}); }catch(_){}
    try{ button.focus({preventScroll:true}); }catch(_){}
    button.click();
    return true;
  }
  function findNumberInputByLabel(labelText){
    const wanted=uiText(labelText);
    const groups=Array.from(doc.querySelectorAll('[data-testid="stNumberInput"]'));
    for(const group of groups){
      const label=group.querySelector("label");
      const labelValue=uiText(label ? (label.innerText || label.textContent || "") : "");
      if(labelValue===wanted){
        const input=group.querySelector('input[type="number"], input');
        if(input) return input;
      }
    }
    // Streamlit DOM 結構若改版，退回以鄰近文字精確比對。
    for(const input of Array.from(doc.querySelectorAll('input[type="number"]'))){
      const group=input.closest('[data-testid="stNumberInput"]') || input.parentElement?.parentElement;
      const content=uiText(group ? (group.innerText || group.textContent || "") : "");
      if(content.startsWith(wanted)) return input;
    }
    return null;
  }
  function setNumberInputValue(labelText,value){
    const input=findNumberInputByLabel(labelText);
    if(!input) return false;
    const next=String(num(value));
    try{ input.focus({preventScroll:true}); }catch(_){}
    try{
      const ownSetter=Object.getOwnPropertyDescriptor(input,"value")?.set;
      const protoSetter=Object.getOwnPropertyDescriptor(win.HTMLInputElement.prototype,"value")?.set;
      (protoSetter || ownSetter)?.call(input,next);
    }catch(_){ input.value=next; }
    input.dispatchEvent(new win.Event("input",{bubbles:true}));
    input.dispatchEvent(new win.Event("change",{bubbles:true}));
    try{ input.blur(); }catch(_){}
    return uiText(input.value)===uiText(next) || String(input.value)===next;
  }
  function applyActualToUi(actual){
    const values=[
      ["實際下車 2.0",actual.unload_bike],
      ["實際上車 2.0",actual.pickup_bike],
      ["實際下車 2.0E",actual.unload_ebike],
      ["實際上車 2.0E",actual.pickup_ebike],
    ];
    let applied=0;
    for(const [label,value] of values){ if(setNumberInputValue(label,value)) applied+=1; }
    return applied===values.length;
  }
  function cn(value) {
    const n=num(value);
    const digits=["零","一","二","三","四","五","六","七","八","九"];
    if(n<10) return digits[n];
    if(n===10) return "十";
    if(n<20) return `十${digits[n-10]}`;
    if(n<100) return `${digits[Math.floor(n/10)]}十${n%10?digits[n%10]:""}`;
    return String(n);
  }
  function countPart(n, kind){ return num(n)>0 ? `${cn(n)}${kind}` : ""; }
  function actionPhrase(plan) {
    if(!plan) return "";
    const down=[countPart(plan.unload_bike,"般"),countPart(plan.unload_ebike,"電")].filter(Boolean).join("");
    const up=[countPart(plan.pickup_bike,"般"),countPart(plan.pickup_ebike,"電")].filter(Boolean).join("");
    const parts=[];
    if(down) parts.push(`下${down}`);
    if(up) parts.push(`上${up}`);
    return parts.join("，") || "目前無需上下車";
  }
  function stationPhrase(plan){ return plan ? `${text(plan.station_name)}，${actionPhrase(plan)}` : "目前沒有可執行場站"; }
  function currentStation(){ return text((currentPlan||{}).station_name || (candidates[0]||{}).station_name); }
  function contextReply(){
    if(contextMessage) return contextMessage;
    if(contextStatus==="updating") return "調度資料更新中，請稍後再說一次";
    if(contextStatus==="unavailable") return "調度資料目前無法使用，請稍後更新";
    if(contextStatus==="blocked") return "目前條件無法執行智慧調度，請先修正畫面提示";
    if(contextStatus==="no_candidates") return "目前沒有可執行場站";
    return "";
  }
  function contextReady(){ return contextStatus==="ready"; }
  function standbyLabel(){
    const station=currentStation();
    if(contextStatus==="ready" && station) return `賈維斯待命｜${station}`;
    if(contextStatus==="no_candidates") return "賈維斯待命｜暫無可執行場站";
    if(contextStatus==="blocked") return "賈維斯待命｜請先修正調度條件";
    if(contextStatus==="unavailable") return "賈維斯待命｜調度服務暫不可用";
    return "賈維斯待命｜調度資料更新中";
  }
  function isIOS(){ return /iPad|iPhone|iPod/.test(win.navigator.userAgent||"") || (win.navigator.platform==="MacIntel" && Number(win.navigator.maxTouchPoints)>1); }
  function isStandalone(){ try{return Boolean(win.navigator.standalone) || win.matchMedia("(display-mode: standalone)").matches;}catch(_){return false;} }
  function preferredZhVoice(){
    try{
      const voices=win.speechSynthesis?.getVoices?.() || [];
      return voices.find(v=>/^zh-TW$/i.test(v.lang)) || voices.find(v=>/^zh/i.test(v.lang)) || null;
    }catch(_){return null;}
  }

  function ensureStyle() {
    if(doc.getElementById("jarvis-voice-style")) return;
    const style=doc.createElement("style");
    style.id="jarvis-voice-style";
    style.textContent=`
      #jarvis-secret-trigger{touch-action:manipulation;user-select:none;-webkit-user-select:none;cursor:pointer}
      #jarvis-voice-indicator{position:fixed;right:12px;top:max(72px,calc(env(safe-area-inset-top,0px) + 58px));bottom:auto;z-index:2147483000;display:none;
        align-items:center;gap:7px;padding:7px 10px;border:1px solid rgba(85,246,255,.35);border-radius:999px;
        background:rgba(4,16,29,.88);color:#dffcff;box-shadow:0 8px 24px rgba(0,0,0,.24);
        font:800 12px/1.2 -apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft JhengHei",sans-serif;backdrop-filter:blur(8px)}
      #jarvis-voice-indicator.on{display:flex} #jarvis-voice-indicator .dot{width:8px;height:8px;border-radius:50%;background:#55f6ff;box-shadow:0 0 10px #55f6ff}
      #jarvis-voice-indicator.busy .dot{background:#ff5da2;box-shadow:0 0 12px #ff5da2}
    `;
    doc.head.appendChild(style);
  }
  function ensureIndicator() {
    ensureStyle();
    let node=doc.getElementById("jarvis-voice-indicator");
    if(!node){ node=doc.createElement("div"); node.id="jarvis-voice-indicator"; node.innerHTML='<span class="dot"></span><span class="label">賈維斯待命</span>'; doc.body.appendChild(node); }
    node.classList.toggle("on",enabled);
    return node;
  }
  function setIndicator(label,busy=false){ const node=ensureIndicator(); node.querySelector(".label").textContent=label; node.classList.toggle("busy",busy); }

  function normalizeSpeechText(value){
    let s=text(value);
    try{s=s.normalize("NFKC");}catch(_){}
    return s
      .replace(/[，,。.!！?？；;：:]/g," ")
      .replace(/\s+/g," ")
      .trim();
  }
  const WAKE_RE=/(?:賈維斯|贾维斯|假維斯|假维斯|加維斯|加维斯|甲維斯|甲维斯|佳維斯|佳维斯|嘉維斯|嘉维斯|賈偉斯|贾伟斯|賈威斯|贾威斯|jarvis)/i;
  function wakeVariants(t){ return WAKE_RE.test(normalizeSpeechText(t)); }
  function stripWake(t){ return normalizeSpeechText(t).replace(new RegExp(`(?:嘿|嗨|hey)?\\s*${WAKE_RE.source}[\\s，,。.!！?？]*`,"ig"),"").trim(); }
  function normalizeActualText(value){
    const nums="零〇一二兩两三四五六七八九十百\\d";
    let s=normalizeSpeechText(value)
      .replace(/(?:麻煩|麻烦)?(?:幫我|帮我)?(?:輸入|输入|填入|填上|設定|设定|記錄|记录)/g," ")
      .replace(/上車|上车/g,"上")
      .replace(/下車|下车/g,"下")
      .replace(/一般/g,"一 般")
      .replace(/兩般|两般/g,"二般")
      .replace(/兩電|两電|兩电|两电/g,"二電");
    s=s.replace(new RegExp(`([${nums}]+)\\s*(?:台)?\\s*(?:班|搬|斑)`,"g"),"$1般");
    s=s.replace(new RegExp(`([${nums}]+)\\s*(?:台)?\\s*(?:电|點|点|店|殿)`,"g"),"$1電");
    return s.replace(/[和跟與与及、]/g," ").replace(/\s+/g," ").trim();
  }
  function commandScore(value, confidence=0){
    const raw=normalizeSpeechText(value); const normalized=normalizeActualText(raw);
    let score=Math.max(0,Number(confidence)||0)*5;
    if(wakeVariants(raw)) score+=100;
    if(/(?:輸入|输入|填入|設定|设定|記錄|记录)/.test(raw)) score+=45;
    if(/分析|下一站|抵達|抵达|完成|確認|确认|跳過|跳过|電池|电池|低電|低电|重複|重复|口令|指令|測試|测试/.test(raw)) score+=35;
    if(/[上下]/.test(normalized)) score+=20;
    if(/般|電/.test(normalized)) score+=25;
    if(/[零〇一二兩两三四五六七八九十百\d]/.test(normalized)) score+=15;
    if((awaitingActual||awaitingConfirm) && /確認|确认|完成|取消|重來|重来|[上下]/.test(normalized)) score+=80;
    return score;
  }
  function pickBestTranscript(result){
    let best=""; let bestScore=-1;
    const count=Math.max(1,Math.min(Number(result?.length)||1,5));
    for(let j=0;j<count;j+=1){
      const alt=result[j]; const transcript=text(alt?.transcript);
      if(!transcript) continue;
      const score=commandScore(transcript,alt?.confidence);
      if(score>bestScore){bestScore=score;best=transcript;}
    }
    return best || text(result?.[0]?.transcript);
  }

  function stopRecognition(){
    if(recognition && listening){ try{ recognition.stop(); }catch(_){} }
    listening=false;
  }
  function startRecognition(){
    if(!enabled || speaking || listening) return;
    if(win.speechSynthesis && win.speechSynthesis.speaking){ win.setTimeout(startRecognition,220); return; }
    const R=win.SpeechRecognition || win.webkitSpeechRecognition;
    if(!R){ setIndicator(isStandalone()?"請改用 Safari 開啟語音":"瀏覽器不支援語音辨識",true); return; }
    if(isStandalone()){ setIndicator("主畫面 App 可能無法語音，建議用 Safari",true); }
    if(!recognition){
      recognition=new R(); recognition.lang="zh-TW"; recognition.continuous=false; recognition.interimResults=false; recognition.maxAlternatives=5;
      recognition.onstart=()=>{listening=true;setIndicator(standbyLabel());};
      recognition.onend=()=>{listening=false;if(enabled&&!speaking)win.setTimeout(startRecognition,isIOS()?220:350);};
      recognition.onerror=(event)=>{
        listening=false; const e=text(event?.error);
        if(e==="not-allowed"||e==="service-not-allowed") setIndicator("請開啟麥克風／Siri 聽寫權限",true);
        else if(e==="audio-capture") setIndicator("找不到可用麥克風",true);
        else if(e==="network") setIndicator("語音辨識網路異常",true);
        else if(enabled) win.setTimeout(startRecognition,isIOS()?350:700);
      };
      recognition.onresult=(event)=>{
        for(let i=event.resultIndex;i<event.results.length;i+=1){
          if(!event.results[i].isFinal) continue;
          const transcript=pickBestTranscript(event.results[i]);
          if(transcript) handleTranscript(transcript);
        }
      };
    }
    try{recognition.start();}catch(_){}
  }
  function speak(message, after=null){
    const phrase=text(message); if(!phrase) return;
    lastSpeech=phrase; speaking=true; stopRecognition();
    win.speechSynthesis.cancel();
    const utterance=new win.SpeechSynthesisUtterance(phrase); utterance.lang="zh-TW"; utterance.rate=1.03;
    const voice=preferredZhVoice(); if(voice) utterance.voice=voice;
    utterance.onstart=()=>setIndicator("賈維斯回覆中",true);
    utterance.onend=()=>{speaking=false;setIndicator(standbyLabel()); if(typeof after==="function") after(); win.setTimeout(startRecognition,250);};
    utterance.onerror=()=>{speaking=false;setIndicator(standbyLabel());if(typeof after==="function") after();win.setTimeout(startRecognition,350);};
    win.speechSynthesis.speak(utterance);
  }

  async function batteryInfo(stationName){
    if(!stationName) return {low:[],ultra:[]};
    try{
      const result=await service.queryStationByName(stationName,{attempts:2,timeoutMs:12000});
      const threshold=Math.max(0,Math.min(100,num(args.threshold)));
      const priority=Math.max(0,Math.min(threshold,num(args.priority_threshold)));
      const bikes=Array.isArray(result?.bikes)?result.bikes:[];
      return {low:bikes.filter(x=>Number.isFinite(x.battery_power)&&x.battery_power<=threshold), ultra:bikes.filter(x=>Number.isFinite(x.battery_power)&&x.battery_power<=priority)};
    }catch(error){ return {low:[],ultra:[],error:text(error?.message||error)}; }
  }
  function batterySuffix(info){
    if(!info || info.error || !info.low?.length) return "";
    if(info.ultra?.length) return `，低電${cn(info.low.length)}台，其中他媽超低電${cn(info.ultra.length)}台`;
    return `，低電${cn(info.low.length)}台`;
  }
  async function speakOne(plan){
    if(!contextReady()){speak(contextReply());return;}
    if(!plan){speak("目前沒有可執行場站");return;}
    setIndicator("正在查電量",true);
    const info=await batteryInfo(text(plan.station_name));
    speak(`${stationPhrase(plan)}${batterySuffix(info)}`);
  }
  async function speakAnalysis(){
    if(!contextReady()){speak(contextReply());return;}
    if(!candidates.length){speak("目前沒有可執行場站");return;}
    setIndicator("正在分析",true);
    const infos=[];
    for(const plan of candidates){ infos.push(await batteryInfo(text(plan.station_name))); }
    const phrases=candidates.map((plan,index)=>`${stationPhrase(plan)}${batterySuffix(infos[index])}`);
    speak(phrases.join("。"));
  }
  async function speakUltra(){
    if(!contextReady()){speak(contextReply());return;}
    const station=currentStation(); if(!station){speak("目前沒有目標場站");return;}
    setIndicator("正在查超低電",true); const info=await batteryInfo(station);
    if(info.error){speak("超低電資料目前讀取失敗");return;}
    if(!info.ultra.length){speak(`${station}，沒有超低電`);return;}
    const pillars=info.ultra.map(x=>text(x.pillar_no)).filter(Boolean);
    speak(`${station}，有${cn(info.ultra.length)}台超低電${pillars.length?`，柱號${pillars.join("、")}`:""}`);
  }
  async function speakBattery(){
    if(!contextReady()){speak(contextReply());return;}
    const station=currentStation(); if(!station){speak("目前沒有目標場站");return;}
    setIndicator("正在查電池",true); const info=await batteryInfo(station);
    if(info.error){speak("電池資料目前讀取失敗");return;}
    if(!info.low.length){speak(`${station}，沒有低電`);return;}
    const lowPillars=info.low.map(x=>text(x.pillar_no)).filter(Boolean);
    const ultraPillars=info.ultra.map(x=>text(x.pillar_no)).filter(Boolean);
    let message=`${station}，低電${cn(info.low.length)}台${lowPillars.length?`，柱號${lowPillars.join("、")}`:""}`;
    if(info.ultra.length) message+=`。其中${ultraPillars.length?`${ultraPillars.join("、")}號柱`:cn(info.ultra.length)+"台"}他媽超低電`;
    speak(message);
  }

  function parseCnNumber(token){
    const raw=text(token); if(/^\d+$/.test(raw)) return Number(raw);
    const map={"零":0,"〇":0,"一":1,"二":2,"兩":2,"两":2,"三":3,"四":4,"五":5,"六":6,"七":7,"八":8,"九":9};
    if(raw==="十") return 10;
    if(raw.includes("十")){
      const [a,b]=raw.split("十"); const tens=a?map[a]??0:1; const ones=b?map[b]??0:0; return tens*10+ones;
    }
    return map[raw] ?? null;
  }
  function emptyActual(){ return {unload_bike:0,unload_ebike:0,pickup_bike:0,pickup_ebike:0}; }
  function plannedActual(){ return currentPlan ? {unload_bike:num(currentPlan.unload_bike),unload_ebike:num(currentPlan.unload_ebike),pickup_bike:num(currentPlan.pickup_bike),pickup_ebike:num(currentPlan.pickup_ebike)} : emptyActual(); }
  function parseActual(raw, allowDirectionFallback=false){
    const normalized=normalizeActualText(raw);
    if(/照原本|照計畫|照计划|一樣|一样/.test(normalized)) return plannedActual();
    if(/都沒有|都没有|零台|沒上下|没上下|清空|歸零|归零/.test(normalized)) return emptyActual();
    const result=emptyActual(); let matched=false; let ambiguous=false;
    const plan=plannedActual();
    const planOnlyDown=(plan.unload_bike+plan.unload_ebike)>0 && (plan.pickup_bike+plan.pickup_ebike)===0;
    const planOnlyUp=(plan.pickup_bike+plan.pickup_ebike)>0 && (plan.unload_bike+plan.unload_ebike)===0;
    const sections=[]; const sectionRe=/([上下])([^上下]*)/g; let sectionMatch;
    while((sectionMatch=sectionRe.exec(normalized))!==null) sections.push({direction:sectionMatch[1]==="上"?"up":"down",body:sectionMatch[2]});
    if(!sections.length){
      let fallback="";
      if(allowDirectionFallback) fallback=planOnlyDown?"down":planOnlyUp?"up":"";
      if(!fallback) return null;
      sections.push({direction:fallback,body:normalized});
    }
    for(const section of sections){
      const re=/([零〇一二兩两三四五六七八九十百\d]+)\s*(?:台)?\s*(般|電)/g; let m; let sectionMatches=0;
      while((m=re.exec(section.body))!==null){
        const n=parseCnNumber(m[1]); if(n===null) continue;
        matched=true; sectionMatches+=1;
        const bike=m[2]==="般";
        if(section.direction==="down") result[bike?"unload_bike":"unload_ebike"]+=n;
        else result[bike?"pickup_bike":"pickup_ebike"]+=n;
      }
      // 有數字／車種字樣卻沒有被完整解析時，視為模糊，不猜。
      const residue=section.body
        .replace(/([零〇一二兩两三四五六七八九十百\d]+)\s*(?:台)?\s*(般|電)/g," ")
        .replace(/(?:台|車|车|輛|辆|要|各|共|總共|总共|然後|然后|再|再來|再来|還有|还有|和|跟|與|与|及|的|請|请|幫我|帮我)/g," ")
        .replace(/\s+/g,"")
        .trim();
      if(/[零〇一二兩两三四五六七八九十百\d般電班电點点店]/.test(residue) || (!sectionMatches && section.body.trim())) ambiguous=true;
    }
    if(!matched || ambiguous) return null;
    return result;
  }
  function actualPhrase(actual){ return actionPhrase(actual); }

  function handleActualInput(command){
    if(awaitingConfirm){
      if(/^(確認|确认|是|對|对|完成|可以|沒錯|没错)/.test(command)){
        const actual=pendingActual||emptyActual(); awaitingConfirm=false; awaitingActual=false; pendingActual=null;
        // 優先直接按下原本的 Streamlit 表單完成鈕，讓畫面原生流程接手。
        if(clickUiButton(["完成本站並安排下一站"])) {
          setIndicator("正在完成本站",true);
        } else {
          emitAction("confirm_completion", {station_name:currentStation(), ...actual, heard:"確認"});
          speak("找不到完成按鈕，已改用後端同步");
        }
        return true;
      }
      if(/取消|重來|重来/.test(command)){ awaitingConfirm=false; awaitingActual=true; pendingActual=null; speak(`${currentStation()}，實際上下多少？`); return true; }
      return false;
    }
    if(awaitingActual){
      const actual=parseActual(command,true);
      if(!actual){ speak("我沒有把數量聽清楚。請說：上幾般幾電，或下幾般幾電"); return true; }
      pendingActual=actual; awaitingActual=false; awaitingConfirm=true;
      const applied=applyActualToUi(actual);
      if(applied){
        setIndicator("畫面已填入，等待確認");
        speak(`${currentStation()}，${actualPhrase(actual)}，畫面已填入，請確認`);
      } else {
        emitAction("fill_actual", {station_name:currentStation(), ...actual, heard:command});
        speak(`${currentStation()}，${actualPhrase(actual)}，畫面欄位沒有找到，已改用後端同步，請確認`);
      }
      return true;
    }
    return false;
  }

  function targetTextFromChange(command){
    return command.replace(/^(?:幫我)?(?:改往|改去|換去|换去|前往|去)\s*/,"").replace(/[。.!！?？]$/g,"").trim();
  }
  function commandHelp(){
    speak("固定口令：賈維斯分析。賈維斯下一站。賈維斯抵達。要改畫面數字說，賈維斯輸入上一般兩電，或賈維斯輸入下兩般一電。賈維斯完成。賈維斯跳過。賈維斯電池。賈維斯超低電。賈維斯重複。");
  }
  function handleTranscript(raw){
    if(!enabled || speaking) return;
    const transcript=normalizeSpeechText(raw);
    if(!transcript) return;
    const previousHeard=lastHeard;
    lastHeard=transcript;
    setIndicator(`聽到｜${transcript.slice(0,18)}`,true);

    // 賈維斯剛主動問「實際上下多少／請確認」時，下一句可直接回答，不必再喊喚醒詞。
    if(awaitingActual || awaitingConfirm){
      const followup=wakeVariants(transcript)?stripWake(transcript):normalizeActualText(transcript);
      if(handleActualInput(followup)) return;
    }

    const hasWake=wakeVariants(transcript);
    // V28.7：除「實際上下多少／請確認」的緊接回答外，所有操作都必須完整喊「賈維斯＋口令」。
    if(!hasWake){ setIndicator(standbyLabel()); return; }
    const command=stripWake(transcript);
    if(!command){ speak("我在。請把口令一起說，例如，賈維斯分析"); return; }

    if(/^(?:口令|指令|怎麼說|怎么说|可以說什麼|可以说什么)/.test(command)){ commandHelp(); return; }
    if(/你聽到什麼|你听到什么|剛剛聽到什麼|刚刚听到什么/.test(command)){ speak(previousHeard?`上一句我聽到，${previousHeard}`:"目前還沒有上一句指令"); return; }
    if(/^(?:測試|测试|語音測試|语音测试)$/.test(command)){ speak(contextReady()?`語音正常。目前場站，${currentStation()||"尚未鎖定"}`:contextReply()); return; }
    if(/請?重複|请?重复|再說一次|再说一次|剛剛說什麼|刚刚说什么/.test(command)){ speak(lastSpeech||"目前沒有可重複的內容"); return; }
    if(/有沒有超低電|有没有超低电|超低電有沒有|超低电有没有|^超低電$|^超低电$/.test(command)){ speakUltra(); return; }
    if(/電池|电池|低電|低电/.test(command)){ speakBattery(); return; }
    if(/忘了|忘記|忘记|這裡.*(?:上下|幾台|多少)|这里.*(?:上下|几台|多少)|(?:上下|作業|作业).*(?:多少|幾台|几台)|上多少|下多少|剛剛.*(?:上|下)|刚刚.*(?:上|下)/.test(command)){ if(!contextReady()){speak(contextReply());return;} speak(stationPhrase(currentPlan||(candidates[0]||null))); return; }

    // 會改動表單的語音，優先使用固定口令：「賈維斯，輸入，上1般2電」。
    const explicitInput=/^(?:麻煩|麻烦)?(?:幫我|帮我)?(?:輸入|输入|填入|填上|設定|设定|記錄|记录)\s*/.test(command);
    const actualCommand=normalizeActualText(command);
    if(explicitInput){
      const actual=parseActual(command,false);
      if(!actual){
        speak("數量沒有聽清楚，沒有執行。請說：賈維斯，輸入，上一般兩電。或：賈維斯，輸入，下兩般一電");
        return;
      }
      if(!contextReady()){ speak(contextReply()); return; }
      const station=currentStation();
      if(!station){ speak("調度資料更新中，請稍後再說一次"); return; }
      if(text(args.mode)!=="active"){ speak(`${station}，請先說賈維斯抵達，鎖定場站後再輸入數量`); return; }
      const phrase=actualPhrase(actual);
      if(applyActualToUi(actual)){
        setIndicator("畫面已填入");
        speak(`${station}，${phrase}，畫面已填入，請確認`);
      } else if(emitAction("fill_actual", {station_name:station, ...actual, heard:transcript})) {
        setIndicator("畫面欄位未找到，改用後端同步",true);
        speak(`${station}，${phrase}，畫面欄位沒有找到，已改用後端同步`);
      }
      return;
    }
    if(!explicitInput && /[上下]/.test(actualCommand) && /(般|電)/.test(actualCommand)){
      speak("這是會修改數字的動作。請用固定口令：賈維斯，輸入，上一般兩電。或：賈維斯，輸入，下兩般一電");
      return;
    }
    if(/[上下].*(?:\d+|[一二兩两三四五六七八九十]).*台/.test(actualCommand) && !/(般|電)/.test(actualCommand)){
      speak("沒有執行。請把車種一起說，例如：賈維斯，輸入，上兩般一電");
      return;
    }
    if(/^(?:清除輸入|清除输入|上下歸零|上下归零)$/.test(command)){
      if(!contextReady()){speak(contextReply());return;}
      if(text(args.mode)!=="active"){speak("請先鎖定目標場站");return;}
      const station=currentStation(); const actual=emptyActual();
      if(applyActualToUi(actual)){
        speak(`${station}，上下車數量已在畫面歸零`);
      } else if(emitAction("fill_actual",{station_name:station,...actual,heard:transcript})) {
        speak(`${station}，畫面欄位沒有找到，已改用後端同步歸零`);
      }
      return;
    }
    if(/完成|做完|處理完|处理完/.test(command)){
      if(!contextReady()){speak(contextReply());return;}
      if(!currentStation()){speak("目前沒有目標場站");return;}
      awaitingActual=true; awaitingConfirm=false; pendingActual=null; speak(`${currentStation()}，實際上下多少？`); return;
    }
    if(/我到了|已抵達|已抵达|抵達|抵达|到目的地|到站了|我到站|^到了$/.test(command)){
      if(!contextReady()){speak(contextReply());return;}
      const station=currentStation(); if(!station){speak("目前沒有目標場站");return;}
      if(text(args.mode)==="candidate") {
        if(clickUiButton(["前往此站"])) {
          setIndicator("正在鎖定場站",true);
          speak(`${station}，已抵達，畫面正在鎖定場站`);
        } else {
          emitAction("lock_station",{station_name:station,heard:transcript});
          speak(`${station}，已抵達，找不到畫面按鈕，已改用後端鎖定`);
        }
      } else speak(`${station}，已抵達`);
      return;
    }
    if(/有人去|有人處理|有人处理|換下一站|换下一站|改往其他|跳過|跳过|不要去/.test(command)){
      if(!contextReady()){speak(contextReply());return;}
      const labels=text(args.mode)==="active" ? ["取消配置"] : ["跳過並找下一站"];
      if(clickUiButton(labels)){
        setIndicator("正在切換下一站",true);
      } else {
        emitAction("skip_current",{station_name:currentStation(),reason:/有人/.test(command)?"有人已前往":"使用者語音改站",heard:transcript});
      }
      return;
    }
    if(/^(?:改往|改去|換去|换去|前往|去)/.test(command)){
      if(!contextReady()){speak(contextReply());return;}
      const target=targetTextFromChange(command); emitAction("change_station",{station_name:currentStation(),target_text:target,heard:transcript}); return;
    }
    if(/分析|全部|候選|候选/.test(command)){ speakAnalysis(); return; }
    if(/下一站|去哪|往哪|該往|该往|前往哪/.test(command)){ speakOne(currentPlan||(candidates[0]||null)); return; }
    speak(`我聽到，${command}，但沒有執行。你可以說，賈維斯口令，聽可用指令`);
  }

  function toggleEnabled(){
    enabled=!enabled; try{win.sessionStorage.setItem(STORAGE_KEY,enabled?"1":"0");}catch(_){}
    ensureIndicator();
    if(enabled){ setIndicator(standbyLabel()); startRecognition(); }
    else { stopRecognition(); win.speechSynthesis.cancel(); setIndicator("賈維斯已關閉"); ensureIndicator().classList.remove("on"); }
  }
  win.addEventListener("taitung:jarvis-enabled-change", event=>{
    enabled=Boolean(event?.detail?.enabled);
    ensureIndicator();
    if(enabled){ setIndicator(standbyLabel()); startRecognition(); }
    else { stopRecognition(); win.speechSynthesis.cancel(); setIndicator("賈維斯已關閉"); ensureIndicator().classList.remove("on"); }
  });
  function hydrate(){
    currentPlan=args.current_plan&&typeof args.current_plan==="object"?args.current_plan:null;
    candidates=Array.isArray(args.candidates)?args.candidates:[];
    contextStatus=text(args.context_status||"updating").toLowerCase();
    contextMessage=text(args.context_message||"");
    try{enabled=win.sessionStorage.getItem(STORAGE_KEY)==="1";}catch(_){enabled=false;}
    ensureIndicator(); if(enabled){ setIndicator(standbyLabel()); startRecognition(); }
    const token=text(args.auto_announce_token);
    if(enabled && contextReady() && args.auto_announce && token && token!==lastAutoAnnounceToken){ lastAutoAnnounceToken=token; win.setTimeout(()=>speakOne(currentPlan||(candidates[0]||null)),450); }
    setHeight();
  }

  win.addEventListener("message", event=>{
    if(!event.data || event.data.type!=="streamlit:render") return;
    args=event.data.args||{}; hydrate();
  });
  send("streamlit:componentReady",{apiVersion:API_VERSION}); setHeight();
})();
</script></body></html>'''

_JARVIS_VOICE_COMPONENT = None


def get_jarvis_voice_component():
    """建立隱藏式賈維斯語音元件；由「測試版」單擊切換啟用狀態。"""
    global _JARVIS_VOICE_COMPONENT
    if _JARVIS_VOICE_COMPONENT is not None:
        return _JARVIS_VOICE_COMPONENT
    component_dir = Path(tempfile.gettempdir()) / "taitung_jarvis_voice_component_v6"
    component_dir.mkdir(parents=True, exist_ok=True)
    index_path = component_dir / "index.html"
    content = JARVIS_BROWSER_COMPONENT_HTML.replace("__LOW_BATTERY_CLIENT_CORE__", LOW_BATTERY_CLIENT_CORE_JS)
    try:
        if not index_path.exists() or index_path.read_text(encoding="utf-8") != content:
            index_path.write_text(content, encoding="utf-8")
    except OSError as exc:
        raise RuntimeError(f"無法建立賈維斯語音元件：{exc}") from exc
    _JARVIS_VOICE_COMPONENT = components.declare_component("taitung_jarvis_voice_v4", path=str(component_dir))
    return _JARVIS_VOICE_COMPONENT



@st.cache_data(show_spinner=False, max_entries=192)
def _build_inline_low_battery_pillars_html(
    station_specs: tuple[tuple[str, str, str], ...],
    threshold: int,
    priority_threshold: int,
    mobile_mode: bool,
    *,
    auto_query: bool,
    force_station: str = "",
) -> str:
    """建立主頁柱號查詢元件；智慧調度自動查，一般分析按站查。"""
    specs = [
        {"name": name, "kind": kind, "target": target}
        for name, kind, target in station_specs
        if str(name).strip()
    ]
    specs_payload = json.dumps(specs, ensure_ascii=False).replace("</", "<\\/")
    display_mode = json.dumps("mobile" if mobile_mode else "desktop")
    force_payload = json.dumps(str(force_station or ""), ensure_ascii=False)
    fingerprint = json.dumps(
        hashlib.sha1(
            (
                f"{APP_VERSION}|{threshold}|{priority_threshold}|{auto_query}|"
                f"{display_mode}|{force_station}|{specs_payload}"
            ).encode("utf-8")
        ).hexdigest()
    )
    component_html = r"""
    <script>
    (() => {
      __LOW_BATTERY_CLIENT_CORE__
      const specs = __STATION_SPECS__;
      const threshold = __THRESHOLD__;
      const priorityThreshold = __PRIORITY_THRESHOLD__;
      const autoQuery = __AUTO_QUERY__;
      const forceStation = __FORCE_STATION__;
      const displayMode = __DISPLAY_MODE__;
      const fingerprint = __FINGERPRINT__;
      const doc = window.parent.document;
      const win = window.parent;
      const service = ensureUbikeBatteryService(win);
      const requestIsMobile = displayMode === "mobile"
        || win.matchMedia("(max-width: 700px)").matches;
      const generation = `${fingerprint}-${Date.now()}-${Math.random().toString(16).slice(2)}`;
      win.__ubikeInlineBatteryGeneration = generation;

      doc.querySelectorAll(".ubike-inline-battery").forEach(node => node.remove());
      doc.getElementById("ubike-inline-battery-style")?.remove();
      const style = doc.createElement("style");
      style.id = "ubike-inline-battery-style";
      style.textContent = `
        .ubike-inline-battery {
          display:flex; align-items:center; flex-wrap:wrap; gap:5px;
          margin-top:7px; padding:7px 8px; border:1px solid rgba(85,246,255,.28);
          border-radius:10px; color:#a9c9d3; background:rgba(4,16,29,.72);
          box-shadow:inset 3px 0 rgba(85,246,255,.42); font-size:12px; line-height:1.45;
        }
        .analysis-station-cell .ubike-inline-battery {max-width:100%;}
        .ubike-inline-battery .ubike-battery-label {color:#8feff5; font-weight:850;}
        .ubike-inline-battery .ubike-pillar {
          display:inline-flex; align-items:center; min-height:25px; padding:3px 7px;
          border:1px solid rgba(85,246,255,.35); border-radius:999px;
          color:#dffcff; background:rgba(85,246,255,.09); font-weight:900;
        }
        .ubike-inline-battery .ubike-pillar.urgent {
          color:#fff; border-color:#ff496c; background:linear-gradient(100deg,#d91d55,#ff3fcf);
          box-shadow:0 0 11px rgba(255,63,207,.38); animation:ubikePillarPulse 1.7s ease-in-out infinite;
        }
        .ubike-inline-battery.is-empty {color:#9cb5bd; box-shadow:inset 3px 0 rgba(126,249,255,.20);}
        .ubike-inline-battery.is-error {color:#ff91a9; border-color:rgba(255,73,108,.46); box-shadow:inset 3px 0 #ff496c;}
        .ubike-inline-battery button {
          min-height:29px; padding:4px 9px; border:1px solid rgba(85,246,255,.42);
          border-radius:8px; color:#061116; background:linear-gradient(105deg,#55f6ff,#f4ff57);
          font:850 12px/1.2 -apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft JhengHei",sans-serif;
          cursor:pointer;
        }
        .ubike-inline-battery .ubike-battery-refresh {
          min-width:29px; padding:3px 7px; margin-left:auto; color:#8feff5;
          background:rgba(85,246,255,.08);
        }
        @keyframes ubikePillarPulse {0%,100%{filter:brightness(1)}50%{filter:brightness(1.28)}}
        @media (max-width:700px) {
          .ubike-inline-battery {gap:4px; padding:6px 7px; font-size:11px;}
          .ubike-inline-battery .ubike-pillar {min-height:24px; padding:3px 6px;}
        }
      `;
      doc.head.appendChild(style);

      const wrappersByName = new Map();
      const normalizedSpecNames = new Set(specs.map(spec => service.normalizeStationName(spec.name)));
      function targetForSpec(spec) {
        if (spec.kind === "analysis") return doc.getElementById(spec.target);
        if (spec.kind === "candidate") {
          return Array.from(doc.querySelectorAll('[class*="st-key-candidate_card_select_"]'))
            .find(node => node.classList.contains(spec.target)) || null;
        }
        return null;
      }
      function registerWrapper(stationName, host, { manual = false } = {}) {
        if (!host || Array.from(host.children || []).some(
          child => child.classList?.contains("ubike-inline-battery"),
        )) return;
        const wrapper = doc.createElement("div");
        wrapper.className = "ubike-inline-battery";
        wrapper.dataset.stationName = stationName;
        if (manual) {
          const button = doc.createElement("button");
          button.type = "button";
          button.textContent = "⚡ 查低電量柱號";
          button.addEventListener("click", event => {
            event.preventDefault();
            event.stopPropagation();
            queryStation(stationName, true);
          });
          wrapper.appendChild(button);
        } else {
          const label = doc.createElement("span");
          label.className = "ubike-battery-label";
          label.textContent = "⚡ 柱號準備中…";
          wrapper.appendChild(label);
        }
        host.appendChild(wrapper);
        const normalized = service.normalizeStationName(stationName);
        if (!wrappersByName.has(normalized)) wrappersByName.set(normalized, []);
        wrappersByName.get(normalized).push(wrapper);
      }
      function attachTargets() {
        for (const spec of specs) {
          const target = targetForSpec(spec);
          if (!target) continue;
          const host = spec.kind === "analysis"
            ? (target.querySelector(".analysis-station-cell") || target)
            : target;
          registerWrapper(spec.name, host, { manual: !autoQuery });
        }
        for (const plan of doc.querySelectorAll(".dispatch-plan-card[data-ubike-station-name]")) {
          const stationName = String(plan.dataset.ubikeStationName || "").trim();
          if (!normalizedSpecNames.has(service.normalizeStationName(stationName))) continue;
          registerWrapper(stationName, plan, { manual: false });
        }
      }
      function wrappersFor(stationName) {
        return wrappersByName.get(service.normalizeStationName(stationName)) || [];
      }
      function setLoading(stationName) {
        for (const wrapper of wrappersFor(stationName)) {
          wrapper.classList.remove("is-error", "is-empty");
          wrapper.replaceChildren();
          const label = doc.createElement("span");
          label.className = "ubike-battery-label";
          label.textContent = "⚡ 正在讀取柱號…";
          wrapper.appendChild(label);
        }
      }
      function formatPillar(value) {
        const raw = String(value || "").trim();
        if (!raw) return "";
        const number = Number.parseInt(raw, 10);
        return Number.isFinite(number) ? String(number).padStart(2, "0") : raw;
      }
      function renderResult(stationName, result) {
        if (win.__ubikeInlineBatteryGeneration !== generation) return;
        const pillarMap = new Map();
        for (const bike of Array.isArray(result?.bikes) ? result.bikes : []) {
          if (!Number.isFinite(Number(bike.battery_power)) || Number(bike.battery_power) > threshold) continue;
          const pillar = formatPillar(bike.pillar_no);
          if (!pillar) continue;
          const previous = pillarMap.get(pillar);
          if (!previous || Number(bike.battery_power) < Number(previous.battery_power)) {
            pillarMap.set(pillar, bike);
          }
        }
        const lowBikes = Array.from(pillarMap.entries())
          .map(([pillar, bike]) => ({ pillar, battery_power: Number(bike.battery_power) }))
          .sort((left, right) => left.pillar.localeCompare(
            right.pillar,
            "zh-Hant",
            { numeric:true, sensitivity:"base" },
          ));
        for (const wrapper of wrappersFor(stationName)) {
          wrapper.replaceChildren();
          wrapper.classList.remove("is-error", "is-empty");
          const label = doc.createElement("span");
          label.className = "ubike-battery-label";
          if (!lowBikes.length) {
            wrapper.classList.add("is-empty");
            label.textContent = `目前沒有低於 ${threshold}% 的柱號`;
            wrapper.appendChild(label);
          } else {
            label.textContent = `≤${threshold}%：`;
            wrapper.appendChild(label);
            for (const bike of lowBikes) {
              const pillar = doc.createElement("span");
              pillar.className = "ubike-pillar";
              const urgent = bike.battery_power <= priorityThreshold;
              pillar.classList.toggle("urgent", urgent);
              pillar.textContent = `${urgent ? "⚠ " : ""}${bike.pillar}柱`;
              wrapper.appendChild(pillar);
            }
          }
          const refresh = doc.createElement("button");
          refresh.type = "button";
          refresh.className = "ubike-battery-refresh";
          refresh.textContent = "↻";
          refresh.title = "重新讀取本站電池資料";
          refresh.addEventListener("click", event => {
            event.preventDefault();
            event.stopPropagation();
            queryStation(stationName, true);
          });
          wrapper.appendChild(refresh);
        }
      }
      function renderError(stationName) {
        if (win.__ubikeInlineBatteryGeneration !== generation) return;
        for (const wrapper of wrappersFor(stationName)) {
          wrapper.replaceChildren();
          wrapper.classList.remove("is-empty");
          wrapper.classList.add("is-error");
          const label = doc.createElement("span");
          label.textContent = "電池資料查詢失敗";
          wrapper.appendChild(label);
          const retry = doc.createElement("button");
          retry.type = "button";
          retry.textContent = "重新查詢";
          retry.addEventListener("click", event => {
            event.preventDefault();
            event.stopPropagation();
            queryStation(stationName, true);
          });
          wrapper.appendChild(retry);
        }
      }
      async function queryStation(stationName, force = false) {
        if (win.__ubikeInlineBatteryGeneration !== generation) return;
        setLoading(stationName);
        try {
          const result = await service.queryStationByName(stationName, {
            force,
            ttlMs: 30000,
            attempts: requestIsMobile ? 3 : 2,
            timeoutMs: requestIsMobile ? 15000 : 12000,
          });
          renderResult(stationName, result);
        } catch (_) {
          renderError(stationName);
        }
      }
      async function runAutomaticQueries() {
        const uniqueStations = [];
        const seen = new Set();
        for (const spec of specs) {
          const normalized = service.normalizeStationName(spec.name);
          if (!normalized || seen.has(normalized)) continue;
          seen.add(normalized);
          uniqueStations.push(spec.name);
        }
        let nextIndex = 0;
        const workerCount = Math.min(
          uniqueStations.length,
          requestIsMobile ? 2 : 4,
        );
        async function worker() {
          while (nextIndex < uniqueStations.length) {
            const stationName = uniqueStations[nextIndex++];
            const force = service.normalizeStationName(stationName)
              === service.normalizeStationName(forceStation);
            await queryStation(stationName, force);
            if (requestIsMobile && nextIndex < uniqueStations.length) {
              await new Promise(resolve => win.setTimeout(resolve, 90));
            }
          }
        }
        await Promise.all(Array.from({ length:workerCount }, worker));
      }

      attachTargets();
      win.setTimeout(attachTargets, 120);
      win.setTimeout(attachTargets, 420);
      if (autoQuery && specs.length) win.setTimeout(runAutomaticQueries, 480);
    })();
    </script>
    """
    return (
        component_html
        .replace("__LOW_BATTERY_CLIENT_CORE__", LOW_BATTERY_CLIENT_CORE_JS)
        .replace("__STATION_SPECS__", specs_payload)
        .replace("__THRESHOLD__", str(int(threshold)))
        .replace("__PRIORITY_THRESHOLD__", str(int(priority_threshold)))
        .replace("__AUTO_QUERY__", "true" if auto_query else "false")
        .replace("__FORCE_STATION__", force_payload)
        .replace("__DISPLAY_MODE__", display_mode)
        .replace("__FINGERPRINT__", fingerprint)
    )


def render_inline_low_battery_pillars(
    station_specs: list[tuple[str, str, str]],
    *,
    threshold: int,
    priority_threshold: int,
    mobile_mode: bool,
    auto_query: bool,
    force_station: str = "",
) -> None:
    """把柱號結果安全注入既有分析列或智慧推薦卡，不改原本卡片格局。"""
    normalized_specs = tuple(
        (str(name), str(kind), str(target))
        for name, kind, target in station_specs
        if str(name).strip()
    )
    if not normalized_specs:
        return
    components.html(
        _build_inline_low_battery_pillars_html(
            normalized_specs,
            min(100, max(0, int(threshold))),
            min(int(threshold), max(0, int(priority_threshold))),
            mobile_mode,
            auto_query=auto_query,
            force_station=force_station,
        ),
        height=0,
        scrolling=False,
    )


@st.cache_data(show_spinner=False, max_entries=32)
def _build_floating_battery_query_html(
    route_station_map: dict[str, list[dict]],
    mobile_mode: bool,
) -> str:
    """快取電量查詢元件 HTML，避免每次互動重複替換大型腳本模板。"""
    route_payload = json.dumps(route_station_map, ensure_ascii=False).replace("</", "<\\/")
    display_mode = json.dumps("mobile" if mobile_mode else "desktop")
    battery_fingerprint = json.dumps(
        hashlib.sha1(
            f"{APP_VERSION}|{display_mode}|{route_payload}".encode("utf-8")
        ).hexdigest()
    )
    component_html = r"""
    <script>
    (() => {
      __LOW_BATTERY_CLIENT_CORE__
      const routeStations = __ROUTE_STATIONS__;
      const displayMode = __DISPLAY_MODE__;
      const fingerprint = __BATTERY_FINGERPRINT__;
      const doc = window.parent.document;
      const win = window.parent;
      const catalogUrl = "https://apis.youbike.com.tw/json/station-min-yb2.json";
      const batteryUrl = "https://apis.youbike.com.tw/api/front/bike/lists";
      const batteryService = ensureUbikeBatteryService(win);
      const defaultThreshold = 89;
      const defaultPriorityThreshold = 40;
      // 換版後使用新偏好鍵，讓本版第一次開啟時確實以 89% 為預設值。
      const preferenceKey = "ubike-battery-query-preferences-v7";
      const previousPageWasOpen = Boolean(
        doc.getElementById("ubike-battery-page")?.classList.contains("open"),
      );
      if (previousPageWasOpen) {
        doc.body.style.overflow = String(win.__ubikeBatteryPreviousBodyOverflow || "");
      }
      win.__ubikeBatteryFingerprint = fingerprint;

      doc.getElementById("ubike-battery-fab")?.remove();
      doc.getElementById("ubike-battery-page")?.remove();
      doc.getElementById("ubike-battery-style")?.remove();

      const style = doc.createElement("style");
      style.id = "ubike-battery-style";
      style.textContent = `
        #ubike-battery-fab {
          position: fixed;
          right: 18px;
          bottom: 278px;
          z-index: 2147483000;
          width: 56px;
          height: 56px;
          border: 0;
          border-radius: 50%;
          color: #15350f;
          background: #9ee56f;
          box-shadow: 0 8px 28px rgba(0,0,0,.28);
          font: 850 13px/1.1 -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft JhengHei", sans-serif;
          cursor: pointer;
        }
        #ubike-battery-fab:hover { transform: translateY(-2px); }
        #ubike-battery-page {
          position: fixed;
          inset: 0;
          z-index: 2147483600;
          display: none;
          overflow-y: auto;
          overflow-x: hidden;
          color: #17212b;
          background: #f4f7f9;
          font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft JhengHei", sans-serif;
          overscroll-behavior: contain;
        }
        #ubike-battery-page.open { display: block; }
        #ubike-battery-page * { box-sizing: border-box; }
        #ubike-battery-page .battery-header {
          position: sticky;
          top: 0;
          z-index: 4;
          display: flex;
          align-items: center;
          gap: 12px;
          min-height: 66px;
          padding: max(10px, env(safe-area-inset-top, 0px)) 16px 10px;
          border-bottom: 1px solid #e2e7eb;
          background: rgba(255,255,255,.96);
          backdrop-filter: blur(10px);
        }
        #ubike-battery-page .battery-back {
          width: 42px;
          height: 42px;
          flex: 0 0 42px;
          border: 1px solid #d9e0e5;
          border-radius: 50%;
          color: #17212b;
          background: #fff;
          font-size: 28px;
          line-height: 1;
          cursor: pointer;
        }
        #ubike-battery-page .battery-title { font-size: 20px; font-weight: 900; }
        #ubike-battery-page .battery-main {
          width: min(100%, 980px);
          margin: 0 auto;
          padding: 16px 16px calc(34px + env(safe-area-inset-bottom, 0px));
        }
        #ubike-battery-page .battery-control-card {
          padding: 15px;
          border: 1px solid #dfe6ea;
          border-radius: 16px;
          background: #fff;
          box-shadow: 0 5px 18px rgba(32,45,55,.07);
        }
        #ubike-battery-page .battery-control-grid {
          display: grid;
          grid-template-columns: minmax(145px, .65fr) minmax(230px, 1.15fr) minmax(245px, 1fr) minmax(135px, .6fr);
          gap: 13px;
          align-items: end;
        }
        #ubike-battery-page .battery-label {
          display: block;
          margin-bottom: 7px;
          color: #4b5965;
          font-size: 13px;
          font-weight: 850;
        }
        #ubike-battery-page .battery-threshold-wrap {
          display: flex;
          align-items: center;
          gap: 8px;
        }
        #ubike-battery-page .battery-threshold {
          width: 100%;
          min-height: 46px;
          padding: 8px 12px;
          border: 2px solid #d9e2e7;
          border-radius: 11px;
          color: #111827;
          background: #fff;
          font-size: 18px;
          font-weight: 800;
        }
        #ubike-battery-page .battery-scope-list { display: flex; flex-wrap: wrap; gap: 8px; }
        #ubike-battery-page .battery-scope {
          position: relative;
          display: inline-flex;
          align-items: center;
          justify-content: center;
          min-width: 70px;
          min-height: 46px;
          padding: 8px 13px;
          border: 2px solid #d9e2e7;
          border-radius: 11px;
          color: #374151;
          background: #fff;
          font-weight: 850;
          cursor: pointer;
        }
        #ubike-battery-page .battery-scope.selected {
          border-color: #55a630;
          color: #234f12;
          background: #e9f8df;
        }
        #ubike-battery-page .battery-scope input { position: absolute; opacity: 0; pointer-events: none; }
        #ubike-battery-page .battery-sort-wrap {
          display: grid;
          grid-template-columns: minmax(0, 1fr) auto;
          gap: 7px;
        }
        #ubike-battery-page .battery-sort {
          width: 100%;
          min-height: 46px;
          padding: 8px 34px 8px 11px;
          border: 2px solid #d9e2e7;
          border-radius: 11px;
          color: #111827;
          background: #fff;
          font-size: 15px;
          font-weight: 800;
          cursor: pointer;
        }
        #ubike-battery-page .battery-locate {
          min-height: 46px;
          padding: 8px 11px;
          border: 2px solid #b9d7ac;
          border-radius: 11px;
          color: #28551a;
          background: #effae9;
          font-size: 13px;
          font-weight: 850;
          white-space: nowrap;
          cursor: pointer;
        }
        #ubike-battery-page .battery-locate:disabled { opacity: .65; cursor: wait; }
        #ubike-battery-page .battery-locate[hidden] { display: none; }
        #ubike-battery-page .battery-refresh {
          width: 100%;
          min-height: 46px;
          border: 0;
          border-radius: 11px;
          color: #fff;
          background: #2f7d1d;
          font-size: 16px;
          font-weight: 850;
          cursor: pointer;
        }
        #ubike-battery-page .battery-refresh:disabled { opacity: .65; cursor: wait; }
        #ubike-battery-page .battery-inclusive {
          margin-top: 9px;
          color: #66727d;
          font-size: 12px;
        }
        #ubike-battery-page .battery-priority-control {
          display: grid;
          grid-template-columns: minmax(205px, .8fr) minmax(165px, .55fr) minmax(250px, 1fr);
          gap: 11px;
          align-items: center;
          margin-top: 12px;
          padding: 12px 14px;
          border: 2px solid #e5e7eb;
          border-radius: 13px;
          background: #f8fafc;
        }
        #ubike-battery-page .battery-priority-control.enabled {
          border-color: #ef4444;
          background: #fff1f2;
          box-shadow: 0 5px 15px rgba(185, 28, 28, .1);
        }
        #ubike-battery-page .battery-priority-toggle {
          display: flex;
          align-items: center;
          gap: 9px;
          color: #7f1d1d;
          font-size: 15px;
          font-weight: 900;
          cursor: pointer;
        }
        #ubike-battery-page .battery-priority-toggle input {
          width: 21px;
          height: 21px;
          accent-color: #dc2626;
        }
        #ubike-battery-page .battery-priority-threshold-wrap {
          display: flex;
          align-items: center;
          gap: 7px;
        }
        #ubike-battery-page .battery-priority-threshold-wrap .battery-label {
          margin: 0;
          white-space: nowrap;
        }
        #ubike-battery-page .battery-priority-threshold {
          width: 88px;
          min-height: 42px;
          padding: 7px 10px;
          border: 2px solid #fca5a5;
          border-radius: 10px;
          color: #7f1d1d;
          background: #fff;
          font-size: 18px;
          font-weight: 900;
        }
        #ubike-battery-page .battery-priority-threshold:disabled {
          border-color: #d1d5db;
          color: #9ca3af;
          background: #f3f4f6;
        }
        #ubike-battery-page .battery-priority-hint {
          color: #7c2d12;
          font-size: 12px;
          font-weight: 750;
          line-height: 1.45;
        }
        #ubike-battery-page .battery-location-status {
          min-height: 18px;
          margin-top: 7px;
          color: #356523;
          font-size: 12px;
          font-weight: 750;
        }
        #ubike-battery-page .battery-location-status.error { color: #b42318; }
        #ubike-battery-page .battery-location-status[hidden] { display: none; }
        #ubike-battery-page .battery-status {
          min-height: 22px;
          margin: 12px 2px 6px;
          color: #5f6b75;
          font-size: 13px;
        }
        #ubike-battery-page .battery-status.error { color: #b42318; font-weight: 750; }
        #ubike-battery-page .battery-summary {
          display: grid;
          grid-template-columns: repeat(auto-fit, minmax(145px, 1fr));
          gap: 9px;
          margin: 10px 0 15px;
        }
        #ubike-battery-page .battery-metric {
          padding: 12px;
          border: 1px solid #dce7d6;
          border-radius: 13px;
          background: #fff;
        }
        #ubike-battery-page .battery-metric.total { border-color: #acd59a; background: #effae9; }
        #ubike-battery-page .battery-metric.priority { border: 2px solid #ef4444; background: #fff1f2; }
        #ubike-battery-page .battery-metric.priority .battery-metric-label,
        #ubike-battery-page .battery-metric.priority .battery-metric-value { color: #b91c1c; }
        #ubike-battery-page .battery-metric-label { color: #66727d; font-size: 12px; font-weight: 800; }
        #ubike-battery-page .battery-metric-value { margin-top: 3px; font-size: 25px; font-weight: 900; }
        #ubike-battery-page .battery-district-summary {
          margin: 10px 0 22px;
          padding: 20px;
          border: 4px solid #b91c1c;
          border-radius: 18px;
          background: linear-gradient(145deg, #fff1f2, #ffedd5);
          box-shadow: 0 12px 30px rgba(185, 28, 28, .24);
        }
        #ubike-battery-page .battery-district-summary[hidden] { display: none; }
        #ubike-battery-page .battery-district-summary-title {
          color: #7f1d1d;
          font-size: 23px;
          line-height: 1.25;
          font-weight: 900;
        }
        #ubike-battery-page .battery-district-summary-hint {
          margin: 5px 0 14px;
          color: #7c2d12;
          font-size: 13px;
          font-weight: 800;
        }
        #ubike-battery-page .battery-district-summary-grid {
          display: grid;
          grid-template-columns: repeat(auto-fit, minmax(155px, 1fr));
          gap: 12px;
        }
        #ubike-battery-page .battery-district-summary .battery-district-metric {
          padding: 17px 12px;
          border: 3px solid #fb923c;
          border-radius: 14px;
          background: #fff;
          box-shadow: 0 5px 13px rgba(154, 52, 18, .12);
          text-align: center;
        }
        #ubike-battery-page .battery-district-summary .battery-district-metric.highest {
          border-color: #dc2626;
          background: #fff1f2;
          box-shadow: 0 7px 17px rgba(185, 28, 28, .2);
        }
        #ubike-battery-page .battery-district-summary .battery-metric-label {
          color: #7c2d12;
          font-size: 18px;
          font-weight: 900;
        }
        #ubike-battery-page .battery-district-summary .battery-metric-value {
          margin-top: 7px;
          color: #b91c1c;
          font-size: 40px;
          line-height: 1.05;
          font-weight: 950;
        }
        #ubike-battery-page .battery-route-title { margin: 18px 2px 8px; font-size: 17px; font-weight: 900; }
        #ubike-battery-page .battery-station {
          margin-bottom: 8px;
          border: 1px solid #dfe5e9;
          border-radius: 13px;
          overflow: hidden;
          background: #fff;
        }
        #ubike-battery-page .battery-station.unmatched { border-color: #f0c36b; }
        #ubike-battery-page .battery-station.priority {
          border: 2px solid #dc2626;
          box-shadow: 0 7px 18px rgba(185, 28, 28, .16);
        }
        #ubike-battery-page .battery-station summary {
          display: flex;
          align-items: center;
          justify-content: space-between;
          gap: 10px;
          min-height: 54px;
          padding: 10px 13px;
          cursor: pointer;
          list-style: none;
        }
        #ubike-battery-page .battery-station summary::-webkit-details-marker { display: none; }
        #ubike-battery-page .battery-station-heading {
          display: flex;
          min-width: 0;
          flex-direction: column;
          gap: 3px;
        }
        #ubike-battery-page .battery-station-name { font-size: 15px; font-weight: 850; }
        #ubike-battery-page .battery-station-meta {
          color: #69757f;
          font-size: 12px;
          font-weight: 720;
        }
        #ubike-battery-page .battery-recommended {
          display: inline-block;
          margin-right: 5px;
          padding: 2px 6px;
          border-radius: 999px;
          color: #244d16;
          background: #dff3d4;
          font-size: 11px;
          font-weight: 900;
        }
        #ubike-battery-page .battery-priority-badge {
          display: inline-block;
          margin-right: 5px;
          padding: 2px 7px;
          border-radius: 999px;
          color: #fff;
          background: #dc2626;
          font-size: 11px;
          font-weight: 900;
        }
        #ubike-battery-page .battery-count {
          flex: 0 0 auto;
          padding: 5px 9px;
          border-radius: 999px;
          color: #236015;
          background: #e7f7df;
          font-size: 13px;
          font-weight: 900;
        }
        #ubike-battery-page .battery-count.needs-change { color: #b42318; background: #fee4e2; }
        #ubike-battery-page .battery-count.unknown { color: #8a5200; background: #fff0c2; }
        #ubike-battery-page .battery-bike-list { border-top: 1px solid #eef1f3; }
        #ubike-battery-page .battery-bike-row {
          display: grid;
          grid-template-columns: 1.2fr .8fr .8fr;
          gap: 8px;
          padding: 10px 13px;
          border-bottom: 1px solid #f0f2f4;
          font-size: 14px;
        }
        #ubike-battery-page .battery-bike-row:last-child { border-bottom: 0; }
        #ubike-battery-page .battery-bike-row.priority { background: #fff1f2; box-shadow: inset 4px 0 #dc2626; }
        #ubike-battery-page .battery-bike-no { font-weight: 850; }
        #ubike-battery-page .battery-power { color: #b42318; font-weight: 900; }
        #ubike-battery-page .battery-empty { padding: 13px; color: #6b7280; font-size: 13px; }

        /* v27.5：電量頁延續原版格局，更新配色、光影與穩定查詢。 */
        #ubike-battery-fab {
          color: #061116;
          background: linear-gradient(135deg, #55f6ff, #f4ff57);
          box-shadow: 0 0 24px rgba(85,246,255,.38), 0 8px 28px rgba(0,0,0,.34);
        }
        #ubike-battery-page {
          color: #effcff;
          background:
            linear-gradient(rgba(85,246,255,.04) 1px, transparent 1px),
            linear-gradient(90deg, rgba(85,246,255,.04) 1px, transparent 1px),
            radial-gradient(circle at 12% 8%, rgba(255,63,207,.16), transparent 28rem),
            radial-gradient(circle at 88% 20%, rgba(85,246,255,.13), transparent 30rem),
            linear-gradient(145deg, #050811, #091526 55%, #100817);
          background-size: 36px 36px, 36px 36px, auto, auto, auto;
        }
        #ubike-battery-page .battery-header {
          color: #effcff;
          border-bottom-color: rgba(85,246,255,.34);
          background: rgba(5,11,22,.96);
          box-shadow: 0 8px 24px rgba(0,0,0,.30), inset 0 -1px rgba(255,63,207,.24);
        }
        #ubike-battery-page .battery-back {
          color: #55f6ff;
          border-color: rgba(85,246,255,.48);
          background: rgba(85,246,255,.08);
          box-shadow: inset 0 0 12px rgba(85,246,255,.04), 0 0 12px rgba(85,246,255,.08);
        }
        #ubike-battery-page .battery-title {color:#f7ffff; text-shadow:0 0 13px rgba(85,246,255,.30);}
        #ubike-battery-page .battery-control-card {
          color:#effcff;
          border-color:rgba(85,246,255,.30);
          background:linear-gradient(135deg, rgba(8,22,38,.96), rgba(28,8,37,.94));
          box-shadow:inset 3px 0 rgba(255,63,207,.45), 0 9px 24px rgba(0,0,0,.26);
        }
        #ubike-battery-page .battery-label {color:#8feff5;}
        #ubike-battery-page .battery-threshold,
        #ubike-battery-page .battery-sort,
        #ubike-battery-page .battery-priority-threshold {
          color:#effcff;
          border-color:rgba(85,246,255,.38);
          background:#071225;
          box-shadow:inset 0 0 12px rgba(85,246,255,.035);
        }
        #ubike-battery-page .battery-sort option {color:#effcff; background:#071225;}
        #ubike-battery-page .battery-threshold-wrap strong,
        #ubike-battery-page .battery-priority-threshold-wrap strong {color:#f4ff57;}
        #ubike-battery-page .battery-scope {
          color:#b2cdd6;
          border-color:rgba(85,246,255,.27);
          background:rgba(6,17,31,.92);
          box-shadow:inset 0 0 10px rgba(85,246,255,.025);
        }
        #ubike-battery-page .battery-scope.selected {
          color:#061116;
          border-color:#f4ff57;
          background:linear-gradient(135deg, #f4ff57, #55f6ff);
          box-shadow:0 0 16px rgba(244,255,87,.18);
        }
        #ubike-battery-page .battery-locate {
          color:#ff91df;
          border-color:rgba(255,63,207,.42);
          background:rgba(255,63,207,.08);
        }
        #ubike-battery-page .battery-refresh {
          color:#061116;
          background:linear-gradient(105deg, #55f6ff, #3ee9c2 68%, #f4ff57);
          box-shadow:0 0 19px rgba(85,246,255,.20), inset 0 -2px rgba(0,0,0,.13);
        }
        #ubike-battery-page .battery-inclusive {color:#8daab5;}
        #ubike-battery-page .battery-priority-control {
          border-color:rgba(255,63,207,.30);
          background:linear-gradient(105deg, rgba(41,7,36,.78), rgba(7,18,32,.92));
          box-shadow:inset 3px 0 rgba(255,63,207,.34);
        }
        #ubike-battery-page .battery-priority-control.enabled {
          border-color:#ff3fcf;
          background:linear-gradient(105deg, rgba(75,5,54,.84), rgba(25,5,27,.94));
          box-shadow:inset 4px 0 #ff3fcf, 0 0 20px rgba(255,63,207,.15);
        }
        #ubike-battery-page .battery-priority-toggle {color:#ff91df;}
        #ubike-battery-page .battery-priority-toggle input {accent-color:#ff3fcf;}
        #ubike-battery-page .battery-priority-threshold {color:#ffb1e8; border-color:rgba(255,63,207,.48);}
        #ubike-battery-page .battery-priority-threshold:disabled {
          color:#607986;
          border-color:rgba(123,151,161,.24);
          background:#0b1622;
        }
        #ubike-battery-page .battery-priority-hint {color:#d6a7c9;}
        #ubike-battery-page .battery-location-status {color:#70f3b4;}
        #ubike-battery-page .battery-location-status.error,
        #ubike-battery-page .battery-status.error {color:#ff7798;}
        #ubike-battery-page .battery-status {color:#9bb7c4;}
        #ubike-battery-page .battery-metric {
          color:#effcff;
          border-color:rgba(85,246,255,.25);
          background:linear-gradient(135deg, rgba(8,22,38,.94), rgba(19,8,29,.94));
          box-shadow:inset 3px 0 rgba(85,246,255,.45), 0 7px 18px rgba(0,0,0,.20);
        }
        #ubike-battery-page .battery-metric.total {
          border-color:rgba(244,255,87,.42);
          background:linear-gradient(135deg, rgba(42,43,7,.80), rgba(8,23,31,.94));
          box-shadow:inset 3px 0 #f4ff57, 0 0 16px rgba(244,255,87,.07);
        }
        #ubike-battery-page .battery-metric.priority {
          border-color:#ff3fcf;
          background:linear-gradient(135deg, rgba(70,5,50,.86), rgba(22,6,25,.96));
          box-shadow:inset 3px 0 #ff3fcf, 0 0 18px rgba(255,63,207,.12);
        }
        #ubike-battery-page .battery-metric-label {color:#8caab5;}
        #ubike-battery-page .battery-metric-value {color:#f7ffff;}
        #ubike-battery-page .battery-metric.priority .battery-metric-label,
        #ubike-battery-page .battery-metric.priority .battery-metric-value {color:#ff91df;}
        #ubike-battery-page .battery-district-summary {
          border-color:#ff3fcf;
          background:linear-gradient(130deg, rgba(74,5,51,.94), rgba(23,7,29,.97) 56%, rgba(5,32,39,.94));
          box-shadow:inset 4px 0 #ff3fcf, 0 0 30px rgba(255,63,207,.15), 0 12px 30px rgba(0,0,0,.30);
        }
        #ubike-battery-page .battery-district-summary-title {color:#ffffff; text-shadow:0 0 13px rgba(255,63,207,.25);}
        #ubike-battery-page .battery-district-summary-hint {color:#efacd8;}
        #ubike-battery-page .battery-district-summary .battery-district-metric {
          border-color:rgba(255,126,215,.55);
          background:linear-gradient(145deg, rgba(42,8,37,.96), rgba(6,20,30,.96));
          box-shadow:0 6px 16px rgba(0,0,0,.22);
        }
        #ubike-battery-page .battery-district-summary .battery-district-metric.highest {
          border-color:#f4ff57;
          background:linear-gradient(145deg, rgba(65,38,7,.94), rgba(40,7,35,.96));
          box-shadow:inset 3px 0 #f4ff57, 0 0 18px rgba(244,255,87,.12);
        }
        #ubike-battery-page .battery-district-summary .battery-metric-label {color:#ff9ce1;}
        #ubike-battery-page .battery-district-summary .battery-metric-value {color:#f4ff57; text-shadow:0 0 14px rgba(244,255,87,.25);}
        #ubike-battery-page .battery-route-title {color:#effcff; text-shadow:0 0 10px rgba(85,246,255,.20);}
        #ubike-battery-page .battery-station {
          color:#effcff;
          border-color:rgba(85,246,255,.24);
          background:linear-gradient(135deg, rgba(8,21,37,.94), rgba(17,7,26,.96));
          box-shadow:inset 3px 0 rgba(85,246,255,.26), 0 7px 18px rgba(0,0,0,.18);
        }
        #ubike-battery-page .battery-station.unmatched {border-color:rgba(244,255,87,.55);}
        #ubike-battery-page .battery-station.priority {
          border-color:#ff3fcf;
          background:linear-gradient(135deg, rgba(67,5,48,.92), rgba(19,6,23,.98));
          box-shadow:inset 4px 0 #ff3fcf, 0 0 20px rgba(255,63,207,.13);
        }
        #ubike-battery-page .battery-station-meta {color:#8caab5;}
        #ubike-battery-page .battery-recommended {color:#7ff9ff; background:rgba(85,246,255,.09);}
        #ubike-battery-page .battery-priority-badge {color:#fff; background:linear-gradient(90deg, #ff3fcf, #ff496c); box-shadow:0 0 10px rgba(255,63,207,.28);}
        #ubike-battery-page .battery-count {color:#7ff9ff; background:rgba(85,246,255,.09);}
        #ubike-battery-page .battery-count.needs-change {color:#ff91df; background:rgba(255,63,207,.12);}
        #ubike-battery-page .battery-count.unknown {color:#f4ff57; background:rgba(244,255,87,.09);}
        #ubike-battery-page .battery-bike-list {border-top-color:rgba(85,246,255,.14);}
        #ubike-battery-page .battery-bike-row {color:#c1d9e1; border-bottom-color:rgba(85,246,255,.10); background:rgba(0,0,0,.10);}
        #ubike-battery-page .battery-bike-row.priority {background:rgba(255,63,207,.13); box-shadow:inset 4px 0 #ff3fcf;}
        #ubike-battery-page .battery-bike-no {color:#effcff;}
        #ubike-battery-page .battery-power {color:#ff75d7; text-shadow:0 0 8px rgba(255,63,207,.16);}
        #ubike-battery-page .battery-empty {color:#829da8;}
        @media (max-width: 700px) {
          #ubike-battery-fab {
            right: 10px;
            bottom: calc(312px + env(safe-area-inset-bottom, 0px));
            width: 52px;
            height: 52px;
          }
          #ubike-battery-page .battery-header { min-height: 60px; padding-left: 10px; padding-right: 10px; }
          #ubike-battery-page .battery-main { padding: 10px 8px calc(28px + env(safe-area-inset-bottom, 0px)); }
          #ubike-battery-page .battery-control-card { padding: 12px; }
          #ubike-battery-page .battery-control-grid { grid-template-columns: 1fr; gap: 11px; }
          #ubike-battery-page .battery-priority-control { grid-template-columns: 1fr; align-items: start; }
          #ubike-battery-page .battery-sort-wrap { grid-template-columns: minmax(0, 1fr) auto; }
          #ubike-battery-page .battery-summary { grid-template-columns: repeat(2, minmax(0,1fr)); }
          #ubike-battery-page .battery-district-summary { padding: 15px 11px; }
          #ubike-battery-page .battery-district-summary-title { font-size: 20px; }
          #ubike-battery-page .battery-district-summary-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 9px; }
          #ubike-battery-page .battery-district-summary .battery-metric-label { font-size: 16px; }
          #ubike-battery-page .battery-district-summary .battery-metric-value { font-size: 34px; }
          #ubike-battery-page .battery-bike-row { grid-template-columns: 1.25fr .72fr .72fr; padding-left: 10px; padding-right: 10px; }
        }
      `;
      doc.head.appendChild(style);

      const fab = doc.createElement("button");
      fab.id = "ubike-battery-fab";
      fab.type = "button";
      fab.title = "查詢 YouBike 2.0E 電量";
      fab.textContent = "⚡電量";
      doc.body.appendChild(fab);

      const page = doc.createElement("section");
      page.id = "ubike-battery-page";
      page.setAttribute("aria-hidden", "true");
      page.innerHTML = `
        <header class="battery-header">
          <button class="battery-back" type="button" aria-label="返回智慧調度">‹</button>
          <div class="battery-title">⚡ 2.0E 電量查詢｜測試版</div>
        </header>
        <main class="battery-main">
          <section class="battery-control-card">
            <div class="battery-control-grid">
              <label>
                <span class="battery-label">更換門檻</span>
                <span class="battery-threshold-wrap">
                  <input class="battery-threshold" type="number" min="0" max="100" step="1" inputmode="numeric" value="89" />
                  <strong>%</strong>
                </span>
              </label>
              <div>
                <span class="battery-label">統計範圍（可複選）</span>
                <div class="battery-scope-list"></div>
              </div>
              <div>
                <span class="battery-label">場站排序</span>
                <div class="battery-sort-wrap">
                  <select class="battery-sort">
                    <option value="nearest">離我最近</option>
                    <option value="district">行政區分區</option>
                  </select>
                  <button class="battery-locate" type="button">📍重新定位</button>
                </div>
              </div>
              <button class="battery-refresh" type="button">更新電量</button>
            </div>
            <div class="battery-inclusive">門檻包含設定值，例如 89% 會列出電量 ≤ 89% 的車輛。</div>
            <div class="battery-priority-control">
              <label class="battery-priority-toggle">
                <input class="battery-priority-enabled" type="checkbox" />
                <span>⚠️ 開啟優先門檻</span>
              </label>
              <label class="battery-priority-threshold-wrap">
                <span class="battery-label">第二門檻</span>
                <input class="battery-priority-threshold" type="number" min="0" max="89" step="1" inputmode="numeric" value="40" disabled />
                <strong>%</strong>
              </label>
              <div class="battery-priority-hint">開啟後，只要場站有電量 ≤ 40% 的電池，就會優先排列；其餘場站仍依原本方式排序。</div>
            </div>
            <div class="battery-location-status" aria-live="polite">尚未取得定位。</div>
          </section>
          <div class="battery-status">請先選擇 D1、D2 或 D3，再按「更新電量」。</div>
          <div class="battery-summary"></div>
          <section class="battery-district-summary" hidden>
            <div class="battery-district-summary-title">📍 各行政區需更換電池</div>
            <div class="battery-district-summary-hint">只顯示至少有 1 顆需要更換的行政區</div>
            <div class="battery-district-summary-grid"></div>
          </section>
          <div class="battery-results"></div>
        </main>
      `;
      doc.body.appendChild(page);

      const backButton = page.querySelector(".battery-back");
      const thresholdInput = page.querySelector(".battery-threshold");
      const priorityControlNode = page.querySelector(".battery-priority-control");
      const priorityEnabledInput = page.querySelector(".battery-priority-enabled");
      const priorityThresholdInput = page.querySelector(".battery-priority-threshold");
      const priorityHintNode = page.querySelector(".battery-priority-hint");
      const scopeList = page.querySelector(".battery-scope-list");
      const sortSelect = page.querySelector(".battery-sort");
      const locateButton = page.querySelector(".battery-locate");
      const locationStatusNode = page.querySelector(".battery-location-status");
      const refreshButton = page.querySelector(".battery-refresh");
      const statusNode = page.querySelector(".battery-status");
      const summaryNode = page.querySelector(".battery-summary");
      const districtSummaryNode = page.querySelector(".battery-district-summary");
      const districtSummaryTitleNode = page.querySelector(".battery-district-summary-title");
      const districtSummaryGridNode = page.querySelector(".battery-district-summary-grid");
      const resultsNode = page.querySelector(".battery-results");
      const routes = ["D1", "D2", "D3"].filter(route => Array.isArray(routeStations[route]) && routeStations[route].length);
      let catalogCache = null;
      let lastResults = [];
      let lastResultRoutes = [];
      let userLocation = null;
      let locationPromise = null;
      let locationError = "";
      let previousBodyOverflow = "";
      const isMobileRequestMode = displayMode === "mobile"
        || win.matchMedia("(max-width: 700px)").matches;

      function clampThreshold(value) {
        const number = Number.parseInt(value, 10);
        return Math.max(0, Math.min(100, Number.isFinite(number) ? number : defaultThreshold));
      }
      function clampPriorityThreshold(value) {
        const replacementThreshold = clampThreshold(thresholdInput.value);
        const number = Number.parseInt(value, 10);
        const fallback = Math.min(defaultPriorityThreshold, replacementThreshold);
        return Math.max(
          0,
          Math.min(replacementThreshold, Number.isFinite(number) ? number : fallback),
        );
      }
      function prioritySortingEnabled() {
        return Boolean(priorityEnabledInput.checked);
      }
      function updatePriorityControls() {
        const enabled = prioritySortingEnabled();
        const priorityThreshold = clampPriorityThreshold(priorityThresholdInput.value);
        priorityThresholdInput.value = String(priorityThreshold);
        priorityThresholdInput.max = String(clampThreshold(thresholdInput.value));
        priorityThresholdInput.disabled = !enabled;
        priorityControlNode.classList.toggle("enabled", enabled);
        priorityHintNode.textContent = enabled
          ? `優先排序已開啟：有電量 ≤ ${priorityThreshold}% 的場站會排在前面。`
          : "此功能目前關閉；開啟後，電量 ≤ 第二門檻的場站會優先排列。";
      }
      function normalizeSortMode(value) {
        return value === "district" ? "district" : "nearest";
      }
      function readPreferences() {
        try {
          const saved = JSON.parse(win.localStorage.getItem(preferenceKey) || "null");
          if (saved && typeof saved === "object") return saved;
        } catch (_) {}
        return {};
      }
      function savePreferences() {
        try {
          win.localStorage.setItem(preferenceKey, JSON.stringify({
            threshold: clampThreshold(thresholdInput.value),
            priority_enabled: prioritySortingEnabled(),
            priority_threshold: clampPriorityThreshold(priorityThresholdInput.value),
            sort_mode: normalizeSortMode(sortSelect.value),
          }));
        } catch (_) {}
      }
      function selectedRoutes() {
        return Array.from(scopeList.querySelectorAll("input:checked")).map(input => input.value);
      }
      function setStatus(message, isError = false) {
        statusNode.textContent = message;
        statusNode.classList.toggle("error", isError);
      }
      function setLocationStatus(message, isError = false) {
        locationStatusNode.textContent = message;
        locationStatusNode.classList.toggle("error", isError);
      }
      function updateSortControls() {
        const usesLocation = normalizeSortMode(sortSelect.value) === "nearest";
        locateButton.hidden = !usesLocation;
        locationStatusNode.hidden = !usesLocation;
        if (!usesLocation) return;
        if (locationPromise) {
          locateButton.disabled = true;
          locateButton.textContent = "定位中…";
          setLocationStatus("正在取得目前位置……");
        } else if (locationError) {
          locateButton.disabled = false;
          locateButton.textContent = "📍再試一次";
          setLocationStatus(
            userLocation
              ? `${locationError}，目前沿用上次定位排序。`
              : `${locationError}，已暫時改用行政區分區顯示。`,
            true,
          );
        } else if (userLocation) {
          locateButton.disabled = false;
          locateButton.textContent = "📍重新定位";
          setLocationStatus(`定位完成（誤差約 ${Math.round(userLocation.accuracy || 0)} 公尺），目前依直線距離由近到遠排列。`);
        } else {
          locateButton.disabled = false;
          locateButton.textContent = "📍取得定位";
          setLocationStatus("尚未取得定位；允許定位後會自動推薦最近場站。");
        }
      }
      function distanceMeters(lat1, lon1, lat2, lon2) {
        const earthRadius = 6371008.8;
        const radians = value => Number(value) * Math.PI / 180;
        const phi1 = radians(lat1);
        const phi2 = radians(lat2);
        const deltaPhi = radians(Number(lat2) - Number(lat1));
        const deltaLambda = radians(Number(lon2) - Number(lon1));
        const a = Math.sin(deltaPhi / 2) ** 2
          + Math.cos(phi1) * Math.cos(phi2) * Math.sin(deltaLambda / 2) ** 2;
        return earthRadius * 2 * Math.atan2(Math.sqrt(a), Math.sqrt(Math.max(0, 1 - a)));
      }
      function formatDistance(meters) {
        if (!Number.isFinite(meters)) return "距離無法計算";
        if (meters < 1000) return `約 ${Math.max(1, Math.round(meters))} 公尺`;
        return `約 ${(meters / 1000).toFixed(meters < 10000 ? 1 : 0)} 公里`;
      }
      function requestCurrentLocation({ force = false } = {}) {
        if (userLocation && !force) {
          updateSortControls();
          return Promise.resolve(userLocation);
        }
        if (locationPromise) return locationPromise;
        const geolocation = (win.navigator && win.navigator.geolocation) || navigator.geolocation;
        if (!geolocation) {
          locationError = "此瀏覽器不支援定位";
          updateSortControls();
          if (lastResults.length) renderResults();
          return Promise.resolve(null);
        }

        locationError = "";
        const previousLocation = userLocation;
        locateButton.disabled = true;
        locateButton.textContent = "定位中…";
        setLocationStatus("正在取得目前位置……");
        locationPromise = new Promise(resolve => {
          const failLocation = error => {
            const errorMessages = {
              1: "定位權限被拒絕",
              2: "目前位置無法取得",
              3: "定位逾時",
            };
            userLocation = previousLocation;
            locationError = errorMessages[Number(error && error.code)]
              || String(error && error.message || "定位失敗");
            resolve(userLocation);
          };
          try {
            geolocation.getCurrentPosition(
              position => {
                const latitude = Number(position.coords.latitude);
                const longitude = Number(position.coords.longitude);
                if (!Number.isFinite(latitude) || !Number.isFinite(longitude)) {
                  failLocation({ message: "定位座標格式錯誤" });
                  return;
                }
                userLocation = {
                  latitude,
                  longitude,
                  accuracy: Number(position.coords.accuracy || 0),
                  updatedAt: Date.now(),
                };
                locationError = "";
                resolve(userLocation);
              },
              failLocation,
              { enableHighAccuracy: true, timeout: 15000, maximumAge: 30000 },
            );
          } catch (error) {
            failLocation(error);
          }
        }).finally(() => {
          locationPromise = null;
          updateSortControls();
          if (lastResults.length) renderResults();
        });
        return locationPromise;
      }
      function normalizeStationName(value) {
        return String(value || "")
          .toLowerCase()
          .replaceAll("臺", "台")
          .replace(/^(?:youbike|ubike)\s*2\s*[.．]?\s*0\s*e?\s*[_\-－—:：]*\s*/i, "")
          .replaceAll("公共自行車租賃站", "")
          .replace(/[^0-9a-z㐀-鿿]/g, "");
      }
      function firstFiniteNumber(...values) {
        for (const value of values) {
          if (value === null || value === undefined || value === "") continue;
          const number = Number(value);
          if (Number.isFinite(number)) return number;
        }
        return null;
      }
      function canonicalDistrict(value) {
        const district = String(value || "").trim().replaceAll("台", "臺");
        return district || "行政區未標示";
      }
      function extractItems(payload) {
        if (Array.isArray(payload)) return payload.filter(item => item && typeof item === "object");
        if (!payload || typeof payload !== "object") return [];
        for (const candidate of [payload.data, payload.result, payload.stations, payload.retVal]) {
          if (Array.isArray(candidate)) return candidate.filter(item => item && typeof item === "object");
          if (candidate && typeof candidate === "object" && Array.isArray(candidate.data)) return candidate.data;
        }
        return [];
      }
      function isTaitungStation(item) {
        const text = [item.county_tw, item.city_tw, item.district_tw, item.address_tw, item.name_tw]
          .map(value => String(value || "")).join(" ").replaceAll("臺", "台");
        return String(item.area_code || "") === "15" || text.includes("台東");
      }
      function waitMilliseconds(milliseconds) {
        return new Promise(resolve => win.setTimeout(resolve, milliseconds));
      }
      async function fetchJson(url) {
        const maxAttempts = isMobileRequestMode ? 3 : 2;
        let lastError = null;
        for (let attempt = 1; attempt <= maxAttempts; attempt += 1) {
          const controller = new AbortController();
          const timeout = win.setTimeout(
            () => controller.abort(),
            isMobileRequestMode ? 18000 : 22000,
          );
          try {
            const response = await fetch(url, {
              cache: "no-store",
              credentials: "omit",
              signal: controller.signal,
            });
            if (!response.ok) {
              const error = new Error(`HTTP ${response.status}`);
              error.retryable = [408, 425, 429, 500, 502, 503, 504].includes(response.status);
              throw error;
            }
            return await response.json();
          } catch (error) {
            lastError = error;
            const networkFailure = error?.name === "AbortError"
              || error instanceof TypeError
              || /load failed|failed to fetch|network/i.test(String(error?.message || error));
            const canRetry = attempt < maxAttempts
              && (networkFailure || error?.retryable === true);
            if (!canRetry) throw error;
            await waitMilliseconds(450 * attempt + Math.floor(Math.random() * 250));
          } finally {
            win.clearTimeout(timeout);
          }
        }
        throw lastError || new Error("網路連線失敗");
      }
      async function getCatalog() {
        if (catalogCache) return catalogCache;
        catalogCache = await batteryService.getCatalog({
          attempts: isMobileRequestMode ? 3 : 2,
          timeoutMs: isMobileRequestMode ? 17000 : 14000,
        });
        return catalogCache;
      }
      function matchCatalogStation(stationName, catalog) {
        const wantedKey = normalizeStationName(stationName);
        const exact = catalog.filter(item => normalizeStationName(item.name_tw || item.sna) === wantedKey);
        if (exact.length === 1) return exact[0];
        const partial = catalog
          .map(item => {
            const candidateKey = normalizeStationName(item.name_tw || item.sna);
            const matches = wantedKey.length >= 4 && candidateKey.length >= 4
              && (wantedKey.includes(candidateKey) || candidateKey.includes(wantedKey));
            const score = matches ? Math.min(wantedKey.length, candidateKey.length) / Math.max(wantedKey.length, candidateKey.length) : 0;
            return { item, score };
          })
          .filter(record => record.score >= .72)
          .sort((left, right) => right.score - left.score);
        if (!partial.length) return null;
        if (partial.length > 1 && partial[1].score >= partial[0].score - .02) return null;
        return partial[0].item;
      }
      async function getBatteryList(stationNo, { force = false } = {}) {
        return batteryService.getBatteryListByStationNo(stationNo, {
          force,
          ttlMs: 45000,
          attempts: isMobileRequestMode ? 3 : 2,
          timeoutMs: isMobileRequestMode ? 16000 : 13000,
        });
      }
      async function mapWithAdaptiveConcurrency(items, initialLimit, task, onWave) {
        const output = new Array(items.length);
        const maximumLimit = isMobileRequestMode ? 3 : 6;
        let limit = Math.max(1, Math.min(maximumLimit, initialLimit));
        let nextIndex = 0;
        let stableWaves = 0;
        while (nextIndex < items.length) {
          const waveIndices = [];
          while (nextIndex < items.length && waveIndices.length < limit) {
            waveIndices.push(nextIndex++);
          }
          const waveResults = await Promise.all(
            waveIndices.map(index => task(items[index], index)),
          );
          waveIndices.forEach((index, position) => {
            output[index] = waveResults[position];
          });
          const failureCount = waveResults.filter(result => result?.fetchError).length;
          if (failureCount) {
            limit = Math.max(1, Math.floor(limit / 2));
            stableWaves = 0;
          } else {
            stableWaves += 1;
            if (stableWaves >= 2 && limit < maximumLimit) {
              limit += 1;
              stableWaves = 0;
            }
          }
          if (typeof onWave === "function") {
            await onWave({ output, completed: nextIndex, limit, failureCount });
          }
          if (nextIndex < items.length) {
            await waitMilliseconds(isMobileRequestMode ? 80 : 25);
          }
        }
        return output;
      }
      function appendTextElement(parent, tagName, className, text) {
        const element = doc.createElement(tagName);
        if (className) element.className = className;
        element.textContent = text;
        parent.appendChild(element);
        return element;
      }
      function lowBikesFor(result, threshold) {
        return (Array.isArray(result.bikes) ? result.bikes : [])
          .filter(bike => bike.battery_power <= threshold);
      }
      function priorityBikesFor(result, priorityThreshold) {
        return (Array.isArray(result.bikes) ? result.bikes : [])
          .filter(bike => bike.battery_power <= priorityThreshold);
      }
      function stationDistanceForResult(result) {
        if (!userLocation) return null;
        const latitude = firstFiniteNumber(result.latitude);
        const longitude = firstFiniteNumber(result.longitude);
        if (
          latitude === null || longitude === null
          || latitude < -90 || latitude > 90 || longitude < -180 || longitude > 180
        ) return null;
        const distance = distanceMeters(
          userLocation.latitude,
          userLocation.longitude,
          latitude,
          longitude,
        );
        return Number.isFinite(distance) ? distance : null;
      }
      function compareStationNames(left, right) {
        return String(left.name || "").localeCompare(
          String(right.name || ""),
          "zh-Hant",
          { numeric: true, sensitivity: "base" },
        );
      }
      function buildDisplayGroups(threshold, replacementResults) {
        // replacementResults 已先排除零顆與未配對場站，行政區標題不會留下空群組。
        const priorityEnabled = prioritySortingEnabled();
        const priorityThreshold = clampPriorityThreshold(priorityThresholdInput.value);
        const decorated = replacementResults.map(result => {
          const priorityBikes = priorityEnabled
            ? priorityBikesFor(result, priorityThreshold)
            : [];
          return {
            result,
            distance: stationDistanceForResult(result),
            priorityCount: priorityBikes.length,
            lowestPriorityPower: priorityBikes.length
              ? Math.min(...priorityBikes.map(bike => bike.battery_power))
              : Number.POSITIVE_INFINITY,
          };
        });
        const comparePriority = (left, right) => {
          if (!priorityEnabled) return 0;
          const countDifference = right.priorityCount - left.priorityCount;
          if (countDifference) return countDifference;
          if (left.priorityCount && right.priorityCount) {
            const powerDifference = left.lowestPriorityPower - right.lowestPriorityPower;
            if (powerDifference) return powerDifference;
          }
          return 0;
        };
        const requestedMode = normalizeSortMode(sortSelect.value);
        if (requestedMode === "nearest" && userLocation) {
          decorated.sort((left, right) => {
            if (left.result.matched !== right.result.matched) return left.result.matched ? -1 : 1;
            const priorityDifference = comparePriority(left, right);
            if (priorityDifference) return priorityDifference;
            const leftDistance = Number.isFinite(left.distance) ? left.distance : Number.POSITIVE_INFINITY;
            const rightDistance = Number.isFinite(right.distance) ? right.distance : Number.POSITIVE_INFINITY;
            const distanceDifference = leftDistance - rightDistance;
            if (distanceDifference) return distanceDifference;
            const batteryDifference = lowBikesFor(right.result, threshold).length
              - lowBikesFor(left.result, threshold).length;
            return batteryDifference || compareStationNames(left.result, right.result);
          });
          return [{
            mode: "nearest",
            title: priorityEnabled
              ? `≤${priorityThreshold}% 優先，再依距離`
              : "離我最近",
            total: decorated.reduce(
              (sum, item) => sum + (item.result.matched ? lowBikesFor(item.result, threshold).length : 0),
              0,
            ),
            priorityBatteryTotal: decorated.reduce((sum, item) => sum + item.priorityCount, 0),
            priorityStationCount: decorated.filter(item => item.priorityCount > 0).length,
            items: decorated,
          }];
        }

        const districtGroups = new Map();
        for (const item of decorated) {
          const district = canonicalDistrict(item.result.district);
          if (!districtGroups.has(district)) districtGroups.set(district, []);
          districtGroups.get(district).push(item);
        }
        return Array.from(districtGroups.entries())
          .map(([district, items]) => {
            items.sort((left, right) => {
              if (left.result.matched !== right.result.matched) return left.result.matched ? -1 : 1;
              const priorityDifference = comparePriority(left, right);
              if (priorityDifference) return priorityDifference;
              const batteryDifference = lowBikesFor(right.result, threshold).length
                - lowBikesFor(left.result, threshold).length;
              return batteryDifference
                || String(left.result.route || "").localeCompare(String(right.result.route || ""), "zh-Hant", { numeric: true })
                || compareStationNames(left.result, right.result);
            });
            return {
              mode: "district",
              title: district,
              total: items.reduce(
                (sum, item) => sum + (item.result.matched ? lowBikesFor(item.result, threshold).length : 0),
                0,
              ),
              priorityBatteryTotal: items.reduce((sum, item) => sum + item.priorityCount, 0),
              priorityStationCount: items.filter(item => item.priorityCount > 0).length,
              items,
            };
          })
          .sort((left, right) => {
            if (priorityEnabled) {
              const batteryDifference = right.priorityBatteryTotal - left.priorityBatteryTotal;
              if (batteryDifference) return batteryDifference;
              const stationDifference = right.priorityStationCount - left.priorityStationCount;
              if (stationDifference) return stationDifference;
            }
            return left.title.localeCompare(
              right.title,
              "zh-Hant",
              { numeric: true, sensitivity: "base" },
            );
          });
      }
      function renderStationCard(item, threshold, mode, isRecommended) {
        const { result, distance } = item;
        const priorityThreshold = clampPriorityThreshold(priorityThresholdInput.value);
        const isPriority = prioritySortingEnabled() && item.priorityCount > 0;
        const lowBikes = lowBikesFor(result, threshold).sort((left, right) => {
          const leftPillar = String(left.pillar_no || "").trim();
          const rightPillar = String(right.pillar_no || "").trim();
          if (!leftPillar && rightPillar) return 1;
          if (leftPillar && !rightPillar) return -1;
          return leftPillar.localeCompare(
            rightPillar,
            "zh-Hant",
            { numeric: true, sensitivity: "base" },
          ) || left.bike_no.localeCompare(right.bike_no, "zh-Hant", { numeric: true });
        });
        // 防禦性檢查：即使上游資料改變，也絕不建立 0 顆場站卡片。
        if (!result.matched || !lowBikes.length) return;

        const details = doc.createElement("details");
        details.className = "battery-station";
        details.classList.toggle("priority", isPriority);
        const summary = doc.createElement("summary");
        const heading = doc.createElement("span");
        heading.className = "battery-station-heading";
        appendTextElement(heading, "span", "battery-station-name", result.name);

        const meta = doc.createElement("span");
        meta.className = "battery-station-meta";
        if (isPriority) {
          appendTextElement(
            meta,
            "span",
            "battery-priority-badge",
            `⚠️ ≤${priorityThreshold}%：${item.priorityCount} 顆`,
          );
        }
        if (isRecommended) {
          appendTextElement(
            meta,
            "span",
            "battery-recommended",
            isPriority ? "優先推薦" : "最近推薦",
          );
        }
        const metaParts = [String(result.route || "").trim(), canonicalDistrict(result.district)].filter(Boolean);
        if (mode === "nearest") metaParts.push(formatDistance(distance));
        meta.appendChild(doc.createTextNode(metaParts.join("｜")));
        heading.appendChild(meta);
        summary.appendChild(heading);

        const count = appendTextElement(
          summary,
          "span",
          "battery-count",
          `需換 ${lowBikes.length} 顆`,
        );
        count.classList.add("needs-change");
        details.appendChild(summary);

        const list = doc.createElement("div");
        list.className = "battery-bike-list";
        for (const bike of lowBikes) {
          const row = doc.createElement("div");
          row.className = "battery-bike-row";
          const isPriorityBike = isPriority && bike.battery_power <= priorityThreshold;
          row.classList.toggle("priority", isPriorityBike);
          appendTextElement(row, "span", "battery-bike-no", `車號 ${bike.bike_no}`);
          appendTextElement(
            row,
            "span",
            "battery-power",
            `${bike.battery_power}%${isPriorityBike ? " ⚠️" : ""}`,
          );
          appendTextElement(row, "span", "", bike.pillar_no ? `柱號 ${bike.pillar_no}` : "柱號 —");
          list.appendChild(row);
        }
        details.appendChild(list);
        resultsNode.appendChild(details);
      }
      function renderResults() {
        const threshold = clampThreshold(thresholdInput.value);
        thresholdInput.value = String(threshold);
        const priorityThreshold = clampPriorityThreshold(priorityThresholdInput.value);
        priorityThresholdInput.value = String(priorityThreshold);
        const priorityEnabled = prioritySortingEnabled();
        summaryNode.replaceChildren();
        districtSummaryGridNode.replaceChildren();
        districtSummaryNode.hidden = true;
        resultsNode.replaceChildren();
        if (!lastResults.length) return;

        // 後續統計與畫面都只使用「確實至少有 1 顆需更換」的場站。
        const replacementResults = lastResults.filter(
          result => result.matched && lowBikesFor(result, threshold).length > 0,
        );
        const totals = Object.fromEntries(lastResultRoutes.map(route => [route, 0]));
        const districtTotals = new Map();
        for (const result of replacementResults) {
          const lowBatteryCount = lowBikesFor(result, threshold).length;
          totals[result.route] += lowBatteryCount;
          const district = canonicalDistrict(result.district);
          districtTotals.set(district, (districtTotals.get(district) || 0) + lowBatteryCount);
        }
        const grandTotal = Object.values(totals).reduce((sum, value) => sum + value, 0);
        const priorityBatteryTotal = priorityEnabled
          ? replacementResults.reduce(
              (sum, result) => sum + priorityBikesFor(result, priorityThreshold).length,
              0,
            )
          : 0;
        const priorityStationTotal = priorityEnabled
          ? replacementResults.filter(
              result => priorityBikesFor(result, priorityThreshold).length > 0,
            ).length
          : 0;
        for (const route of lastResultRoutes) {
          const metric = doc.createElement("div");
          metric.className = "battery-metric";
          appendTextElement(metric, "div", "battery-metric-label", `${route} 需更換`);
          appendTextElement(metric, "div", "battery-metric-value", `${totals[route] || 0} 顆`);
          summaryNode.appendChild(metric);
        }
        const totalMetric = doc.createElement("div");
        totalMetric.className = "battery-metric total";
        appendTextElement(totalMetric, "div", "battery-metric-label", "所選範圍合計");
        appendTextElement(totalMetric, "div", "battery-metric-value", `${grandTotal} 顆`);
        summaryNode.appendChild(totalMetric);
        if (priorityEnabled) {
          const priorityMetric = doc.createElement("div");
          priorityMetric.className = "battery-metric priority";
          appendTextElement(
            priorityMetric,
            "div",
            "battery-metric-label",
            `⚠️ ≤${priorityThreshold}% 優先`,
          );
          appendTextElement(
            priorityMetric,
            "div",
            "battery-metric-value",
            `${priorityBatteryTotal} 顆／${priorityStationTotal} 站`,
          );
          summaryNode.appendChild(priorityMetric);
        }

        const sortedDistrictTotals = Array.from(districtTotals.entries()).sort(
          ([leftDistrict], [rightDistrict]) => leftDistrict.localeCompare(
            rightDistrict,
            "zh-Hant",
            { numeric: true, sensitivity: "base" },
          ),
        );
        const highestDistrictTotal = Math.max(0, ...sortedDistrictTotals.map(([, total]) => total));
        for (const [district, total] of sortedDistrictTotals) {
          const metric = doc.createElement("div");
          metric.className = "battery-metric battery-district-metric";
          if (total === highestDistrictTotal) metric.classList.add("highest");
          appendTextElement(metric, "div", "battery-metric-label", `${district}｜需更換`);
          appendTextElement(metric, "div", "battery-metric-value", `${total} 顆`);
          districtSummaryGridNode.appendChild(metric);
        }
        districtSummaryTitleNode.textContent = `📍 各行政區需更換電池｜共 ${grandTotal} 顆`;
        districtSummaryNode.hidden = !sortedDistrictTotals.length;

        const groups = buildDisplayGroups(threshold, replacementResults);
        if (!groups.length || groups.every(group => !group.items.length)) {
          appendTextElement(
            resultsNode,
            "div",
            "battery-empty",
            `目前所選範圍沒有電量 ≤ ${threshold}%、需要更換的電池。`,
          );
          savePreferences();
          return;
        }
        for (const group of groups) {
          const prioritySuffix = priorityEnabled && group.priorityBatteryTotal
            ? `｜⚠️ 優先 ${group.priorityBatteryTotal} 顆／${group.priorityStationCount} 站`
            : "";
          const title = group.mode === "nearest"
            ? `${group.title}｜${group.items.length} 個場站${prioritySuffix}`
            : `${group.title}｜${group.total} 顆需更換${prioritySuffix}`;
          appendTextElement(resultsNode, "div", "battery-route-title", title);
          const recommendedIndex = group.mode === "nearest"
            ? group.items.findIndex(item => Number.isFinite(item.distance) && item.result.matched)
            : -1;
          group.items.forEach((item, index) => {
            renderStationCard(item, threshold, group.mode, index === recommendedIndex);
          });
        }
        savePreferences();
      }
      async function runQuery() {
        const selected = selectedRoutes();
        if (!selected.length) {
          setStatus("請至少選擇一個統計範圍。", true);
          return;
        }
        const queryToken = `${fingerprint}-${Date.now()}-${Math.random().toString(16).slice(2)}`;
        win.__ubikeBatteryActiveQueryToken = queryToken;
        refreshButton.disabled = true;
        refreshButton.textContent = "查詢中…";
        summaryNode.replaceChildren();
        districtSummaryGridNode.replaceChildren();
        districtSummaryNode.hidden = true;
        resultsNode.replaceChildren();
        setStatus("正在取得臺東場站與 2.0E 車輛資料……");
        try {
          const catalog = await getCatalog();
          const entries = selected.flatMap(route => (routeStations[route] || []).map(station => ({ ...station, route })));
          let completed = 0;
          let lastPartialRenderCount = 0;
          lastResultRoutes = selected;
          lastResults = await mapWithAdaptiveConcurrency(
            entries,
            isMobileRequestMode ? 3 : 6,
            async (entry, _index) => {
              const matched = matchCatalogStation(entry.name, catalog);
              if (!matched) {
                completed += 1;
                setStatus(`已完成 ${completed}／${entries.length} 個場站……`);
                return { ...entry, matched: false, bikes: [] };
              }
              const stationNo = String(matched.station_no || matched.sno || matched.station_id || "").trim();
              const enrichedEntry = {
                ...entry,
                district: String(
                  entry.district || matched.district_tw || matched.sarea || matched.district || "",
                ).trim(),
                latitude: firstFiniteNumber(matched.lat, matched.latitude),
                longitude: firstFiniteNumber(matched.lng, matched.longitude),
              };
              let bikes = [];
              try { bikes = stationNo ? await getBatteryList(stationNo, { force: true }) : []; }
              catch (_) { return { ...enrichedEntry, matched: false, bikes: [], fetchError: true }; }
              finally {
                completed += 1;
                setStatus(`已完成 ${completed}／${entries.length} 個場站……`);
              }
              return { ...enrichedEntry, matched: true, stationNo, bikes };
            },
            async ({ output, completed: waveCompleted, limit, failureCount }) => {
              if (win.__ubikeBatteryActiveQueryToken !== queryToken) return;
              const renderStep = isMobileRequestMode ? 6 : 12;
              const shouldRender = waveCompleted >= entries.length
                || waveCompleted - lastPartialRenderCount >= renderStep;
              if (shouldRender) {
                lastResults = output.filter(Boolean);
                renderResults();
                lastPartialRenderCount = waveCompleted;
              }
              setStatus(
                `已完成 ${waveCompleted}／${entries.length} 個場站｜自動並行 ${limit} 站`
                + `${failureCount ? "｜偵測到連線不穩，已自動降速" : ""}`,
                false,
              );
            },
          );
          if (win.__ubikeBatteryActiveQueryToken !== queryToken) return;
          renderResults();
          const unmatchedCount = lastResults.filter(
            result => !result.matched && !result.fetchError,
          ).length;
          const fetchErrorCount = lastResults.filter(result => result.fetchError).length;
          const timeText = new Intl.DateTimeFormat("zh-TW", {
            timeZone: "Asia/Taipei", hour: "2-digit", minute: "2-digit", second: "2-digit", hour12: false,
          }).format(new Date());
          setStatus(
            `更新時間 ${timeText}｜${lastResults.length - unmatchedCount - fetchErrorCount} 個場站完成`
            + `${unmatchedCount ? `｜${unmatchedCount} 個未配對` : ""}`
            + `${fetchErrorCount ? `｜${fetchErrorCount} 個場站連線失敗，可再按一次更新` : ""}`
            + "｜速度已依裝置與網路自動調整",
            Boolean(fetchErrorCount),
          );
        } catch (error) {
          const rawMessage = String(error?.message || error);
          const isLoadFailure = error?.name === "AbortError"
            || /load failed|failed to fetch|networkerror|network request failed/i.test(rawMessage);
          const message = isLoadFailure
            ? "手機網路連線中斷或查詢逾時，請縮小查詢範圍後再試。"
            : rawMessage;
          setStatus(`電量查詢失敗：${message}`, true);
        } finally {
          if (win.__ubikeBatteryActiveQueryToken === queryToken) {
            refreshButton.disabled = false;
            refreshButton.textContent = "更新電量";
          }
        }
      }
      function openPage() {
        if (!page.classList.contains("open")) {
          previousBodyOverflow = doc.body.style.overflow;
          win.__ubikeBatteryPreviousBodyOverflow = previousBodyOverflow;
        }
        doc.body.style.overflow = "hidden";
        page.classList.add("open");
        page.setAttribute("aria-hidden", "false");
        page.scrollTop = 0;
        if (normalizeSortMode(sortSelect.value) === "nearest") {
          const locationIsStale = !userLocation
            || Date.now() - Number(userLocation.updatedAt || 0) > 5 * 60 * 1000;
          if (locationIsStale) requestCurrentLocation({ force: Boolean(userLocation) });
        }
      }
      function closePage() {
        page.classList.remove("open");
        page.setAttribute("aria-hidden", "true");
        doc.body.style.overflow = String(
          win.__ubikeBatteryPreviousBodyOverflow ?? previousBodyOverflow ?? "",
        );
      }

      const preferences = readPreferences();
      thresholdInput.value = String(clampThreshold(preferences.threshold ?? defaultThreshold));
      priorityEnabledInput.checked = Boolean(preferences.priority_enabled);
      priorityThresholdInput.value = String(
        clampPriorityThreshold(preferences.priority_threshold ?? defaultPriorityThreshold),
      );
      sortSelect.value = normalizeSortMode(preferences.sort_mode);
      for (const route of routes) {
        const label = doc.createElement("label");
        label.className = "battery-scope";
        const input = doc.createElement("input");
        input.type = "checkbox";
        input.value = route;
        input.checked = false;
        label.classList.toggle("selected", input.checked);
        label.append(input, doc.createTextNode(route));
        input.addEventListener("change", () => {
          label.classList.toggle("selected", input.checked);
          savePreferences();
        });
        scopeList.appendChild(label);
      }
      updatePriorityControls();
      updateSortControls();

      fab.addEventListener("click", openPage);
      backButton.addEventListener("click", closePage);
      refreshButton.addEventListener("click", runQuery);
      locateButton.addEventListener("click", () => requestCurrentLocation({ force: true }));
      sortSelect.addEventListener("change", () => {
        sortSelect.value = normalizeSortMode(sortSelect.value);
        savePreferences();
        updateSortControls();
        if (sortSelect.value === "nearest" && !userLocation) requestCurrentLocation();
        else if (lastResults.length) renderResults();
      });
      thresholdInput.addEventListener("change", () => {
        thresholdInput.value = String(clampThreshold(thresholdInput.value));
        priorityThresholdInput.value = String(
          clampPriorityThreshold(priorityThresholdInput.value),
        );
        updatePriorityControls();
        if (lastResults.length) renderResults();
        else savePreferences();
      });
      priorityEnabledInput.addEventListener("change", () => {
        updatePriorityControls();
        if (lastResults.length) renderResults();
        else savePreferences();
      });
      priorityThresholdInput.addEventListener("change", () => {
        priorityThresholdInput.value = String(
          clampPriorityThreshold(priorityThresholdInput.value),
        );
        updatePriorityControls();
        if (lastResults.length) renderResults();
        else savePreferences();
      });
      win.addEventListener("keydown", event => {
        if (event.key === "Escape" && page.classList.contains("open")) closePage();
      });

      if (displayMode === "mobile" || win.matchMedia("(max-width: 700px)").matches) {
        fab.style.bottom = "calc(312px + env(safe-area-inset-bottom, 0px))";
      }
      if (previousPageWasOpen) openPage();
    })();
    </script>
    """
    component_html = (
        component_html
        .replace("__LOW_BATTERY_CLIENT_CORE__", LOW_BATTERY_CLIENT_CORE_JS)
        .replace("__ROUTE_STATIONS__", route_payload)
        .replace("__DISPLAY_MODE__", display_mode)
        .replace("__BATTERY_FINGERPRINT__", battery_fingerprint)
    )
    return component_html


def render_floating_battery_query(
    route_station_map: dict[str, list[dict]],
    mobile_mode: bool,
) -> None:
    """建立懸浮電量入口與全螢幕查詢頁；資料由使用者瀏覽器直讀 YouBike。"""
    components.html(
        _build_floating_battery_query_html(route_station_map, mobile_mode),
        height=0,
        scrolling=False,
    )

def build_battery_route_station_map(status_df: pd.DataFrame) -> dict[str, list[dict]]:
    """由目前選定配置整理 D1／D2／D3 場站，供 2.0E 電量查詢使用。"""
    route_station_map: dict[str, list[dict]] = {}
    if status_df.empty or "場站名稱" not in status_df.columns or "路線區域" not in status_df.columns:
        return route_station_map

    seen_by_route: dict[str, set[str]] = {}
    station_rows = status_df.reindex(
        columns=["路線區域", "場站名稱", "行政區"]
    ).itertuples(index=False, name=None)
    for route_raw, station_name_raw, district_raw in station_rows:
        route = "" if pd.isna(route_raw) else str(route_raw).strip().upper()
        station_name = (
            "" if pd.isna(station_name_raw) else str(station_name_raw).strip()
        )
        station_key = normalize_youbike_station_key(station_name)
        if route not in ("D1", "D2", "D3") or not station_key:
            continue
        route_seen = seen_by_route.setdefault(route, set())
        if station_key in route_seen:
            continue
        route_seen.add(station_key)
        route_station_map.setdefault(route, []).append(
            {
                "name": station_name,
                "district": "" if pd.isna(district_raw) else str(district_raw).strip(),
            }
        )

    return {
        route: route_station_map[route]
        for route in ("D1", "D2", "D3")
        if route_station_map.get(route)
    }



def get_dispatch_geolocation_component():
    """建立可由手機瀏覽器回傳目前經緯度的 Streamlit 雙向元件。"""
    global _DISPATCH_GEOLOCATION_COMPONENT
    if _DISPATCH_GEOLOCATION_COMPONENT is not None:
        return _DISPATCH_GEOLOCATION_COMPONENT

    component_dir = Path(tempfile.gettempdir()) / "dispatch_geolocation_component_v3"
    component_dir.mkdir(parents=True, exist_ok=True)
    index_path = component_dir / "index.html"
    try:
        if not index_path.exists() or index_path.read_text(encoding="utf-8") != DISPATCH_GEOLOCATION_COMPONENT_HTML:
            index_path.write_text(DISPATCH_GEOLOCATION_COMPONENT_HTML, encoding="utf-8")
    except OSError as exc:
        raise RuntimeError(f"無法建立定位元件：{exc}") from exc

    _DISPATCH_GEOLOCATION_COMPONENT = components.declare_component(
        "dispatch_geolocation_v3",
        path=str(component_dir),
    )
    return _DISPATCH_GEOLOCATION_COMPONENT


def normalize_coordinate(value, minimum: float, maximum: float) -> float | None:
    try:
        coordinate = float(value)
    except (TypeError, ValueError, OverflowError):
        return None
    if not math.isfinite(coordinate) or not minimum <= coordinate <= maximum:
        return None
    return coordinate


LONG_DISTANCE_START_POINTS = {
    "台東維調": {
        "description": "大忠路30號附近",
        # 使用約略中心點；實際執行時會優先採用每 30 秒更新的 GPS 位置。
        "latitude": 22.7418,
        "longitude": 121.1266,
    },
    "池上維調": {
        "description": "池上火車站旁轉角",
        "latitude": 23.1260174,
        "longitude": 121.219459,
    },
}
ALL_DISPATCH_ZONES = ("D1", "D2", "D3")
LONG_DISTANCE_ROUTE_ZONES = ("D2", "D3")
LONG_DISTANCE_LOOP_DIRECTION_OPTIONS = ("AI 自動選擇", "D2 先行", "D3 先行")
LONG_DISTANCE_TRANSFER_LABEL = "玉長公路"
SHARED_GEOLOCATION_REFRESH_SECONDS = 30


def normalize_dispatch_zone(value) -> str | None:
    """把配置中的路線名稱辨識為 D1／D2／D3。"""
    normalized = re.sub(r"\s+", "", str(value or "").upper())
    for zone in ALL_DISPATCH_ZONES:
        if zone in normalized:
            return zone
    return None


def normalize_long_distance_zone(value) -> str | None:
    """長途環狀邏輯只接受 D2／D3。"""
    zone = normalize_dispatch_zone(value)
    return zone if zone in LONG_DISTANCE_ROUTE_ZONES else None


def configuration_options_for_type(
    options: list[tuple[str, str]],
    configuration_type: str,
) -> tuple[list[tuple[str, str]], bool]:
    """依工作表名稱找出指定配置；找不到時安全退回全部工作表。"""
    keywords = CONFIGURATION_TYPE_KEYWORDS.get(configuration_type, ())
    matched = [
        option for option in options
        if any(keyword.lower() in str(option[0]).lower() for keyword in keywords)
    ]
    return (matched, True) if matched else (list(options), False)


def preferred_configuration_sheet(
    options: list[tuple[str, str]],
    configuration_type: str = "",
) -> str:
    """優先選指定配置類型中，同時涵蓋 D1／D2／D3 最完整的工作表。"""
    scoped_options, _matched = configuration_options_for_type(options, configuration_type)
    sheet_order: list[str] = []
    coverage: dict[str, set[str]] = {}
    for sheet_name, route in scoped_options:
        if sheet_name not in coverage:
            coverage[sheet_name] = set()
            sheet_order.append(sheet_name)
        zone = normalize_dispatch_zone(route)
        if zone:
            coverage[sheet_name].add(zone)
    if not sheet_order:
        return ""
    return max(sheet_order, key=lambda sheet: (len(coverage.get(sheet, set())), -sheet_order.index(sheet)))


def location_payload_is_valid(location) -> bool:
    if not isinstance(location, dict):
        return False
    return (
        normalize_coordinate(location.get("latitude"), -90.0, 90.0) is not None
        and normalize_coordinate(location.get("longitude"), -180.0, 180.0) is not None
    )


def newest_valid_location(*locations) -> dict | None:
    """在 GPS、上一完成場站與固定起點中取最新且有效的座標。"""
    valid_locations = [dict(location) for location in locations if location_payload_is_valid(location)]
    if not valid_locations:
        return None
    return max(valid_locations, key=lambda item: float(item.get("updated_at") or 0))


def render_shared_geolocation(active_base: dict) -> dict | None:
    """配置表載入後即啟動背景定位，之後每 30 秒更新一次。"""
    token = str(active_base.get("token") or "").strip()
    if not token:
        return None

    prefix = f"shared_geolocation::{token}"
    state_key = f"{prefix}::state"
    request_key = f"{prefix}::request_token"
    processed_event_key = f"{prefix}::processed_event"

    payload = None
    try:
        geolocation_component = get_dispatch_geolocation_component()
        payload = geolocation_component(
            key=f"dispatch_geolocation_background::{token}",
            default=None,
            request_token=str(st.session_state.get(request_key) or ""),
            auto_start=True,
            auto_refresh=True,
            auto_refresh_seconds=SHARED_GEOLOCATION_REFRESH_SECONDS,
            compact=True,
        )
    except Exception as exc:
        st.session_state[f"{prefix}::error"] = str(exc)

    if isinstance(payload, dict):
        event_id = str(payload.get("event_id") or "").strip()
        if event_id and st.session_state.get(processed_event_key) != event_id:
            st.session_state[processed_event_key] = event_id
            response_request_token = str(payload.get("request_token") or "").strip()
            active_request_token = str(st.session_state.get(request_key) or "").strip()
            if response_request_token and response_request_token == active_request_token:
                st.session_state.pop(request_key, None)

            if payload.get("ok"):
                latitude = normalize_coordinate(payload.get("latitude"), -90.0, 90.0)
                longitude = normalize_coordinate(payload.get("longitude"), -180.0, 180.0)
                if latitude is not None and longitude is not None:
                    st.session_state[state_key] = {
                        "latitude": latitude,
                        "longitude": longitude,
                        "accuracy": max(0.0, float(payload.get("accuracy") or 0)),
                        "updated_at": time.time(),
                        "source": "gps",
                    }
                    st.session_state.pop(f"{prefix}::error", None)
            else:
                st.session_state[f"{prefix}::error"] = str(payload.get("error") or "定位失敗")

    location = st.session_state.get(state_key)
    return dict(location) if location_payload_is_valid(location) else None


def request_shared_geolocation_refresh(active_base: dict) -> None:
    token = str(active_base.get("token") or "").strip()
    if token:
        st.session_state[f"shared_geolocation::{token}::request_token"] = uuid.uuid4().hex


def render_shared_location_summary(active_base: dict, location: dict | None) -> None:
    """顯示精簡定位狀態；定位元件本體在背景運作，不佔主畫面。"""
    token = str(active_base.get("token") or "").strip()
    error_text = str(st.session_state.get(f"shared_geolocation::{token}::error") or "").strip()
    if location_payload_is_valid(location):
        updated_at = datetime.fromtimestamp(
            float(location.get("updated_at") or time.time()),
            TAIPEI_TIMEZONE,
        ).strftime("%H:%M:%S")
        accuracy = float(location.get("accuracy") or 0)
        accuracy_text = f"｜誤差約 {accuracy:.0f} 公尺" if accuracy else ""
        st.caption(f"📍 定位正常｜每 {SHARED_GEOLOCATION_REFRESH_SECONDS} 秒更新｜最後更新 {updated_at}{accuracy_text}")
    elif error_text:
        st.caption(f"🔴 定位尚未啟用：{error_text}")
    else:
        st.caption(f"🟡 正在取得定位｜取得後每 {SHARED_GEOLOCATION_REFRESH_SECONDS} 秒更新")


def build_long_distance_status_dataframe(
    *,
    active_base: dict,
    options: list[tuple[str, str]],
    selected_sheet: str,
    selected_shift: str,
    selected_zones: list[str],
    status_cache: dict,
) -> tuple[pd.DataFrame, dict[str, dict]]:
    """整合指定 D1／D2／D3 配置、既有現況、最新即時車數與官方座標。"""
    selected_zone_set = {zone for zone in selected_zones if zone in ALL_DISPATCH_ZONES}
    if not selected_zone_set:
        return pd.DataFrame(), {}

    # D1、D2、D3 可能分別放在不同工作表。每一區先找自動選定的主工作表，
    # 找不到時再到其他可用資料中補齊，確保三區都會被讀取。
    chosen_by_zone: dict[str, tuple[str, str]] = {}
    for zone in ALL_DISPATCH_ZONES:
        if zone not in selected_zone_set:
            continue
        same_sheet_match = next(
            (
                (sheet_name, route)
                for sheet_name, route in options
                if sheet_name == selected_sheet and normalize_dispatch_zone(route) == zone
            ),
            None,
        )
        fallback_match = next(
            (
                (sheet_name, route)
                for sheet_name, route in options
                if normalize_dispatch_zone(route) == zone
            ),
            None,
        )
        chosen = same_sheet_match or fallback_match
        if chosen is not None:
            chosen_by_zone[zone] = chosen

    latest_live_records = st.session_state.get(f"latest_live_records::{active_base['token']}")
    if not isinstance(latest_live_records, list):
        latest_live_records = []
    latest_live_event_id = str(
        st.session_state.get(f"latest_live_event_id::{active_base['token']}") or ""
    ).strip()
    latest_live_fetched_at = str(
        st.session_state.get(f"latest_live_fetched_at::{active_base['token']}") or ""
    ).strip()
    latest_live_match_index: dict[str, object] | None = None
    latest_live_match_cache: dict[str, tuple[dict | None, float, bool]] = {}

    frames: list[pd.DataFrame] = []
    combined_locations: dict[str, dict] = {}
    cache_changed = False

    for zone in ALL_DISPATCH_ZONES:
        if zone not in selected_zone_set or zone not in chosen_by_zone:
            continue
        sheet_name, route = chosen_by_zone[zone]
        context_key = status_context_key(sheet_name, route, selected_shift)
        route_df = cached_parse_route(active_base["bytes"], sheet_name, route, selected_shift)
        route_df = blank_current_status(route_df)
        saved_records = status_cache.get("contexts", {}).get(context_key)
        if saved_records is not None:
            route_df = restore_current_status(route_df, saved_records)

        metadata = status_cache.setdefault("metadata", {}).setdefault(context_key, {})
        should_apply_live_event = bool(
            latest_live_records
            and latest_live_event_id
            and str(metadata.get("last_live_event_id") or "") != latest_live_event_id
        )
        if should_apply_live_event:
            if latest_live_match_index is None:
                latest_live_match_index = build_youbike_match_index(latest_live_records)
            route_df, _report_df, summary = apply_youbike_updates_to_dataframe(
                route_df,
                latest_live_records,
                match_index=latest_live_match_index,
                match_cache=latest_live_match_cache,
            )
            if safe_nonnegative_int(summary.get("matched_count")) > 0:
                records = dataframe_to_status_records(route_df)
                if records != status_cache.setdefault("contexts", {}).get(context_key):
                    status_cache["contexts"][context_key] = records
                    cache_changed = True
                location_map = build_youbike_station_location_map(
                    route_df,
                    latest_live_records,
                    match_index=latest_live_match_index,
                    match_cache=latest_live_match_cache,
                )
                previous_locations = metadata.get("station_locations", {})
                if not isinstance(previous_locations, dict):
                    previous_locations = {}
                merged_locations = dict(previous_locations)
                merged_locations.update(location_map)
                metadata["station_locations"] = merged_locations
                metadata["last_live_event_id"] = latest_live_event_id
                if latest_live_fetched_at:
                    metadata["fetched_at"] = latest_live_fetched_at
                combined_locations.update(merged_locations)
                cache_changed = True

        metadata = status_cache.get("metadata", {}).get(context_key, {})
        if isinstance(metadata, dict) and isinstance(metadata.get("station_locations"), dict):
            combined_locations.update(metadata["station_locations"])

        route_df = route_df.copy()
        route_df["路線區域"] = zone
        route_df["配置來源"] = f"{route}｜{sheet_name}"
        route_df["_狀態內容鍵"] = context_key
        frames.append(route_df)

    if cache_changed:
        save_cached_status(active_base["token"], active_base["expires_at"], status_cache)

    if not frames:
        return pd.DataFrame(), combined_locations

    combined_df = pd.concat(frames, ignore_index=True)
    combined_df = combined_df.drop_duplicates(
        subset=["行政區", "場站名稱"],
        keep="first",
    ).reset_index(drop=True)
    return combined_df, combined_locations


def save_dispatch_dataframe_contexts(
    updated_df: pd.DataFrame,
    *,
    status_cache: dict,
    active_base: dict,
    default_context_key: str,
) -> None:
    """一般配置存單一內容鍵；D2／D3 合併頁則分別寫回原本的內容鍵。"""
    if "_狀態內容鍵" in updated_df.columns:
        for context_key, context_df in updated_df.groupby("_狀態內容鍵", sort=False):
            normalized_context_key = str(context_key or "").strip()
            if not normalized_context_key:
                continue
            status_cache.setdefault("contexts", {})[normalized_context_key] = dataframe_to_status_records(context_df)
    else:
        status_cache.setdefault("contexts", {})[default_context_key] = dataframe_to_status_records(updated_df)
    save_cached_status(active_base["token"], active_base["expires_at"], status_cache)


def adjust_candidates_for_trip_mode(
    candidates: list[dict],
    *,
    trip_mode: str,
    endpoint_location: dict | None,
) -> list[dict]:
    """以實際道路成本納入單趟終點、來回返程與環狀回維調方向。"""
    endpoint_valid = location_payload_is_valid(endpoint_location)
    endpoint_metrics = (
        road_metrics_to_endpoint(candidates, endpoint_location)
        if endpoint_valid and candidates and isinstance(endpoint_location, dict)
        else {}
    )

    adjusted: list[dict] = []
    for original in candidates:
        candidate = dict(original)
        endpoint_distance_km = 0.0
        endpoint_drive_minutes = 0.0
        endpoint_metric = endpoint_metrics.get(str(candidate.get("station_name") or ""), {})

        if endpoint_valid:
            if endpoint_metric.get("road_route_available") is False:
                # 指定要返回／抵達某處時，無道路連通的候選站不納入。
                continue
            endpoint_distance_km = float(endpoint_metric.get("road_distance_km") or 0.0)
            endpoint_drive_minutes = float(endpoint_metric.get("drive_minutes") or 0.0)

        # 一般模式只評估「目前位置 → 單一場站」的即時效益，不安排後續路線。
        if trip_mode == "一般模式":
            endpoint_weight = 0.0
        elif trip_mode == "來回":
            endpoint_weight = 1.0
        elif trip_mode == "環狀一圈":
            endpoint_weight = 0.20 if endpoint_valid else 0.0
        else:
            endpoint_weight = 0.35 if endpoint_valid else 0.0

        route_total_minutes = (
            float(candidate["estimated_total_minutes"])
            + endpoint_drive_minutes * endpoint_weight
        )
        candidate["endpoint_distance_km"] = endpoint_distance_km
        candidate["endpoint_drive_minutes"] = endpoint_drive_minutes
        candidate["endpoint_routing_fallback"] = bool(endpoint_metric.get("routing_fallback"))
        candidate["route_total_minutes"] = route_total_minutes
        candidate["base_score"] = float(candidate.get("score") or 0)
        candidate["score"] = (
            safe_nonnegative_int(candidate.get("dispatch_count"))
            / max(1.0, route_total_minutes)
            * float(candidate.get("reason_multiplier") or 1.0)
        )
        adjusted.append(candidate)

    return sorted(
        adjusted,
        key=lambda item: (
            float(item.get("score") or 0),
            safe_nonnegative_int(item.get("dispatch_count")),
            -float(item.get("estimated_distance_km") or 0),
        ),
        reverse=True,
    )

def loop_zone_order_from_preference(preference: str) -> list[str] | None:
    """將畫面上的環狀方向轉成實際 D2／D3 執行順序；AI 選擇會回傳 None。"""
    normalized = str(preference or "").strip()
    if normalized == "D2 先行":
        return ["D2", "D3"]
    if normalized == "D3 先行":
        return ["D3", "D2"]
    return None


def loop_movement_direction(start_name: str, phase_index: int) -> int:
    """相容舊版呼叫；v24.1 起不再以南北緯度限制場站。"""
    del start_name, phase_index
    return 0


def adjust_candidates_for_loop_direction(
    candidates: list[dict],
    *,
    current_location: dict,
    movement_direction: int,
) -> list[dict]:
    """相容舊版資料；路線改由實際道路時間決定，不再因緯度方向降權。"""
    del current_location, movement_direction
    return candidates

def summarize_loop_preview(preview: list[dict]) -> tuple[float, int, float, int]:
    """回傳環狀預覽比較鍵：效益、調度量、負時間與涵蓋區域數。"""
    if not preview:
        return (0.0, 0, float("-inf"), 0)
    total_dispatch = sum(safe_nonnegative_int(item.get("dispatch_count")) for item in preview)
    total_minutes = sum(float(item.get("estimated_total_minutes") or 0) for item in preview)
    zones = {str(item.get("route_zone") or "") for item in preview if item.get("route_zone")}
    efficiency = total_dispatch / max(1.0, total_minutes)
    return (efficiency, total_dispatch, -total_minutes, len(zones))


def build_long_distance_route_preview(
    dispatch_df: pd.DataFrame,
    *,
    station_locations: dict[str, dict],
    current_location: dict,
    truck_bike: int,
    truck_ebike: int,
    max_capacity: int,
    cooldowns: dict[str, dict],
    rejection_history: list[dict],
    now_timestamp: float,
    current_round: int,
    trip_mode: str,
    endpoint_location: dict | None,
    forced_first_station: str = "",
    max_stops: int = 6,
    loop_zone_order: list[str] | None = None,
    loop_start_name: str = "",
    active_loop_phase: str = "",
) -> list[dict]:
    """逐站模擬路線；環狀模式依 D2／D3 階段前進，中途只經玉長公路跨區一次。"""
    work_df = dispatch_df.copy().reset_index(drop=True)
    simulated_location = dict(current_location)
    simulated_truck_bike = safe_nonnegative_int(truck_bike)
    simulated_truck_ebike = safe_nonnegative_int(truck_ebike)
    preview: list[dict] = []

    if trip_mode == "環狀一圈" and loop_zone_order:
        zone_sequence = [zone for zone in loop_zone_order if zone in LONG_DISTANCE_ROUTE_ZONES]
        if active_loop_phase in zone_sequence:
            zone_sequence = zone_sequence[zone_sequence.index(active_loop_phase):]
    else:
        zone_sequence = [""]

    original_zone_order = [zone for zone in (loop_zone_order or []) if zone in LONG_DISTANCE_ROUTE_ZONES]
    for zone in zone_sequence:
        while len(preview) < max(1, max_stops):
            phase_df = work_df
            phase_index = 0
            if zone:
                phase_df = work_df[
                    work_df["路線區域"].astype(str).map(normalize_long_distance_zone).eq(zone)
                ].copy()
                phase_index = original_zone_order.index(zone) if zone in original_zone_order else 0
            if phase_df.empty:
                break

            candidates = calculate_dispatch_candidates(
                phase_df,
                station_locations=station_locations,
                current_location=simulated_location,
                truck_bike=simulated_truck_bike,
                truck_ebike=simulated_truck_ebike,
                max_capacity=max_capacity,
                cooldowns=cooldowns,
                rejection_history=rejection_history,
                now_timestamp=now_timestamp,
                current_round=current_round,
            )
            candidates = adjust_candidates_for_trip_mode(
                candidates,
                trip_mode=trip_mode,
                endpoint_location=endpoint_location,
            )
            if not candidates:
                break

            chosen = candidates[0]
            if not preview and forced_first_station:
                forced = next(
                    (candidate for candidate in candidates if candidate["station_name"] == forced_first_station),
                    None,
                )
                if forced is not None:
                    chosen = forced

            chosen = dict(chosen)
            chosen["preview_order"] = len(preview) + 1
            chosen["route_zone"] = zone or str(chosen.get("route_zone") or "")
            if preview and zone and str(preview[-1].get("route_zone") or "") != zone:
                chosen["crossing_before"] = LONG_DISTANCE_TRANSFER_LABEL
            preview.append(chosen)

            station_mask = (
                work_df["場站名稱"].astype(str).eq(str(chosen["station_name"]))
                & work_df["行政區"].astype(str).eq(str(chosen.get("region") or ""))
            )
            if not station_mask.any():
                break
            work_df.loc[station_mask, "2.0 現況"] = (
                safe_nonnegative_int(chosen.get("current_bike"))
                + safe_nonnegative_int(chosen.get("unload_bike"))
                - safe_nonnegative_int(chosen.get("pickup_bike"))
            )
            work_df.loc[station_mask, "2.0E 現況"] = (
                safe_nonnegative_int(chosen.get("current_ebike"))
                + safe_nonnegative_int(chosen.get("unload_ebike"))
                - safe_nonnegative_int(chosen.get("pickup_ebike"))
            )
            work_df = work_df.loc[~station_mask].reset_index(drop=True)
            simulated_truck_bike = safe_nonnegative_int(chosen.get("truck_after_bike"))
            simulated_truck_ebike = safe_nonnegative_int(chosen.get("truck_after_ebike"))
            simulated_location = {
                "latitude": float(chosen["latitude"]),
                "longitude": float(chosen["longitude"]),
                "updated_at": now_timestamp + len(preview),
                "source": "route_preview",
            }

        if len(preview) >= max(1, max_stops):
            break

    return preview

def rerank_candidates_with_road_lookahead(
    candidates: list[dict],
    *,
    dispatch_df: pd.DataFrame,
    station_locations: dict[str, dict],
    current_location: dict,
    truck_bike: int,
    truck_ebike: int,
    max_capacity: int,
    cooldowns: dict[str, dict],
    rejection_history: list[dict],
    now_timestamp: float,
    current_round: int,
    trip_mode: str,
    endpoint_location: dict | None,
    loop_zone_order: list[str] | None,
    loop_start_name: str,
    active_loop_phase: str,
) -> list[dict]:
    """預看後續三站再決定第一站；只評估前四名，兼顧品質與速度。"""
    if len(candidates) <= 1:
        return candidates

    option_count = min(ROAD_ROUTER_LOOKAHEAD_OPTIONS, len(candidates))
    evaluated: list[dict] = []
    for candidate in candidates[:option_count]:
        preview = build_long_distance_route_preview(
            dispatch_df,
            station_locations=station_locations,
            current_location=current_location,
            truck_bike=truck_bike,
            truck_ebike=truck_ebike,
            max_capacity=max_capacity,
            cooldowns=cooldowns,
            rejection_history=rejection_history,
            now_timestamp=now_timestamp,
            current_round=current_round,
            trip_mode=trip_mode,
            endpoint_location=endpoint_location,
            forced_first_station=str(candidate["station_name"]),
            max_stops=ROAD_ROUTER_LOOKAHEAD_STOPS,
            loop_zone_order=loop_zone_order,
            loop_start_name=loop_start_name,
            active_loop_phase=active_loop_phase,
        )
        total_dispatch = sum(safe_nonnegative_int(item.get("dispatch_count")) for item in preview)
        total_minutes = sum(float(item.get("estimated_total_minutes") or 0.0) for item in preview)
        lookahead_efficiency = total_dispatch / max(1.0, total_minutes)

        updated = dict(candidate)
        immediate_score = float(candidate.get("score") or 0.0)
        updated["immediate_score"] = immediate_score
        updated["lookahead_score"] = lookahead_efficiency
        updated["lookahead_dispatch_count"] = total_dispatch
        updated["lookahead_total_minutes"] = total_minutes
        # 即時本站占 55%，後續道路連續性占 45%；偏遠單站會因下一段成本自然降權。
        updated["score"] = immediate_score * 0.55 + lookahead_efficiency * 0.45
        evaluated.append(updated)

    evaluated.sort(
        key=lambda item: (
            float(item.get("score") or 0.0),
            safe_nonnegative_int(item.get("lookahead_dispatch_count")),
            -float(item.get("lookahead_total_minutes") or 0.0),
        ),
        reverse=True,
    )
    return [*evaluated, *candidates[option_count:]]


def render_long_distance_route_preview(
    preview: list[dict],
    *,
    trip_mode: str,
    endpoint_label: str,
    loop_zone_order: list[str] | None = None,
) -> None:
    if not preview:
        return
    total_dispatch = sum(safe_nonnegative_int(plan.get("dispatch_count")) for plan in preview)
    total_distance = sum(float(plan.get("estimated_distance_km") or 0) for plan in preview)
    total_minutes = sum(float(plan.get("estimated_total_minutes") or 0) for plan in preview)
    if endpoint_label and location_payload_is_valid(preview[-1]) and float(preview[-1].get("endpoint_distance_km") or 0) > 0:
        total_distance += float(preview[-1].get("endpoint_distance_km") or 0)
        total_minutes += float(preview[-1].get("endpoint_drive_minutes") or 0)

    st.markdown(
        f"**路線預覽｜{len(preview)} 站｜預計調度 {total_dispatch} 台｜約 {total_distance:.1f} km｜約 {total_minutes:.0f} 分鐘**"
    )
    if trip_mode == "環狀一圈" and loop_zone_order:
        st.info(
            f"環狀方向：{loop_zone_order[0]} 先行 → 經 {LONG_DISTANCE_TRANSFER_LABEL} → "
            f"{loop_zone_order[1]} → 返回出發維調"
        )

    preview_rows = []
    previous_zone = ""
    for plan in preview:
        route_zone = str(plan.get("route_zone") or "")
        if previous_zone and route_zone and route_zone != previous_zone:
            preview_rows.append(
                {
                    "順序": "↔",
                    "區域": "轉場",
                    "場站": f"經 {LONG_DISTANCE_TRANSFER_LABEL} 前往 {route_zone}",
                    "作業": "跨越海岸山脈，不在兩區間反覆折返",
                    "調度量": "—",
                    "距離(km)": "估算",
                    "預估(分)": "依路況",
                }
            )
        preview_rows.append(
            {
                "順序": safe_nonnegative_int(plan.get("preview_order")),
                "區域": route_zone or plan.get("region", ""),
                "場站": plan.get("station_name", ""),
                "作業": dispatch_action_text(plan),
                "調度量": safe_nonnegative_int(plan.get("dispatch_count")),
                "路網距離(km)": round(float(plan.get("estimated_distance_km") or 0), 1),
                "預估(分)": round(float(plan.get("estimated_total_minutes") or 0)),
                "道路資料": "備援估算" if plan.get("routing_fallback") else "實際路網",
            }
        )
        if route_zone:
            previous_zone = route_zone
    st.dataframe(pd.DataFrame(preview_rows), hide_index=True, use_container_width=True)
    if trip_mode == "來回":
        st.caption(f"最後一站後會返回：{endpoint_label}")
    elif trip_mode == "環狀一圈":
        st.caption(f"完成第二區後會返回：{endpoint_label}")
    elif endpoint_label:
        st.caption(f"單趟路線會逐步朝終點方向安排：{endpoint_label}")
    st.caption("此為目前即時資料與道路路網的路線預覽；每完成一站或資料變動後，只重排未鎖定的目前階段與後續路線。")

def build_youbike_station_location_map(
    base_df: pd.DataFrame,
    live_records: list[dict],
    *,
    match_index: dict[str, object] | None = None,
    match_cache: dict[str, tuple[dict | None, float, bool]] | None = None,
) -> dict[str, dict]:
    """把 Excel 場站名稱安全配對到官方站號及經緯度，供智慧調度計算距離。"""
    location_map: dict[str, dict] = {}
    resolved_match_index = (
        match_index
        if isinstance(match_index, dict)
        else build_youbike_match_index(live_records)
    )
    resolved_match_cache = match_cache if isinstance(match_cache, dict) else {}
    for station_name in base_df["場站名稱"].astype(str).drop_duplicates():
        cached_match = resolved_match_cache.get(station_name)
        if cached_match is None:
            cached_match = match_youbike_station(
                station_name,
                live_records,
                resolved_match_index,
            )
            resolved_match_cache[station_name] = cached_match
        matched_record, _score, ambiguous = cached_match
        if matched_record is None or ambiguous:
            continue
        latitude = normalize_coordinate(matched_record.get("latitude"), -90.0, 90.0)
        longitude = normalize_coordinate(matched_record.get("longitude"), -180.0, 180.0)
        if latitude is None or longitude is None:
            continue
        location_map[station_name] = {
            "station_id": str(matched_record.get("station_id") or "").strip(),
            "official_name": str(matched_record.get("station_name") or station_name).strip(),
            "latitude": latitude,
            "longitude": longitude,
        }
    return location_map


def haversine_distance_km(
    origin_latitude: float,
    origin_longitude: float,
    destination_latitude: float,
    destination_longitude: float,
) -> float:
    """計算兩個座標間的大圓直線距離。"""
    earth_radius_km = 6371.0088
    lat1 = math.radians(origin_latitude)
    lat2 = math.radians(destination_latitude)
    delta_lat = math.radians(destination_latitude - origin_latitude)
    delta_lon = math.radians(destination_longitude - origin_longitude)
    value = (
        math.sin(delta_lat / 2) ** 2
        + math.cos(lat1) * math.cos(lat2) * math.sin(delta_lon / 2) ** 2
    )
    return earth_radius_km * 2 * math.atan2(math.sqrt(value), math.sqrt(max(0.0, 1 - value)))


class RoadRoutingError(RuntimeError):
    """道路路網服務連線或資料格式異常。"""


def _rounded_coordinate(longitude: float, latitude: float, *, precision: int) -> tuple[float, float]:
    return (round(float(longitude), precision), round(float(latitude), precision))


@st.cache_data(
    show_spinner=False,
    ttl=ROAD_ROUTER_CACHE_TTL_SECONDS,
    max_entries=512,
)
def fetch_road_table_cached(
    coordinates: tuple[tuple[float, float], ...],
    sources: tuple[int, ...],
    destinations: tuple[int, ...],
) -> dict:
    """批次取得道路行車時間／距離矩陣；同一座標組合五分鐘內直接使用快取。"""
    if len(coordinates) < 2 or not sources or not destinations:
        raise RoadRoutingError("道路矩陣缺少來源或目的地座標。")

    coordinate_text = ";".join(
        f"{longitude:.5f},{latitude:.5f}" for longitude, latitude in coordinates
    )
    source_text = ";".join(str(index) for index in sources)
    destination_text = ";".join(str(index) for index in destinations)
    url = (
        f"{ROAD_ROUTER_BASE_URL}/table/v1/{ROAD_ROUTER_PROFILE}/{coordinate_text}"
        f"?annotations=duration,distance&sources={source_text}&destinations={destination_text}"
    )

    last_error: Exception | None = None
    for attempt in range(1, ROAD_ROUTER_MAX_ATTEMPTS + 1):
        request = Request(
            url,
            headers={
                "Accept": "application/json",
                "User-Agent": f"Taitung-YouBike-Dispatch/{APP_VERSION}",
            },
            method="GET",
        )
        try:
            with urlopen(request, timeout=ROAD_ROUTER_TIMEOUT_SECONDS) as response:
                payload = json.loads(response.read().decode("utf-8"))
            if not isinstance(payload, dict) or payload.get("code") != "Ok":
                message = payload.get("message") if isinstance(payload, dict) else "格式錯誤"
                raise RoadRoutingError(f"道路服務回傳失敗：{message or '未知原因'}")
            durations = payload.get("durations")
            distances = payload.get("distances")
            if not isinstance(durations, list) or not isinstance(distances, list):
                raise RoadRoutingError("道路服務未回傳時間／距離矩陣。")
            return {
                "durations": durations,
                "distances": distances,
                "data_version": str(payload.get("data_version") or ""),
            }
        except HTTPError as exc:
            last_error = exc
            retryable = exc.code == 429 or 500 <= exc.code <= 599
            if retryable and attempt < ROAD_ROUTER_MAX_ATTEMPTS:
                time.sleep(0.35 * attempt)
                continue
            raise RoadRoutingError(f"道路服務 HTTP {exc.code}") from exc
        except (URLError, TimeoutError, json.JSONDecodeError, RoadRoutingError) as exc:
            last_error = exc
            if attempt < ROAD_ROUTER_MAX_ATTEMPTS:
                time.sleep(0.35 * attempt)
                continue
            break

    raise RoadRoutingError(f"道路服務暫時無法使用：{last_error or '未知錯誤'}")


ROAD_PAIR_CACHE_STATE_KEY = "road_pair_metric_cache::v24.1"
ROAD_PAIR_CACHE_CLEANUP_STATE_KEY = "road_pair_metric_cache_cleanup::v26.3"
ROAD_ROUTER_STATUS_STATE_KEY = "road_router_status::v24.1"
ROAD_PAIR_CACHE_MAX_AGE_SECONDS = 1800
ROAD_PAIR_CACHE_MAX_ENTRIES = 5000
ROAD_PAIR_CACHE_CLEANUP_INTERVAL_SECONDS = 30


def _road_pair_cache_key(
    origin_latitude: float,
    origin_longitude: float,
    destination_latitude: float,
    destination_longitude: float,
) -> str:
    origin = _rounded_coordinate(
        origin_longitude,
        origin_latitude,
        precision=ROAD_ROUTER_ORIGIN_PRECISION,
    )
    destination = _rounded_coordinate(
        destination_longitude,
        destination_latitude,
        precision=ROAD_ROUTER_STATION_PRECISION,
    )
    return f"{origin[0]:.4f},{origin[1]:.4f}>{destination[0]:.5f},{destination[1]:.5f}"


def _road_pair_cache() -> dict[str, dict]:
    """保存已取得的單段道路結果，讓不同預看方案共用，不重複呼叫路網服務。"""
    now = time.time()
    raw_cache = st.session_state.get(ROAD_PAIR_CACHE_STATE_KEY, {})
    if not isinstance(raw_cache, dict):
        raw_cache = {}
    last_cleanup = float(
        st.session_state.get(ROAD_PAIR_CACHE_CLEANUP_STATE_KEY) or 0.0
    )
    if (
        now - last_cleanup < ROAD_PAIR_CACHE_CLEANUP_INTERVAL_SECONDS
        and len(raw_cache) <= ROAD_PAIR_CACHE_MAX_ENTRIES
    ):
        return raw_cache

    cache = {
        str(key): value
        for key, value in raw_cache.items()
        if isinstance(value, dict)
        and now - float(value.get("cached_at") or 0) <= ROAD_PAIR_CACHE_MAX_AGE_SECONDS
    }
    if len(cache) > ROAD_PAIR_CACHE_MAX_ENTRIES:
        newest = sorted(
            cache.items(),
            key=lambda item: float(item[1].get("cached_at") or 0),
            reverse=True,
        )[:ROAD_PAIR_CACHE_MAX_ENTRIES]
        cache = dict(newest)
    st.session_state[ROAD_PAIR_CACHE_STATE_KEY] = cache
    st.session_state[ROAD_PAIR_CACHE_CLEANUP_STATE_KEY] = now
    return cache


def _cache_road_pair(
    cache: dict[str, dict],
    *,
    origin_latitude: float,
    origin_longitude: float,
    destination_latitude: float,
    destination_longitude: float,
    metric: dict,
) -> None:
    key = _road_pair_cache_key(
        origin_latitude,
        origin_longitude,
        destination_latitude,
        destination_longitude,
    )
    cache[key] = {**metric, "cached_at": time.time()}


def _fallback_road_metric(
    origin_latitude: float,
    origin_longitude: float,
    destination_latitude: float,
    destination_longitude: float,
) -> dict:
    """道路服務失效時保持系統可操作；畫面會明確標示這不是道路路網結果。"""
    straight_distance_km = haversine_distance_km(
        origin_latitude,
        origin_longitude,
        destination_latitude,
        destination_longitude,
    )
    estimated_distance_km = max(0.05, straight_distance_km * DISPATCH_ROAD_DISTANCE_FACTOR)
    estimated_drive_minutes = max(
        1.0,
        estimated_distance_km / DISPATCH_ESTIMATED_SPEED_KMH * 60,
    )
    return {
        "straight_distance_km": straight_distance_km,
        "road_distance_km": estimated_distance_km,
        "drive_minutes": estimated_drive_minutes,
        "detour_ratio": estimated_distance_km / max(0.05, straight_distance_km),
        "routing_source": "備援估算",
        "routing_fallback": True,
        "road_route_available": None,
        "routing_data_version": "",
    }


def road_metrics_from_origin(
    *,
    origin_latitude: float,
    origin_longitude: float,
    destinations: list[tuple[str, float, float]],
) -> dict[str, dict]:
    """由目前位置批次計算道路距離；已查過的道路段直接共用快取。"""
    output: dict[str, dict] = {}
    pair_cache = _road_pair_cache()
    missing: list[tuple[str, float, float]] = []

    for station_name, destination_latitude, destination_longitude in destinations:
        cache_key = _road_pair_cache_key(
            origin_latitude,
            origin_longitude,
            destination_latitude,
            destination_longitude,
        )
        cached = pair_cache.get(cache_key)
        if isinstance(cached, dict):
            metric = dict(cached)
            metric.pop("cached_at", None)
            metric["straight_distance_km"] = haversine_distance_km(
                origin_latitude,
                origin_longitude,
                destination_latitude,
                destination_longitude,
            )
            output[station_name] = metric
        else:
            missing.append((station_name, destination_latitude, destination_longitude))

    rounded_origin = _rounded_coordinate(
        origin_longitude,
        origin_latitude,
        precision=ROAD_ROUTER_ORIGIN_PRECISION,
    )
    for start in range(0, len(missing), ROAD_ROUTER_BATCH_SIZE):
        batch = missing[start : start + ROAD_ROUTER_BATCH_SIZE]
        rounded_destinations = [
            _rounded_coordinate(longitude, latitude, precision=ROAD_ROUTER_STATION_PRECISION)
            for _name, latitude, longitude in batch
        ]
        coordinates = tuple([rounded_origin, *rounded_destinations])
        try:
            matrix = fetch_road_table_cached(
                coordinates,
                (0,),
                tuple(range(1, len(coordinates))),
            )
            st.session_state[ROAD_ROUTER_STATUS_STATE_KEY] = {
                "ok": True,
                "updated_at": time.time(),
                "message": "",
            }
            durations = matrix["durations"][0]
            distances = matrix["distances"][0]
            for index, (station_name, destination_latitude, destination_longitude) in enumerate(batch):
                duration_seconds = durations[index] if index < len(durations) else None
                distance_meters = distances[index] if index < len(distances) else None
                straight_distance_km = haversine_distance_km(
                    origin_latitude,
                    origin_longitude,
                    destination_latitude,
                    destination_longitude,
                )
                if duration_seconds is None or distance_meters is None:
                    metric = {
                        "straight_distance_km": straight_distance_km,
                        "road_route_available": False,
                        "routing_source": "道路路網無可行駛路線",
                        "routing_fallback": False,
                        "routing_data_version": matrix.get("data_version", ""),
                    }
                else:
                    road_distance_km = max(0.05, float(distance_meters) / 1000.0)
                    metric = {
                        "straight_distance_km": straight_distance_km,
                        "road_distance_km": road_distance_km,
                        "drive_minutes": max(1.0, float(duration_seconds) / 60.0),
                        "detour_ratio": road_distance_km / max(0.05, straight_distance_km),
                        "routing_source": "OSRM／OpenStreetMap 道路路網",
                        "routing_fallback": False,
                        "road_route_available": True,
                        "routing_data_version": matrix.get("data_version", ""),
                    }
                output[station_name] = metric
                _cache_road_pair(
                    pair_cache,
                    origin_latitude=origin_latitude,
                    origin_longitude=origin_longitude,
                    destination_latitude=destination_latitude,
                    destination_longitude=destination_longitude,
                    metric=metric,
                )
        except RoadRoutingError as exc:
            st.session_state[ROAD_ROUTER_STATUS_STATE_KEY] = {
                "ok": False,
                "updated_at": time.time(),
                "message": str(exc),
            }
            for station_name, destination_latitude, destination_longitude in batch:
                output[station_name] = {
                    "straight_distance_km": haversine_distance_km(
                        origin_latitude,
                        origin_longitude,
                        destination_latitude,
                        destination_longitude,
                    ),
                    "road_route_available": False,
                    "routing_source": "道路服務暫時無法使用",
                    "routing_fallback": False,
                    "routing_service_error": True,
                    "routing_data_version": "",
                }

    st.session_state[ROAD_PAIR_CACHE_STATE_KEY] = pair_cache
    return output


def attach_analysis_road_metrics(
    result_df: pd.DataFrame,
    *,
    station_locations: dict[str, dict],
    current_location: dict | None,
) -> tuple[pd.DataFrame, bool, str]:
    """把目前 GPS 到各場站的道路距離與時間附加到一般分析結果。"""
    enriched = result_df.copy()
    enriched["距離目前位置 (km)"] = pd.Series(np.nan, index=enriched.index, dtype="float64")
    enriched["預估行車時間 (分)"] = pd.Series(np.nan, index=enriched.index, dtype="float64")

    if not location_payload_is_valid(current_location):
        return enriched, False, "尚未取得有效 GPS，距離與行車時間暫時無法計算。"

    origin_latitude = float(current_location["latitude"])
    origin_longitude = float(current_location["longitude"])
    destinations: list[tuple[str, float, float]] = []
    for station_name in enriched["場站名稱"].astype(str).drop_duplicates():
        location = station_locations.get(station_name)
        if not isinstance(location, dict):
            continue
        latitude = normalize_coordinate(location.get("latitude"), -90.0, 90.0)
        longitude = normalize_coordinate(location.get("longitude"), -180.0, 180.0)
        if latitude is None or longitude is None:
            continue
        destinations.append((station_name, latitude, longitude))

    if not destinations:
        return enriched, False, "目前分析場站沒有可用座標，無法依距離或行車時間排序。"

    road_metrics = road_metrics_from_origin(
        origin_latitude=origin_latitude,
        origin_longitude=origin_longitude,
        destinations=destinations,
    )
    valid_count = 0
    for station_name, metric in road_metrics.items():
        if not isinstance(metric, dict) or metric.get("road_route_available") is False:
            continue
        distance = metric.get("road_distance_km")
        drive_minutes = metric.get("drive_minutes")
        if distance is None or drive_minutes is None:
            continue
        station_mask = enriched["場站名稱"].astype(str).eq(str(station_name))
        enriched.loc[station_mask, "距離目前位置 (km)"] = float(distance)
        enriched.loc[station_mask, "預估行車時間 (分)"] = float(drive_minutes)
        valid_count += int(station_mask.sum())

    if valid_count <= 0:
        return enriched, False, "道路路網目前沒有回傳可用路線，已保留原始分析順序。"
    return enriched, True, f"已取得 {valid_count} 個場站的道路距離與預估行車時間。"


def build_station_alert_records(status_df: pd.DataFrame) -> list[dict]:
    """建立總車數／空位數為 0 或 1 的場站警示。"""
    alerts: list[dict] = []
    for row in status_df.to_dict(orient="records"):
        station_name = str(row.get("場站名稱") or "").strip()
        if not station_name:
            continue
        service_status = normalize_current_status(row.get("服務狀態"))
        if service_status is not None and service_status != 1:
            continue
        total_bikes = station_total_bikes(row)
        empty_spaces = station_empty_spaces(row)
        alert_parts: list[str] = []
        severity = 0
        if total_bikes == 0:
            alert_parts.append("總車數 0 台")
            severity = max(severity, 2)
        elif total_bikes == 1:
            alert_parts.append("總車數僅 1 台")
            severity = max(severity, 1)
        if empty_spaces == 0:
            alert_parts.append("空位 0 格")
            severity = max(severity, 2)
        elif empty_spaces == 1:
            alert_parts.append("空位僅 1 格")
            severity = max(severity, 1)
        if severity <= 0:
            continue
        alerts.append(
            {
                "station_name": station_name,
                "region": str(row.get("行政區") or "").strip(),
                "route_zone": str(row.get("路線區域") or "").strip(),
                "total_bikes": total_bikes,
                "empty_spaces": empty_spaces,
                "parking_spaces": normalize_current_status(row.get("總柱數")),
                "current_bike": normalize_current_status(row.get("2.0 現況")),
                "current_ebike": normalize_current_status(row.get("2.0E 現況")),
                "standard_bike": safe_nonnegative_int(row.get("2.0 標準")),
                "standard_ebike": safe_nonnegative_int(row.get("2.0E 標準")),
                "severity": severity,
                "severity_label": "緊急" if severity >= 2 else "注意",
                "message": "、".join(alert_parts),
            }
        )
    return sorted(
        alerts,
        key=lambda item: (
            -safe_nonnegative_int(item.get("severity")),
            999 if item.get("total_bikes") is None else safe_nonnegative_int(item.get("total_bikes")),
            999 if item.get("empty_spaces") is None else safe_nonnegative_int(item.get("empty_spaces")),
            str(item.get("station_name") or ""),
        ),
    )


def notify_station_alert_changes(
    status_df: pd.DataFrame,
    *,
    token: str,
    context_key: str,
    alerts: list[dict] | None = None,
) -> None:
    """只有警示首次出現或嚴重度／內容改變時通知；恢復後會清除舊狀態。"""
    alert_records = alerts if isinstance(alerts, list) else build_station_alert_records(status_df)
    state_key = f"station_alert_state::{token}::{hashlib.sha1(context_key.encode('utf-8')).hexdigest()[:12]}"
    previous = st.session_state.get(state_key, {})
    if not isinstance(previous, dict):
        previous = {}
    current: dict[str, str] = {}
    changed: list[dict] = []
    for alert in alert_records:
        station_name = str(alert["station_name"])
        signature = f"{safe_nonnegative_int(alert['severity'])}|{alert['message']}"
        current[station_name] = signature
        if previous.get(station_name) != signature:
            changed.append(alert)
    st.session_state[state_key] = current

    for alert in changed[:4]:
        icon = "🚨" if safe_nonnegative_int(alert.get("severity")) >= 2 else "⚠️"
        st.toast(f"{alert['station_name']}：{alert['message']}", icon=icon)
    if len(changed) > 4:
        st.toast(f"另有 {len(changed) - 4} 個場站出現新的車柱警示", icon="📣")


def add_road_metrics_to_alerts(
    alerts: list[dict],
    *,
    station_locations: dict[str, dict],
    current_location: dict | None,
) -> list[dict]:
    """批次加入警示場站的道路距離與時間；GPS 不可用時保留空值。"""
    output = [dict(alert) for alert in alerts]
    if not output or not location_payload_is_valid(current_location):
        return output
    destinations: list[tuple[str, float, float]] = []
    for alert in output:
        location = station_locations.get(str(alert["station_name"]))
        if not isinstance(location, dict):
            continue
        latitude = normalize_coordinate(location.get("latitude"), -90.0, 90.0)
        longitude = normalize_coordinate(location.get("longitude"), -180.0, 180.0)
        if latitude is None or longitude is None:
            continue
        destinations.append((str(alert["station_name"]), latitude, longitude))
    if not destinations:
        return output
    metrics = road_metrics_from_origin(
        origin_latitude=float(current_location["latitude"]),
        origin_longitude=float(current_location["longitude"]),
        destinations=destinations,
    )
    for alert in output:
        metric = metrics.get(str(alert["station_name"]))
        if not isinstance(metric, dict) or metric.get("road_route_available") is False:
            continue
        alert["estimated_distance_km"] = metric.get("road_distance_km")
        alert["estimated_drive_minutes"] = metric.get("drive_minutes")
    return output


def render_station_alert_summary(
    status_df: pd.DataFrame,
    *,
    station_locations: dict[str, dict],
    current_location: dict | None,
    alerts: list[dict] | None = None,
) -> list[dict]:
    """在工作頁頂端保留車柱警示區，回傳同一份資料供智慧調度按鈕使用。"""
    alert_records = add_road_metrics_to_alerts(
        alerts if isinstance(alerts, list) else build_station_alert_records(status_df),
        station_locations=station_locations,
        current_location=current_location,
    )
    if not alert_records:
        return []
    urgent_count = sum(1 for item in alert_records if safe_nonnegative_int(item.get("severity")) >= 2)
    warning_count = len(alert_records) - urgent_count
    with st.expander(
        f"🚨 車柱警示：緊急 {urgent_count} 站｜注意 {warning_count} 站",
        expanded=urgent_count > 0,
    ):
        st.caption("警示只提供決策參考，不會改變 AI 分數；同一狀態只通知一次，惡化時會再次通知。")
        for alert in alert_records[:15]:
            distance = alert.get("estimated_distance_km")
            drive_minutes = alert.get("estimated_drive_minutes")
            road_text = "GPS／道路資料待取得"
            if distance is not None and drive_minutes is not None:
                road_text = f"{float(distance):.1f} km｜約 {float(drive_minutes):.0f} 分"
            level_icon = "🚨" if safe_nonnegative_int(alert.get("severity")) >= 2 else "⚠️"
            st.markdown(
                f"**{level_icon} {alert['route_zone']}｜{alert['station_name']}｜{alert['message']}**  \n"
                f"2.0 {optional_count_text(alert.get('current_bike'))}／標準 {safe_nonnegative_int(alert.get('standard_bike'))}｜"
                f"2.0E {optional_count_text(alert.get('current_ebike'))}／標準 {safe_nonnegative_int(alert.get('standard_ebike'))}｜"
                f"總柱 {optional_count_text(alert.get('parking_spaces'))}｜{road_text}"
            )
        if len(alert_records) > 15:
            st.caption(f"另有 {len(alert_records) - 15} 站警示，可使用全場站搜尋查看。")
    return alert_records


def road_metrics_to_endpoint(
    candidates: list[dict],
    endpoint_location: dict,
) -> dict[str, dict]:
    """批次計算候選站到終點的道路成本；與路線預看共用單段快取。"""
    endpoint_latitude = normalize_coordinate(endpoint_location.get("latitude"), -90.0, 90.0)
    endpoint_longitude = normalize_coordinate(endpoint_location.get("longitude"), -180.0, 180.0)
    if endpoint_latitude is None or endpoint_longitude is None:
        return {}

    output: dict[str, dict] = {}
    pair_cache = _road_pair_cache()
    missing: list[dict] = []
    for candidate in candidates:
        cache_key = _road_pair_cache_key(
            float(candidate["latitude"]),
            float(candidate["longitude"]),
            endpoint_latitude,
            endpoint_longitude,
        )
        cached = pair_cache.get(cache_key)
        if isinstance(cached, dict):
            metric = dict(cached)
            metric.pop("cached_at", None)
            output[str(candidate["station_name"])] = metric
        else:
            missing.append(candidate)

    for start in range(0, len(missing), ROAD_ROUTER_BATCH_SIZE):
        batch = missing[start : start + ROAD_ROUTER_BATCH_SIZE]
        origins = [
            _rounded_coordinate(
                float(candidate["longitude"]),
                float(candidate["latitude"]),
                precision=ROAD_ROUTER_STATION_PRECISION,
            )
            for candidate in batch
        ]
        endpoint = _rounded_coordinate(
            endpoint_longitude,
            endpoint_latitude,
            precision=ROAD_ROUTER_STATION_PRECISION,
        )
        coordinates = tuple([*origins, endpoint])
        endpoint_index = len(coordinates) - 1
        try:
            matrix = fetch_road_table_cached(
                coordinates,
                tuple(range(len(batch))),
                (endpoint_index,),
            )
            st.session_state[ROAD_ROUTER_STATUS_STATE_KEY] = {
                "ok": True,
                "updated_at": time.time(),
                "message": "",
            }
            for index, candidate in enumerate(batch):
                duration_seconds = matrix["durations"][index][0]
                distance_meters = matrix["distances"][index][0]
                if duration_seconds is None or distance_meters is None:
                    metric = {
                        "road_route_available": False,
                        "routing_fallback": False,
                    }
                else:
                    metric = {
                        "road_route_available": True,
                        "routing_fallback": False,
                        "road_distance_km": max(0.05, float(distance_meters) / 1000.0),
                        "drive_minutes": max(1.0, float(duration_seconds) / 60.0),
                    }
                output[str(candidate["station_name"])] = metric
                _cache_road_pair(
                    pair_cache,
                    origin_latitude=float(candidate["latitude"]),
                    origin_longitude=float(candidate["longitude"]),
                    destination_latitude=endpoint_latitude,
                    destination_longitude=endpoint_longitude,
                    metric=metric,
                )
        except RoadRoutingError as exc:
            st.session_state[ROAD_ROUTER_STATUS_STATE_KEY] = {
                "ok": False,
                "updated_at": time.time(),
                "message": str(exc),
            }
            for candidate in batch:
                output[str(candidate["station_name"])] = {
                    "road_route_available": False,
                    "routing_fallback": False,
                    "routing_service_error": True,
                }

    st.session_state[ROAD_PAIR_CACHE_STATE_KEY] = pair_cache
    return output


def build_dispatch_plan_for_station(
    row: pd.Series | dict,
    *,
    truck_bike: int,
    truck_ebike: int,
    max_capacity: int,
    global_bike_shortage: int,
    global_ebike_shortage: int,
) -> dict | None:
    """依車上存量及總容量，模擬本站先下車後上車的最大可行調度量。"""
    current_bike = normalize_current_status(row.get("2.0 現況"))
    current_ebike = normalize_current_status(row.get("2.0E 現況"))
    if current_bike is None or current_ebike is None:
        return None

    service_status = normalize_current_status(row.get("服務狀態"))
    if service_status is not None and service_status != 1:
        return None

    standard_bike = safe_nonnegative_int(row.get("2.0 標準"))
    standard_ebike = safe_nonnegative_int(row.get("2.0E 標準"))

    bike_shortage = max(0, standard_bike - current_bike)
    ebike_shortage = max(0, standard_ebike - current_ebike)
    bike_extra = max(0, current_bike - standard_bike)
    ebike_extra = max(0, current_ebike - standard_ebike)

    unload_bike = min(bike_shortage, truck_bike)
    unload_ebike = min(ebike_shortage, truck_ebike)
    bike_after_unload = truck_bike - unload_bike
    ebike_after_unload = truck_ebike - unload_ebike
    free_capacity = max(0, max_capacity - bike_after_unload - ebike_after_unload)

    pickup = {"bike": 0, "ebike": 0}
    pickup_needs = [
        ("bike", bike_extra, global_bike_shortage),
        ("ebike", ebike_extra, global_ebike_shortage),
    ]
    # 空間不足時，優先將整體缺口較大的車種上車帶走；仍以本站總調度量最大化為第一目標。
    pickup_needs.sort(key=lambda item: (item[2], item[1]), reverse=True)
    for vehicle_type, extra_count, _network_shortage in pickup_needs:
        amount = min(extra_count, free_capacity)
        pickup[vehicle_type] = amount
        free_capacity -= amount

    pickup_bike = pickup["bike"]
    pickup_ebike = pickup["ebike"]
    dispatch_count = unload_bike + unload_ebike + pickup_bike + pickup_ebike
    if dispatch_count <= 0:
        return None

    final_bike = bike_after_unload + pickup_bike
    final_ebike = ebike_after_unload + pickup_ebike
    if final_bike + final_ebike > max_capacity:
        return None

    return {
        "station_name": str(row.get("場站名稱") or "").strip(),
        "region": str(row.get("行政區") or "").strip(),
        "route_zone": normalize_long_distance_zone(row.get("路線區域")) or "",
        "status_context_key": str(row.get("_狀態內容鍵") or "").strip(),
        "current_bike": current_bike,
        "current_ebike": current_ebike,
        "standard_bike": standard_bike,
        "standard_ebike": standard_ebike,
        "total_bikes": station_total_bikes(row),
        "empty_spaces": station_empty_spaces(row),
        "parking_spaces": normalize_current_status(row.get("總柱數")),
        "service_status": service_status,
        "unload_bike": unload_bike,
        "unload_ebike": unload_ebike,
        "pickup_bike": pickup_bike,
        "pickup_ebike": pickup_ebike,
        "dispatch_count": dispatch_count,
        "truck_before_bike": safe_nonnegative_int(truck_bike),
        "truck_before_ebike": safe_nonnegative_int(truck_ebike),
        "truck_after_bike": final_bike,
        "truck_after_ebike": final_ebike,
        "max_capacity": safe_nonnegative_int(max_capacity),
    }


def calculate_dispatch_candidates(
    dispatch_df: pd.DataFrame,
    *,
    station_locations: dict[str, dict],
    current_location: dict,
    truck_bike: int,
    truck_ebike: int,
    max_capacity: int,
    cooldowns: dict[str, dict],
    rejection_history: list[dict],
    now_timestamp: float,
    current_round: int,
) -> list[dict]:
    """依實際道路路網、可調度量與作業時間建立候選排名。"""
    del rejection_history, now_timestamp
    valid_df = dispatch_df.copy()
    current_bike_series = pd.to_numeric(valid_df["2.0 現況"], errors="coerce")
    current_ebike_series = pd.to_numeric(valid_df["2.0E 現況"], errors="coerce")
    standard_bike_series = pd.to_numeric(valid_df["2.0 標準"], errors="coerce").fillna(0)
    standard_ebike_series = pd.to_numeric(valid_df["2.0E 標準"], errors="coerce").fillna(0)
    global_bike_shortage = int((standard_bike_series - current_bike_series).clip(lower=0).fillna(0).sum())
    global_ebike_shortage = int((standard_ebike_series - current_ebike_series).clip(lower=0).fillna(0).sum())

    origin_lat = normalize_coordinate(current_location.get("latitude"), -90.0, 90.0)
    origin_lon = normalize_coordinate(current_location.get("longitude"), -180.0, 180.0)
    if origin_lat is None or origin_lon is None:
        return []

    prepared: list[tuple[dict, dict, float, float]] = []
    destinations: list[tuple[str, float, float]] = []
    # to_dict(records) 比 iterrows 更輕；先排除無調度量或無座標場站，再送道路矩陣。
    for row in valid_df.to_dict(orient="records"):
        station_name = str(row.get("場站名稱") or "").strip()
        if not station_name:
            continue
        cooldown = cooldowns.get(station_name)
        if isinstance(cooldown, dict) and safe_nonnegative_int(cooldown.get("resume_after_round")) >= current_round:
            continue

        location = station_locations.get(station_name)
        if not isinstance(location, dict):
            continue
        destination_lat = normalize_coordinate(location.get("latitude"), -90.0, 90.0)
        destination_lon = normalize_coordinate(location.get("longitude"), -180.0, 180.0)
        if destination_lat is None or destination_lon is None:
            continue

        plan = build_dispatch_plan_for_station(
            row,
            truck_bike=truck_bike,
            truck_ebike=truck_ebike,
            max_capacity=max_capacity,
            global_bike_shortage=global_bike_shortage,
            global_ebike_shortage=global_ebike_shortage,
        )
        if plan is None:
            continue

        # 庫存場站規則：圖書館／轉運站平常不進入自動候選；
        # 只有本站本身已非常少車（總車 0/1）或非常多車（空位 0/1）才解鎖。
        if is_inventory_station_name(station_name):
            plan["inventory_station"] = True
            plan["inventory_attention"] = inventory_station_needs_attention(plan)
            if not plan["inventory_attention"]:
                continue

        prepared.append((plan, location, destination_lat, destination_lon))
        destinations.append((station_name, destination_lat, destination_lon))

    if not prepared:
        return []

    road_metrics = road_metrics_from_origin(
        origin_latitude=origin_lat,
        origin_longitude=origin_lon,
        destinations=destinations,
    )

    candidates: list[dict] = []
    for plan, location, destination_lat, destination_lon in prepared:
        station_name = str(plan["station_name"])
        metric = road_metrics.get(station_name)
        if not isinstance(metric, dict) or metric.get("road_route_available") is False:
            # 無道路可行駛路線就不列候選，避免直線橫切山脈。
            continue

        road_distance_km = float(metric.get("road_distance_km") or 0.0)
        estimated_drive_minutes = float(metric.get("drive_minutes") or 0.0)
        if road_distance_km <= 0 or estimated_drive_minutes <= 0:
            continue

        estimated_operation_minutes = (
            DISPATCH_OPERATION_BASE_MINUTES
            + plan["dispatch_count"] * DISPATCH_OPERATION_MINUTES_PER_BIKE
        )
        estimated_total_minutes = estimated_drive_minutes + estimated_operation_minutes
        raw_efficiency = plan["dispatch_count"] / max(1.0, estimated_total_minutes)

        score = raw_efficiency

        plan.update(
            {
                "station_id": str(location.get("station_id") or ""),
                "official_name": str(location.get("official_name") or station_name),
                "latitude": destination_lat,
                "longitude": destination_lon,
                "straight_distance_km": float(metric.get("straight_distance_km") or 0.0),
                "estimated_distance_km": road_distance_km,
                "estimated_drive_minutes": estimated_drive_minutes,
                "estimated_operation_minutes": estimated_operation_minutes,
                "estimated_total_minutes": estimated_total_minutes,
                "road_detour_ratio": float(metric.get("detour_ratio") or 1.0),
                "routing_source": str(metric.get("routing_source") or "道路路網"),
                "routing_fallback": bool(metric.get("routing_fallback")),
                "road_route_available": metric.get("road_route_available"),
                "routing_data_version": str(metric.get("routing_data_version") or ""),
                "raw_efficiency": raw_efficiency,
                "score": score,
            }
        )
        candidates.append(plan)

    return sorted(
        candidates,
        key=lambda item: (
            float(item["score"]),
            safe_nonnegative_int(item["dispatch_count"]),
            -float(item["estimated_distance_km"]),
        ),
        reverse=True,
    )

def inspect_dispatch_station(
    dispatch_df: pd.DataFrame,
    *,
    station_name: str,
    station_locations: dict[str, dict],
    current_location: dict,
    truck_bike: int,
    truck_ebike: int,
    max_capacity: int,
    cooldowns: dict[str, dict],
    current_round: int,
    allowed_route_zone: str = "",
) -> dict:
    """搜尋任意場站並說明目前能否執行，不受前 10 名候選限制。"""
    matched = dispatch_df[dispatch_df["場站名稱"].astype(str).eq(str(station_name))]
    if matched.empty:
        return {"station_name": station_name, "processable": False, "reason": "配置表找不到此場站。"}
    row = matched.iloc[0].to_dict()
    route_zone = normalize_long_distance_zone(row.get("路線區域")) or str(row.get("路線區域") or "")
    base = {
        "station_name": str(row.get("場站名稱") or station_name),
        "region": str(row.get("行政區") or ""),
        "route_zone": route_zone,
        "current_bike": normalize_current_status(row.get("2.0 現況")),
        "current_ebike": normalize_current_status(row.get("2.0E 現況")),
        "standard_bike": safe_nonnegative_int(row.get("2.0 標準")),
        "standard_ebike": safe_nonnegative_int(row.get("2.0E 標準")),
        "total_bikes": station_total_bikes(row),
        "empty_spaces": station_empty_spaces(row),
        "parking_spaces": normalize_current_status(row.get("總柱數")),
        "service_status": normalize_current_status(row.get("服務狀態")),
        "processable": False,
        "reason": "",
    }
    if base["current_bike"] is None or base["current_ebike"] is None:
        base["reason"] = "2.0／2.0E 現況資料不完整。"
        return base
    if base["service_status"] is not None and base["service_status"] != 1:
        base["reason"] = "YouBike 顯示此站目前暫停服務。"
        return base
    if allowed_route_zone and route_zone != allowed_route_zone:
        base["reason"] = f"環狀一圈目前正在執行 {allowed_route_zone}，完成本區後才會轉往 {route_zone or '其他區域'}。"
        return base
    cooldown = cooldowns.get(str(station_name))
    if isinstance(cooldown, dict) and safe_nonnegative_int(cooldown.get("resume_after_round")) >= current_round:
        remaining = max(1, safe_nonnegative_int(cooldown.get("resume_after_round")) - current_round + 1)
        base["reason"] = f"此站仍在暫時忽略中，尚餘 {remaining} 回。"
        return base
    location = station_locations.get(str(station_name))
    if not isinstance(location, dict):
        base["reason"] = "缺少場站座標。"
        return base
    latitude = normalize_coordinate(location.get("latitude"), -90.0, 90.0)
    longitude = normalize_coordinate(location.get("longitude"), -180.0, 180.0)
    if latitude is None or longitude is None:
        base["reason"] = "場站座標格式無效。"
        return base
    if not location_payload_is_valid(current_location):
        base["reason"] = "尚未取得有效 GPS，無法計算道路時間。"
        return base

    current_bike_series = pd.to_numeric(dispatch_df["2.0 現況"], errors="coerce")
    current_ebike_series = pd.to_numeric(dispatch_df["2.0E 現況"], errors="coerce")
    standard_bike_series = pd.to_numeric(dispatch_df["2.0 標準"], errors="coerce").fillna(0)
    standard_ebike_series = pd.to_numeric(dispatch_df["2.0E 標準"], errors="coerce").fillna(0)
    plan = build_dispatch_plan_for_station(
        row,
        truck_bike=truck_bike,
        truck_ebike=truck_ebike,
        max_capacity=max_capacity,
        global_bike_shortage=int((standard_bike_series - current_bike_series).clip(lower=0).fillna(0).sum()),
        global_ebike_shortage=int((standard_ebike_series - current_ebike_series).clip(lower=0).fillna(0).sum()),
    )
    if plan is None:
        bike_diff = int(base["current_bike"]) - int(base["standard_bike"])
        ebike_diff = int(base["current_ebike"]) - int(base["standard_ebike"])
        if bike_diff == 0 and ebike_diff == 0:
            base["reason"] = "此站目前符合配置，不需要調度。"
        elif bike_diff < 0 and truck_bike <= 0 and ebike_diff >= 0:
            base["reason"] = "此站缺 2.0，但貨車目前沒有可下車的 2.0。"
        elif ebike_diff < 0 and truck_ebike <= 0 and bike_diff >= 0:
            base["reason"] = "此站缺 2.0E，但貨車目前沒有可下車的 2.0E。"
        elif bike_diff > 0 or ebike_diff > 0:
            base["reason"] = "本站有多車，但貨車剩餘容量不足或目前無可行上下車組合。"
        else:
            base["reason"] = "依目前貨車車種與容量，本站沒有可執行的調度量。"
        return base

    metrics = road_metrics_from_origin(
        origin_latitude=float(current_location["latitude"]),
        origin_longitude=float(current_location["longitude"]),
        destinations=[(str(station_name), latitude, longitude)],
    )
    metric = metrics.get(str(station_name))
    if not isinstance(metric, dict) or metric.get("road_route_available") is False:
        base["reason"] = "道路路網查無可行駛路線，或道路服務暫時不可用。"
        return base
    distance = float(metric.get("road_distance_km") or 0.0)
    drive_minutes = float(metric.get("drive_minutes") or 0.0)
    if distance <= 0 or drive_minutes <= 0:
        base["reason"] = "道路距離／行車時間資料不完整。"
        return base
    operation_minutes = DISPATCH_OPERATION_BASE_MINUTES + plan["dispatch_count"] * DISPATCH_OPERATION_MINUTES_PER_BIKE
    total_minutes = drive_minutes + operation_minutes
    plan.update(base)
    plan.update(
        {
            "processable": True,
            "reason": "可執行；按『設為目前選擇場站』後，仍需再按『前往此站』才會鎖定。",
            "station_id": str(location.get("station_id") or ""),
            "official_name": str(location.get("official_name") or station_name),
            "latitude": latitude,
            "longitude": longitude,
            "straight_distance_km": float(metric.get("straight_distance_km") or 0.0),
            "estimated_distance_km": distance,
            "estimated_drive_minutes": drive_minutes,
            "estimated_operation_minutes": operation_minutes,
            "estimated_total_minutes": total_minutes,
            "road_detour_ratio": float(metric.get("detour_ratio") or 1.0),
            "routing_source": str(metric.get("routing_source") or "道路路網"),
            "routing_fallback": bool(metric.get("routing_fallback")),
            "road_route_available": metric.get("road_route_available"),
            "routing_data_version": str(metric.get("routing_data_version") or ""),
            "raw_efficiency": plan["dispatch_count"] / max(1.0, total_minutes),
            "score": plan["dispatch_count"] / max(1.0, total_minutes),
        }
    )
    return plan


def render_station_inspection_card(inspection: dict) -> None:
    """顯示搜尋場站的完整現況與可執行原因。"""
    processable = bool(inspection.get("processable"))
    status_icon = "✅" if processable else "⛔"
    road_text = "道路資料待取得"
    if inspection.get("estimated_distance_km") is not None and inspection.get("estimated_drive_minutes") is not None:
        road_text = (
            f"{float(inspection['estimated_distance_km']):.1f} km｜"
            f"約 {float(inspection['estimated_drive_minutes']):.0f} 分"
        )
    action_text = dispatch_action_text(inspection) if processable else "目前不可執行"
    st.markdown(
        f"""
        <section class="dispatch-plan-card" style="margin:.45rem 0 .7rem;">
          <div class="dispatch-plan-header">
            <div>
              <div class="dispatch-plan-kicker">{status_icon} 全場站搜尋結果</div>
              <div class="dispatch-plan-title">{html.escape(str(inspection.get('station_name') or ''))}</div>
              <div class="dispatch-plan-region">{html.escape(str(inspection.get('route_zone') or ''))}｜{html.escape(str(inspection.get('region') or ''))}</div>
            </div>
            <div class="dispatch-plan-badge">{html.escape('可執行' if processable else '不可執行')}</div>
          </div>
          <div class="dispatch-plan-action">
            2.0：{html.escape(vehicle_balance_text(inspection.get('current_bike'), inspection.get('standard_bike')))}<br>
            2.0E：{html.escape(vehicle_balance_text(inspection.get('current_ebike'), inspection.get('standard_ebike')))}<br>
            總車 {optional_count_text(inspection.get('total_bikes'))} 台｜空位 {optional_count_text(inspection.get('empty_spaces'))} 格｜總柱 {optional_count_text(inspection.get('parking_spaces'))} 格<br>
            {html.escape(road_text)}｜{html.escape(action_text)}
          </div>
          <div class="dispatch-plan-note">{html.escape(str(inspection.get('reason') or ''))}</div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def dispatch_action_text(plan: dict) -> str:
    parts: list[str] = []
    if safe_nonnegative_int(plan.get("unload_bike")):
        parts.append(f"下車 2.0 × {safe_nonnegative_int(plan['unload_bike'])}")
    if safe_nonnegative_int(plan.get("unload_ebike")):
        parts.append(f"下車 2.0E × {safe_nonnegative_int(plan['unload_ebike'])}")
    if safe_nonnegative_int(plan.get("pickup_bike")):
        parts.append(f"上車 2.0 × {safe_nonnegative_int(plan['pickup_bike'])}")
    if safe_nonnegative_int(plan.get("pickup_ebike")):
        parts.append(f"上車 2.0E × {safe_nonnegative_int(plan['pickup_ebike'])}")
    return "｜".join(parts) if parts else "無可行調度"


def jarvis_plan_payload(plan: dict | None) -> dict:
    """把智慧調度已算好的必要 context 提供給賈維斯，不重送整張 DataFrame。

    V28.6 延續既有執行中資料，不要求配置表新增欄位。這份 payload 之後可直接
    作為真正 GPT／Realtime 助理的候選場站 context。
    """
    if not isinstance(plan, dict):
        return {}
    keys = (
        # 場站身分／區域
        "station_name", "station_id", "official_name", "region", "route_zone",
        "latitude", "longitude",
        # 場站現況與配置
        "current_bike", "current_ebike", "standard_bike", "standard_ebike",
        "total_bikes", "empty_spaces", "parking_spaces", "service_status",
        # 本站建議作業
        "unload_bike", "unload_ebike", "pickup_bike", "pickup_ebike", "dispatch_count",
        # 貨車作業前後載量
        "truck_before_bike", "truck_before_ebike", "truck_after_bike", "truck_after_ebike",
        "max_capacity",
        # GPS／道路路網衍生結果
        "straight_distance_km", "estimated_distance_km", "estimated_drive_minutes",
        "estimated_operation_minutes", "estimated_total_minutes", "road_detour_ratio",
        "routing_source", "routing_fallback", "road_route_available", "routing_data_version",
        # 排名與特殊規則
        "raw_efficiency", "score", "inventory_station", "inventory_attention",
    )
    return {key: plan.get(key) for key in keys if key in plan}


def render_jarvis_voice_assistant(
    *,
    dispatch_prefix: str,
    mode: str,
    candidates: list[dict],
    current_plan: dict | None,
    threshold: int,
    priority_threshold: int,
    auto_announce: bool = False,
    context_status: str = "ready",
    context_message: str = "",
) -> dict | None:
    """渲染隱藏式語音助理，並以 Streamlit session state 保存調度 context。

    V28.5 起瀏覽器不再自行保存 120 秒候選資料。ready 狀態由 Python 寫入
    session state；短暫 rerun／定位更新時可沿用最後一份快照，但語音會先提示更新中。
    """
    if mode not in {"candidate", "active"}:
        return None

    server_context_key = f"{dispatch_prefix}::jarvis_server_context_v1"
    status = str(context_status or "ready").strip().lower()
    incoming_candidates = [
        jarvis_plan_payload(plan)
        for plan in candidates[:SMART_DISPATCH_CANDIDATE_LIMIT]
        if isinstance(plan, dict)
    ]
    incoming_plan = jarvis_plan_payload(current_plan)
    has_incoming_station = bool(str(incoming_plan.get("station_name") or "").strip()) or any(
        bool(str(plan.get("station_name") or "").strip()) for plan in incoming_candidates
    )

    if status == "ready" and has_incoming_station:
        snapshot = {
            "mode": mode,
            "current_plan": incoming_plan,
            "candidates": incoming_candidates,
            "updated_at": time.time(),
        }
        st.session_state[server_context_key] = snapshot
    elif status in {"no_candidates", "blocked", "unavailable"}:
        # 這些是已確認的終態，不應沿用舊目標。
        st.session_state.pop(server_context_key, None)
    elif status == "updating" and not has_incoming_station:
        # rerun / GPS / 即時車數更新期間保留最後一份伺服器快照，
        # 但前端會依 context_status 阻止使用者誤用舊資料。
        saved = st.session_state.get(server_context_key)
        if isinstance(saved, dict):
            saved_plan = saved.get("current_plan")
            saved_candidates = saved.get("candidates")
            if isinstance(saved_plan, dict):
                incoming_plan = dict(saved_plan)
            if isinstance(saved_candidates, list):
                incoming_candidates = [dict(item) for item in saved_candidates if isinstance(item, dict)]

    try:
        component = get_jarvis_voice_component()
        payload = component(
            key=f"jarvis_voice::{dispatch_prefix}",
            default=None,
            mode=mode,
            app_version=APP_VERSION,
            context_schema_version=5,
            context_status=status,
            context_message=str(context_message or "").strip(),
            candidate_count=min(len(incoming_candidates), SMART_DISPATCH_CANDIDATE_LIMIT),
            candidates=incoming_candidates,
            current_plan=incoming_plan,
            threshold=min(100, max(0, int(threshold))),
            priority_threshold=min(int(threshold), max(0, int(priority_threshold))),
            auto_announce=bool(auto_announce and status == "ready"),
            auto_announce_token=uuid.uuid4().hex if auto_announce and status == "ready" else "",
        )
        return payload if isinstance(payload, dict) else None
    except Exception as exc:
        # 語音是測試外掛，任何失敗都不能拖垮原 v27.5 調度流程。
        st.caption(f"賈維斯語音測試元件未啟用：{exc}")
        return None


def resolve_voice_station_name(status_df: pd.DataFrame, target_text: str) -> str:
    """把語音中的場站名稱安全對到目前配置表；先精確，再做唯一包含比對。"""
    wanted = normalize_station_key(target_text)
    if not wanted or "場站名稱" not in status_df.columns:
        return ""
    names = [str(name).strip() for name in status_df["場站名稱"].dropna().astype(str).drop_duplicates()]
    exact = [name for name in names if normalize_station_key(name) == wanted]
    if len(exact) == 1:
        return exact[0]
    partial = [name for name in names if wanted in normalize_station_key(name) or normalize_station_key(name) in wanted]
    return partial[0] if len(partial) == 1 else ""


def _dispatch_truck_before_counts(plan: dict) -> tuple[int, int]:
    """取得推薦前貨車載量；相容舊版已鎖定但尚未含 before 欄位的行程。"""
    before_bike = plan.get("truck_before_bike")
    before_ebike = plan.get("truck_before_ebike")
    if before_bike is None:
        before_bike = (
            safe_nonnegative_int(plan.get("truck_after_bike"))
            + safe_nonnegative_int(plan.get("unload_bike"))
            - safe_nonnegative_int(plan.get("pickup_bike"))
        )
    if before_ebike is None:
        before_ebike = (
            safe_nonnegative_int(plan.get("truck_after_ebike"))
            + safe_nonnegative_int(plan.get("unload_ebike"))
            - safe_nonnegative_int(plan.get("pickup_ebike"))
        )
    return max(0, int(before_bike)), max(0, int(before_ebike))


def render_dispatch_plan_card(plan: dict, *, title: str) -> None:
    """呈現單一推薦場站，並用作業前後對照快速確認貨車載量。"""
    station_name = str(plan.get("station_name") or "").strip()
    station_name_attr = html.escape(station_name, quote=True)
    before_bike, before_ebike = _dispatch_truck_before_counts(plan)
    after_bike = safe_nonnegative_int(plan.get("truck_after_bike"))
    after_ebike = safe_nonnegative_int(plan.get("truck_after_ebike"))
    max_capacity = max(
        1,
        safe_nonnegative_int(plan.get("max_capacity"))
        or before_bike + before_ebike
        or after_bike + after_ebike
        or 1,
    )
    before_total = before_bike + before_ebike
    after_total = after_bike + after_ebike
    after_free = max(0, max_capacity - after_total)

    routing_fallback = bool(plan.get("routing_fallback"))
    detour_ratio = max(1.0, float(plan.get("road_detour_ratio") or 1.0))
    if routing_fallback:
        routing_note = (
            '<div class="dispatch-plan-note dispatch-road-warning">⚠️ 道路服務暫時無法連線；'
            '本次資料為舊版相容備援；新版道路規劃在服務失效時會停止產生 AI 路線，避免橫切山脈。</div>'
        )
    else:
        detour_text = f"｜道路／直線約 {detour_ratio:.1f} 倍" if detour_ratio >= 1.35 else ""
        routing_note = (
            '<div class="dispatch-plan-note dispatch-road-ok">🛣️ 已依實際可行駛道路計算'
            f'{html.escape(detour_text)}；可自然判斷市區短折返、偏遠道路與繞山路線。</div>'
        )

    st.markdown(
        f"""
        <section class="dispatch-plan-card" data-ubike-station-name="{station_name_attr}">
          <div class="dispatch-plan-header">
            <div>
              <div class="dispatch-plan-kicker">{html.escape(title)}</div>
              <div class="dispatch-plan-title">{html.escape(station_name)}</div>
              <div class="dispatch-plan-region">{html.escape(str(plan.get('region') or ''))}</div>
            </div>
            <div class="dispatch-plan-badge">可調度 <strong>{safe_nonnegative_int(plan['dispatch_count'])}</strong> 台</div>
          </div>

          <div class="dispatch-plan-grid">
            <div><span>路網距離</span><strong>{float(plan['estimated_distance_km']):.1f} km</strong></div>
            <div><span>行車時間</span><strong>{float(plan['estimated_drive_minutes']):.0f} 分</strong></div>
            <div><span>預估總時間</span><strong>{float(plan['estimated_total_minutes']):.0f} 分</strong></div>
            <div><span>綜合效益</span><strong>{float(plan['score']):.2f}</strong><small>台／分</small></div>
          </div>

          <div class="dispatch-plan-action-label">場站現況／標準</div>
          <div class="dispatch-plan-action">
            2.0：{html.escape(vehicle_balance_text(plan.get('current_bike'), plan.get('standard_bike')))}<br>
            2.0E：{html.escape(vehicle_balance_text(plan.get('current_ebike'), plan.get('standard_ebike')))}<br>
            總車 {optional_count_text(plan.get('total_bikes'))} 台｜空位 {optional_count_text(plan.get('empty_spaces'))} 格｜總柱 {optional_count_text(plan.get('parking_spaces'))} 格
          </div>
          <div class="dispatch-plan-action-label">本站作業</div>
          <div class="dispatch-plan-action">{html.escape(dispatch_action_text(plan))}</div>

          <div class="dispatch-truck-compare">
            <div class="dispatch-truck-row dispatch-truck-before">
              <span>目前車上數量</span>
              <strong>2.0＝{before_bike} 台｜2.0E＝{before_ebike} 台</strong>
              <small>合計 {before_total}／{max_capacity} 台</small>
            </div>
            <div class="dispatch-truck-arrow" aria-hidden="true">↓</div>
            <div class="dispatch-truck-row dispatch-truck-after">
              <span>完成後貨車</span>
              <strong>2.0＝{after_bike} 台｜2.0E＝{after_ebike} 台</strong>
              <small>合計 {after_total}／{max_capacity} 台・剩餘 {after_free} 格</small>
            </div>
          </div>
          {routing_note}
        </section>
        """,
        unsafe_allow_html=True,
    )


def render_dispatch_truck_status(
    *,
    truck_bike: int,
    truck_ebike: int,
    max_capacity: int,
    locked: bool,
) -> None:
    """在調度區頂端顯示精簡貨車狀態，避免反覆打開設定確認。"""
    total = truck_bike + truck_ebike
    remaining = max(0, max_capacity - total)
    load_percent = min(100.0, total / max(1, max_capacity) * 100)
    lock_text = "🔒 行程中已鎖定" if locked else "可調整"
    st.markdown(
        f"""
        <section class="dispatch-truck-status">
          <div class="dispatch-truck-status-head">
            <span>🚚 目前貨車</span><small>{lock_text}</small>
          </div>
          <div class="dispatch-truck-status-grid">
            <div><span>2.0</span><strong>{truck_bike}<small>台</small></strong></div>
            <div><span>2.0E</span><strong>{truck_ebike}<small>台</small></strong></div>
            <div><span>合計</span><strong>{total}<small>／{max_capacity}</small></strong></div>
            <div><span>剩餘空位</span><strong>{remaining}<small>格</small></strong></div>
          </div>
          <div class="dispatch-load-track"><i style="width:{load_percent:.1f}%"></i></div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def render_dispatch_auxiliary_panels(
    *,
    dispatch_prefix: str,
    cooldown_key: str,
    cooldowns: dict[str, dict],
    history: list[dict],
    now_timestamp: float,
    current_round: int,
) -> None:
    """把暫時忽略場站與紀錄移到主要決策區後方，讓下一站卡片優先出現。"""
    if cooldowns:
        cooldown_lines = []
        for station_name, data in sorted(
            cooldowns.items(),
            key=lambda item: safe_nonnegative_int(item[1].get("resume_after_round")),
        ):
            remaining_rounds = max(
                1,
                safe_nonnegative_int(data.get("resume_after_round")) - current_round + 1,
            )
            cooldown_lines.append(
                f"{station_name}：尚忽略 {remaining_rounds} 回"
            )
        with st.expander(f"暫時忽略的場站（{len(cooldowns)}）", expanded=False):
            st.write("  \n".join(cooldown_lines))
            if st.button(
                "清除全部忽略",
                key=f"{dispatch_prefix}::clear_cooldowns",
                use_container_width=True,
            ):
                st.session_state[cooldown_key] = {}
                rerun_app()

    if history:
        with st.expander(f"本班次調度紀錄（{len(history)}）", expanded=False):
            history_rows = []
            for event in reversed(history[-30:]):
                event_time = datetime.fromtimestamp(
                    float(event.get("timestamp") or now_timestamp),
                    TAIPEI_TIMEZONE,
                ).strftime("%H:%M:%S")
                if event.get("action") == "rejected":
                    detail = str(event.get("note") or "使用者跳過")
                    result_text = f"跳過／忽略{DISPATCH_IGNORE_ROUNDS}回"
                elif event.get("action") == "cancelled":
                    detail = str(event.get("note") or "未執行本站作業，已重新定位")
                    result_text = f"取消配置／忽略{DISPATCH_IGNORE_ROUNDS}回"
                else:
                    detail = (
                        f"下車2.0 {safe_nonnegative_int(event.get('unload_bike'))}、"
                        f"上車2.0 {safe_nonnegative_int(event.get('pickup_bike'))}、"
                        f"下車2.0E {safe_nonnegative_int(event.get('unload_ebike'))}、"
                        f"上車2.0E {safe_nonnegative_int(event.get('pickup_ebike'))}"
                    )
                    result_text = "已完成"
                history_rows.append(
                    {
                        "時間": event_time,
                        "場站": event.get("station_name", ""),
                        "結果": result_text,
                        "原因／作業": detail,
                    }
                )
            st.dataframe(pd.DataFrame(history_rows), hide_index=True, use_container_width=True)

def render_smart_dispatch(
    *,
    full_status_df: pd.DataFrame,
    selected_region: str,
    status_cache: dict,
    current_context_key: str,
    active_base: dict,
    page_title: str = "智慧動態調度",
    page_caption: str | None = None,
    dispatch_scope: str = "standard",
    external_location: dict | None = None,
    fallback_location: dict | None = None,
    location_label: str = "",
    trip_mode: str = "單趟",
    endpoint_location: dict | None = None,
    endpoint_label: str = "",
    loop_direction_preference: str = "",
    loop_start_name: str = "",
    station_locations_override: dict[str, dict] | None = None,
    allow_manual_station_choice: bool = False,
    show_route_preview: bool = False,
    require_external_location: bool = False,
) -> None:
    """逐站詢問、載量限制、動態重算；一般模式可強制只採用背景 GPS。"""
    st.markdown('<div id="smart-dispatch-anchor"></div>', unsafe_allow_html=True)
    st.subheader(page_title)
    st.caption(
        page_caption
        or "每次只安排一站；同意後鎖定目的地，完成本站才重新計算。距離與時間優先使用實際可行駛道路路網，路網無路徑的場站不會用直線橫切補上。"
    )

    st.markdown(
        """
        <style>
        .dispatch-plan-card {
            border: 1px solid rgba(22,119,255,.25); border-radius: 20px;
            padding: 1rem; margin: .45rem 0 .7rem;
            background: linear-gradient(145deg, rgba(22,119,255,.085), rgba(16,185,129,.055));
            box-shadow: 0 10px 28px rgba(15,23,42,.07);
        }
        .dispatch-plan-header {display:flex; justify-content:space-between; align-items:flex-start; gap:.8rem;}
        .dispatch-plan-kicker {font-size:.76rem; font-weight:850; letter-spacing:.06em; color:#1677ff;}
        .dispatch-plan-title {font-size:1.55rem; line-height:1.2; font-weight:900; margin-top:.18rem;}
        .dispatch-plan-region {font-size:.8rem; opacity:.7; margin-top:.25rem;}
        .dispatch-plan-badge {flex:0 0 auto; padding:.42rem .62rem; border-radius:999px; background:rgba(22,119,255,.12); color:#0b63ce; font-size:.72rem; font-weight:750; white-space:nowrap;}
        .dispatch-plan-badge strong {font-size:1.02rem;}
        .dispatch-plan-grid {display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:.42rem; margin-top:.78rem;}
        .dispatch-plan-grid div {padding:.58rem .38rem; border-radius:12px; background:rgba(255,255,255,.76); text-align:center; min-width:0;}
        .dispatch-plan-grid span {display:block; font-size:.66rem; opacity:.66; white-space:nowrap;}
        .dispatch-plan-grid strong {display:inline-block; font-size:.98rem; margin-top:.16rem;}
        .dispatch-plan-grid small {font-size:.62rem; margin-left:.12rem; opacity:.7;}
        .dispatch-plan-action-label {margin-top:.68rem; font-size:.67rem; font-weight:800; opacity:.66;}
        .dispatch-plan-action {margin-top:.22rem; padding:.72rem .78rem; border-radius:12px; font-size:1.02rem; font-weight:900; background:rgba(22,119,255,.12);}
        .dispatch-truck-compare {margin-top:.58rem; padding:.55rem; border-radius:14px; background:rgba(255,255,255,.56);}
        .dispatch-truck-row {display:grid; grid-template-columns:6.2rem 1fr auto; gap:.45rem; align-items:center; padding:.44rem .5rem; border-radius:10px;}
        .dispatch-truck-row span {font-size:.72rem; font-weight:800; opacity:.68;}
        .dispatch-truck-row strong {font-size:.82rem;}
        .dispatch-truck-row small {font-size:.66rem; opacity:.66; white-space:nowrap;}
        .dispatch-truck-before {background:rgba(148,163,184,.08);}
        .dispatch-truck-after {background:rgba(16,185,129,.11);}
        .dispatch-truck-arrow {height:.55rem; line-height:.55rem; text-align:center; font-size:.7rem; opacity:.5;}
        .dispatch-plan-note {margin-top:.48rem; padding:.48rem .58rem; border-radius:10px; font-size:.72rem; line-height:1.45; background:rgba(245,158,11,.1);}
        .dispatch-road-ok {background:rgba(16,185,129,.10);}
        .dispatch-road-warning {background:rgba(245,158,11,.12);}
        .dispatch-truck-status {margin:.35rem 0 .7rem; padding:.72rem .78rem; border:1px solid rgba(148,163,184,.22); border-radius:16px; background:rgba(148,163,184,.06);}
        .dispatch-truck-status-head {display:flex; justify-content:space-between; align-items:center; margin-bottom:.48rem;}
        .dispatch-truck-status-head span {font-weight:850; font-size:.84rem;}
        .dispatch-truck-status-head small {font-size:.66rem; opacity:.62;}
        .dispatch-truck-status-grid {display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:.35rem;}
        .dispatch-truck-status-grid div {padding:.38rem .3rem; text-align:center; border-radius:10px; background:rgba(255,255,255,.68);}
        .dispatch-truck-status-grid span {display:block; font-size:.63rem; opacity:.62;}
        .dispatch-truck-status-grid strong {font-size:.98rem;}
        .dispatch-truck-status-grid strong small {font-size:.62rem; margin-left:.08rem; opacity:.7;}
        .dispatch-load-track {height:5px; margin-top:.5rem; border-radius:999px; overflow:hidden; background:rgba(148,163,184,.24);}
        .dispatch-load-track i {display:block; height:100%; border-radius:999px; background:linear-gradient(90deg,#1677ff,#10b981);}
        .dispatch-candidate-grid {display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:.48rem; margin:.55rem 0 .7rem;}
        .dispatch-candidate-card {padding:.68rem .72rem; border:1px solid rgba(148,163,184,.22); border-radius:14px; background:rgba(255,255,255,.72); min-width:0;}
        .dispatch-candidate-card.is-active {border-color:rgba(22,119,255,.52); background:rgba(22,119,255,.08); box-shadow:inset 0 0 0 1px rgba(22,119,255,.12);}
        .dispatch-candidate-head {display:flex; justify-content:space-between; gap:.55rem; align-items:flex-start;}
        .dispatch-candidate-rank {font-size:.67rem; font-weight:850; color:#1677ff;}
        .dispatch-candidate-name {font-size:.94rem; font-weight:900; line-height:1.3; margin-top:.12rem;}
        .dispatch-candidate-count {flex:0 0 auto; font-size:.7rem; font-weight:850; padding:.28rem .45rem; border-radius:999px; background:rgba(16,185,129,.12); color:#087f5b; white-space:nowrap;}
        .dispatch-candidate-action {margin-top:.45rem; font-size:.75rem; font-weight:800; line-height:1.45;}
        .dispatch-candidate-meta {display:flex; flex-wrap:wrap; gap:.28rem .55rem; margin-top:.4rem; font-size:.67rem; opacity:.7;}
        [class*="st-key-candidate_card_select_"] {margin-bottom:.48rem;}
        [class*="st-key-candidate_card_select_"] button {
            min-height:96px; height:auto; padding:.72rem .82rem; border-radius:16px;
            justify-content:flex-start; text-align:left; white-space:normal; line-height:1.48;
            box-shadow:0 4px 14px rgba(15,23,42,.035);
        }
        [class*="st-key-candidate_card_select_"] button p {
            width:100%; margin:0; text-align:left; white-space:normal; line-height:1.48;
        }
        @media (max-width: 700px) {
          .dispatch-plan-card {padding:.82rem; border-radius:18px;}
          .dispatch-plan-title {font-size:1.42rem;}
          .dispatch-plan-grid {grid-template-columns:repeat(2,minmax(0,1fr));}
          .dispatch-truck-row {grid-template-columns:5.7rem 1fr; gap:.2rem .4rem;}
          .dispatch-truck-row small {grid-column:2; white-space:normal;}
          .dispatch-truck-status {padding:.65rem;}
          .dispatch-truck-status-grid {grid-template-columns:repeat(4,minmax(0,1fr)); gap:.24rem;}
          .dispatch-truck-status-grid strong {font-size:.9rem;}
          .dispatch-candidate-grid {grid-template-columns:1fr;}
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    dispatch_prefix = f"smart_dispatch::{dispatch_scope}::{active_base['token']}::{current_context_key}"
    location_state_key = f"{dispatch_prefix}::location"
    active_trip_key = f"{dispatch_prefix}::active_trip"
    cooldown_key = f"{dispatch_prefix}::cooldowns"
    dispatch_round_key = f"{dispatch_prefix}::decision_round"
    history_key = f"{dispatch_prefix}::history"
    location_request_key = f"{dispatch_prefix}::location_request_token"
    location_request_pending_key = f"{dispatch_prefix}::location_request_pending"
    manual_station_key = f"{dispatch_prefix}::manual_next_station"
    floating_station_request_key = f"{dispatch_prefix}::floating_station_request"
    loop_order_key = f"{dispatch_prefix}::loop_zone_order"
    loop_phase_key = f"{dispatch_prefix}::loop_active_phase"
    battery_force_refresh_key = f"{dispatch_prefix}::battery_force_refresh_station"
    jarvis_processed_event_key = f"{dispatch_prefix}::jarvis_processed_event"
    jarvis_auto_announce_key = f"{dispatch_prefix}::jarvis_auto_announce"
    battery_threshold, battery_priority_threshold = get_low_battery_thresholds(
        active_base["token"]
    )
    battery_mobile_mode = is_mobile_browser()

    incoming_floating_station = pop_floating_dispatch_station_request()
    if incoming_floating_station:
        st.session_state[floating_station_request_key] = incoming_floating_station
    requested_station = str(
        st.session_state.get(floating_station_request_key, "") or ""
    ).strip()

    st.session_state.setdefault(cooldown_key, {})
    st.session_state.setdefault(dispatch_round_key, 0)
    st.session_state.setdefault(history_key, [])
    pending_truck_key = f"{dispatch_prefix}::pending_truck_counts"
    pending_truck_counts = st.session_state.pop(pending_truck_key, None)
    if isinstance(pending_truck_counts, dict):
        st.session_state[f"{dispatch_prefix}::truck_bike"] = safe_nonnegative_int(
            pending_truck_counts.get("bike")
        )
        st.session_state[f"{dispatch_prefix}::truck_ebike"] = safe_nonnegative_int(
            pending_truck_counts.get("ebike")
        )
    now_timestamp = time.time()
    current_round = safe_nonnegative_int(st.session_state.get(dispatch_round_key))

    # 場站只忽略接下來 2 個調度回合；第 3 回起自動恢復候選資格。
    raw_cooldowns = dict(st.session_state.get(cooldown_key, {}))
    cooldowns: dict[str, dict] = {}
    for station_name, data in raw_cooldowns.items():
        if not isinstance(data, dict):
            continue
        normalized_data = dict(data)
        resume_after_round = normalized_data.get("resume_after_round")
        if resume_after_round is None:
            # 相容舊版 10 分鐘冷卻資料：尚未到期者改為從現在起忽略 2 回。
            if float(normalized_data.get("until") or 0) <= now_timestamp:
                continue
            resume_after_round = current_round + DISPATCH_IGNORE_ROUNDS - 1
            normalized_data["resume_after_round"] = resume_after_round
            normalized_data.pop("until", None)
        if safe_nonnegative_int(resume_after_round) >= current_round:
            cooldowns[station_name] = normalized_data
    st.session_state[cooldown_key] = cooldowns
    history = list(st.session_state.get(history_key, []))

    active_trip = st.session_state.get(active_trip_key)
    trip_locked = isinstance(active_trip, dict)
    if trip_locked and requested_station:
        st.session_state.pop(floating_station_request_key, None)
        requested_station = ""
        st.info("目前已有鎖定中的目的地；完成或取消本站後，才能用懸浮搜尋改選下一站。")
    max_capacity_input_key = f"{dispatch_prefix}::max_capacity"
    truck_bike_input_key = f"{dispatch_prefix}::truck_bike"
    truck_ebike_input_key = f"{dispatch_prefix}::truck_ebike"
    has_saved_truck_settings = all(
        key in st.session_state
        for key in (max_capacity_input_key, truck_bike_input_key, truck_ebike_input_key)
    )

    settings_title = "🚚 貨車載量設定" + ("（前往中已鎖定）" if trip_locked else "")
    with st.expander(settings_title, expanded=not has_saved_truck_settings and not trip_locked):
        control_col_1, control_col_2, control_col_3 = st.columns(3)
        with control_col_1:
            max_capacity = int(st.number_input(
                "最高載量",
                min_value=1,
                max_value=100,
                value=14,
                step=1,
                key=max_capacity_input_key,
                disabled=trip_locked,
            ))
        with control_col_2:
            truck_bike = int(st.number_input(
                "車上 2.0",
                min_value=0,
                max_value=100,
                value=0,
                step=1,
                key=truck_bike_input_key,
                disabled=trip_locked,
            ))
        with control_col_3:
            truck_ebike = int(st.number_input(
                "車上 2.0E",
                min_value=0,
                max_value=100,
                value=0,
                step=1,
                key=truck_ebike_input_key,
                disabled=trip_locked,
            ))

    total_on_truck = truck_bike + truck_ebike
    render_dispatch_truck_status(
        truck_bike=truck_bike,
        truck_ebike=truck_ebike,
        max_capacity=max_capacity,
        locked=trip_locked,
    )
    if total_on_truck > max_capacity:
        st.error(
            f"目前車上合計 {total_on_truck} 台，已超過最高載量 {max_capacity} 台。請先修正數量，系統不會安排路線。"
        )
        if trip_mode == "一般模式":
            render_jarvis_voice_assistant(
                dispatch_prefix=dispatch_prefix, mode="candidate", candidates=[], current_plan=None,
                threshold=battery_threshold, priority_threshold=battery_priority_threshold,
                context_status="blocked", context_message="目前貨車載量超過上限，請先修正車上數量",
            )
        return

    shared_location_mode = require_external_location or fallback_location is not None or external_location is not None
    location_payload = None
    if not shared_location_mode:
        try:
            geolocation_component = get_dispatch_geolocation_component()
            location_payload = geolocation_component(
                key=f"dispatch_geolocation::{dispatch_prefix}",
                default=None,
                request_token=str(st.session_state.get(location_request_key) or ""),
                auto_start=False,
                auto_refresh=False,
                compact=False,
            )
        except Exception as exc:
            st.error(f"定位功能建立失敗：{exc}")

        if isinstance(location_payload, dict):
            location_event_id = str(location_payload.get("event_id") or "").strip()
            processed_location_event_key = f"{dispatch_prefix}::processed_location_event"
            if location_event_id and st.session_state.get(processed_location_event_key) != location_event_id:
                st.session_state[processed_location_event_key] = location_event_id
                response_request_token = str(location_payload.get("request_token") or "").strip()
                active_request_token = str(st.session_state.get(location_request_key) or "").strip()
                if response_request_token and response_request_token == active_request_token:
                    st.session_state[location_request_pending_key] = False
                    st.session_state.pop(location_request_key, None)
                if location_payload.get("ok"):
                    latitude = normalize_coordinate(location_payload.get("latitude"), -90.0, 90.0)
                    longitude = normalize_coordinate(location_payload.get("longitude"), -180.0, 180.0)
                    if latitude is not None and longitude is not None:
                        st.session_state[location_state_key] = {
                            "latitude": latitude,
                            "longitude": longitude,
                            "accuracy": max(0.0, float(location_payload.get("accuracy") or 0)),
                            "updated_at": now_timestamp,
                            "source": "gps",
                        }
                        st.success("目前位置已更新，下一站排名已重新計算。")
                else:
                    st.warning(f"目前位置尚未更新：{location_payload.get('error') or '定位失敗'}")

    stored_location = st.session_state.get(location_state_key)
    if require_external_location:
        current_location = dict(external_location) if location_payload_is_valid(external_location) else None
    else:
        current_location = newest_valid_location(stored_location, external_location, fallback_location)
    location_request_pending = bool(st.session_state.get(location_request_pending_key, False))
    if location_request_pending and not shared_location_mode:
        st.info("正在讀取取消配置後的目前位置；定位完成後會自動重新安排下一個場站。")
        if trip_mode == "一般模式":
            render_jarvis_voice_assistant(
                dispatch_prefix=dispatch_prefix, mode="candidate", candidates=[], current_plan=None,
                threshold=battery_threshold, priority_threshold=battery_priority_threshold,
                context_status="updating", context_message="正在更新目前位置，請稍後再說一次",
            )
        return
    if isinstance(current_location, dict):
        source_lookup = {
            "gps": "GPS定位",
            "completed_station": "上一個完成場站",
            "dispatch_start": location_label or "維調出發點",
        }
        source_text = source_lookup.get(str(current_location.get("source") or ""), location_label or "目前位置")
        accuracy = float(current_location.get("accuracy") or 0)
        accuracy_text = f"｜誤差約 {accuracy:.0f} 公尺" if accuracy and current_location.get("source") == "gps" else ""
        st.caption(
            f"計算起點：{source_text}｜{float(current_location['latitude']):.6f}, "
            f"{float(current_location['longitude']):.6f}{accuracy_text}"
        )
    else:
        st.info("尚未取得有效位置，系統無法把距離與行車時間納入下一站評估。")
        if trip_mode == "一般模式":
            render_jarvis_voice_assistant(
                dispatch_prefix=dispatch_prefix, mode="candidate", candidates=[], current_plan=None,
                threshold=battery_threshold, priority_threshold=battery_priority_threshold,
                context_status="updating", context_message="尚未取得有效 GPS 位置，正在等待定位",
            )
        return

    metadata = status_cache.get("metadata", {}).get(current_context_key, {})
    station_locations = (
        dict(station_locations_override)
        if isinstance(station_locations_override, dict)
        else (metadata.get("station_locations", {}) if isinstance(metadata, dict) else {})
    )
    if not isinstance(station_locations, dict) or not station_locations:
        st.info("尚未取得場站官方座標。請先在上方執行一次「高速取得全部 YouBike 場站車數」。")
        if trip_mode == "一般模式":
            render_jarvis_voice_assistant(
                dispatch_prefix=dispatch_prefix, mode="candidate", candidates=[], current_plan=None,
                threshold=battery_threshold, priority_threshold=battery_priority_threshold,
                context_status="updating", context_message="場站座標資料尚未完成，請稍後再說一次",
            )
        return


    dispatch_df = full_status_df.copy()
    if selected_region != "全部":
        dispatch_df = dispatch_df[dispatch_df["行政區"] == selected_region].copy()

    resolved_loop_order: list[str] = []
    active_loop_phase = ""
    if trip_mode == "環狀一圈":
        stored_loop_order = st.session_state.get(loop_order_key)
        if isinstance(stored_loop_order, (list, tuple)):
            stored_loop_order = [
                str(zone) for zone in stored_loop_order if str(zone) in LONG_DISTANCE_ROUTE_ZONES
            ]
        else:
            stored_loop_order = []

        explicit_loop_order = loop_zone_order_from_preference(loop_direction_preference)
        if len(stored_loop_order) == 2:
            resolved_loop_order = list(stored_loop_order)
            loop_resolution_text = "已鎖定"
        elif explicit_loop_order:
            resolved_loop_order = explicit_loop_order
            loop_resolution_text = "手動選擇"
        else:
            direction_previews: dict[tuple[str, str], list[dict]] = {}
            for candidate_order in (("D2", "D3"), ("D3", "D2")):
                direction_previews[candidate_order] = build_long_distance_route_preview(
                    dispatch_df,
                    station_locations=station_locations,
                    current_location=current_location,
                    truck_bike=truck_bike,
                    truck_ebike=truck_ebike,
                    max_capacity=max_capacity,
                    cooldowns=cooldowns,
                    rejection_history=history,
                    now_timestamp=now_timestamp,
                    current_round=current_round,
                    trip_mode=trip_mode,
                    endpoint_location=endpoint_location,
                    max_stops=8,
                    loop_zone_order=list(candidate_order),
                    loop_start_name=loop_start_name,
                )
            best_order = max(
                direction_previews,
                key=lambda order: summarize_loop_preview(direction_previews[order]),
            )
            resolved_loop_order = list(best_order)
            loop_resolution_text = "AI 自動選擇"

        stored_phase = str(st.session_state.get(loop_phase_key) or "").strip()
        active_loop_phase = (
            stored_phase if stored_phase in resolved_loop_order else resolved_loop_order[0]
        )
        st.info(
            f"環狀方向（{loop_resolution_text}）：{resolved_loop_order[0]} 先行 → "
            f"經 {LONG_DISTANCE_TRANSFER_LABEL} → {resolved_loop_order[1]} → 返回 {loop_start_name or '出發維調'}｜"
            f"目前階段：{active_loop_phase}"
        )

    if isinstance(active_trip, dict):
        trip_id = str(active_trip.get("trip_id") or normalize_station_key(active_trip.get("station_name")))
        render_dispatch_plan_card(active_trip, title="已同意前往／目的地已鎖定")
        active_station_name = str(active_trip.get("station_name") or "").strip()
        render_inline_low_battery_pillars(
            [(active_station_name, "plan", "")],
            threshold=battery_threshold,
            priority_threshold=battery_priority_threshold,
            mobile_mode=battery_mobile_mode,
            auto_query=True,
            force_station=str(st.session_state.pop(battery_force_refresh_key, "") or ""),
        )
        maps_query = urlencode(
            {
                "api": 1,
                "destination": f"{active_trip['latitude']},{active_trip['longitude']}",
                "travelmode": "driving",
            }
        )
        st.markdown(f"[🧭 開啟 Google Maps 導航](https://www.google.com/maps/dir/?{maps_query})")
        st.info("到場作業後，可依現場變數修改實際上下車數量；只有按下完成本站，系統才會安排下一站。")

        voice_completion_values: dict | None = None
        voice_cancel_requested = False
        voice_next_station = ""
        if trip_mode == "一般模式":
            voice_event = render_jarvis_voice_assistant(
                dispatch_prefix=dispatch_prefix,
                mode="active",
                candidates=[active_trip],
                current_plan=active_trip,
                threshold=battery_threshold,
                priority_threshold=battery_priority_threshold,
                auto_announce=False,
            )
            if isinstance(voice_event, dict):
                event_id = str(voice_event.get("event_id") or "").strip()
                if event_id and st.session_state.get(jarvis_processed_event_key) != event_id:
                    st.session_state[jarvis_processed_event_key] = event_id
                    event_type = str(voice_event.get("type") or "").strip()
                    if event_type == "fill_actual":
                        fill_values = {
                            "unload_bike": safe_nonnegative_int(voice_event.get("unload_bike")),
                            "unload_ebike": safe_nonnegative_int(voice_event.get("unload_ebike")),
                            "pickup_bike": safe_nonnegative_int(voice_event.get("pickup_bike")),
                            "pickup_ebike": safe_nonnegative_int(voice_event.get("pickup_ebike")),
                        }
                        st.session_state[f"{dispatch_prefix}::actual_unload_bike::{trip_id}"] = fill_values["unload_bike"]
                        st.session_state[f"{dispatch_prefix}::actual_unload_ebike::{trip_id}"] = fill_values["unload_ebike"]
                        st.session_state[f"{dispatch_prefix}::actual_pickup_bike::{trip_id}"] = fill_values["pickup_bike"]
                        st.session_state[f"{dispatch_prefix}::actual_pickup_ebike::{trip_id}"] = fill_values["pickup_ebike"]
                    elif event_type == "confirm_completion":
                        voice_completion_values = {
                            "unload_bike": safe_nonnegative_int(voice_event.get("unload_bike")),
                            "unload_ebike": safe_nonnegative_int(voice_event.get("unload_ebike")),
                            "pickup_bike": safe_nonnegative_int(voice_event.get("pickup_bike")),
                            "pickup_ebike": safe_nonnegative_int(voice_event.get("pickup_ebike")),
                        }
                    elif event_type in {"skip_current", "change_station"}:
                        voice_cancel_requested = True
                        if event_type == "change_station":
                            voice_next_station = resolve_voice_station_name(
                                full_status_df, str(voice_event.get("target_text") or "")
                            )

        with st.expander("現場變數／修改實際上下車數量", expanded=True):
            with st.form(key=f"{dispatch_prefix}::complete_trip_form::{trip_id}", clear_on_submit=False):
                action_col_1, action_col_2 = st.columns(2)
                with action_col_1:
                    actual_unload_bike = int(st.number_input(
                        "實際下車 2.0",
                        min_value=0,
                        value=safe_nonnegative_int(active_trip.get("unload_bike")),
                        step=1,
                        key=f"{dispatch_prefix}::actual_unload_bike::{trip_id}",
                    ))
                    actual_pickup_bike = int(st.number_input(
                        "實際上車 2.0",
                        min_value=0,
                        value=safe_nonnegative_int(active_trip.get("pickup_bike")),
                        step=1,
                        key=f"{dispatch_prefix}::actual_pickup_bike::{trip_id}",
                    ))
                with action_col_2:
                    actual_unload_ebike = int(st.number_input(
                        "實際下車 2.0E",
                        min_value=0,
                        value=safe_nonnegative_int(active_trip.get("unload_ebike")),
                        step=1,
                        key=f"{dispatch_prefix}::actual_unload_ebike::{trip_id}",
                    ))
                    actual_pickup_ebike = int(st.number_input(
                        "實際上車 2.0E",
                        min_value=0,
                        value=safe_nonnegative_int(active_trip.get("pickup_ebike")),
                        step=1,
                        key=f"{dispatch_prefix}::actual_pickup_ebike::{trip_id}",
                    ))

                completed = st.form_submit_button(
                    "✅ 完成本站並安排下一站",
                    type="primary",
                    use_container_width=True,
                )
                cancelled = st.form_submit_button(
                    "❌ 取消配置",
                    use_container_width=True,
                )

            if voice_completion_values is not None:
                actual_unload_bike = safe_nonnegative_int(voice_completion_values.get("unload_bike"))
                actual_unload_ebike = safe_nonnegative_int(voice_completion_values.get("unload_ebike"))
                actual_pickup_bike = safe_nonnegative_int(voice_completion_values.get("pickup_bike"))
                actual_pickup_ebike = safe_nonnegative_int(voice_completion_values.get("pickup_ebike"))
                completed = True
            if voice_cancel_requested:
                cancelled = True

            if cancelled:
                cancelled_at = time.time()
                station_name = str(active_trip.get("station_name") or "").strip()
                if station_name:
                    cooldowns[station_name] = {
                        "resume_after_round": current_round + DISPATCH_IGNORE_ROUNDS,
                        "reason": "取消配置",
                        "note": "已取消已鎖定配置，忽略2回後恢復評估",
                        "rejected_at": cancelled_at,
                    }
                    st.session_state[cooldown_key] = cooldowns
                st.session_state[dispatch_round_key] = current_round + 1
                history.append(
                    {
                        "action": "cancelled",
                        "station_name": station_name,
                        "timestamp": cancelled_at,
                        "note": "未執行本站上下車作業；保留目前貨車數量並重新定位",
                    }
                )
                st.session_state[history_key] = history[-100:]
                st.session_state.pop(active_trip_key, None)
                st.session_state.pop(manual_station_key, None)
                if voice_next_station:
                    st.session_state[manual_station_key] = voice_next_station
                st.session_state[jarvis_auto_announce_key] = True
                if shared_location_mode:
                    st.session_state.pop(location_state_key, None)
                else:
                    st.session_state.pop(location_state_key, None)
                    st.session_state[location_request_key] = uuid.uuid4().hex
                    st.session_state[location_request_pending_key] = True
                rerun_app()

            if completed:
                operation_start_bike, operation_start_ebike = _dispatch_truck_before_counts(active_trip)
                final_truck_bike = operation_start_bike - actual_unload_bike + actual_pickup_bike
                final_truck_ebike = operation_start_ebike - actual_unload_ebike + actual_pickup_ebike
                final_total = final_truck_bike + final_truck_ebike
                station_final_bike = (
                    safe_nonnegative_int(active_trip.get("current_bike"))
                    + actual_unload_bike
                    - actual_pickup_bike
                )
                station_final_ebike = (
                    safe_nonnegative_int(active_trip.get("current_ebike"))
                    + actual_unload_ebike
                    - actual_pickup_ebike
                )

                errors: list[str] = []
                if actual_unload_bike > operation_start_bike:
                    errors.append("實際下車的 2.0 超過車上現有數量")
                if actual_unload_ebike > operation_start_ebike:
                    errors.append("實際下車的 2.0E 超過車上現有數量")
                if final_truck_bike < 0 or final_truck_ebike < 0:
                    errors.append("作業後車上數量不可為負數")
                if final_total > max_capacity:
                    errors.append(f"作業後合計 {final_total} 台，超過最高載量 {max_capacity} 台")
                if station_final_bike < 0 or station_final_ebike < 0:
                    errors.append("實際上車數量超過本站作業前可用車數")

                if errors:
                    for error in errors:
                        st.error(error)
                else:
                    updated_full_df = full_status_df.copy()
                    station_mask = (
                        updated_full_df["場站名稱"].astype(str).eq(str(active_trip["station_name"]))
                        & updated_full_df["行政區"].astype(str).eq(str(active_trip.get("region") or ""))
                    )
                    if not station_mask.any():
                        st.error("找不到目前目的地在配置表中的資料，未寫入完成結果。")
                    else:
                        updated_full_df.loc[station_mask, "2.0 現況"] = station_final_bike
                        updated_full_df.loc[station_mask, "2.0E 現況"] = station_final_ebike
                        station_final_total = station_final_bike + station_final_ebike
                        updated_full_df.loc[station_mask, "總車數"] = station_final_total
                        previous_station_total = station_total_bikes(active_trip)
                        previous_empty_spaces = station_empty_spaces(active_trip)
                        if previous_station_total is not None and previous_empty_spaces is not None:
                            total_change = station_final_total - previous_station_total
                            updated_full_df.loc[station_mask, "空位數"] = max(
                                0,
                                previous_empty_spaces - total_change,
                            )
                        save_dispatch_dataframe_contexts(
                            updated_full_df,
                            status_cache=status_cache,
                            active_base=active_base,
                            default_context_key=current_context_key,
                        )
                        st.session_state[pending_truck_key] = {
                            "bike": final_truck_bike,
                            "ebike": final_truck_ebike,
                        }
                        st.session_state[location_state_key] = {
                            "latitude": float(active_trip["latitude"]),
                            "longitude": float(active_trip["longitude"]),
                            "accuracy": 0.0,
                            "updated_at": time.time(),
                            "source": "completed_station",
                        }
                        history.append(
                            {
                                "action": "completed",
                                "station_name": active_trip["station_name"],
                                "route_zone": str(active_trip.get("route_zone") or ""),
                                "timestamp": time.time(),
                                "unload_bike": actual_unload_bike,
                                "unload_ebike": actual_unload_ebike,
                                "pickup_bike": actual_pickup_bike,
                                "pickup_ebike": actual_pickup_ebike,
                            }
                        )
                        st.session_state[history_key] = history[-100:]
                        st.session_state[dispatch_round_key] = current_round + 1
                        st.session_state.pop(active_trip_key, None)
                        st.session_state.pop(manual_station_key, None)
                        st.session_state[jarvis_auto_announce_key] = True
                        rerun_app()
        render_dispatch_auxiliary_panels(
            dispatch_prefix=dispatch_prefix,
            cooldown_key=cooldown_key,
            cooldowns=cooldowns,
            history=history,
            now_timestamp=now_timestamp,
            current_round=current_round,
        )
        return

    candidate_df = dispatch_df
    phase_index = 0
    if trip_mode == "環狀一圈" and resolved_loop_order:
        phase_index = resolved_loop_order.index(active_loop_phase)
        candidate_df = dispatch_df[
            dispatch_df["路線區域"].astype(str).map(normalize_long_distance_zone).eq(active_loop_phase)
        ].copy()

    candidates = calculate_dispatch_candidates(
        candidate_df,
        station_locations=station_locations,
        current_location=current_location,
        truck_bike=truck_bike,
        truck_ebike=truck_ebike,
        max_capacity=max_capacity,
        cooldowns=cooldowns,
        rejection_history=history,
        now_timestamp=now_timestamp,
        current_round=current_round,
    )
    candidates = adjust_candidates_for_trip_mode(
        candidates,
        trip_mode=trip_mode,
        endpoint_location=endpoint_location,
    )
    if trip_mode == "環狀一圈" and resolved_loop_order:
        # 區域內不再以南北緯度限制；實際道路時間會自然決定短距離折返或遠距離順行。
        # 第一區已無可執行場站時，只跨越一次海岸山脈，轉往第二區後不再折返。
        if not candidates and phase_index == 0:
            active_loop_phase = resolved_loop_order[1]
            st.session_state[loop_phase_key] = active_loop_phase
            st.session_state.pop(manual_station_key, None)
            st.info(
                f"{resolved_loop_order[0]} 目前已無可執行場站，接下來經 {LONG_DISTANCE_TRANSFER_LABEL} "
                f"轉往 {resolved_loop_order[1]}；之後不會再自動折返 {resolved_loop_order[0]}。"
            )
            phase_index = 1
            candidate_df = dispatch_df[
                dispatch_df["路線區域"].astype(str).map(normalize_long_distance_zone).eq(active_loop_phase)
            ].copy()
            candidates = calculate_dispatch_candidates(
                candidate_df,
                station_locations=station_locations,
                current_location=current_location,
                truck_bike=truck_bike,
                truck_ebike=truck_ebike,
                max_capacity=max_capacity,
                cooldowns=cooldowns,
                rejection_history=history,
                now_timestamp=now_timestamp,
                current_round=current_round,
            )
            candidates = adjust_candidates_for_trip_mode(
                candidates,
                trip_mode=trip_mode,
                endpoint_location=endpoint_location,
            )

    # 智慧調度頁不再顯示重複的搜尋欄；右側懸浮搜尋會把點選結果送到這裡。
    if requested_station:
        st.session_state.pop(floating_station_request_key, None)
        normalized_request = normalize_station_key(requested_station)
        matched_station = next(
            (
                str(name)
                for name in dispatch_df["場站名稱"].dropna().astype(str).drop_duplicates()
                if normalize_station_key(name) == normalized_request
            ),
            "",
        )
        if not matched_station:
            st.warning(f"懸浮搜尋找不到「{requested_station}」，請重新搜尋。")
        else:
            requested_candidate = next(
                (
                    candidate for candidate in candidates
                    if str(candidate.get("station_name")) == matched_station
                ),
                None,
            )
            if requested_candidate is not None and allow_manual_station_choice:
                if matched_station == str(candidates[0]["station_name"]):
                    st.session_state.pop(manual_station_key, None)
                else:
                    st.session_state[manual_station_key] = matched_station
                st.session_state[battery_force_refresh_key] = matched_station
                persist_runtime_state(active_base["token"])
                st.success(f"已由懸浮搜尋套用：{matched_station}；仍需按『前往此站』才會正式鎖定。")
            else:
                inspection = inspect_dispatch_station(
                    dispatch_df,
                    station_name=matched_station,
                    station_locations=station_locations,
                    current_location=current_location,
                    truck_bike=truck_bike,
                    truck_ebike=truck_ebike,
                    max_capacity=max_capacity,
                    cooldowns=cooldowns,
                    current_round=current_round,
                    allowed_route_zone=(
                        active_loop_phase
                        if trip_mode == "環狀一圈" and resolved_loop_order
                        else ""
                    ),
                )
                st.warning(
                    f"{matched_station} 目前無法設為下一站："
                    f"{inspection.get('reason') or '目前沒有可執行的調度量。'}"
                )
                render_station_inspection_card(inspection)

    if not candidates:
        if trip_mode == "環狀一圈" and resolved_loop_order and active_loop_phase == resolved_loop_order[-1]:
            st.success(
                f"環狀路線的 {resolved_loop_order[0]}、{resolved_loop_order[1]} 目前都沒有可執行場站，"
                f"請返回 {loop_start_name or '出發維調'}。"
            )
            if location_payload_is_valid(endpoint_location):
                return_query = urlencode(
                    {
                        "api": 1,
                        "destination": f"{endpoint_location['latitude']},{endpoint_location['longitude']}",
                        "travelmode": "driving",
                    }
                )
                st.markdown(f"[🏁 導航返回 {loop_start_name or '出發維調'}](https://www.google.com/maps/dir/?{return_query})")
        else:
            road_status = st.session_state.get(ROAD_ROUTER_STATUS_STATE_KEY, {})
            if isinstance(road_status, dict) and road_status.get("ok") is False:
                if trip_mode == "一般模式":
                    st.error(
                        "道路路網服務暫時無法使用。為避免用直線距離誤判效率，本次不產生場站推薦；"
                        "請稍後按更新重新計算。"
                    )
                else:
                    st.error(
                        "道路路網服務暫時無法使用。為避免用直線誤判、橫切山脈，本次不產生 AI 路線；"
                        "請稍後按更新重新計算。"
                    )
            else:
                st.warning(
                    "目前找不到可執行的下一站。可能原因：全部符合配置、車上車種不足、貨車已滿、"
                    "場站仍在忽略回合、道路路網查無可行駛路線，或部分場站缺少座標／現況資料。"
                )
        render_dispatch_auxiliary_panels(
            dispatch_prefix=dispatch_prefix,
            cooldown_key=cooldown_key,
            cooldowns=cooldowns,
            history=history,
            now_timestamp=now_timestamp,
            current_round=current_round,
        )
        if trip_mode == "一般模式":
            road_status = st.session_state.get(ROAD_ROUTER_STATUS_STATE_KEY, {})
            road_failed = isinstance(road_status, dict) and road_status.get("ok") is False
            render_jarvis_voice_assistant(
                dispatch_prefix=dispatch_prefix,
                mode="candidate",
                candidates=[],
                current_plan=None,
                threshold=battery_threshold,
                priority_threshold=battery_priority_threshold,
                context_status="unavailable" if road_failed else "no_candidates",
                context_message=(
                    "道路路網服務暫時無法使用，請稍後更新重新計算"
                    if road_failed else
                    "目前沒有可執行場站"
                ),
            )
        return

    if show_route_preview and len(candidates) > 1:
        candidates = rerank_candidates_with_road_lookahead(
            candidates,
            dispatch_df=dispatch_df,
            station_locations=station_locations,
            current_location=current_location,
            truck_bike=truck_bike,
            truck_ebike=truck_ebike,
            max_capacity=max_capacity,
            cooldowns=cooldowns,
            rejection_history=history,
            now_timestamp=now_timestamp,
            current_round=current_round,
            trip_mode=trip_mode,
            endpoint_location=endpoint_location,
            loop_zone_order=resolved_loop_order,
            loop_start_name=loop_start_name,
            active_loop_phase=active_loop_phase,
        )

    manual_station_name = str(st.session_state.get(manual_station_key) or "").strip()
    if manual_station_name and not any(
        candidate["station_name"] == manual_station_name for candidate in candidates
    ):
        st.session_state.pop(manual_station_key, None)
        manual_station_name = ""

    recommended = next(
        (candidate for candidate in candidates if candidate["station_name"] == manual_station_name),
        candidates[0],
    )

    voice_rejected = False
    if trip_mode == "一般模式":
        auto_announce = bool(st.session_state.pop(jarvis_auto_announce_key, False))
        voice_event = render_jarvis_voice_assistant(
            dispatch_prefix=dispatch_prefix,
            mode="candidate",
            candidates=candidates[:SMART_DISPATCH_CANDIDATE_LIMIT],
            current_plan=recommended,
            threshold=battery_threshold,
            priority_threshold=battery_priority_threshold,
            auto_announce=auto_announce,
        )
        if isinstance(voice_event, dict):
            event_id = str(voice_event.get("event_id") or "").strip()
            if event_id and st.session_state.get(jarvis_processed_event_key) != event_id:
                st.session_state[jarvis_processed_event_key] = event_id
                event_type = str(voice_event.get("type") or "").strip()
                if event_type == "lock_station":
                    locked_station = str(voice_event.get("station_name") or recommended.get("station_name") or "").strip()
                    locked_candidate = next(
                        (candidate for candidate in candidates if str(candidate.get("station_name")) == locked_station),
                        recommended,
                    )
                    locked_trip = dict(locked_candidate)
                    locked_trip["trip_id"] = uuid.uuid4().hex
                    locked_trip["accepted_at"] = time.time()
                    st.session_state[active_trip_key] = locked_trip
                    st.session_state[battery_force_refresh_key] = str(locked_trip.get("station_name") or "").strip()
                    rerun_app()
                elif event_type == "skip_current":
                    voice_rejected = True
                elif event_type == "change_station":
                    target_station = resolve_voice_station_name(
                        dispatch_df, str(voice_event.get("target_text") or "")
                    )
                    target_candidate = next(
                        (candidate for candidate in candidates if str(candidate.get("station_name")) == target_station),
                        None,
                    )
                    if target_candidate is not None:
                        if target_station == str(candidates[0].get("station_name") or ""):
                            st.session_state.pop(manual_station_key, None)
                        else:
                            st.session_state[manual_station_key] = target_station
                        st.session_state[battery_force_refresh_key] = target_station
                        st.session_state[jarvis_auto_announce_key] = True
                        rerun_app()
                    else:
                        voice_rejected = True

    recommendation_title = "使用者指定下一站" if manual_station_name else "下一站最高效益推薦"
    render_dispatch_plan_card(recommended, title=recommendation_title)

    smart_alerts = build_station_alert_records(dispatch_df)
    if smart_alerts:
        with st.expander(f"🚨 警示場站快捷操作（{len(smart_alerts)}）", expanded=False):
            st.caption("警示不影響 AI 分數；可查看、設為目前選擇，或暫時略過 2 回。")
            for alert_index, alert in enumerate(smart_alerts[:10], start=1):
                alert_station = str(alert["station_name"])
                alert_candidate = next(
                    (
                        candidate for candidate in candidates
                        if str(candidate.get("station_name")) == alert_station
                    ),
                    None,
                )
                st.markdown(
                    f"**{'🚨' if safe_nonnegative_int(alert.get('severity')) >= 2 else '⚠️'} "
                    f"{alert.get('route_zone') or ''}｜{alert_station}｜{alert.get('message') or ''}**"
                )
                alert_col_1, alert_col_2, alert_col_3 = st.columns(3)
                detail_key = f"{dispatch_prefix}::alert_detail::{normalize_station_key(alert_station)}"
                with alert_col_1:
                    if st.button(
                        "查看詳情",
                        use_container_width=True,
                        key=f"{detail_key}::toggle::{alert_index}",
                    ):
                        st.session_state[detail_key] = not bool(st.session_state.get(detail_key))
                        rerun_app()
                with alert_col_2:
                    if st.button(
                        "設為下一站",
                        type="primary",
                        use_container_width=True,
                        disabled=alert_candidate is None,
                        key=f"{dispatch_prefix}::alert_use::{alert_index}::{normalize_station_key(alert_station)}",
                        help=(
                            "套用為目前選擇，仍需按『前往此站』才會鎖定。"
                            if alert_candidate is not None
                            else "此站依目前貨車載量、道路、配置或環狀階段無法執行。"
                        ),
                    ):
                        if alert_station == str(candidates[0]["station_name"]):
                            st.session_state.pop(manual_station_key, None)
                        else:
                            st.session_state[manual_station_key] = alert_station
                        st.session_state[battery_force_refresh_key] = alert_station
                        persist_runtime_state(active_base["token"])
                        rerun_app()
                with alert_col_3:
                    if st.button(
                        "暫時略過",
                        use_container_width=True,
                        key=f"{dispatch_prefix}::alert_skip::{alert_index}::{normalize_station_key(alert_station)}",
                    ):
                        skipped_at = time.time()
                        cooldowns[alert_station] = {
                            "resume_after_round": current_round + DISPATCH_IGNORE_ROUNDS,
                            "reason": "警示區暫時略過",
                            "note": "",
                            "rejected_at": skipped_at,
                        }
                        st.session_state[cooldown_key] = cooldowns
                        st.session_state[dispatch_round_key] = current_round + 1
                        history.append(
                            {
                                "action": "rejected",
                                "station_name": alert_station,
                                "timestamp": skipped_at,
                                "reason": "使用者跳過",
                                "note": "由警示區暫時略過",
                            }
                        )
                        st.session_state[history_key] = history[-100:]
                        if str(st.session_state.get(manual_station_key) or "") == alert_station:
                            st.session_state.pop(manual_station_key, None)
                        rerun_app()
                if st.session_state.get(detail_key):
                    if alert_candidate is not None:
                        render_station_inspection_card({**alert_candidate, "processable": True, "reason": "此站為目前可執行候選。"})
                    else:
                        alert_inspection = inspect_dispatch_station(
                            dispatch_df,
                            station_name=alert_station,
                            station_locations=station_locations,
                            current_location=current_location,
                            truck_bike=truck_bike,
                            truck_ebike=truck_ebike,
                            max_capacity=max_capacity,
                            cooldowns=cooldowns,
                            current_round=current_round,
                            allowed_route_zone=(
                                active_loop_phase
                                if trip_mode == "環狀一圈" and resolved_loop_order
                                else ""
                            ),
                        )
                        render_station_inspection_card(alert_inspection)

    # 候選場站直接以整張按鈕卡呈現；點一下即套用，不再使用折疊面板與上方下拉選單。
    visible_candidates = candidates[:SMART_DISPATCH_CANDIDATE_LIMIT]
    if visible_candidates:
        st.markdown(
            f"""
            <div style="display:flex;justify-content:space-between;align-items:end;gap:.7rem;margin:.82rem 0 .42rem;">
              <div>
                <div style="font-size:1rem;font-weight:900;">直接點選下一站</div>
                <div style="font-size:.72rem;opacity:.64;margin-top:.12rem;">點選任一場站後立即重新計算，不需要再按確認。</div>
              </div>
              <div style="font-size:.7rem;font-weight:800;opacity:.58;white-space:nowrap;">共 {len(visible_candidates)} 站</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        candidate_key_scope = hashlib.sha1(dispatch_prefix.encode("utf-8")).hexdigest()[:10]
        candidate_battery_specs: list[tuple[str, str, str]] = [
            (str(recommended["station_name"]), "plan", "")
        ]
        for rank, candidate in enumerate(visible_candidates, start=1):
            station_name = str(candidate["station_name"])
            is_active = station_name == str(recommended["station_name"])
            rank_text = "🤖 AI 首選" if rank == 1 else f"第 {rank} 名"
            route_zone = str(candidate.get("route_zone") or "").strip()
            rank_zone_text = f"{rank_text}｜{route_zone}" if route_zone else rank_text
            total_minutes = float(
                candidate.get("route_total_minutes", candidate.get("estimated_total_minutes") or 0)
            )
            active_text = "｜✅ 目前採用" if is_active else ""
            card_label = (
                f"**{rank_zone_text}｜{station_name}｜可調度 "
                f"{safe_nonnegative_int(candidate.get('dispatch_count'))} 台{active_text}**  \n"
                f"2.0：{vehicle_balance_text(candidate.get('current_bike'), candidate.get('standard_bike'))}｜"
                f"2.0E：{vehicle_balance_text(candidate.get('current_ebike'), candidate.get('standard_ebike'))}  \n"
                f"總車 {optional_count_text(candidate.get('total_bikes'))} 台｜"
                f"空位 {optional_count_text(candidate.get('empty_spaces'))} 格｜"
                f"總柱 {optional_count_text(candidate.get('parking_spaces'))} 格  \n"
                f"{dispatch_action_text(candidate)}  \n"
                f"🛣️ {float(candidate.get('estimated_distance_km') or 0):.1f} km　"
                f"⏱️ {total_minutes:.0f} 分　⚡ {float(candidate.get('score') or 0):.2f} 台／分"
            )

            candidate_widget_key = (
                f"candidate_card_select_{candidate_key_scope}_{rank}_"
                f"{normalize_station_key(station_name)}"
            )
            candidate_battery_specs.append(
                (station_name, "candidate", f"st-key-{candidate_widget_key}")
            )
            candidate_clicked = st.button(
                card_label,
                type="primary" if is_active else "secondary",
                use_container_width=True,
                disabled=not allow_manual_station_choice,
                key=candidate_widget_key,
                help=(
                    "目前已採用此場站；點擊後維持此選擇。"
                    if is_active
                    else f"直接改選 {station_name} 並重新計算。"
                ),
            )

            if candidate_clicked:
                if station_name == str(candidates[0]["station_name"]):
                    st.session_state.pop(manual_station_key, None)
                else:
                    st.session_state[manual_station_key] = station_name
                st.session_state[battery_force_refresh_key] = station_name
                persist_runtime_state(active_base["token"])
                rerun_app()

        render_inline_low_battery_pillars(
            candidate_battery_specs,
            threshold=battery_threshold,
            priority_threshold=battery_priority_threshold,
            mobile_mode=battery_mobile_mode,
            auto_query=True,
            force_station=str(st.session_state.pop(battery_force_refresh_key, "") or ""),
        )

        if allow_manual_station_choice and manual_station_name:
            if st.button(
                "↩️ 恢復 AI 最高效益首選",
                use_container_width=True,
                key=f"{dispatch_prefix}::clear_manual_station",
            ):
                st.session_state.pop(manual_station_key, None)
                st.session_state[battery_force_refresh_key] = str(
                    candidates[0].get("station_name") or ""
                ).strip()
                persist_runtime_state(active_base["token"])
                rerun_app()

    if show_route_preview:
        preview = build_long_distance_route_preview(
            dispatch_df,
            station_locations=station_locations,
            current_location=current_location,
            truck_bike=truck_bike,
            truck_ebike=truck_ebike,
            max_capacity=max_capacity,
            cooldowns=cooldowns,
            rejection_history=history,
            now_timestamp=now_timestamp,
            current_round=current_round,
            trip_mode=trip_mode,
            endpoint_location=endpoint_location,
            forced_first_station=manual_station_name,
            max_stops=8 if trip_mode == "環狀一圈" else 6,
            loop_zone_order=resolved_loop_order,
            loop_start_name=loop_start_name,
            active_loop_phase=active_loop_phase,
        )
        with st.expander("查看完整路線預覽", expanded=True):
            render_long_distance_route_preview(
                preview,
                trip_mode=trip_mode,
                endpoint_label=endpoint_label,
                loop_zone_order=resolved_loop_order,
            )

    decision_col_1, decision_col_2 = st.columns(2)
    with decision_col_1:
        accepted = st.button(
            "✅ 前往此站",
            type="primary",
            use_container_width=True,
            key=f"{dispatch_prefix}::accept_recommendation",
        )
    with decision_col_2:
        rejected = st.button(
            "⏭️ 跳過並找下一站",
            use_container_width=True,
            key=f"{dispatch_prefix}::skip_recommendation",
        )

    if voice_rejected:
        rejected = True

    if accepted:
        locked_trip = dict(recommended)
        locked_trip["trip_id"] = uuid.uuid4().hex
        locked_trip["accepted_at"] = time.time()
        st.session_state[battery_force_refresh_key] = str(
            recommended.get("station_name") or ""
        ).strip()
        if trip_mode == "環狀一圈" and resolved_loop_order:
            st.session_state[loop_order_key] = list(resolved_loop_order)
            st.session_state[loop_phase_key] = active_loop_phase or resolved_loop_order[0]
            locked_trip["loop_zone_order"] = list(resolved_loop_order)
            locked_trip["loop_active_phase"] = active_loop_phase or resolved_loop_order[0]
        st.session_state[active_trip_key] = locked_trip
        rerun_app()

    if rejected:
        rejected_at = time.time()
        cooldowns[recommended["station_name"]] = {
            "resume_after_round": current_round + DISPATCH_IGNORE_ROUNDS,
            "reason": "使用者跳過",
            "note": "",
            "rejected_at": rejected_at,
        }
        st.session_state[cooldown_key] = cooldowns
        st.session_state[dispatch_round_key] = current_round + 1
        history.append(
            {
                "action": "rejected",
                "station_name": recommended["station_name"],
                "timestamp": rejected_at,
                "reason": "使用者跳過",
                "note": "",
                "score_before_rejection": recommended["score"],
                "dispatch_count": recommended["dispatch_count"],
            }
        )
        st.session_state[history_key] = history[-100:]
        st.session_state.pop(manual_station_key, None)
        st.session_state[jarvis_auto_announce_key] = True
        rerun_app()

    render_dispatch_auxiliary_panels(
        dispatch_prefix=dispatch_prefix,
        cooldown_key=cooldown_key,
        cooldowns=cooldowns,
        history=history,
        now_timestamp=now_timestamp,
        current_round=current_round,
    )


def render_long_distance_route_page(
    *,
    active_base: dict,
    options: list[tuple[str, str]],
    selected_sheet: str,
    selected_shift: str,
    configuration_type: str,
    status_cache: dict,
    shared_location: dict | None,
    prebuilt_status_df: pd.DataFrame | None = None,
    prebuilt_station_locations: dict[str, dict] | None = None,
) -> pd.DataFrame | None:
    """整合一般即時推薦與 D2／D3 長途路線的智慧調度頁。"""
    st.markdown(
        """
        <section style="padding:1rem 1.05rem;border:1px solid rgba(22,119,255,.22);border-radius:20px;
        background:linear-gradient(135deg,rgba(22,119,255,.08),rgba(16,185,129,.05));margin:.35rem 0 .9rem;">
          <div style="font-size:1.45rem;font-weight:900;">🚚 智慧調度</div>
          <div style="margin-top:.3rem;font-size:.86rem;opacity:.72;line-height:1.55;">
            一般模式只依最新定位、即時車數、道路時間與貨車載量推薦最高效率單站；
            單趟、來回與環狀模式才會建立後續路線。
          </div>
        </section>
        """,
        unsafe_allow_html=True,
    )

    settings_prefix = (
        f"long_distance_settings::{active_base['token']}::{configuration_type}::{selected_shift}"
    )
    long_context_key = f"全區智慧調度｜{configuration_type}｜{selected_sheet}｜{selected_shift}"
    dispatch_prefix = f"smart_dispatch::long_distance::{active_base['token']}::{long_context_key}"
    active_trip_key = f"{dispatch_prefix}::active_trip"
    loop_order_key = f"{dispatch_prefix}::loop_zone_order"
    loop_phase_key = f"{dispatch_prefix}::loop_active_phase"
    active_trip = st.session_state.get(active_trip_key)
    trip_locked = isinstance(active_trip, dict)
    stored_loop_order = st.session_state.get(loop_order_key)
    loop_started = (
        isinstance(stored_loop_order, (list, tuple))
        and [str(zone) for zone in stored_loop_order if str(zone) in LONG_DISTANCE_ROUTE_ZONES]
        in (["D2", "D3"], ["D3", "D2"])
    )
    settings_locked = trip_locked or loop_started

    if loop_started and not trip_locked:
        reset_col_1, reset_col_2 = st.columns([2, 1])
        with reset_col_1:
            st.caption(
                f"環狀方向已鎖定：{stored_loop_order[0]} 先行 → 經 {LONG_DISTANCE_TRANSFER_LABEL} → "
                f"{stored_loop_order[1]}。完成整圈或需要改方向時再重新規劃。"
            )
        with reset_col_2:
            if st.button(
                "重新規劃整圈",
                use_container_width=True,
                key=f"{settings_prefix}::reset_loop_route",
            ):
                st.session_state.pop(loop_order_key, None)
                st.session_state.pop(loop_phase_key, None)
                st.session_state.pop(f"{dispatch_prefix}::manual_next_station", None)
                rerun_app()

    with st.expander(
        "本次智慧調度設定" + ("（任務執行中已鎖定）" if settings_locked else ""),
        expanded=not settings_locked,
    ):
        trip_mode = st.radio(
            "路線模式",
            ["一般模式", "單趟", "來回", "環狀一圈"],
            horizontal=True,
            key=f"{settings_prefix}::trip_mode",
            disabled=settings_locked,
        )

        loop_direction_preference = ""
        start_name = ""
        endpoint_name = ""

        if trip_mode == "一般模式":
            selected_zones = list(ALL_DISPATCH_ZONES)
            st.caption(
                "一般模式固定讀取 D1、D2、D3；只使用每 30 秒更新的即時定位安排最高效率場站，"
                "不建立完整路線或後續站序。"
            )
        else:
            start_name = st.selectbox(
                "出發維調",
                list(LONG_DISTANCE_START_POINTS.keys()),
                key=f"{settings_prefix}::start_name",
                disabled=settings_locked,
            )

            if trip_mode == "環狀一圈":
                selected_zones = list(LONG_DISTANCE_ROUTE_ZONES)
                st.multiselect(
                    "執行範圍",
                    list(LONG_DISTANCE_ROUTE_ZONES),
                    default=list(LONG_DISTANCE_ROUTE_ZONES),
                    key=f"{settings_prefix}::loop_zones",
                    disabled=True,
                    help="環狀一圈固定同時載入 D2 與 D3。",
                )
                loop_direction_preference = st.radio(
                    "環狀方向",
                    list(LONG_DISTANCE_LOOP_DIRECTION_OPTIONS),
                    horizontal=True,
                    key=f"{settings_prefix}::loop_direction",
                    disabled=settings_locked,
                )
                endpoint_name = start_name
                st.caption(
                    f"環狀規則：從 {start_name} 出發 → 先完成第一區 → 經 {LONG_DISTANCE_TRANSFER_LABEL} "
                    "跨越海岸山脈 → 完成第二區 → 返回出發維調。開始後不會在 D2、D3 之間反覆折返。"
                )
            else:
                selected_zones = st.multiselect(
                    "執行範圍",
                    list(LONG_DISTANCE_ROUTE_ZONES),
                    default=list(LONG_DISTANCE_ROUTE_ZONES),
                    key=f"{settings_prefix}::zones",
                    disabled=trip_locked,
                )
                if trip_mode == "單趟":
                    endpoint_choice = st.selectbox(
                        "單趟結束方向",
                        ["最後一站結束", "台東維調", "池上維調"],
                        key=f"{settings_prefix}::single_endpoint",
                        disabled=trip_locked,
                    )
                    if endpoint_choice != "最後一站結束":
                        endpoint_name = endpoint_choice
                else:
                    endpoint_name = start_name

            start_data = LONG_DISTANCE_START_POINTS[start_name]
            st.caption(f"{start_name}｜{start_data['description']}（固定位置為約略點，實際計算優先採用最新 GPS）")

    if not selected_zones:
        st.warning("請至少選擇一個執行區域。")
        return

    if trip_mode != "環狀一圈" and not trip_locked:
        st.session_state.pop(loop_order_key, None)
        st.session_state.pop(loop_phase_key, None)

    render_shared_location_summary(active_base, shared_location)
    refresh_col_1, refresh_col_2 = st.columns([1, 2])
    with refresh_col_1:
        if st.button(
            "立即更新定位",
            use_container_width=True,
            key=f"{settings_prefix}::refresh_location",
        ):
            request_shared_geolocation_refresh(active_base)
            rerun_app()
    with refresh_col_2:
        live_meta = status_cache.get("metadata", {})
        latest_times = [
            str(meta.get("fetched_at") or "")
            for meta in live_meta.values()
            if isinstance(meta, dict) and meta.get("fetched_at")
        ]
        if latest_times:
            suffix = "重新計算未鎖定推薦" if trip_mode == "一般模式" else "重新計算未鎖定路線"
            st.caption(f"即時車數最近同步：{max(latest_times)}｜資料變動後會{suffix}")
        else:
            st.caption("即時車數尚未完成第一次同步；上方同步元件完成後會自動帶入。")

    if isinstance(prebuilt_status_df, pd.DataFrame) and not prebuilt_status_df.empty:
        zone_mask = prebuilt_status_df["路線區域"].astype(str).isin(selected_zones)
        dispatch_status_df = prebuilt_status_df.loc[zone_mask].copy().reset_index(drop=True)
        station_locations = (
            dict(prebuilt_station_locations)
            if isinstance(prebuilt_station_locations, dict)
            else {}
        )
    else:
        dispatch_status_df, station_locations = build_long_distance_status_dataframe(
            active_base=active_base,
            options=options,
            selected_sheet=selected_sheet,
            selected_shift=selected_shift,
            selected_zones=selected_zones,
            status_cache=status_cache,
        )
    if dispatch_status_df.empty:
        st.warning("目前配置中找不到可用的場站。")
        return

    start_location = None
    if start_name:
        start_location = {
            "latitude": float(LONG_DISTANCE_START_POINTS[start_name]["latitude"]),
            "longitude": float(LONG_DISTANCE_START_POINTS[start_name]["longitude"]),
            "accuracy": 0.0,
            "updated_at": 0.0,
            "source": "dispatch_start",
        }

    endpoint_location = None
    endpoint_label = ""
    if endpoint_name:
        endpoint_data = LONG_DISTANCE_START_POINTS[endpoint_name]
        endpoint_location = {
            "latitude": float(endpoint_data["latitude"]),
            "longitude": float(endpoint_data["longitude"]),
            "updated_at": 0.0,
            "source": "dispatch_start",
        }
        endpoint_label = f"{endpoint_name}（{endpoint_data['description']}）"

    if trip_mode == "一般模式":
        page_title = "全區最高效益場站"
        page_caption = (
            "系統只評估目前 GPS 位置到各場站的可行駛道路時間、可調度台數與貨車剩餘載量，"
            "每次只推薦一站，不產生完整路線。即時車數或定位更新後，未鎖定推薦會立即重算。"
        )
    else:
        zone_text = "＋".join(selected_zones)
        page_title = f"{zone_text} AI 動態路線"
        page_caption = (
            "AI 會依可行駛道路時間預看後續三站，再決定真正的下一站；每完成一站、即時數據更新或手動改選後，"
            "都會重排未鎖定路線。市區短折返會自然保留，偏遠場站與繞山成本則會完整計入。"
        )

    render_smart_dispatch(
        full_status_df=dispatch_status_df,
        selected_region="全部",
        status_cache=status_cache,
        current_context_key=long_context_key,
        active_base=active_base,
        page_title=page_title,
        page_caption=page_caption,
        dispatch_scope="long_distance",
        external_location=shared_location,
        fallback_location=None if trip_mode == "一般模式" else start_location,
        location_label="即時 GPS" if trip_mode == "一般模式" else start_name,
        trip_mode=trip_mode,
        endpoint_location=endpoint_location,
        endpoint_label=endpoint_label,
        loop_direction_preference=loop_direction_preference,
        loop_start_name=start_name,
        station_locations_override=station_locations,
        allow_manual_station_choice=True,
        show_route_preview=trip_mode != "一般模式",
        require_external_location=trip_mode == "一般模式",
    )
    return dispatch_status_df


@fragment_if_available
def render_smart_dispatch_page_fragment(
    *,
    active_base: dict,
    options: list[tuple[str, str]],
    selected_sheet: str,
    selected_shift: str,
    configuration_type: str,
    status_cache: dict,
    shared_location: dict | None,
    base_df: pd.DataFrame,
    station_locations: dict[str, dict],
    battery_route_map: dict[str, list[dict]],
    mobile_mode: bool,
) -> None:
    """將智慧調度互動限制在局部重跑，並直接沿用本輪已整合的全區資料。"""
    search_source = render_long_distance_route_page(
        active_base=active_base,
        options=options,
        selected_sheet=selected_sheet,
        selected_shift=selected_shift,
        configuration_type=configuration_type,
        status_cache=status_cache,
        shared_location=shared_location,
        prebuilt_status_df=base_df,
        prebuilt_station_locations=station_locations,
    )
    if not isinstance(search_source, pd.DataFrame) or search_source.empty:
        search_source = base_df
    render_floating_station_search(
        build_analysis_result(search_source),
        mobile_mode,
        page_mode="智慧調度",
    )
    render_floating_battery_query(
        battery_route_map,
        mobile_mode,
    )
    persist_runtime_state(active_base["token"])
    st.caption(
        "即時車數來源：YouBike 官網；智慧推薦與道路時間為動態估算，"
        "仍請以現場與實際路況為準。"
    )


@st.cache_resource(show_spinner=False, max_entries=8)
def cached_load_workbook(source: bytes | str) -> dict[str, pd.DataFrame]:
    """只開啟 Excel 一次；共用唯讀工作簿，避免每次快取命中仍反序列化整本檔案。"""
    excel_source = BytesIO(source) if isinstance(source, bytes) else source
    with pd.ExcelFile(excel_source, engine="openpyxl") as book:
        workbook = book.book
        visible_sheet_names = [
            worksheet.title
            for worksheet in workbook.worksheets
            if worksheet.sheet_state == "visible"
            and worksheet.title in book.sheet_names
        ]
        return {
            sheet_name: pd.read_excel(
                book,
                sheet_name=sheet_name,
                header=None,
                dtype=object,
                engine="openpyxl",
            )
            for sheet_name in visible_sheet_names
        }


@st.cache_data(show_spinner=False, max_entries=16)
def cached_available_sources(source: bytes | str) -> list[tuple[str, str]]:
    """只保存輕量的工作表／分區選項，平常重跑不複製整本 Excel。"""
    return list(available_sources(cached_load_workbook(source)))


@st.cache_data(show_spinner=False, max_entries=48)
def cached_parse_route(source: bytes | str, sheet_name: str, route: str, shift: str) -> pd.DataFrame:
    """快取已解析的分區資料，避免每次輸入都重新掃描 Excel 工作表。"""
    workbook = cached_load_workbook(source)
    # 資源快取中的工作表視為唯讀；只在首次解析此路線時複製單一工作表。
    return parse_route(workbook[sheet_name].copy(deep=True), route, shift)


@fragment_if_available
def render_general_analysis_results_fragment(
    *,
    edited_df: pd.DataFrame,
    analysis_result_df: pd.DataFrame,
    road_metrics_available: bool,
    road_metrics_message: str,
    battery_route_map: dict[str, list[dict]],
    mobile_mode: bool,
    current_context_key: str,
    selected_region: str,
    configuration_type: str,
    selected_shift: str,
    active_base_token: str,
) -> None:
    """局部重跑排序、報表及懸浮工具，避免一般分析操作重新載入整個系統。"""
    st.markdown('<div id="analysis-results-anchor"></div>', unsafe_allow_html=True)
    st.markdown("---")
    st.subheader("調度結果")

    result_df = analysis_result_df

    with st.expander("排序設定", expanded=False):
        sort_control_1, sort_control_2 = st.columns([2, 1])
        with sort_control_1:
            selected_sort_field = st.selectbox(
                "排序",
                list(SORT_FIELD_OPTIONS.keys()),
                index=0,
                key=f"analysis_sort_field::{current_context_key}::{selected_region}",
            )
        with sort_control_2:
            selected_sort_direction = st.selectbox(
                "排序方向",
                ["由大到小／Z→A／倒序", "由小到大／A→Z／正序"],
                index=0,
                key=f"analysis_sort_direction::{current_context_key}::{selected_region}",
            )

    sort_descending = selected_sort_direction.startswith("由大到小")
    if selected_sort_field in ("距離目前位置最近", "預估行車時間最短"):
        sort_descending = False
        if road_metrics_available:
            st.caption("距離／時間排序固定由近到遠；道路資訊已依目前 GPS 計算。")
        else:
            st.info(road_metrics_message)

    sorted_region_frames = [
        sort_dispatch_results(region_rows, selected_sort_field, sort_descending)
        for _region, region_rows in result_df.groupby("行政區", sort=False)
    ]
    if sorted_region_frames:
        result_df = pd.concat(sorted_region_frames, ignore_index=True)

    bike_summary = calculate_inventory_summary(
        edited_df, "2.0 現況", "2.0 標準"
    )
    ebike_summary = calculate_inventory_summary(
        edited_df, "2.0E 現況", "2.0E 標準"
    )
    render_inventory_summary_card("2.0 總覽", "bike", bike_summary)
    render_inventory_summary_card("2.0E 總覽", "ebike", ebike_summary)

    missing_bike_count = safe_nonnegative_int(bike_summary["missing_count"])
    missing_ebike_count = safe_nonnegative_int(ebike_summary["missing_count"])
    if missing_bike_count or missing_ebike_count:
        render_missing_data_notice(missing_bike_count, missing_ebike_count)

    if result_df.empty:
        st.success("✨ 所有場站皆符合配置，目前不需要調度。")
    else:
        render_dispatch_legend()

    source_region_groups = {
        str(region): region_df
        for region, region_df in edited_df.groupby(
            edited_df["行政區"].astype(str),
            sort=False,
        )
    }
    result_region_groups = {
        str(region): region_df
        for region, region_df in result_df.groupby(
            result_df["行政區"].astype(str),
            sort=False,
        )
    }
    empty_result_df = result_df.iloc[0:0]
    for region, region_source_df in source_region_groups.items():
        region_result_df = result_region_groups.get(region, empty_result_df)
        zone_values = []
        if "路線區域" in region_source_df.columns:
            zone_values = [
                zone
                for zone in region_source_df["路線區域"].astype(str).drop_duplicates().tolist()
                if zone in ALL_DISPATCH_ZONES
            ]
        zone_prefix = "／".join(zone_values)
        heading = f"{zone_prefix}｜{region}" if zone_prefix else region
        st.markdown(f"#### {heading}")
        render_region_inventory_overview(region, region_source_df)
        if region_result_df.empty:
            st.success("此行政區目前全部符合配置。")
        else:
            render_analysis_result_table(region_result_df)

    if not result_df.empty:
        battery_threshold, battery_priority_threshold = get_low_battery_thresholds(
            active_base_token
        )
        analysis_battery_specs = [
            (
                str(station_name),
                "analysis",
                f"analysis-result-anchor-{int(row_index)}",
            )
            for row_index, station_name in result_df["場站名稱"].items()
            if str(station_name).strip()
        ]
        render_inline_low_battery_pillars(
            analysis_battery_specs,
            threshold=battery_threshold,
            priority_threshold=battery_priority_threshold,
            mobile_mode=mobile_mode,
            auto_query=False,
        )

    if not result_df.empty:
        export_tools_open = on_demand_toggle(
            "⬇️ 開啟報表下載",
            key=f"export_tools_open::{current_context_key}::{selected_region}",
            help_text="只有開啟時才產生 CSV 與彩色 Excel，避免平常操作反覆建立檔案。",
        )
        if export_tools_open:
            with st.container(border=True):
                export_df = make_colored_export_df(result_df)
                csv_data = export_df.to_csv(index=False).encode("utf-8-sig")
                excel_data = build_colored_excel(export_df)
                render_new_window_download_panel(
                    csv_data=csv_data,
                    csv_filename=(
                        f"{configuration_type}_D1_D2_D3_{selected_shift}_調度分析_彩色標記.csv"
                    ),
                    excel_data=excel_data,
                    excel_filename=(
                        f"{configuration_type}_D1_D2_D3_{selected_shift}_調度分析_彩色.xlsx"
                    ),
                )

    st.caption("即時車數來源：YouBike 官網；以現場狀況為準。")
    render_floating_station_search(result_df, mobile_mode, page_mode="一般分析")
    render_floating_battery_query(
        battery_route_map,
        mobile_mode,
    )
    persist_runtime_state(active_base_token)


st.markdown(
    '<div id="ubike-page-top-anchor" aria-hidden="true"></div>',
    unsafe_allow_html=True,
)
render_app_hero()
render_jarvis_trigger_bootstrap()

mobile_mode = is_mobile_browser()

url_base_token = get_base_token()
clear_stored_base_token = bool(st.session_state.pop("clear_browser_token_pending", False))
base_token = recover_base_token_from_browser(
    url_base_token,
    clear_stored=clear_stored_base_token,
)
if base_token and not url_base_token:
    # iOS 重新建立頁面但遺失查詢參數時，自動把找回的 token 補回網址。
    set_base_token(base_token)
cached_base, cache_expired = load_cached_base(base_token)

if "base_uploader_version" not in st.session_state:
    st.session_state["base_uploader_version"] = 0

if cache_expired:
    clear_base_token()
    st.session_state["base_uploader_version"] += 1
    st.session_state["base_expired_notice"] = True
    base_token = None
    cached_base = None

with st.sidebar:
    st.header("配置")
    st.caption("系統版本：測試版")

    if st.session_state.pop("base_expired_notice", False):
        st.warning("原配置無法讀取，請重新上傳。")

    uploaded_excel = st.file_uploader(
        "配置表",
        type=["xlsx"],
        key=f"base_uploader_{st.session_state['base_uploader_version']}",
    )
active_base = cached_base

if uploaded_excel is not None:
    uploaded_bytes = uploaded_excel.getvalue()
    uploaded_digest = hashlib.sha256(uploaded_bytes).hexdigest()

    if active_base is None or active_base["sha256"] != uploaded_digest:
        active_base = save_cached_base(uploaded_excel.name, uploaded_bytes)

if st.session_state.pop("full_reset_notice", False):
    st.success("已全部重置。")

if st.session_state.pop("data_zero_notice", False):
    st.success("現況已清空。")

if active_base is None:
    if base_token:
        st.warning("瀏覽器已找回先前的配置識別碼，但伺服器暫存檔已不存在；這通常發生於重新部署或主機重啟，請重新上傳一次配置表。")
    else:
        st.info("請上傳配置表。")
    st.stop()

# Streamlit 手機連線被系統回收後會建立新 session；在任何 widget 建立前先恢復關鍵狀態。
restore_runtime_state(active_base["token"])

# 配置表一載入就開始定位，之後每 30 秒在背景更新一次。
shared_location = render_shared_geolocation(active_base)
with st.sidebar:
    render_shared_location_summary(active_base, shared_location)
    if st.button(
        "立即更新定位",
        use_container_width=True,
        key=f"sidebar_refresh_location::{active_base['token']}",
    ):
        request_shared_geolocation_refresh(active_base)
        rerun_app()

try:
    options = cached_available_sources(active_base["bytes"])
except Exception as exc:
    st.error(f"Excel 讀取失敗：{exc}")
    st.stop()

if not options:
    st.error("找不到可使用的 D1／D2／D3 配置。")
    st.stop()

with st.sidebar:
    st.caption(active_base["name"])
    with st.expander("資料管理", expanded=False):
        if st.button(
            "清空現況",
            use_container_width=True,
        ):
            blank_status_cache = build_blank_status_cache(active_base["bytes"], options)
            save_cached_status(
                active_base["token"],
                active_base["expires_at"],
                blank_status_cache,
            )
            st.session_state["data_zero_notice"] = True
            rerun_app()

        if st.button(
            "全部重置",
            use_container_width=True,
            type="primary",
        ):
            reset_token = active_base.get("token")
            if reset_token:
                delete_cached_base(reset_token)
            clear_base_token()
            st.session_state["clear_browser_token_pending"] = True
            st.session_state["base_uploader_version"] += 1
            st.session_state["full_reset_notice"] = True
            rerun_app()

with st.container(border=True):
    selected_configuration_type = st.selectbox(
        "配置類型",
        list(CONFIGURATION_TYPES),
        key=f"configuration_type::{active_base['token']}",
        help="採手動切換，不會依星期自動變更，適合夜班跨日作業。",
    )
    selected_shift = st.selectbox(
        "班別",
        list(SHIFT_COLUMNS.keys()),
        key=f"shift::{active_base['token']}",
    )
    page_mode = st.radio(
        "工作模式",
        ["一般分析", "智慧調度"],
        horizontal=True,
        key=(
            f"page_mode::{active_base['token']}::{selected_configuration_type}::"
            f"{selected_shift}"
        ),
    )

configuration_options, configuration_matched = configuration_options_for_type(
    options,
    selected_configuration_type,
)
selected_sheet = preferred_configuration_sheet(
    configuration_options,
    selected_configuration_type,
)
if not configuration_matched:
    st.info(
        f"目前 Excel 工作表名稱沒有辨識到『{selected_configuration_type}』關鍵字，"
        "因此先安全沿用現有 D1／D2／D3 配置。日後只要在工作表名稱加入暑假、平日或假日，"
        "系統就會自動切換到對應工作表。"
    )

available_zone_set = {
    zone
    for _sheet_name, route in configuration_options
    if (zone := normalize_dispatch_zone(route)) is not None
}
missing_zones = [zone for zone in ALL_DISPATCH_ZONES if zone not in available_zone_set]
if missing_zones:
    st.error(
        f"{selected_configuration_type} 缺少以下區域：{'、'.join(missing_zones)}。"
        "系統必須同時讀取 D1、D2、D3。"
    )
    st.stop()

selected_route = "D1＋D2＋D3"
current_context_key = f"{selected_configuration_type}｜D1D2D3整合｜{selected_sheet}｜{selected_shift}"
status_cache = load_cached_status(active_base["token"], active_base["expires_at"])
base_df, combined_station_locations = build_long_distance_status_dataframe(
    active_base=active_base,
    options=configuration_options,
    selected_sheet=selected_sheet,
    selected_shift=selected_shift,
    selected_zones=list(ALL_DISPATCH_ZONES),
    status_cache=status_cache,
)

if base_df.empty:
    st.warning("D1、D2、D3 沒有可用場站。")
    st.stop()

aggregate_meta = status_cache.setdefault("metadata", {}).setdefault(current_context_key, {})
if combined_station_locations:
    previous_locations = aggregate_meta.get("station_locations", {})
    if not isinstance(previous_locations, dict):
        previous_locations = {}
    merged_locations = dict(previous_locations)
    merged_locations.update(combined_station_locations)
    aggregate_meta["station_locations"] = merged_locations

previous_live_meta = status_cache.get("metadata", {}).get(current_context_key, {})
render_low_battery_threshold_controls(
    active_base["token"],
    page_mode=page_mode,
)
render_context_strip(
    route=f"{selected_configuration_type}｜D1／D2／D3",
    shift=selected_shift,
    station_count=len(base_df),
    page_mode=page_mode,
    live_meta=previous_live_meta,
)

with st.expander("即時資料狀態與配對明細", expanded=False):
    if isinstance(previous_live_meta, dict) and previous_live_meta.get("fetched_at"):
        previous_source_time = str(previous_live_meta.get("latest_source_time") or "").strip()
        source_time_text = f"｜官方資料時間 {previous_source_time}" if previous_source_time else ""
        st.caption(
            f"上次同步：{previous_live_meta['fetched_at']}"
            f"{source_time_text}｜{safe_nonnegative_int(previous_live_meta.get('matched_count'))} 站"
        )

    st.caption(
        "即時數據預設每 1 分鐘自動更新一次；右側懸浮「更新」按鈕可隨時手動更新。"
        "目的地一旦同意前往會保持鎖定，不會因即時數據變動自行換站。"
    )

    browser_payload = None
    try:
        browser_sync_component = get_youbike_browser_sync_component()
        browser_payload = browser_sync_component(
            catalog_url=YOUBIKE_STATION_CATALOG_URL,
            parking_url=YOUBIKE_PARKING_INFO_URL,
            # 第一輪每次最多查 20 站，最多 4 個請求並行；後續只重查漏站。
            batch_size=20,
            request_concurrency=4,
            max_batch_rounds=8,
            max_single_rounds=3,
            wave_delay_ms=70,
            button_label="🔄 手動更新即時車數",
            auto_refresh=True,
            auto_refresh_seconds=60,
            signature_scope=active_base["token"],
            force_initial_delivery=not bool(
                isinstance(previous_live_meta, dict) and previous_live_meta.get("fetched_at")
            ),
            key=f"browser_youbike_sync::{current_context_key}",
            default=None,
        )
    except YouBikeDataError as exc:
        st.error(f"瀏覽器同步元件建立失敗：{exc}")
    except Exception as exc:
        st.error(f"瀏覽器同步元件發生未預期錯誤：{exc}")

    if isinstance(browser_payload, dict):
        browser_event_id = str(browser_payload.get("event_id") or "").strip()
        processed_event_key = f"processed_browser_youbike_event::{current_context_key}"
        already_processed = bool(
            browser_event_id
            and st.session_state.get(processed_event_key) == browser_event_id
        )

        if not already_processed:
            try:
                if browser_event_id:
                    # 先登記事件，避免 Streamlit 元件保留上次回傳值時重複寫入。
                    st.session_state[processed_event_key] = browser_event_id

                with st.spinner("正在配對臺東場站名稱並寫入 2.0／2.0E 現況……"):
                    live_payload = normalize_browser_live_payload(browser_payload)
                    st.session_state[f"latest_live_records::{active_base['token']}"] = list(live_payload["records"])
                    st.session_state[f"latest_live_event_id::{active_base['token']}"] = str(
                        live_payload.get("event_id") or browser_event_id or uuid.uuid4().hex
                    )
                    st.session_state[f"latest_live_fetched_at::{active_base['token']}"] = str(
                        live_payload.get("fetched_at") or ""
                    )
                    previous_location_map = (
                        dict(previous_live_meta.get("station_locations", {}))
                        if isinstance(previous_live_meta, dict)
                        and isinstance(previous_live_meta.get("station_locations"), dict)
                        else {}
                    )
                    live_match_index = build_youbike_match_index(live_payload["records"])
                    live_match_cache: dict[str, tuple[dict | None, float, bool]] = {}
                    previous_location_map.update(
                        build_youbike_station_location_map(
                            base_df,
                            live_payload["records"],
                            match_index=live_match_index,
                            match_cache=live_match_cache,
                        )
                    )
                    live_updated_df, live_report_df, live_summary = apply_youbike_updates_to_dataframe(
                        base_df,
                        live_payload["records"],
                        match_index=live_match_index,
                        match_cache=live_match_cache,
                    )

                    if live_summary["matched_count"] <= 0:
                        st.error("沒有任何場站通過安全配對，因此未修改現況資料。")
                    else:
                        base_df = live_updated_df
                        live_event_id = str(live_payload.get("event_id") or browser_event_id or "")
                        common_live_meta = {
                            "source": live_payload.get("source", "YouBike 官網公開接口（瀏覽器直連，免 TDX）"),
                            "fetched_at": live_payload["fetched_at"],
                            "latest_source_time": live_payload.get("latest_source_time", ""),
                            "matched_count": live_summary["matched_count"],
                            "skipped_count": live_summary["skipped_count"],
                            "unmatched_count": live_summary["unmatched_count"],
                            "station_count": live_payload.get("station_count", 0),
                            "requested_station_count": live_payload.get("requested_station_count", 0),
                            "missing_station_count": live_payload.get("missing_station_count", 0),
                            "request_count": live_payload.get("request_count", 0),
                            "batch_round_count": live_payload.get("batch_round_count", 0),
                            "single_round_count": live_payload.get("single_round_count", 0),
                            "station_locations": previous_location_map,
                            "last_live_event_id": live_event_id,
                        }
                        status_cache.setdefault("metadata", {})[current_context_key] = dict(common_live_meta)
                        if "_狀態內容鍵" in base_df.columns:
                            for source_context_key in base_df["_狀態內容鍵"].dropna().astype(str).unique():
                                status_cache.setdefault("metadata", {})[source_context_key] = dict(common_live_meta)
                        save_dispatch_dataframe_contexts(
                            base_df,
                            status_cache=status_cache,
                            active_base=active_base,
                            default_context_key=current_context_key,
                        )
                        official_time = str(live_payload.get("latest_source_time") or "").strip()
                        official_time_text = f"｜官方資料時間：{official_time}" if official_time else ""
                        returned_count = safe_nonnegative_int(live_payload.get("station_count"))
                        requested_count = safe_nonnegative_int(live_payload.get("requested_station_count"))
                        missing_count = safe_nonnegative_int(live_payload.get("missing_station_count"))
                        request_count = safe_nonnegative_int(live_payload.get("request_count"))
                        failed_request_count = safe_nonnegative_int(live_payload.get("failed_request_count"))
                        batch_round_count = safe_nonnegative_int(live_payload.get("batch_round_count"))
                        single_round_count = safe_nonnegative_int(live_payload.get("single_round_count"))
                        fetch_text = (
                            f"｜官網即時資料：{returned_count}／{requested_count} 站"
                            if requested_count else f"｜官網即時資料：{returned_count} 站"
                        )
                        elapsed_seconds = safe_nonnegative_int(live_payload.get("elapsed_ms")) / 1000
                        request_text = (
                            f"｜共 {request_count} 次請求（批次 {batch_round_count} 輪、"
                            f"單站補查 {single_round_count} 輪）｜耗時 {elapsed_seconds:.1f} 秒"
                        )
                        st.success(
                            f"✅ 高速同步完成：已寫入 {live_summary['matched_count']}／"
                            f"{live_summary['total_count']} 個 Excel 場站{fetch_text}{request_text}｜系統取得時間："
                            f"{live_payload['fetched_at']}{official_time_text}"
                        )
                        if missing_count:
                            st.warning(
                                f"官網本次仍有 {missing_count} 個站號未回傳即時資料；系統已完成多輪批次與"
                                "單站補查，未取得者會保留原本數字，不會用 0 覆蓋。"
                            )
                        elif failed_request_count:
                            st.info(
                                f"所有場站皆已取得；過程中有 {failed_request_count} 次暫時失敗，"
                                "已由自動重試或後續補查補齊。"
                            )

                    if not live_report_df.empty:
                        problem_df = live_report_df[live_report_df["結果"] != "已寫入"]
                        with st.expander(
                            f"查看配對明細（未寫入 {len(problem_df)} 筆）",
                            expanded=not problem_df.empty,
                        ):
                            report_height = min(520, max(110, 42 + len(live_report_df) * 35))
                            st.dataframe(
                                live_report_df,
                                hide_index=True,
                                use_container_width=True,
                                height=report_height,
                                row_height=35,
                            )
            except YouBikeDataError as exc:
                st.error(f"YouBike 官網同步失敗：{exc}")
            except Exception as exc:
                st.error(f"YouBike 官網同步發生未預期錯誤：{exc}")


station_alerts = build_station_alert_records(base_df)
notify_station_alert_changes(
    base_df,
    token=active_base["token"],
    context_key=current_context_key,
    alerts=station_alerts,
)
render_station_alert_summary(
    base_df,
    station_locations=combined_station_locations,
    current_location=shared_location,
    alerts=station_alerts,
)

if page_mode == "智慧調度":
    render_smart_dispatch_page_fragment(
        active_base=active_base,
        options=configuration_options,
        selected_sheet=selected_sheet,
        selected_shift=selected_shift,
        configuration_type=selected_configuration_type,
        status_cache=status_cache,
        shared_location=shared_location,
        base_df=base_df,
        station_locations=combined_station_locations,
        battery_route_map=build_battery_route_station_map(base_df),
        mobile_mode=mobile_mode,
    )
    st.stop()


# 先選 D1／D2／D3，再依該區域提供可選的行政區，避免一次看到全部 120 個場站。
zone_filter_options = ["全部"] + [
    zone for zone in ALL_DISPATCH_ZONES
    if zone in set(base_df["路線區域"].astype(str))
]
selected_zone = st.selectbox(
    "調度區域",
    zone_filter_options,
    key=f"analysis_zone::{active_base['token']}::{selected_configuration_type}::{selected_shift}",
)

zone_df = (
    base_df
    if selected_zone == "全部"
    else base_df[base_df["路線區域"].astype(str).eq(selected_zone)]
)
regions = ["全部"] + list(dict.fromkeys(zone_df["行政區"].astype(str).tolist()))
selected_region = st.selectbox(
    "行政區",
    regions,
    key=(
        f"analysis_region::{active_base['token']}::{selected_configuration_type}::"
        f"{selected_shift}::{selected_zone}"
    ),
)
working_df = (
    zone_df
    if selected_region == "全部"
    else zone_df[zone_df["行政區"].astype(str).eq(selected_region)]
)
working_df = working_df.reset_index(drop=True)
edited_df = working_df.copy()
analysis_result_df = build_analysis_result(edited_df)
analysis_result_df = analysis_result_df[
    (analysis_result_df["2.0 缺／多幾台"] != "符合")
    | (analysis_result_df["2.0E 缺／多幾台"] != "符合")
].reset_index(drop=True)
analysis_result_df, road_metrics_available, road_metrics_message = attach_analysis_road_metrics(
    analysis_result_df,
    station_locations=combined_station_locations,
    current_location=shared_location,
)
battery_route_map = build_battery_route_station_map(base_df)

render_general_analysis_results_fragment(
    edited_df=edited_df,
    analysis_result_df=analysis_result_df,
    road_metrics_available=road_metrics_available,
    road_metrics_message=road_metrics_message,
    battery_route_map=battery_route_map,
    mobile_mode=mobile_mode,
    current_context_key=current_context_key,
    selected_region=selected_region,
    configuration_type=selected_configuration_type,
    selected_shift=selected_shift,
    active_base_token=active_base["token"],
)
