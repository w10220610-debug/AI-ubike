from __future__ import annotations

import hashlib
import html
import json
import os
import re
import time
import unicodedata
import uuid
from datetime import datetime
from difflib import SequenceMatcher
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image, ImageEnhance, ImageOps
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dispatch_core import (
    SHIFT_COLUMNS,
    available_sources,
    build_result,
    calculate_totals,
    load_workbook,
    parse_route,
)


def is_mobile_browser() -> bool:
    """依瀏覽器 User-Agent 判斷是否為手機或平板。"""
    try:
        user_agent = st.context.headers.get("User-Agent", "")
    except Exception:
        # 舊版 Streamlit 沒有 st.context 時，改由側邊欄手動切換。
        return False

    mobile_tokens = ("android", "iphone", "ipad", "ipod", "mobile", "windows phone")
    normalized_user_agent = user_agent.lower()
    return any(token in normalized_user_agent for token in mobile_tokens)


def safe_nonnegative_int(value) -> int:
    """把 Excel 儲存格值安全轉成不小於 0 的整數。"""
    try:
        return max(0, int(value))
    except (TypeError, ValueError, OverflowError):
        return 0


CURRENT_STATUS_COLUMNS = ("2.0 現況", "2.0E 現況")
RECOGNITION_ERROR_TEXT = "識別錯誤"


def normalize_current_status(value) -> int | None:
    """將現況值正規化；空白與無法轉換的內容保留為 None，不再自動變成 0。"""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    text = str(value).strip()
    if not text or text == RECOGNITION_ERROR_TEXT:
        return None

    try:
        return max(0, int(float(text)))
    except (TypeError, ValueError, OverflowError):
        return None


def blank_current_status(status_df: pd.DataFrame) -> pd.DataFrame:
    """建立現況預設為空白的資料表。"""
    blank_df = status_df.copy()
    for column in CURRENT_STATUS_COLUMNS:
        blank_df[column] = pd.Series([pd.NA] * len(blank_df), dtype="Int64")
    return blank_df



def coerce_nullable_current_status(status_df: pd.DataFrame) -> pd.DataFrame:
    """讓現況欄維持可輸入空白的 nullable integer 型別；使用向量化避免逐格 Python 迴圈。"""
    normalized_df = status_df.copy()
    for column in CURRENT_STATUS_COLUMNS:
        if column not in normalized_df.columns:
            normalized_df[column] = pd.Series(pd.NA, index=normalized_df.index, dtype="Int64")
            continue
        numeric = pd.to_numeric(
            normalized_df[column].replace(RECOGNITION_ERROR_TEXT, pd.NA),
            errors="coerce",
        )
        normalized_df[column] = pd.Series(
            np.trunc(numeric).clip(lower=0),
            index=normalized_df.index,
        ).astype("Int64")
    return normalized_df


def format_dispatch_status(current_value, standard_value) -> str:
    """現況為空白時顯示識別錯誤，其餘依標準值計算缺車或多車。"""
    current = normalize_current_status(current_value)
    if current is None:
        return RECOGNITION_ERROR_TEXT

    standard = safe_nonnegative_int(standard_value)
    difference = current - standard
    if difference > 0:
        return f"多 {difference} 台"
    if difference < 0:
        return f"缺 {abs(difference)} 台"
    return "符合"



def build_result_with_recognition_errors(status_df: pd.DataFrame) -> pd.DataFrame:
    """建立分析結果；以向量化計算取代逐列 iterrows。"""
    result = status_df[["行政區", "場站名稱"]].astype(str).copy()

    def build_status_series(current_column: str, standard_column: str) -> pd.Series:
        current = pd.Series(
            np.trunc(pd.to_numeric(status_df[current_column], errors="coerce")),
            index=status_df.index,
        ).clip(lower=0)
        standard = pd.Series(
            np.trunc(pd.to_numeric(status_df[standard_column], errors="coerce")),
            index=status_df.index,
        ).fillna(0).clip(lower=0)
        difference = current - standard
        output = pd.Series(RECOGNITION_ERROR_TEXT, index=status_df.index, dtype="object")
        valid = current.notna()
        output.loc[valid & difference.eq(0)] = "符合"
        output.loc[valid & difference.gt(0)] = (
            "多 " + difference.loc[valid & difference.gt(0)].astype(int).astype(str) + " 台"
        )
        output.loc[valid & difference.lt(0)] = (
            "缺 " + difference.loc[valid & difference.lt(0)].abs().astype(int).astype(str) + " 台"
        )
        return output

    result["2.0 缺／多幾台"] = build_status_series("2.0 現況", "2.0 標準")
    result["2.0E 缺／多幾台"] = build_status_series("2.0E 現況", "2.0E 標準")
    return result



def calculate_totals_ignoring_missing(
    status_df: pd.DataFrame,
    current_column: str,
    standard_column: str,
) -> tuple[int, int]:
    """缺／多合計只計算有效現況；使用向量化加速。"""
    current = pd.Series(
        np.trunc(pd.to_numeric(status_df[current_column], errors="coerce")),
        index=status_df.index,
    ).clip(lower=0)
    standard = pd.Series(
        np.trunc(pd.to_numeric(status_df[standard_column], errors="coerce")),
        index=status_df.index,
    ).fillna(0).clip(lower=0)
    difference = (current - standard).dropna()
    short_total = int((-difference[difference.lt(0)]).sum())
    extra_total = int(difference[difference.gt(0)].sum())
    return short_total, extra_total


def add_dispatch_indicator(value) -> str:
    """在分析結果後方加上容易辨識的缺車／多車圖案。"""
    text = str(value)
    if "多" in text:
        return f"{text} 🔴"
    if "缺" in text or "少" in text:
        return f"{text} 🟠"
    return text


def extract_dispatch_count(value) -> int | None:
    """只擷取缺車／多車的台數；「符合」不納入排序條件。"""
    text = str(value)
    if not any(status in text for status in ("多", "缺", "少")):
        return None

    match = re.search(r"\d+", text)
    return int(match.group()) if match else 0


SORT_FIELD_OPTIONS = {
    "最大缺／多台數": "max",
    "缺／多總台數": "total",
    "2.0 缺／多台數": "bike",
    "2.0E 缺／多台數": "ebike",
    "場站名稱": "station",
    "Excel 原始順序": "original",
}



def sort_dispatch_results(result_df, sort_field: str, descending: bool):
    """依使用者選擇排序；以向量化字串擷取取代逐格正規表示式。"""
    if result_df.empty:
        return result_df

    sorted_df = result_df.copy()
    bike_counts = pd.to_numeric(
        sorted_df["2.0 缺／多幾台"].astype(str).str.extract(r"(\d+)", expand=False),
        errors="coerce",
    )
    ebike_counts = pd.to_numeric(
        sorted_df["2.0E 缺／多幾台"].astype(str).str.extract(r"(\d+)", expand=False),
        errors="coerce",
    )
    bike_valid = sorted_df["2.0 缺／多幾台"].astype(str).str.contains("多|缺|少", regex=True)
    ebike_valid = sorted_df["2.0E 缺／多幾台"].astype(str).str.contains("多|缺|少", regex=True)
    bike_counts = bike_counts.where(bike_valid)
    ebike_counts = ebike_counts.where(ebike_valid)

    count_frame = pd.concat([bike_counts, ebike_counts], axis=1)
    sorted_df["_排序最大台數"] = count_frame.max(axis=1, skipna=True).fillna(-1).astype(int)
    sorted_df["_排序總台數"] = count_frame.fillna(0).sum(axis=1).astype(int)
    sorted_df["_排序2.0台數"] = bike_counts.fillna(-1).astype(int)
    sorted_df["_排序2.0E台數"] = ebike_counts.fillna(-1).astype(int)
    sorted_df["_原始順序"] = np.arange(len(sorted_df), dtype=int)

    sort_key = SORT_FIELD_OPTIONS.get(sort_field, "max")
    ascending = not descending

    if sort_key == "max":
        by = ["_排序最大台數", "_排序總台數", "_排序2.0台數", "_排序2.0E台數", "_原始順序"]
        ascending_values = [ascending, ascending, ascending, ascending, True]
    elif sort_key == "total":
        by = ["_排序總台數", "_排序最大台數", "_排序2.0台數", "_排序2.0E台數", "_原始順序"]
        ascending_values = [ascending, ascending, ascending, ascending, True]
    elif sort_key == "bike":
        by = ["_排序2.0台數", "_排序2.0E台數", "_排序總台數", "_原始順序"]
        ascending_values = [ascending, ascending, ascending, True]
    elif sort_key == "ebike":
        by = ["_排序2.0E台數", "_排序2.0台數", "_排序總台數", "_原始順序"]
        ascending_values = [ascending, ascending, ascending, True]
    elif sort_key == "station":
        by = ["場站名稱", "_原始順序"]
        ascending_values = [ascending, True]
    else:
        by = ["_原始順序"]
        ascending_values = [ascending]

    sorted_df = sorted_df.sort_values(by=by, ascending=ascending_values, kind="mergesort")
    return sorted_df.drop(
        columns=[
            "_排序最大台數", "_排序總台數", "_排序2.0台數",
            "_排序2.0E台數", "_原始順序",
        ]
    ).reset_index(drop=True)


def make_colored_export_df(result_df: pd.DataFrame) -> pd.DataFrame:
    """建立含紅／橘圖案的匯出資料，CSV 可直接辨識缺車與多車。"""
    export_df = result_df[
        ["行政區", "場站名稱", "2.0 缺／多幾台", "2.0E 缺／多幾台"]
    ].copy()
    for status_column in ("2.0 缺／多幾台", "2.0E 缺／多幾台"):
        export_df[status_column] = export_df[status_column].map(add_dispatch_indicator)
    return export_df


@st.cache_data(show_spinner=False, max_entries=32)
def build_colored_excel(export_df: pd.DataFrame) -> bytes:
    """輸出真正帶有儲存格底色的 Excel 分析表。"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "調度分析"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    orange_fill = PatternFill("solid", fgColor="FCE5CD")
    green_fill = PatternFill("solid", fgColor="D9EAD3")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")

    headers = list(export_df.columns)
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for record in export_df.itertuples(index=False, name=None):
        worksheet.append(list(record))

    status_columns = {3, 4}
    for row in worksheet.iter_rows(min_row=2):
        for column_index, cell in enumerate(row, start=1):
            cell.alignment = Alignment(
                horizontal="center" if column_index in status_columns else "left",
                vertical="center",
            )
            if column_index not in status_columns:
                continue
            text = str(cell.value or "")
            if "多" in text:
                cell.fill = red_fill
            elif "缺" in text or "少" in text:
                cell.fill = orange_fill
            elif RECOGNITION_ERROR_TEXT in text:
                cell.fill = yellow_fill
            elif "符合" in text:
                cell.fill = green_fill

    widths = [14, 34, 20, 20]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()



OCR_MATCH_THRESHOLD = 0.62
OCR_MIN_TEXT_CONFIDENCE = 0.35
OCR_MAX_COUNT = 200


def normalize_ocr_text(value) -> str:
    """統一全形字、空白與常見 OCR 符號，便於後續比對。"""
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("｜", "|").replace("—", "-").replace("–", "-")
    return re.sub(r"\s+", " ", text).strip()



@lru_cache(maxsize=8192)
def _normalize_station_key_cached(text: str) -> str:
    normalized = normalize_ocr_text(text).lower()
    return re.sub(r"[^0-9a-z\u3400-\u9fff]", "", normalized)


def normalize_station_key(value) -> str:
    """建立只保留中英數字的場站比對鍵，並快取重複字串。"""
    return _normalize_station_key_cached(str(value or ""))




YOUBIKE_STATION_CATALOG_URL = "https://apis.youbike.com.tw/json/station-min-yb2.json"
YOUBIKE_PARKING_INFO_URL = "https://apis.youbike.com.tw/tw2/parkingInfo"
YOUBIKE_REQUEST_TIMEOUT_SECONDS = 25
YOUBIKE_HTTP_MAX_ATTEMPTS = 3
YOUBIKE_STATION_BATCH_SIZE = 100
YOUBIKE_MATCH_THRESHOLD = 0.82
TAIPEI_TIMEZONE = ZoneInfo("Asia/Taipei")


class YouBikeDataError(RuntimeError):
    """YouBike 官網公開資料連線或格式異常。"""


def _first_nonempty(*values):
    """回傳第一個不是 None／空字串的值。"""
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return None


def _youbike_http_json(
    url: str,
    *,
    method: str = "GET",
    json_body: dict | None = None,
):
    """讀取 YouBike 官網公開 JSON；免 TDX、免帳號、免 API 金鑰。"""
    request_headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.7",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/126.0 Safari/537.36"
        ),
        "Referer": "https://www.youbike.com.tw/region/taitung/stations/",
        "Origin": "https://www.youbike.com.tw",
    }

    encoded_body = None
    if json_body is not None:
        encoded_body = json.dumps(json_body, ensure_ascii=False).encode("utf-8")
        request_headers["Content-Type"] = "application/json;charset=UTF-8"

    last_error: Exception | None = None
    for attempt in range(1, YOUBIKE_HTTP_MAX_ATTEMPTS + 1):
        request = Request(
            url,
            data=encoded_body,
            headers=request_headers,
            method=method,
        )

        try:
            with urlopen(request, timeout=YOUBIKE_REQUEST_TIMEOUT_SECONDS) as response:
                raw_body = response.read().decode("utf-8-sig")

            try:
                payload = json.loads(raw_body)
            except json.JSONDecodeError as exc:
                raise YouBikeDataError("YouBike 官網回傳內容不是有效的 JSON。") from exc

            if isinstance(payload, dict):
                ret_code = payload.get("retCode")
                if ret_code not in (None, 1, "1", True):
                    ret_message = str(payload.get("retMsg") or "官方資料服務回傳失敗").strip()
                    raise YouBikeDataError(f"YouBike 官網資料服務錯誤：{ret_message}")
            return payload

        except HTTPError as exc:
            last_error = exc
            try:
                detail = exc.read().decode("utf-8", errors="replace")[:240]
            except Exception:
                detail = ""

            # 429 與 5xx 通常是暫時性錯誤，先短暫退避後自動重試。
            if exc.code == 429 or 500 <= exc.code <= 599:
                if attempt < YOUBIKE_HTTP_MAX_ATTEMPTS:
                    time.sleep(0.8 * attempt)
                    continue
                if exc.code == 429:
                    raise YouBikeDataError(
                        "官方資料請求過於頻繁，請等候約 1 分鐘再試。"
                    ) from exc

            if exc.code in (401, 403):
                raise YouBikeDataError(
                    "YouBike 官網暫時拒絕此主機連線；請稍後再試，照片辨識功能仍可使用。"
                ) from exc
            raise YouBikeDataError(
                f"YouBike 官網資料回傳 HTTP {exc.code}。{detail or '請稍後再試。'}"
            ) from exc
        except (URLError, TimeoutError) as exc:
            last_error = exc
            if attempt < YOUBIKE_HTTP_MAX_ATTEMPTS:
                time.sleep(0.8 * attempt)
                continue

    if isinstance(last_error, URLError):
        reason = getattr(last_error, "reason", last_error)
        raise YouBikeDataError(f"無法連線至 YouBike 官網資料服務：{reason}") from last_error
    if isinstance(last_error, TimeoutError):
        raise YouBikeDataError("連線 YouBike 官網資料服務逾時，請稍後重試。") from last_error
    raise YouBikeDataError("YouBike 官網資料服務暫時無法使用，請稍後重試。")


def _extract_youbike_station_items(payload) -> list[dict]:
    """相容清單陣列、data 包裝，以及 retVal.data 等常見官方格式。"""
    containers = [payload]
    if isinstance(payload, dict):
        containers.append(payload.get("retVal"))

    for container in containers:
        if isinstance(container, list):
            return [item for item in container if isinstance(item, dict)]
        if not isinstance(container, dict):
            continue
        for key in ("data", "result", "stations", "retVal"):
            values = container.get(key)
            if isinstance(values, list):
                return [item for item in values if isinstance(item, dict)]
            if isinstance(values, dict):
                nested_data = values.get("data")
                if isinstance(nested_data, list):
                    return [item for item in nested_data if isinstance(item, dict)]
    return []


def normalize_youbike_station_key(value) -> str:
    """正規化 Excel 與 YouBike 官網站名，處理臺／台與車種前綴差異。"""
    text = normalize_ocr_text(value).lower().replace("臺", "台")
    text = re.sub(
        r"^(?:youbike|ubike)\s*2\s*[.．]?\s*0\s*e?\s*[_\-－—:：]*\s*",
        "",
        text,
        flags=re.IGNORECASE,
    )
    text = text.replace("公共自行車租賃站", "")
    return re.sub(r"[^0-9a-z\u3400-\u9fff]", "", text)


def _youbike_station_similarity(excel_name: str, api_name: str) -> float:
    excel_key = normalize_youbike_station_key(excel_name)
    api_key = normalize_youbike_station_key(api_name)
    if not excel_key or not api_key:
        return 0.0
    if excel_key == api_key:
        return 1.0
    if min(len(excel_key), len(api_key)) >= 4 and (
        excel_key in api_key or api_key in excel_key
    ):
        return 0.96
    return SequenceMatcher(None, excel_key, api_key, autojunk=False).ratio()


def _looks_like_taitung_station(record: dict) -> bool:
    """以官方欄位及經緯度範圍篩出臺東縣候選場站。"""
    location_text = " ".join(
        str(record.get(key) or "")
        for key in (
            "county_tw", "city_tw", "scity", "district_tw", "address_tw",
            "name_tw", "sarea", "ar", "sna",
        )
    ).replace("臺", "台")
    if "台東縣" in location_text:
        return True

    try:
        latitude = float(_first_nonempty(record.get("lat"), record.get("latitude")))
        longitude = float(_first_nonempty(record.get("lng"), record.get("longitude")))
    except (TypeError, ValueError):
        return False

    # 包含臺東本島、綠島與蘭嶼。後續仍會以場站名稱做一對一安全配對。
    return 21.85 <= latitude <= 23.60 and 120.70 <= longitude <= 122.20


@st.cache_data(show_spinner=False, ttl=21600, max_entries=4)
def fetch_youbike_taitung_station_catalog() -> list[dict]:
    """取得 YouBike 全臺站點清單並留下臺東縣候選場站。"""
    payload = _youbike_http_json(YOUBIKE_STATION_CATALOG_URL)
    items = _extract_youbike_station_items(payload)
    records: list[dict] = []

    for item in items:
        if not _looks_like_taitung_station(item):
            continue
        station_no = str(
            _first_nonempty(item.get("station_no"), item.get("sno"), item.get("station_id"))
            or ""
        ).strip()
        station_name = str(
            _first_nonempty(item.get("name_tw"), item.get("sna"), item.get("station_name"))
            or ""
        ).strip()
        if not station_no or not station_name:
            continue

        raw_status = _first_nonempty(item.get("status"), item.get("act"), 1)
        records.append(
            {
                "station_uid": station_no,
                "station_id": station_no,
                "station_name": station_name,
                "station_key": normalize_youbike_station_key(station_name),
                "service_status": safe_nonnegative_int(raw_status),
                "source_update_time": str(
                    _first_nonempty(item.get("updated_at"), item.get("mday"), item.get("time"))
                    or ""
                ).strip(),
                "latitude": _first_nonempty(item.get("lat"), item.get("latitude")),
                "longitude": _first_nonempty(item.get("lng"), item.get("longitude")),
            }
        )

    if not records:
        raise YouBikeDataError("YouBike 官網沒有回傳可辨識的臺東場站清單。")
    return records


def _batched_station_numbers(station_numbers: list[str]):
    """將場站編號分批，避免單次 POST 過大而被官方服務拒絕。"""
    for start_index in range(0, len(station_numbers), YOUBIKE_STATION_BATCH_SIZE):
        yield station_numbers[start_index : start_index + YOUBIKE_STATION_BATCH_SIZE]


@st.cache_data(show_spinner=False, ttl=60, max_entries=8)
def fetch_youbike_taitung_bike_data(refresh_bucket: int) -> dict:
    """取得臺東縣 YouBike 2.0／2.0E 即時可借車數；完全不使用 TDX。"""
    del refresh_bucket  # 讓快取依傳入分鐘批次更新，同分鐘內避免重複打官方接口。
    catalog = fetch_youbike_taitung_station_catalog()
    station_numbers = [record["station_id"] for record in catalog]

    parking_items: list[dict] = []
    station_batches = list(_batched_station_numbers(station_numbers))
    for batch_index, station_batch in enumerate(station_batches):
        payload = _youbike_http_json(
            YOUBIKE_PARKING_INFO_URL,
            method="POST",
            json_body={"station_no": station_batch},
        )
        parking_items.extend(_extract_youbike_station_items(payload))
        if batch_index < len(station_batches) - 1:
            time.sleep(0.15)

    parking_by_station = {
        str(_first_nonempty(item.get("station_no"), item.get("sno")) or "").strip(): item
        for item in parking_items
        if str(_first_nonempty(item.get("station_no"), item.get("sno")) or "").strip()
    }

    records: list[dict] = []
    source_times: list[str] = []
    for station in catalog:
        parking = parking_by_station.get(station["station_id"])
        if not isinstance(parking, dict):
            continue

        detail = _first_nonempty(
            parking.get("available_spaces_detail"),
            parking.get("sbi_detail"),
        )
        if not isinstance(detail, dict):
            detail = {}

        general_bikes = normalize_current_status(detail.get("yb2"))
        electric_bikes = normalize_current_status(detail.get("eyb"))
        source_update_time = str(
            _first_nonempty(
                parking.get("updated_at"),
                parking.get("mday"),
                parking.get("time"),
                station.get("source_update_time"),
            )
            or ""
        ).strip()
        if source_update_time:
            source_times.append(source_update_time)

        raw_service_status = _first_nonempty(
            parking.get("status"),
            parking.get("act"),
            station.get("service_status"),
            1,
        )
        records.append(
            {
                **station,
                "service_status": safe_nonnegative_int(raw_service_status),
                "general_bikes": general_bikes,
                "electric_bikes": electric_bikes,
                "available_spaces": normalize_current_status(
                    _first_nonempty(parking.get("available_spaces"), parking.get("sbi"))
                ),
                "empty_spaces": normalize_current_status(
                    _first_nonempty(parking.get("empty_spaces"), parking.get("bemp"))
                ),
                "parking_spaces": normalize_current_status(
                    _first_nonempty(parking.get("parking_spaces"), parking.get("tot"))
                ),
                "source_update_time": source_update_time,
            }
        )

    if not records:
        raise YouBikeDataError(
            "YouBike 官網沒有回傳臺東場站即時車數，可能是官方資料服務暫時異常。"
        )

    fetched_at = datetime.now(TAIPEI_TIMEZONE).strftime("%Y/%m/%d %H:%M:%S")
    return {
        "records": records,
        "fetched_at": fetched_at,
        "latest_source_time": max(source_times) if source_times else "",
        "station_count": len(records),
        "request_batch_count": len(station_batches),
        "source": "YouBike 官網公開接口（免 TDX）",
    }


def match_youbike_station(
    excel_name: str,
    live_records: list[dict],
) -> tuple[dict | None, float, bool]:
    """配對 Excel 與 YouBike 官網站名；過於接近時不自動寫入。"""
    excel_key = normalize_youbike_station_key(excel_name)
    if not excel_key:
        return None, 0.0, False

    exact_matches = [record for record in live_records if record.get("station_key") == excel_key]
    if len(exact_matches) == 1:
        return exact_matches[0], 1.0, False
    if len(exact_matches) > 1:
        return None, 1.0, True

    ranked = sorted(
        (
            (_youbike_station_similarity(excel_name, record.get("station_name", "")), record)
            for record in live_records
        ),
        key=lambda item: item[0],
        reverse=True,
    )
    if not ranked:
        return None, 0.0, False

    best_score, best_record = ranked[0]
    if best_score < YOUBIKE_MATCH_THRESHOLD:
        return None, best_score, False

    second_score = ranked[1][0] if len(ranked) > 1 else 0.0
    ambiguous = best_score < 0.96 and second_score >= best_score - 0.035
    if ambiguous:
        return None, best_score, True
    return best_record, best_score, False


def apply_youbike_updates_to_dataframe(
    base_df: pd.DataFrame,
    live_records: list[dict],
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """將 YouBike 官網即時車數寫入目前配置，只寫入安全且資料完整的配對。"""
    updated_df = base_df.copy()
    report_rows: list[dict] = []
    used_station_ids: set[str] = set()
    matched_count = 0
    skipped_count = 0
    unmatched_count = 0

    for row_index, row in updated_df.iterrows():
        excel_name = str(row.get("場站名稱", "") or "").strip()
        matched_record, score, ambiguous = match_youbike_station(excel_name, live_records)

        if matched_record is None:
            unmatched_count += 1
            report_rows.append(
                {
                    "Excel 場站": excel_name,
                    "YouBike 場站": "",
                    "2.0": pd.NA,
                    "2.0E": pd.NA,
                    "結果": "名稱可能重複，未寫入" if ambiguous else "找不到安全配對",
                    "相似度": round(score, 3),
                }
            )
            continue

        station_id = str(matched_record.get("station_id", "") or "")
        if station_id in used_station_ids:
            unmatched_count += 1
            report_rows.append(
                {
                    "Excel 場站": excel_name,
                    "YouBike 場站": matched_record.get("station_name", ""),
                    "2.0": pd.NA,
                    "2.0E": pd.NA,
                    "結果": "同一官網場站已配對，未重複寫入",
                    "相似度": round(score, 3),
                }
            )
            continue

        bike_count = normalize_current_status(matched_record.get("general_bikes"))
        ebike_count = normalize_current_status(matched_record.get("electric_bikes"))
        service_status = safe_nonnegative_int(matched_record.get("service_status"))
        if service_status != 1 or bike_count is None or ebike_count is None:
            skipped_count += 1
            reason = "場站目前非正常服務" if service_status != 1 else "官網車種明細不完整"
            report_rows.append(
                {
                    "Excel 場站": excel_name,
                    "YouBike 場站": matched_record.get("station_name", ""),
                    "2.0": pd.NA if bike_count is None else bike_count,
                    "2.0E": pd.NA if ebike_count is None else ebike_count,
                    "結果": f"{reason}，未寫入",
                    "相似度": round(score, 3),
                }
            )
            continue

        used_station_ids.add(station_id)
        updated_df.at[row_index, "2.0 現況"] = bike_count
        updated_df.at[row_index, "2.0E 現況"] = ebike_count
        matched_count += 1
        report_rows.append(
            {
                "Excel 場站": excel_name,
                "YouBike 場站": matched_record.get("station_name", ""),
                "2.0": bike_count,
                "2.0E": ebike_count,
                "結果": "已寫入",
                "相似度": round(score, 3),
            }
        )

    report_df = pd.DataFrame(report_rows)
    summary = {
        "matched_count": matched_count,
        "skipped_count": skipped_count,
        "unmatched_count": unmatched_count,
        "total_count": len(updated_df),
    }
    return coerce_nullable_current_status(updated_df), report_df, summary


def partial_similarity(needle: str, haystack: str) -> float:
    """計算局部相似度；正規化結果會快取，並對高相似結果提前結束。"""
    needle_key = normalize_station_key(needle)
    haystack_key = normalize_station_key(haystack)
    if not needle_key or not haystack_key:
        return 0.0
    if needle_key in haystack_key:
        return 1.0

    direct = SequenceMatcher(None, needle_key, haystack_key, autojunk=False).ratio()
    if direct >= 0.86 or len(haystack_key) <= len(needle_key):
        return direct

    best = direct
    minimum = max(1, len(needle_key) - 2)
    maximum = min(len(haystack_key), len(needle_key) + 2)
    for window_size in range(minimum, maximum + 1):
        for start in range(0, len(haystack_key) - window_size + 1):
            window = haystack_key[start : start + window_size]
            score = SequenceMatcher(None, needle_key, window, autojunk=False).ratio()
            if score > best:
                best = score
                if best >= 0.96:
                    return best
    return best


def parse_count_token(value) -> int | None:
    """只接受獨立的整數儲存格，排除 D1、站名門牌與 2.0 欄位標題。"""
    text = normalize_ocr_text(value)
    if not text or any(symbol in text for symbol in (":", "/", "%")):
        return None

    # 2.0、2.0E 是欄位名稱，不是車數。
    if re.search(r"2\s*[.．,，]\s*0\s*[Ee]?", text):
        return None

    compact = re.sub(r"\s+", "", text)
    compact = compact.translate(str.maketrans({"O": "0", "o": "0", "〇": "0", "○": "0"}))
    compact = compact.strip("|[](){}<>，,。.;；")

    # 必須是單獨數字；這可避免把 D1 或「238巷」誤認成車數。
    match = re.fullmatch(r"([0-9]{1,3})(?:台)?", compact)
    if not match:
        return None

    number = int(match.group(1))
    return number if 0 <= number <= OCR_MAX_COUNT else None


def extract_numbers_from_row(row_items: list[dict]) -> list[tuple[int, float]]:
    """依畫面由左到右，取出一列中的獨立整數儲存格與水平位置。"""
    numbers: list[tuple[int, float]] = []
    for item in sorted(row_items, key=lambda current: current["x"]):
        number = parse_count_token(item["text"])
        if number is not None:
            numbers.append((number, float(item["x"])))
    return numbers


def detect_numeric_column_centers(rows: list[list[dict]]) -> list[float]:
    """從整張表格的重複數字位置推定「單位、在站2.0、在站2.0E」欄位。"""
    all_items = [item for row_items in rows for item in row_items]
    if not all_items:
        return []

    right_edge = max(float(item["x"]) + float(item.get("width", 1.0)) for item in all_items)
    unit_headers = [
        item
        for item in all_items
        if "單位" in normalize_ocr_text(item.get("text", ""))
    ]

    if unit_headers:
        # 以「單位」標題為起點，排除左側責任區 D1 與站名中的門牌號碼。
        unit_header_x = float(np.median([float(item["x"]) for item in unit_headers]))
        minimum_numeric_x = unit_header_x - max(20.0, right_edge * 0.025)
    else:
        # 固定版面中，單位欄位位於畫面約三分之一之後。
        minimum_numeric_x = right_edge * 0.30

    x_positions: list[float] = []
    widths: list[float] = []
    for item in all_items:
        if float(item["x"]) < minimum_numeric_x:
            continue
        if parse_count_token(item.get("text")) is None:
            continue
        x_positions.append(float(item["x"]))
        widths.append(max(1.0, float(item.get("width", 1.0))))

    if not x_positions:
        return []

    median_width = float(np.median(widths)) if widths else 12.0
    tolerance = max(12.0, median_width * 1.55)
    clusters: list[dict] = []

    for x_value in sorted(x_positions):
        nearest = None
        nearest_distance = None
        for cluster in clusters:
            distance = abs(x_value - float(cluster["center"]))
            if distance <= tolerance and (nearest_distance is None or distance < nearest_distance):
                nearest = cluster
                nearest_distance = distance
        if nearest is None:
            clusters.append({"values": [x_value], "center": x_value})
        else:
            nearest["values"].append(x_value)
            nearest["center"] = float(np.median(nearest["values"]))

    # 表格欄位會在多列重複；單次出現的時間、總數、座標等雜訊不採用。
    repeated = [cluster for cluster in clusters if len(cluster["values"]) >= 2]
    repeated.sort(key=lambda cluster: float(cluster["center"]))
    return [float(cluster["center"]) for cluster in repeated]


def choose_station_counts(
    numbers: list[tuple[int, float]],
    numeric_column_centers: list[float],
) -> tuple[int | None, int | None, int | None]:
    """固定讀取「單位」右側的在站 2.0、2.0E，而不是最右側綁車欄位。"""
    if len(numbers) < 3:
        return None, None, None

    # 表格由左至右前三個重複數字欄，依序是：單位、在站2.0、在站2.0E。
    if len(numeric_column_centers) >= 3:
        targets = numeric_column_centers[:3]
        gaps = [targets[index + 1] - targets[index] for index in range(2)]
        max_distance = max(14.0, min(gaps) * 0.46)
        selected: list[int] = []
        used_indexes: set[int] = set()

        for target in targets:
            candidates = [
                (abs(x_value - target), index, number)
                for index, (number, x_value) in enumerate(numbers)
                if index not in used_indexes
            ]
            if not candidates:
                return None, None, None
            distance, selected_index, selected_number = min(candidates, key=lambda item: item[0])
            if distance > max_distance:
                return None, None, None
            used_indexes.add(selected_index)
            selected.append(selected_number)

        unit_count, bike_count, ebike_count = selected
    else:
        # 後備規則：嚴格數字框由左至右前三個欄位。
        unit_count, bike_count, ebike_count = [number for number, _ in numbers[:3]]

    # 單位是場站總柱數；在站兩車種合計不應超過單位，異常列直接忽略。
    if unit_count <= 0 or bike_count < 0 or ebike_count < 0:
        return None, None, None
    if bike_count + ebike_count > unit_count:
        return None, None, None

    return unit_count, bike_count, ebike_count

def group_ocr_items_into_rows(items: list[dict]) -> list[list[dict]]:
    """依文字框垂直位置合併成表格列。"""
    if not items:
        return []

    heights = [max(1.0, float(item["height"])) for item in items]
    median_height = float(np.median(heights)) if heights else 20.0
    tolerance = max(10.0, median_height * 0.72)
    rows: list[dict] = []

    for item in sorted(items, key=lambda current: (current["y"], current["x"])):
        target = None
        target_distance = None
        for row in rows:
            distance = abs(float(item["y"]) - float(row["center_y"]))
            if distance <= tolerance and (target_distance is None or distance < target_distance):
                target = row
                target_distance = distance
        if target is None:
            rows.append({"center_y": float(item["y"]), "items": [item]})
        else:
            target["items"].append(item)
            target["center_y"] = sum(float(current["y"]) for current in target["items"]) / len(target["items"])

    return [sorted(row["items"], key=lambda current: current["x"]) for row in sorted(rows, key=lambda current: current["center_y"])]


@st.cache_resource(show_spinner=False)
def get_local_ocr_engine():
    """載入免費本機 RapidOCR；模型只在第一次載入。"""
    from rapidocr import RapidOCR

    return RapidOCR()


def decode_rapidocr_output(result) -> list[dict]:
    """同時相容 RapidOCR 新版物件輸出與較舊版列表輸出。"""
    decoded: list[dict] = []

    boxes = getattr(result, "boxes", None)
    texts = getattr(result, "txts", None)
    scores = getattr(result, "scores", None)
    if boxes is not None and texts is not None:
        scores = scores if scores is not None else [1.0] * len(texts)
        for box, text, score in zip(boxes, texts, scores):
            points = np.asarray(box, dtype=float)
            if points.size < 8:
                continue
            x_min, x_max = float(points[:, 0].min()), float(points[:, 0].max())
            y_min, y_max = float(points[:, 1].min()), float(points[:, 1].max())
            decoded.append(
                {
                    "text": normalize_ocr_text(text),
                    "score": float(score or 0.0),
                    "x": x_min,
                    "y": (y_min + y_max) / 2,
                    "width": max(1.0, x_max - x_min),
                    "height": max(1.0, y_max - y_min),
                }
            )
        return decoded

    legacy = result[0] if isinstance(result, tuple) and result else result
    if isinstance(legacy, list):
        for entry in legacy:
            if not isinstance(entry, (list, tuple)) or len(entry) < 3:
                continue
            box, text, score = entry[0], entry[1], entry[2]
            points = np.asarray(box, dtype=float)
            if points.size < 8:
                continue
            x_min, x_max = float(points[:, 0].min()), float(points[:, 0].max())
            y_min, y_max = float(points[:, 1].min()), float(points[:, 1].max())
            decoded.append(
                {
                    "text": normalize_ocr_text(text),
                    "score": float(score or 0.0),
                    "x": x_min,
                    "y": (y_min + y_max) / 2,
                    "width": max(1.0, x_max - x_min),
                    "height": max(1.0, y_max - y_min),
                }
            )
    return decoded



def prepare_image_for_ocr(file_bytes: bytes) -> np.ndarray:
    """校正方向與清晰度；先限制手機超高解析度照片，降低 OCR 計算量。"""
    image = Image.open(BytesIO(file_bytes))
    image = ImageOps.exif_transpose(image).convert("RGB")

    # 手機原圖常超過 4000px；縮到 2400px 長邊通常仍足以辨識表格文字，速度明顯較快。
    max_long_edge = 2400
    long_edge = max(image.width, image.height)
    if long_edge > max_long_edge:
        scale = max_long_edge / long_edge
        image = image.resize(
            (max(1, int(image.width * scale)), max(1, int(image.height * scale))),
            Image.Resampling.LANCZOS,
        )
    elif image.width < 1500:
        scale = min(1.75, 1500 / max(1, image.width))
        target_size = (int(image.width * scale), int(image.height * scale))
        if max(target_size) > max_long_edge:
            scale = max_long_edge / long_edge
            target_size = (int(image.width * scale), int(image.height * scale))
        image = image.resize(target_size, Image.Resampling.LANCZOS)

    image = ImageOps.autocontrast(image, cutoff=1)
    image = ImageEnhance.Sharpness(image).enhance(1.25)
    return np.asarray(image)



@st.cache_data(show_spinner=False, max_entries=48)
def _cached_ocr_geometry(file_bytes: bytes) -> tuple[list[list[dict]], list[dict]]:
    """依照片內容快取 OCR 結果；相同照片再次辨識時不重跑模型。"""
    engine = get_local_ocr_engine()
    image_array = prepare_image_for_ocr(file_bytes)
    result = engine(image_array)
    items = [
        item for item in decode_rapidocr_output(result)
        if item["text"] and item["score"] >= OCR_MIN_TEXT_CONFIDENCE
    ]
    return group_ocr_items_into_rows(items), items


def run_ocr_on_photo(file_name: str, file_bytes: bytes) -> tuple[list[list[dict]], list[dict]]:
    """辨識單張照片；幾何結果依照片內容快取，檔名在回傳時再附加。"""
    cached_rows, cached_items = _cached_ocr_geometry(file_bytes)
    items = [{**item, "file_name": file_name} for item in cached_items]
    item_lookup = {
        (item["text"], item["x"], item["y"], item["width"], item["height"]): item
        for item in items
    }
    rows = []
    for cached_row in cached_rows:
        rows.append([
            item_lookup.get(
                (item["text"], item["x"], item["y"], item["width"], item["height"]),
                {**item, "file_name": file_name},
            )
            for item in cached_row
        ])
    return rows, items



def pick_station_for_row(
    row_text: str,
    station_index: list[tuple[str, str]],
) -> tuple[str | None, float]:
    """從預先正規化的場站索引中找最接近的場站，避免每列重算所有站名。"""
    row_key = normalize_station_key(row_text)
    best_station = None
    best_score = 0.0
    for station_name, station_key in station_index:
        if not station_key:
            continue
        if station_key in row_key:
            return station_name, 1.0
        score = partial_similarity(station_key, row_key)
        if score > best_score:
            best_station = station_name
            best_score = score
            if best_score >= 0.96:
                break
    if best_score < OCR_MATCH_THRESHOLD:
        return None, best_score
    return best_station, best_score


def uploaded_photo_name_and_bytes(uploaded_photo) -> tuple[str, bytes]:
    """同時支援 Streamlit UploadedFile 與暫存在 session_state 的照片紀錄。"""
    if isinstance(uploaded_photo, dict):
        file_name = str(uploaded_photo.get("name") or "未命名照片")
        file_bytes = uploaded_photo.get("bytes", b"")
    else:
        file_name = str(getattr(uploaded_photo, "name", "未命名照片"))
        file_bytes = uploaded_photo.getvalue()

    if not isinstance(file_bytes, bytes):
        file_bytes = bytes(file_bytes)
    return file_name, file_bytes


def analyze_station_photos(uploaded_photos, station_df: pd.DataFrame) -> dict:
    """辨識照片中的在站數；成功資料直接更新，失敗資料明確標示識別錯誤。"""
    station_names = [
        str(value).strip()
        for value in station_df["場站名稱"].tolist()
        if str(value).strip()
    ]
    station_index = [(name, normalize_station_key(name)) for name in station_names]
    best_by_station: dict[str, dict] = {}
    errors_by_key: dict[tuple[str, str], dict] = {}
    scanned_row_count = 0

    for uploaded_photo in uploaded_photos:
        file_name, file_bytes = uploaded_photo_name_and_bytes(uploaded_photo)
        rows, _ = run_ocr_on_photo(file_name, file_bytes)
        numeric_column_centers = detect_numeric_column_centers(rows)
        photo_has_station_match = False
        photo_has_success = False

        for row_items in rows:
            row_text = " ".join(item["text"] for item in row_items).strip()
            if not row_text:
                continue
            scanned_row_count += 1

            station_name, match_score = pick_station_for_row(row_text, station_index)
            if not station_name:
                continue

            photo_has_station_match = True
            numbers = extract_numbers_from_row(row_items)
            unit_count, bike_count, ebike_count = choose_station_counts(
                numbers, numeric_column_centers
            )
            error_key = (file_name, station_name)

            if bike_count is None or ebike_count is None:
                errors_by_key[error_key] = {
                    "場站名稱": station_name,
                    "2.0 現況": RECOGNITION_ERROR_TEXT,
                    "2.0E 現況": RECOGNITION_ERROR_TEXT,
                    "來源": file_name,
                }
                continue

            photo_has_success = True
            errors_by_key.pop(error_key, None)
            row_score = float(np.mean([item["score"] for item in row_items]))
            evidence_score = (row_score * 0.45) + (match_score * 0.55)
            record = {
                "場站名稱": station_name,
                "單位": unit_count,
                "2.0 現況": bike_count,
                "2.0E 現況": ebike_count,
                "來源": file_name,
                "信心": round(evidence_score, 3),
            }

            # 重疊照片出現相同場站時，採用綜合信心較高的一列。
            previous = best_by_station.get(station_name)
            if previous is None or evidence_score > float(previous["信心"]):
                best_by_station[station_name] = record

        if not photo_has_station_match and not photo_has_success:
            errors_by_key[(file_name, "未識別場站")] = {
                "場站名稱": "未識別場站",
                "2.0 現況": RECOGNITION_ERROR_TEXT,
                "2.0E 現況": RECOGNITION_ERROR_TEXT,
                "來源": file_name,
            }

    recognized_rows = sorted(
        best_by_station.values(),
        key=lambda record: station_names.index(record["場站名稱"]),
    )
    updates = {
        record["場站名稱"]: (int(record["2.0 現況"]), int(record["2.0E 現況"]))
        for record in recognized_rows
    }

    return {
        "updates": updates,
        "recognized_df": pd.DataFrame(recognized_rows),
        "error_df": pd.DataFrame(list(errors_by_key.values())),
        "photo_count": len(uploaded_photos),
        "scanned_row_count": scanned_row_count,
    }



def apply_ocr_updates_to_dataframe(base_df: pd.DataFrame, updates: dict[str, tuple[int, int]]) -> pd.DataFrame:
    """只更新 OCR 成功辨識的場站；使用 map 避免逐列寫入。"""
    updated_df = base_df.copy()
    if not updates:
        return updated_df

    bike_updates = {name: values[0] for name, values in updates.items()}
    ebike_updates = {name: values[1] for name, values in updates.items()}
    station_series = updated_df["場站名稱"].astype(str)
    bike_mapped = station_series.map(bike_updates)
    ebike_mapped = station_series.map(ebike_updates)
    bike_mask = bike_mapped.notna()
    ebike_mask = ebike_mapped.notna()
    updated_df.loc[bike_mask, "2.0 現況"] = bike_mapped.loc[bike_mask].astype(int)
    updated_df.loc[ebike_mask, "2.0E 現況"] = ebike_mapped.loc[ebike_mask].astype(int)
    return coerce_nullable_current_status(updated_df)


BASE_CACHE_DIR = Path(__file__).resolve().parent / ".base_cache"


def get_base_token() -> str | None:
    """從網址參數取得本次瀏覽器使用的基底識別碼。"""
    try:
        token = st.query_params.get("base")
    except Exception:
        params = st.experimental_get_query_params()
        values = params.get("base", [])
        token = values[0] if values else None

    if isinstance(token, list):
        token = token[0] if token else None

    token = str(token or "").strip()
    if len(token) != 32 or any(char not in "0123456789abcdef" for char in token.lower()):
        return None
    return token.lower()


def set_base_token(token: str) -> None:
    """把基底識別碼寫入網址，讓重新整理後仍能找到同一份基底。"""
    try:
        st.query_params["base"] = token
    except Exception:
        params = st.experimental_get_query_params()
        params["base"] = token
        st.experimental_set_query_params(**params)


def clear_base_token() -> None:
    """移除已失效的基底識別碼。"""
    try:
        if "base" in st.query_params:
            del st.query_params["base"]
    except Exception:
        params = st.experimental_get_query_params()
        params.pop("base", None)
        st.experimental_set_query_params(**params)


def base_cache_paths(token: str) -> tuple[Path, Path]:
    return BASE_CACHE_DIR / f"{token}.xlsx", BASE_CACHE_DIR / f"{token}.json"


def current_status_cache_path(token: str) -> Path:
    """取得指定基底所對應的現況暫存檔路徑。"""
    return BASE_CACHE_DIR / f"{token}.status.json"


def delete_cached_base(token: str) -> None:
    """刪除指定的暫存基底，以及與它綁定的現況資料。"""
    excel_path, metadata_path = base_cache_paths(token)
    status_path = current_status_cache_path(token)
    for path in (excel_path, metadata_path, status_path):
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass


def load_cached_base(token: str | None) -> tuple[dict | None, bool]:
    """讀取已保存的配置基底；不設定自動失效時間。"""
    if not token:
        return None, False

    excel_path, metadata_path = base_cache_paths(token)
    if not excel_path.exists() or not metadata_path.exists():
        return None, False

    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        file_bytes = excel_path.read_bytes()
        return {
            "token": token,
            "name": str(metadata.get("name") or "配置基底.xlsx"),
            "bytes": file_bytes,
            "sha256": str(
                metadata.get("sha256") or hashlib.sha256(file_bytes).hexdigest()
            ),
            "uploaded_at": float(metadata.get("uploaded_at") or 0.0),
            "expires_at": None,
        }, False
    except (OSError, ValueError, TypeError, json.JSONDecodeError):
        return None, False


def load_cached_status(token: str, expires_at: float | None = None) -> dict:
    """讀取與基底綁定的現況資料與同步資訊；不設定自動失效時間。"""
    status_path = current_status_cache_path(token)
    if not status_path.exists():
        return {"contexts": {}, "metadata": {}}

    try:
        payload = json.loads(status_path.read_text(encoding="utf-8"))
        contexts = payload.get("contexts", {})
        metadata = payload.get("metadata", {})
        if not isinstance(contexts, dict):
            contexts = {}
        if not isinstance(metadata, dict):
            metadata = {}
        return {"contexts": contexts, "metadata": metadata}
    except (OSError, TypeError, ValueError, json.JSONDecodeError):
        return {"contexts": {}, "metadata": {}}


def save_cached_status(
    token: str,
    expires_at: float | None,
    payload: dict,
) -> None:
    """保存現況資料；不設定自動失效時間。"""
    BASE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    status_path = current_status_cache_path(token)
    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))

    try:
        if status_path.exists() and status_path.read_text(encoding="utf-8") == encoded:
            return
    except OSError:
        pass

    temporary_path = status_path.with_suffix(".status.tmp")
    try:
        temporary_path.write_text(encoded, encoding="utf-8")
        temporary_path.replace(status_path)
    except OSError:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass


def status_context_key(sheet_name: str, route: str, shift: str) -> str:
    """為每個工作表、分區與班別建立獨立的現況保存區。"""
    return "｜".join((str(sheet_name), str(route), str(shift)))


def restore_current_status(base_df: pd.DataFrame, saved_records) -> pd.DataFrame:
    """把先前保存的現況套回剛解析完成的基底資料。"""
    restored_df = base_df.copy()
    if not isinstance(saved_records, list):
        return restored_df

    saved_values = {}
    for record in saved_records:
        if not isinstance(record, dict):
            continue
        key = (str(record.get("行政區", "")), str(record.get("場站名稱", "")))
        saved_values[key] = (
            normalize_current_status(record.get("2.0 現況")),
            normalize_current_status(record.get("2.0E 現況")),
        )

    for row_index, row in restored_df.iterrows():
        key = (str(row.get("行政區", "")), str(row.get("場站名稱", "")))
        if key not in saved_values:
            continue
        bike_now, ebike_now = saved_values[key]
        restored_df.at[row_index, "2.0 現況"] = pd.NA if bike_now is None else bike_now
        restored_df.at[row_index, "2.0E 現況"] = pd.NA if ebike_now is None else ebike_now

    return coerce_nullable_current_status(restored_df)


def merge_current_status(full_df: pd.DataFrame, edited_df: pd.DataFrame) -> pd.DataFrame:
    """把目前畫面修改的現況合併回完整分區資料，避免切換行政區時遺失。"""
    merged_df = full_df.copy()
    edited_values = {
        (str(row.get("行政區", "")), str(row.get("場站名稱", ""))): (
            normalize_current_status(row.get("2.0 現況")),
            normalize_current_status(row.get("2.0E 現況")),
        )
        for _, row in edited_df.iterrows()
    }

    for row_index, row in merged_df.iterrows():
        key = (str(row.get("行政區", "")), str(row.get("場站名稱", "")))
        if key not in edited_values:
            continue
        bike_now, ebike_now = edited_values[key]
        merged_df.at[row_index, "2.0 現況"] = pd.NA if bike_now is None else bike_now
        merged_df.at[row_index, "2.0E 現況"] = pd.NA if ebike_now is None else ebike_now

    return coerce_nullable_current_status(merged_df)



def dataframe_to_status_records(status_df: pd.DataFrame) -> list[dict]:
    """將完整現況轉成 JSON 紀錄；以 DataFrame 向量化處理。"""
    columns = ["行政區", "場站名稱", "2.0 現況", "2.0E 現況"]
    records_df = coerce_nullable_current_status(status_df[columns].copy())
    records_df["行政區"] = records_df["行政區"].astype(str)
    records_df["場站名稱"] = records_df["場站名稱"].astype(str)
    object_df = records_df.astype(object).where(records_df.notna(), None)
    return object_df.to_dict(orient="records")


def save_cached_base(file_name: str, file_bytes: bytes) -> dict:
    """將新上傳的 Excel 保存於伺服器，不設定自動失效時間。"""
    BASE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    token = uuid.uuid4().hex
    uploaded_at = time.time()
    digest = hashlib.sha256(file_bytes).hexdigest()
    excel_path, metadata_path = base_cache_paths(token)

    excel_path.write_bytes(file_bytes)
    metadata_path.write_text(
        json.dumps(
            {
                "name": file_name,
                "uploaded_at": uploaded_at,
                "sha256": digest,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    set_base_token(token)

    return {
        "token": token,
        "name": file_name,
        "bytes": file_bytes,
        "sha256": digest,
        "uploaded_at": uploaded_at,
        "expires_at": None,
    }


def format_remaining_time(expires_at: float | None) -> str:
    """相容舊介面；目前保存期限為無時間限制。"""
    return "無時間限制"


def rerun_app() -> None:
    """相容新舊版 Streamlit 的重新執行方法。"""
    try:
        st.rerun()
    except AttributeError:
        st.experimental_rerun()


def clear_editor_session_state(base_token: str | None = None) -> None:
    """清除資料編輯器與手機數字欄位的暫存，避免舊值蓋回歸零結果。"""
    token_prefix = f"editor::{base_token}::" if base_token else "editor::"
    keys_to_remove = [
        key
        for key in list(st.session_state.keys())
        if isinstance(key, str) and key.startswith(token_prefix)
    ]
    for key in keys_to_remove:
        del st.session_state[key]


def build_blank_status_cache(workbook_data: dict, options: list[tuple[str, str]]) -> dict:
    """把所有工作表、分區與班別的現況清成空白。"""
    blank_contexts: dict[str, list[dict]] = {}

    for sheet_name, route in options:
        for shift in SHIFT_COLUMNS.keys():
            try:
                route_df = parse_route(workbook_data[sheet_name], route, shift)
            except Exception:
                continue

            if route_df.empty:
                continue

            context_key = status_context_key(sheet_name, route, shift)
            blank_contexts[context_key] = dataframe_to_status_records(
                blank_current_status(route_df)
            )

    return {"contexts": blank_contexts, "metadata": {}}


def dispatch_status_class(value) -> str:
    """回傳分析結果儲存格的視覺狀態類別。"""
    text = str(value)
    if "多" in text:
        return "analysis-status-extra"
    if "缺" in text or "少" in text:
        return "analysis-status-short"
    if RECOGNITION_ERROR_TEXT in text:
        return "analysis-status-error"
    return "analysis-status-ok"


def render_analysis_result_table(region_df: pd.DataFrame) -> None:
    """以響應式 HTML 表格完整展開分析結果，避免表格內水平或垂直捲動。"""
    rows = []
    for row_index, row in region_df.iterrows():
        station_name = html.escape(str(row.get("場站名稱", "")))
        bike_status = add_dispatch_indicator(row.get("2.0 缺／多幾台", ""))
        ebike_status = add_dispatch_indicator(row.get("2.0E 缺／多幾台", ""))
        bike_class = dispatch_status_class(bike_status)
        ebike_class = dispatch_status_class(ebike_status)
        rows.append(
            f'<tr id="analysis-result-anchor-{int(row_index)}" class="analysis-result-row">'
            f'<td class="analysis-station-cell">{station_name}</td>'
            f'<td class="analysis-status-cell {bike_class}">{html.escape(bike_status)}</td>'
            f'<td class="analysis-status-cell {ebike_class}">{html.escape(ebike_status)}</td>'
            "</tr>"
        )

    table_html = (
        '<div class="analysis-result-table-wrap">'
        '<table class="analysis-result-table">'
        '<colgroup>'
        '<col class="analysis-col-station" />'
        '<col class="analysis-col-status" />'
        '<col class="analysis-col-status" />'
        '</colgroup>'
        '<thead><tr>'
        '<th>場站名稱</th>'
        '<th>2.0 缺／多</th>'
        '<th>2.0E 缺／多</th>'
        '</tr></thead>'
        f'<tbody>{"".join(rows)}</tbody>'
        '</table></div>'
    )
    st.markdown(table_html, unsafe_allow_html=True)


def render_floating_station_search(result_df: pd.DataFrame, mobile_mode: bool) -> None:
    """建立搜尋分析結果、回頁首與跳至分析區的垂直懸浮按鍵。"""
    stations = []
    for row_index, row in result_df.reset_index(drop=True).iterrows():
        station_name = str(row.get("場站名稱", "")).strip()
        if not station_name:
            continue
        stations.append(
            {
                "name": station_name,
                "region": str(row.get("行政區", "")).strip(),
                "bike": str(row.get("2.0 缺／多幾台", "")).strip(),
                "ebike": str(row.get("2.0E 缺／多幾台", "")).strip(),
                "anchor": f"analysis-result-anchor-{row_index}",
            }
        )

    station_payload = json.dumps(stations, ensure_ascii=False).replace("</", "<\\/")
    display_mode = "mobile" if mobile_mode else "desktop"

    components.html(
        f"""
        <script>
        (() => {{
            const stations = {station_payload};
            const displayMode = {json.dumps(display_mode)};
            const doc = window.parent.document;
            const win = window.parent;

            const oldRoot = doc.getElementById("ubike-float-tools");
            if (oldRoot) oldRoot.remove();
            const oldStyle = doc.getElementById("ubike-float-tools-style");
            if (oldStyle) oldStyle.remove();

            const style = doc.createElement("style");
            style.id = "ubike-float-tools-style";
            style.textContent = `
                #ubike-float-tools {{
                    position: fixed;
                    right: 18px;
                    bottom: 22px;
                    z-index: 2147483000;
                    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI",
                                 "Microsoft JhengHei", sans-serif;
                }}
                #ubike-float-tools * {{ box-sizing: border-box; }}
                #ubike-float-tools .uft-actions {{
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    gap: 8px;
                }}
                #ubike-float-tools .uft-button {{
                    width: 56px;
                    height: 56px;
                    flex: 0 0 56px;
                    border: 0;
                    border-radius: 50%;
                    color: #151515;
                    font-weight: 850;
                    cursor: pointer;
                    box-shadow: 0 8px 28px rgba(0, 0, 0, 0.28);
                    transition: transform 0.16s ease, box-shadow 0.16s ease;
                }}
                #ubike-float-tools .uft-button:hover {{
                    transform: translateY(-2px);
                    box-shadow: 0 11px 32px rgba(0, 0, 0, 0.34);
                }}
                #ubike-float-tools .uft-top {{ background: #f1f3f5; font-size: 12px; }}
                #ubike-float-tools .uft-analysis {{ background: #8bd3ff; font-size: 13px; }}
                #ubike-float-tools .uft-search {{ background: #ffbf00; font-size: 24px; }}
                #ubike-float-tools .uft-panel {{
                    position: absolute;
                    right: 66px;
                    bottom: 0;
                    width: min(340px, calc(100vw - 96px));
                    padding: 13px;
                    border: 1px solid rgba(0, 0, 0, 0.13);
                    border-radius: 16px;
                    background: rgba(255, 255, 255, 0.98);
                    color: #171717;
                    box-shadow: 0 16px 46px rgba(0, 0, 0, 0.28);
                    backdrop-filter: blur(8px);
                }}
                #ubike-float-tools .uft-panel[hidden] {{ display: none !important; }}
                #ubike-float-tools .uft-title-row {{
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    margin-bottom: 9px;
                }}
                #ubike-float-tools .uft-title {{ font-size: 15px; font-weight: 800; }}
                #ubike-float-tools .uft-close {{
                    border: 0;
                    background: transparent;
                    color: #555;
                    font-size: 20px;
                    cursor: pointer;
                }}
                #ubike-float-tools .uft-input {{
                    width: 100%;
                    min-height: 44px;
                    padding: 10px 12px;
                    border: 2px solid #e0e0e0;
                    border-radius: 11px;
                    outline: none;
                    font-size: 16px;
                    color: #111;
                    background: #fff;
                }}
                #ubike-float-tools .uft-input:focus {{
                    border-color: #ffbf00;
                    box-shadow: 0 0 0 3px rgba(255, 191, 0, 0.18);
                }}
                #ubike-float-tools .uft-hint {{
                    margin: 7px 2px 8px;
                    color: #666;
                    font-size: 12px;
                }}
                #ubike-float-tools .uft-results {{
                    display: flex;
                    flex-direction: column;
                    gap: 5px;
                    max-height: min(310px, 48vh);
                    overflow-y: auto;
                    overflow-x: hidden;
                }}
                #ubike-float-tools .uft-result {{
                    width: 100%;
                    padding: 9px 10px;
                    border: 1px solid #e4e4e4;
                    border-radius: 10px;
                    background: #fff;
                    text-align: left;
                    cursor: pointer;
                }}
                #ubike-float-tools .uft-result:hover,
                #ubike-float-tools .uft-result:focus {{
                    border-color: #ffbf00;
                    background: #fff9df;
                    outline: none;
                }}
                #ubike-float-tools .uft-result-name {{
                    display: block;
                    color: #111;
                    font-size: 14px;
                    font-weight: 750;
                }}
                #ubike-float-tools .uft-result-region,
                #ubike-float-tools .uft-result-status {{
                    display: block;
                    margin-top: 2px;
                    color: #777;
                    font-size: 12px;
                    line-height: 1.35;
                }}
                #ubike-float-tools .uft-result-status {{ color: #444; }}
                #ubike-float-tools .uft-empty {{
                    padding: 12px 4px 5px;
                    color: #777;
                    font-size: 13px;
                    text-align: center;
                }}
                #ubike-search-toast {{
                    position: fixed;
                    left: 50%;
                    bottom: var(--ubike-float-toast-bottom, 210px);
                    z-index: 2147483001;
                    transform: translateX(-50%);
                    padding: 10px 15px;
                    border-radius: 999px;
                    background: rgba(20, 20, 20, 0.92);
                    color: #fff;
                    font: 700 14px/1.25 -apple-system, BlinkMacSystemFont, "Segoe UI",
                          "Microsoft JhengHei", sans-serif;
                    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.28);
                    pointer-events: none;
                }}
                .ubike-analysis-row-focus {{
                    position: relative;
                    z-index: 2;
                    outline: 3px solid #ff9f00;
                    outline-offset: -3px;
                    animation: ubikeAnalysisPulse 0.72s ease-in-out 3;
                }}
                @keyframes ubikeAnalysisPulse {{
                    0%, 100% {{ filter: brightness(1); }}
                    50% {{ filter: brightness(1.16); }}
                }}
                @media (max-width: 700px) {{
                    #ubike-float-tools {{
                        right: 10px;
                        bottom: calc(72px + env(safe-area-inset-bottom, 0px));
                    }}
                    #ubike-float-tools .uft-button {{
                        width: 52px;
                        height: 52px;
                        flex-basis: 52px;
                    }}
                    #ubike-float-tools .uft-panel {{
                        right: 60px;
                        bottom: 0;
                        width: min(305px, calc(100vw - 82px));
                        max-height: 72vh;
                    }}
                }}
            `;
            doc.head.appendChild(style);

            const root = doc.createElement("div");
            root.id = "ubike-float-tools";
            root.innerHTML = `
                <div class="uft-panel" hidden>
                    <div class="uft-title-row">
                        <div class="uft-title">🔎 搜尋分析結果</div>
                        <button class="uft-close" type="button" aria-label="關閉搜尋">×</button>
                    </div>
                    <input class="uft-input" type="search"
                           placeholder="輸入場站名稱或行政區" autocomplete="off" />
                    <div class="uft-hint">只搜尋目前的調度分析結果｜快捷鍵 Ctrl + K</div>
                    <div class="uft-results"></div>
                </div>
                <div class="uft-actions">
                    <button class="uft-button uft-top" type="button" title="回到頁面最上方">TOP</button>
                    <button class="uft-button uft-analysis" type="button" title="跳到調度分析結果">分析</button>
                    <button class="uft-button uft-search" type="button" title="搜尋分析結果（Ctrl + K）">🔎</button>
                </div>
            `;
            doc.body.appendChild(root);

            const topButton = root.querySelector(".uft-top");
            const analysisButton = root.querySelector(".uft-analysis");
            const searchButton = root.querySelector(".uft-search");
            const panel = root.querySelector(".uft-panel");
            const closeButton = root.querySelector(".uft-close");
            const input = root.querySelector(".uft-input");
            const results = root.querySelector(".uft-results");

            function isMobileLayout() {{
                return displayMode === "mobile" || win.matchMedia("(max-width: 700px)").matches;
            }}

            function updateFloatingPosition() {{
                // 固定成本：不再掃描整個頁面 DOM。手機保留底部安全距離即可。
                const bottomGap = isMobileLayout() ? 72 : 22;
                root.style.bottom = isMobileLayout()
                    ? `calc(${{bottomGap}}px + env(safe-area-inset-bottom, 0px))`
                    : `${{bottomGap}}px`;
                doc.documentElement.style.setProperty(
                    "--ubike-float-toast-bottom",
                    `${{bottomGap + 188}}px`,
                );
            }}

            function getScrollTargets() {{
                const candidates = [
                    doc.scrollingElement,
                    doc.documentElement,
                    doc.body,
                    doc.querySelector('[data-testid="stAppViewContainer"]'),
                    doc.querySelector('[data-testid="stMain"]'),
                    doc.querySelector('section.main'),
                    doc.querySelector('.stApp'),
                ];
                return Array.from(new Set(candidates.filter(Boolean)));
            }}

            function scrollPageToTop() {{
                setOpen(false);
                const topAnchor = doc.getElementById("ubike-page-top-anchor");
                if (topAnchor) topAnchor.scrollIntoView({{ behavior: "smooth", block: "start" }});

                for (const target of getScrollTargets()) {{
                    try {{
                        target.scrollTo({{ top: 0, left: 0, behavior: "smooth" }});
                    }} catch (_error) {{
                        target.scrollTop = 0;
                        target.scrollLeft = 0;
                    }}
                }}
                try {{
                    win.scrollTo({{ top: 0, left: 0, behavior: "smooth" }});
                }} catch (_error) {{
                    win.scrollTo(0, 0);
                }}
                win.setTimeout(() => {{
                    for (const target of getScrollTargets()) {{
                        target.scrollTop = 0;
                        target.scrollLeft = 0;
                    }}
                    win.scrollTo(0, 0);
                }}, 420);
                showToast("已回到頁面最上方");
            }}

            function showToast(message) {{
                const previous = doc.getElementById("ubike-search-toast");
                if (previous) previous.remove();
                const toast = doc.createElement("div");
                toast.id = "ubike-search-toast";
                toast.textContent = message;
                doc.body.appendChild(toast);
                win.setTimeout(() => toast.remove(), 2100);
            }}

            function setOpen(open) {{
                panel.hidden = !open;
                if (open) {{
                    input.value = "";
                    renderResults("");
                    win.setTimeout(() => input.focus(), 30);
                }}
            }}

            function jumpToStation(station) {{
                setOpen(false);
                const anchor = doc.getElementById(station.anchor);
                if (!anchor) {{
                    showToast("找不到這個分析結果，請重新整理後再試");
                    return;
                }}

                anchor.scrollIntoView({{ behavior: "smooth", block: "center" }});
                anchor.classList.remove("ubike-analysis-row-focus");
                void anchor.offsetWidth;
                anchor.classList.add("ubike-analysis-row-focus");
                win.setTimeout(() => anchor.classList.remove("ubike-analysis-row-focus"), 2400);
                showToast(`已找到分析結果：${{station.name}}`);
            }}

            function createResultButton(station) {{
                const button = doc.createElement("button");
                button.type = "button";
                button.className = "uft-result";

                const name = doc.createElement("span");
                name.className = "uft-result-name";
                name.textContent = station.name;
                button.appendChild(name);

                if (station.region) {{
                    const region = doc.createElement("span");
                    region.className = "uft-result-region";
                    region.textContent = station.region;
                    button.appendChild(region);
                }}

                const status = doc.createElement("span");
                status.className = "uft-result-status";
                status.textContent = `2.0：${{station.bike || "—"}}｜2.0E：${{station.ebike || "—"}}`;
                button.appendChild(status);

                button.addEventListener("click", () => jumpToStation(station));
                return button;
            }}

            function renderResults(query) {{
                const keyword = String(query || "").trim().toLocaleLowerCase("zh-TW");
                const matched = stations
                    .filter((station) => {{
                        const haystack = `${{station.name}} ${{station.region}} ${{station.bike}} ${{station.ebike}}`
                            .toLocaleLowerCase("zh-TW");
                        return !keyword || haystack.includes(keyword);
                    }})
                    .slice(0, 12);

                results.replaceChildren();
                if (!matched.length) {{
                    const empty = doc.createElement("div");
                    empty.className = "uft-empty";
                    empty.textContent = stations.length
                        ? "分析結果中查無符合的場站"
                        : "目前沒有需要調度的分析結果";
                    results.appendChild(empty);
                    return;
                }}
                matched.forEach((station) => results.appendChild(createResultButton(station)));
            }}

            topButton.addEventListener("click", scrollPageToTop);
            analysisButton.addEventListener("click", () => {{
                setOpen(false);
                const anchor = doc.getElementById("analysis-results-anchor");
                if (!anchor) {{
                    showToast("找不到分析結果區");
                    return;
                }}
                anchor.scrollIntoView({{ behavior: "smooth", block: "start" }});
            }});
            searchButton.addEventListener("click", () => setOpen(panel.hidden));
            closeButton.addEventListener("click", () => setOpen(false));
            input.addEventListener("input", (event) => renderResults(event.target.value));
            input.addEventListener("keydown", (event) => {{
                if (event.key === "Enter") {{
                    const firstResult = results.querySelector(".uft-result");
                    if (firstResult) firstResult.click();
                }}
                if (event.key === "Escape") setOpen(false);
            }});

            if (win.__ubikeSearchKeyHandler) {{
                win.removeEventListener("keydown", win.__ubikeSearchKeyHandler);
            }}
            win.__ubikeSearchKeyHandler = (event) => {{
                if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === "k") {{
                    event.preventDefault();
                    setOpen(true);
                }}
                if (event.key === "Escape" && !panel.hidden) setOpen(false);
            }};
            win.addEventListener("keydown", win.__ubikeSearchKeyHandler);

            if (win.__ubikeFloatingResizeHandler) {{
                win.removeEventListener("resize", win.__ubikeFloatingResizeHandler);
                if (win.visualViewport) {{
                    win.visualViewport.removeEventListener("resize", win.__ubikeFloatingResizeHandler);
                    win.visualViewport.removeEventListener("scroll", win.__ubikeFloatingResizeHandler);
                }}
            }}
            win.__ubikeFloatingResizeHandler = () => updateFloatingPosition();
            win.addEventListener("resize", win.__ubikeFloatingResizeHandler, {{ passive: true }});
            if (win.visualViewport) {{
                win.visualViewport.addEventListener("resize", win.__ubikeFloatingResizeHandler, {{ passive: true }});
                win.visualViewport.addEventListener("scroll", win.__ubikeFloatingResizeHandler, {{ passive: true }});
            }}

            // 舊版若已建立 MutationObserver，先關閉；之後只在視窗尺寸改變時更新。
            if (win.__ubikeFloatingObserver) {{
                win.__ubikeFloatingObserver.disconnect();
                win.__ubikeFloatingObserver = null;
            }}

            updateFloatingPosition();
            win.setTimeout(updateFloatingPosition, 350);
            renderResults("");
        }})();
        </script>
        """,
        height=0,
        scrolling=False,
    )

st.set_page_config(
    page_title="臺東 YouBike 智慧調度決策系統",
    page_icon="🚚",
    layout="wide",
)

st.markdown(
    """
    <style>
    html, body, .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"] {
        max-width: 100%;
        overflow-x: hidden !important;
    }
    .block-container {
        width: 100%;
        max-width: 100%;
        padding-top: 1.3rem;
        padding-bottom: 3rem;
        overflow-x: clip;
    }
    [data-testid="stMetricValue"] {font-size: 1.65rem;}
    div[data-testid="stDataEditor"] {border: 1px solid #dddddd; border-radius: 10px;}
    div[data-testid="stNumberInput"] input {text-align: center;}

    .analysis-result-table-wrap {
        width: 100%;
        max-width: 100%;
        margin: 0.15rem 0 0.75rem;
        overflow: visible;
    }
    .analysis-result-table {
        width: 100%;
        max-width: 100%;
        table-layout: fixed;
        border-collapse: separate;
        border-spacing: 0;
        border: 1px solid #d9dee5;
        border-radius: 10px;
        overflow: hidden;
        background: white;
    }
    .analysis-result-table .analysis-col-station {width: 42%;}
    .analysis-result-table .analysis-col-status {width: 29%;}
    .analysis-result-table th,
    .analysis-result-table td {
        padding: 8px 6px;
        border-right: 1px solid #e4e8ed;
        border-bottom: 1px solid #e4e8ed;
        vertical-align: middle;
        line-height: 1.3;
        white-space: normal;
        overflow-wrap: anywhere;
        word-break: break-word;
    }
    .analysis-result-table th:last-child,
    .analysis-result-table td:last-child {border-right: 0;}
    .analysis-result-table tbody tr:last-child td {border-bottom: 0;}
    .analysis-result-table th {
        background: #f4f6f8;
        color: #222;
        font-size: 0.88rem;
        font-weight: 800;
        text-align: center;
    }
    .analysis-result-table td {
        font-size: 0.9rem;
    }
    .analysis-station-cell {
        font-weight: 750;
        text-align: left;
    }
    .analysis-status-cell {
        text-align: center;
        font-weight: 750;
    }
    .analysis-status-extra {background: #fce1e1;}
    .analysis-status-short {background: #fff0df;}
    .analysis-status-error {background: #fff7cc;}
    .analysis-status-ok {background: #e7f5e9;}

    /* 手機版：縮小頁面留白，並放大數字輸入框，方便直接叫出九宮格鍵盤。 */
    @media (max-width: 900px) {
        .block-container {
            padding-left: 0.22rem;
            padding-right: 0.22rem;
            padding-top: 0.7rem;
        }

        .analysis-result-table-wrap {
            margin-left: 0;
            margin-right: 0;
        }
        .analysis-result-table .analysis-col-station {width: 40%;}
        .analysis-result-table .analysis-col-status {width: 30%;}
        .analysis-result-table th,
        .analysis-result-table td {
            padding: 7px 3px;
            font-size: 0.78rem;
            line-height: 1.24;
        }
        .analysis-result-table th {
            font-size: 0.75rem;
        }

        div[data-testid="stDataEditor"] {
            font-size: 0.78rem;
        }

        div[data-testid="stDataEditor"] [role="columnheader"],
        div[data-testid="stDataEditor"] [role="gridcell"] {
            padding-left: 2px !important;
            padding-right: 2px !important;
        }

        div[data-testid="stNumberInput"] input {
            min-height: 44px;
            font-size: 18px !important;
            font-weight: 700;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


@st.cache_data(show_spinner=False)
def cached_load_workbook(source: bytes | str):
    return load_workbook(source)


@st.cache_data(show_spinner=False, max_entries=48)
def cached_parse_route(source: bytes | str, sheet_name: str, route: str, shift: str) -> pd.DataFrame:
    """快取已解析的分區資料，避免每次輸入都重新掃描 Excel 工作表。"""
    workbook = cached_load_workbook(source)
    return parse_route(workbook[sheet_name], route, shift)


st.markdown(
    '<div id="ubike-page-top-anchor" aria-hidden="true"></div>',
    unsafe_allow_html=True,
)
st.title("🚚 臺東 YouBike 智慧調度工具")
# 效能優化版：批次表單、OCR 快取、解析快取、向量化計算、低成本懸浮工具。
st.caption("Excel 作為配置基底｜自動區分 D1、D2、D3 與行政區｜分開計算 2.0、2.0E")

mobile_detected = is_mobile_browser()

base_token = get_base_token()
cached_base, cache_expired = load_cached_base(base_token)

if "base_uploader_version" not in st.session_state:
    st.session_state["base_uploader_version"] = 0

if cache_expired:
    clear_base_token()
    st.session_state["base_uploader_version"] += 1
    st.session_state["base_expired_notice"] = True
    base_token = None
    cached_base = None

with st.sidebar:
    st.header("配置基底")

    if st.session_state.pop("base_expired_notice", False):
        st.warning("原配置基底無法讀取，請重新上傳。")

    uploaded_excel = st.file_uploader(
        "上傳配置表（必須）",
        type=["xlsx"],
        key=f"base_uploader_{st.session_state['base_uploader_version']}",
        help="上傳後會持續保留；除非按下全部重置，否則重新整理頁面不必重傳。",
    )
    st.caption("📋 分析結果固定只顯示缺車／多車場站；兩項皆符合者不顯示。")
    mobile_input_mode = st.checkbox(
        "📱 手機九宮格輸入模式",
        value=mobile_detected,
        help="手機通常會自動啟用；若沒有自動偵測，可手動勾選。",
    )
    st.info("現況欄與配置基底會自動保存，程式不設定到期時間。")

active_base = cached_base

if uploaded_excel is not None:
    uploaded_bytes = uploaded_excel.getvalue()
    uploaded_digest = hashlib.sha256(uploaded_bytes).hexdigest()

    # 同一個上傳元件在每次互動時都會重新執行；相同檔案不重複建立保存副本。
    if active_base is None or active_base["sha256"] != uploaded_digest:
        active_base = save_cached_base(uploaded_excel.name, uploaded_bytes)

if st.session_state.pop("full_reset_notice", False):
    st.success("✅ 已完成全部重置：配置基底與所有保留數據都已清除。")

if st.session_state.pop("data_zero_notice", False):
    st.success("✅ 現況已清空：配置基底保留，所有分區與班別的現況均恢復為空白。")

if active_base is None:
    st.info("📤 請先從左側上傳 Excel 配置表；上傳後會持續保留，重新整理頁面不必重傳。")
    st.stop()

try:
    workbook_data = cached_load_workbook(active_base["bytes"])
    source_caption = f"已使用上傳檔案：{active_base['name']}"
except Exception as exc:
    st.error(f"Excel 讀取失敗：{exc}")
    st.stop()

options = available_sources(workbook_data)
if not options:
    st.error("這份 Excel 找不到 D1／D2／D3 的標準配置區塊。")
    st.stop()

# 頁面最上方的兩種重置功能。
reset_col_1, reset_col_2 = st.columns(2)
with reset_col_1:
    if st.button(
        "🗑️ 全部重置",
        use_container_width=True,
        type="primary",
        help="刪除目前配置基底、網址識別碼，以及所有已保存的現況數據。",
    ):
        reset_token = active_base.get("token")
        if reset_token:
            delete_cached_base(reset_token)
        clear_base_token()
        clear_editor_session_state(reset_token)
        st.session_state["base_uploader_version"] += 1
        st.session_state["full_reset_notice"] = True
        rerun_app()

with reset_col_2:
    if st.button(
        "🧹 清空現況",
        use_container_width=True,
        help="保留 Excel 配置基底，將所有 D1／D2／D3、所有班別的 2.0 與 2.0E 現況清成空白。",
    ):
        blank_status_cache = build_blank_status_cache(workbook_data, options)
        save_cached_status(
            active_base["token"],
            active_base["expires_at"],
            blank_status_cache,
        )
        clear_editor_session_state(active_base["token"])
        st.session_state["data_zero_notice"] = True
        rerun_app()

st.caption(
    "🗑️ 全部重置：清除基底＋所有保留數據　｜　"
    "🧹 清空現況：基底不動，所有現況恢復成空白"
)

with st.sidebar:
    st.success("✅ 基底已保存：無時間限制")
    st.caption(f"目前基底：{active_base['name']}")

option_labels = [f"{route}｜{sheet_name}" for sheet_name, route in options]

selector_1, selector_2 = st.columns([2, 1])
with selector_1:
    default_option = "D1｜115_D1暑假配置表"
    default_index = option_labels.index(default_option) if default_option in option_labels else 0
    selected_label = st.selectbox("📚 配置表版本與分區", option_labels, index=default_index)
with selector_2:
    selected_shift = st.selectbox("⏰ 班別", list(SHIFT_COLUMNS.keys()))

selected_index = option_labels.index(selected_label)
selected_sheet, selected_route = options[selected_index]
base_df = cached_parse_route(active_base["bytes"], selected_sheet, selected_route, selected_shift)

status_cache = load_cached_status(active_base["token"], active_base["expires_at"])
current_context_key = status_context_key(selected_sheet, selected_route, selected_shift)
saved_current_status = status_cache["contexts"].get(current_context_key)
base_df = blank_current_status(base_df)
if saved_current_status is not None:
    base_df = restore_current_status(base_df, saved_current_status)

if base_df.empty:
    st.warning("目前選擇的區塊沒有可用場站資料，請改選其他版本或班別。")
    st.stop()



with st.expander("🌐 YouBike 官網即時車數（免 TDX 金鑰）", expanded=True):
    st.caption(
        "按下同步後，系統會直接讀取 YouBike 官網站點資料，取得臺東縣各場站目前可借的 "
        "YouBike 2.0 與 2.0E，再把名稱安全配對成功的資料寫入目前選擇的分區與班別。"
    )
    st.success(
        "✅ 不需要 Client ID、Client Secret，也不需要設定 Streamlit Secrets；"
        "按下按鈕即可取得官網最新一批 2.0／2.0E 車數。"
    )

    previous_live_meta = status_cache.get("metadata", {}).get(current_context_key, {})
    if isinstance(previous_live_meta, dict) and previous_live_meta.get("fetched_at"):
        previous_source_time = str(previous_live_meta.get("latest_source_time") or "").strip()
        source_time_text = f"｜官方資料時間 {previous_source_time}" if previous_source_time else ""
        st.info(
            f"此分區上次官網同步：{previous_live_meta['fetched_at']}"
            f"{source_time_text}｜成功寫入 "
            f"{safe_nonnegative_int(previous_live_meta.get('matched_count'))} 個場站"
        )

    sync_youbike_now = st.button(
        "🔄 取得 YouBike 官網即時車數並寫入現況",
        type="primary",
        use_container_width=True,
        help="為避免對官方服務造成負擔，同一批資料會快取約 60 秒。",
        key=f"sync_youbike::{current_context_key}",
    )

    if sync_youbike_now:
        try:
            with st.spinner("正在讀取 YouBike 官網臺東場站車數並進行名稱配對……"):
                live_payload = fetch_youbike_taitung_bike_data(int(time.time() // 60))
                live_updated_df, live_report_df, live_summary = apply_youbike_updates_to_dataframe(
                    base_df,
                    live_payload["records"],
                )

                if live_summary["matched_count"] <= 0:
                    st.error("沒有任何場站通過安全配對，因此未修改現況資料。")
                else:
                    base_df = live_updated_df
                    status_cache["contexts"][current_context_key] = dataframe_to_status_records(base_df)
                    status_cache.setdefault("metadata", {})[current_context_key] = {
                        "source": live_payload.get("source", "YouBike 官網公開接口（免 TDX）"),
                        "fetched_at": live_payload["fetched_at"],
                        "latest_source_time": live_payload.get("latest_source_time", ""),
                        "matched_count": live_summary["matched_count"],
                        "skipped_count": live_summary["skipped_count"],
                        "unmatched_count": live_summary["unmatched_count"],
                    }
                    save_cached_status(
                        active_base["token"],
                        active_base["expires_at"],
                        status_cache,
                    )
                    clear_editor_session_state(active_base["token"])
                    official_time = str(live_payload.get("latest_source_time") or "").strip()
                    official_time_text = f"｜官方資料時間：{official_time}" if official_time else ""
                    st.success(
                        f"✅ 官網同步完成：已寫入 {live_summary['matched_count']}／"
                        f"{live_summary['total_count']} 個場站｜系統取得時間："
                        f"{live_payload['fetched_at']}{official_time_text}"
                    )

                if not live_report_df.empty:
                    problem_df = live_report_df[live_report_df["結果"] != "已寫入"]
                    with st.expander(
                        f"查看配對明細（未寫入 {len(problem_df)} 筆）",
                        expanded=not problem_df.empty,
                    ):
                        report_height = min(520, max(110, 42 + len(live_report_df) * 35))
                        st.dataframe(
                            live_report_df,
                            hide_index=True,
                            use_container_width=True,
                            height=report_height,
                            row_height=35,
                        )
        except YouBikeDataError as exc:
            st.error(f"YouBike 官網同步失敗：{exc}")
        except Exception as exc:
            st.error(f"YouBike 官網同步發生未預期錯誤：{exc}")

    st.caption(
        "資料來源：YouBike 微笑單車官網公開 JSON 接口（不是 TDX，也不是擷取網頁畫面）。"
        "官方資料可能有約 1 分鐘更新延遲；場站非正常服務、車種明細不完整或名稱配對"
        "不確定時，系統不會自動覆蓋。"
    )


# 多張平板翻拍照片辨識：只讀取「單位」後方的在站 2.0／2.0E。
# 照片先累積在 session_state，因此手機即使每次只能加入一張，也能分批加入後一次辨識。
photo_context_id = (
    f"{active_base['token']}::{selected_sheet}::{selected_route}::{selected_shift}"
)
photo_pool_key = f"station_photo_pool::{photo_context_id}"
photo_uploader_version_key = f"station_photo_uploader_version::{photo_context_id}"

if photo_pool_key not in st.session_state:
    st.session_state[photo_pool_key] = []
if photo_uploader_version_key not in st.session_state:
    st.session_state[photo_uploader_version_key] = 0

with st.expander("📷 多張照片直接更新現況", expanded=True):
    st.caption(
        "照片格式請保持與站點列表相同。可一次選取多張，也可分批加入；"
        "系統只讀取『在站』下方的 2.0、2.0E。"
    )

    newly_uploaded_photos = st.file_uploader(
        "上傳照片（可一次選多張／可重複分批加入）",
        type=["jpg", "jpeg", "png", "webp"],
        accept_multiple_files=True,
        key=(
            f"station_photo_uploader::{photo_context_id}::"
            f"{st.session_state[photo_uploader_version_key]}"
        ),
        help=(
            "電腦與手機照片圖庫可一次選取多張；若手機『直接拍照』一次只能回傳一張，"
            "可拍完一張後再次加入，系統會累積到同一批。"
        ),
    )

    # 將本次選取加入照片池；以內容雜湊去重，避免 Streamlit 每次重跑重複累加。
    if newly_uploaded_photos:
        photo_pool = list(st.session_state.get(photo_pool_key, []))
        known_digests = {str(record.get("sha256", "")) for record in photo_pool}

        for uploaded_photo in newly_uploaded_photos:
            photo_bytes = uploaded_photo.getvalue()
            photo_digest = hashlib.sha256(photo_bytes).hexdigest()
            if photo_digest in known_digests:
                continue
            photo_pool.append(
                {
                    "name": str(uploaded_photo.name),
                    "bytes": photo_bytes,
                    "sha256": photo_digest,
                }
            )
            known_digests.add(photo_digest)

        st.session_state[photo_pool_key] = photo_pool

    queued_station_photos = list(st.session_state.get(photo_pool_key, []))

    if queued_station_photos:
        queued_names = "、".join(record["name"] for record in queued_station_photos)
        st.info(f"目前已加入 {len(queued_station_photos)} 張照片：{queued_names}")
    else:
        st.info("目前尚未加入照片。")

    photo_action_col_1, photo_action_col_2 = st.columns(2)
    with photo_action_col_1:
        run_photo_ocr = st.button(
            "🔍 辨識並直接寫入現況",
            type="primary",
            use_container_width=True,
            disabled=not queued_station_photos,
            key=f"run_ocr::{photo_context_id}",
        )

    with photo_action_col_2:
        clear_all_photos = st.button(
            "🗑️ 一鍵清除所有照片",
            use_container_width=True,
            disabled=not queued_station_photos,
            key=f"clear_photos::{photo_context_id}",
        )

    if clear_all_photos:
        st.session_state[photo_pool_key] = []
        st.session_state[photo_uploader_version_key] += 1
        rerun_app()

    if run_photo_ocr:
        try:
            with st.spinner(f"正在辨識 {len(queued_station_photos)} 張照片並更新現況……"):
                analysis_result = analyze_station_photos(queued_station_photos, base_df)
                recognized_updates = analysis_result.get("updates", {})

                error_df = analysis_result.get("error_df", pd.DataFrame())

                if recognized_updates:
                    updated_full_df = apply_ocr_updates_to_dataframe(base_df, recognized_updates)
                    status_cache["contexts"][current_context_key] = dataframe_to_status_records(updated_full_df)
                    save_cached_status(active_base["token"], active_base["expires_at"], status_cache)
                    clear_editor_session_state(active_base["token"])
                    base_df = updated_full_df

                    st.success(f"✅ 已直接更新 {len(recognized_updates)} 個場站。")
                    recognized_df = analysis_result.get("recognized_df", pd.DataFrame())
                    if not recognized_df.empty:
                        recognized_table_height = max(105, 40 + (len(recognized_df) * 35) + 4)
                        st.dataframe(
                            recognized_df[["場站名稱", "單位", "2.0 現況", "2.0E 現況", "來源"]],
                            hide_index=True,
                            use_container_width=True,
                            height=recognized_table_height,
                            row_height=35,
                        )
                else:
                    st.error("識別錯誤：這批照片沒有讀到可安全寫入的場站資料。")

                if not error_df.empty:
                    st.error(f"⚠️ 共有 {len(error_df)} 筆識別錯誤，未寫入現況。")
                    error_table_height = max(105, 40 + (len(error_df) * 35) + 4)
                    st.dataframe(
                        error_df[["場站名稱", "2.0 現況", "2.0E 現況", "來源"]],
                        hide_index=True,
                        use_container_width=True,
                        height=error_table_height,
                        row_height=35,
                    )
        except ImportError:
            st.error(
                "尚未安裝免費 OCR 套件。請確認 requirements.txt 內含 rapidocr 與 onnxruntime。"
            )
        except Exception as exc:
            st.error(f"照片辨識失敗：{exc}")

regions = ["全部"] + list(dict.fromkeys(base_df["行政區"].tolist()))
selected_region = st.selectbox("📍 行政區", regions)
working_df = base_df if selected_region == "全部" else base_df[base_df["行政區"] == selected_region]
working_df = working_df.reset_index(drop=True)


st.caption(f"{source_caption}｜工作表：{selected_sheet}｜分區：{selected_route}｜共 {len(working_df)} 個場站")
st.markdown('<div id="current-status-input-anchor"></div>', unsafe_allow_html=True)

editor_key = (
    f"editor::{active_base['token']}::{selected_sheet}::{selected_route}::"
    f"{selected_shift}::{selected_region}"
)

with st.expander("✏️ 輸入現場現況（點此展開）", expanded=False):
    with st.form(key=f"status_form::{editor_key}", clear_on_submit=False):
        working_df = coerce_nullable_current_status(working_df)
    
        if mobile_input_mode:
            # value=None 會保留空白；手機點選後仍會叫出數字鍵盤。
            st.caption("📱 手機九宮格輸入已啟用：空白代表尚未輸入或未成功識別。")
            edited_df = working_df.copy()
    
            header_station, header_bike, header_ebike = st.columns([1.55, 0.9, 0.95])
            header_station.markdown("**場站／標準**")
            header_bike.markdown("**2.0 現況**")
            header_ebike.markdown("**2.0E 現況**")
    
            for row_index, row in working_df.iterrows():
                st.markdown(
                    f'<div id="station-anchor-{row_index}" class="station-anchor"></div>',
                    unsafe_allow_html=True,
                )
                station_col, bike_col, ebike_col = st.columns([1.55, 0.9, 0.95])
    
                with station_col:
                    st.markdown(f"**{row['場站名稱']}**")
                    st.caption(
                        f"標準：2.0＝{safe_nonnegative_int(row['2.0 標準'])}｜"
                        f"2.0E＝{safe_nonnegative_int(row['2.0E 標準'])}"
                    )
    
                with bike_col:
                    bike_now = st.number_input(
                        f"{row['場站名稱']} 2.0 現況",
                        min_value=0,
                        value=normalize_current_status(row["2.0 現況"]),
                        step=1,
                        format="%d",
                        placeholder="空白",
                        key=f"{editor_key}::mobile::2.0::{row_index}",
                        label_visibility="collapsed",
                    )
    
                with ebike_col:
                    ebike_now = st.number_input(
                        f"{row['場站名稱']} 2.0E 現況",
                        min_value=0,
                        value=normalize_current_status(row["2.0E 現況"]),
                        step=1,
                        format="%d",
                        placeholder="空白",
                        key=f"{editor_key}::mobile::2.0E::{row_index}",
                        label_visibility="collapsed",
                    )
    
                edited_df.at[row_index, "2.0 現況"] = (
                    pd.NA if bike_now is None else int(bike_now)
                )
                edited_df.at[row_index, "2.0E 現況"] = (
                    pd.NA if ebike_now is None else int(ebike_now)
                )
    
            edited_df = coerce_nullable_current_status(edited_df)
        else:
            # 電腦版維持可快速批次編輯的資料表，空白儲存格可直接輸入。
            editor_row_height = 38
            editor_header_height = 40
            editor_height = editor_header_height + (len(working_df) * editor_row_height) + 4
    
            edited_df = st.data_editor(
                working_df,
                key=editor_key,
                hide_index=True,
                use_container_width=True,
                height=editor_height,
                row_height=editor_row_height,
                column_order=["場站名稱", "2.0 現況", "2.0E 現況", "2.0 標準", "2.0E 標準"],
                disabled=["行政區", "場站名稱", "2.0 標準", "2.0E 標準"],
                column_config={
                    "行政區": None,
                    "場站名稱": st.column_config.TextColumn("場站", width=120),
                    "2.0 現況": st.column_config.NumberColumn(
                        "2.0現", width=52, min_value=0, step=1, format="%d", required=False
                    ),
                    "2.0E 現況": st.column_config.NumberColumn(
                        "2.0E現", width=58, min_value=0, step=1, format="%d", required=False
                    ),
                    "2.0 標準": st.column_config.NumberColumn("2.0標", width=52, format="%d"),
                    "2.0E 標準": st.column_config.NumberColumn("2.0E標", width=58, format="%d"),
                },
            )
            edited_df = coerce_nullable_current_status(edited_df)

        status_form_submitted = st.form_submit_button(
            "✅ 套用並儲存現況",
            type="primary",
            use_container_width=True,
            help="可先連續輸入多個場站，按下後才一次分析與儲存，避免每格輸入都重跑整個系統。",
        )

# 表單送出時才會重跑；只有資料真的改變才序列化並寫入磁碟。
full_status_df = merge_current_status(base_df, edited_df)
current_records = dataframe_to_status_records(full_status_df)
previous_records = status_cache["contexts"].get(current_context_key)
if current_records != previous_records:
    status_cache["contexts"][current_context_key] = current_records
    save_cached_status(active_base["token"], active_base["expires_at"], status_cache)

if status_form_submitted:
    st.success("✅ 現況已套用並儲存，分析結果已更新。")
st.caption("💾 輸入期間不重跑；按下『套用並儲存現況』後一次完成分析與保存。")

st.markdown('<div id="analysis-results-anchor"></div>', unsafe_allow_html=True)
st.markdown("---")
st.subheader("📋 調度分析結果")

result_df = build_result_with_recognition_errors(edited_df)

# 分析結果固定排除「2.0、2.0E 都符合」的場站。
# 若只有其中一種車型符合，該場站仍保留，但「符合」不參與排序。
result_df = result_df[
    (result_df["2.0 缺／多幾台"] != "符合")
    | (result_df["2.0E 缺／多幾台"] != "符合")
].reset_index(drop=True)

sort_control_1, sort_control_2 = st.columns([2, 1])
with sort_control_1:
    selected_sort_field = st.selectbox(
        "↕️ 分析結果排序欄位",
        list(SORT_FIELD_OPTIONS.keys()),
        index=0,
        help="排序會同步套用到畫面、CSV 與彩色 Excel。",
    )
with sort_control_2:
    selected_sort_direction = st.selectbox(
        "排序方向",
        ["由大到小／Z→A／倒序", "由小到大／A→Z／正序"],
        index=0,
    )

sort_descending = selected_sort_direction.startswith("由大到小")

# 保留行政區分段；每個行政區內依使用者選擇的方式排序。
sorted_region_frames = []
for region_name in result_df["行政區"].drop_duplicates():
    region_rows = result_df[result_df["行政區"] == region_name]
    sorted_region_frames.append(
        sort_dispatch_results(region_rows, selected_sort_field, sort_descending)
    )

if sorted_region_frames:
    result_df = pd.concat(sorted_region_frames, ignore_index=True)

bike_short, bike_extra = calculate_totals_ignoring_missing(
    edited_df, "2.0 現況", "2.0 標準"
)
ebike_short, ebike_extra = calculate_totals_ignoring_missing(
    edited_df, "2.0E 現況", "2.0E 標準"
)

m1, m2, m3, m4 = st.columns(4)
m1.metric("2.0 缺車合計", f"{bike_short} 台")
m2.metric("2.0 多車合計", f"{bike_extra} 台")
m3.metric("2.0E 缺車合計", f"{ebike_short} 台")
m4.metric("2.0E 多車合計", f"{ebike_extra} 台")

missing_bike_count = int(edited_df["2.0 現況"].isna().sum())
missing_ebike_count = int(edited_df["2.0E 現況"].isna().sum())
if missing_bike_count or missing_ebike_count:
    st.warning(
        f"⚠️ 尚有識別錯誤／空白資料：2.0 共 {missing_bike_count} 筆，"
        f"2.0E 共 {missing_ebike_count} 筆；缺／多合計暫不計入這些欄位。"
    )

if result_df.empty:
    st.success("✨ 所有場站皆符合配置，目前不需要調度。")
else:
    st.caption("🔴 多車｜🟠 缺車｜識別錯誤＝現況空白或 OCR 無法判讀｜下載檔沿用目前排序")

    # 使用固定寬度的響應式表格：完整向下展開，不在表格內產生左右或上下捲動。
    for region, region_df in result_df.groupby("行政區", sort=False):
        st.markdown(f"#### {selected_route}｜{region}")
        render_analysis_result_table(region_df)

    export_df = make_colored_export_df(result_df)
    csv_data = export_df.to_csv(index=False).encode("utf-8-sig")
    excel_data = build_colored_excel(export_df)

    download_col_1, download_col_2 = st.columns(2)
    with download_col_1:
        st.download_button(
            "⬇️ 下載彩色標記 CSV",
            data=csv_data,
            file_name=f"{selected_route}_{selected_shift}_調度分析_彩色標記.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with download_col_2:
        st.download_button(
            "⬇️ 下載彩色 Excel",
            data=excel_data,
            file_name=f"{selected_route}_{selected_shift}_調度分析_彩色.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.caption("CSV 以 🔴／🟠 圖案表示缺／多；Excel 另以淡黃色標示識別錯誤。")

st.caption(
    "資料來源標示：YouBike 微笑單車官網站點資料（即時同步功能）。"
    "資料僅供調度參考，實際車數與服務狀態以現場及官方系統為準。"
)

# 懸浮搜尋只讀取目前排序完成、實際顯示的分析結果。
render_floating_station_search(result_df, mobile_input_mode)
